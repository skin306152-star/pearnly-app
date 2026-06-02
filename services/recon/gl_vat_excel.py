# -*- coding: utf-8 -*-
"""
services/recon/gl_vat_excel.py · Pearnly

GL-vs-VAT reconciliation Excel export (openpyxl, i18n th/en/zh/ja). Pure
presentation — consumes ReconRow / GlVatSummary from gl_vat_types.
"""

import io
from typing import List

from services.recon.gl_vat_types import ReconRow, GlVatSummary

# ─────────────────────────────────────────────────────────────────────
# i18n 表头字典
# ─────────────────────────────────────────────────────────────────────
_I18N = {
    "th": {
        "sheet1": "รายละเอียดผลต่าง",
        "sheet2": "กระทบยอด",
        "sheet3": "วิธีใช้งาน",
        "h_status": "สถานะ",
        "h_doc_no": "เลขที่เอกสาร",
        "h_date": "วันที่",
        "h_customer": "ชื่อลูกค้า",
        "h_vat_amt": "ยอดตามรายงานภาษีขาย",
        "h_gl_amt": "ยอดตาม GL",
        "h_diff": "ผลต่าง",
        "h_acct": "รหัสบัญชีรายได้",
        "not_found": "ไม่พบข้อมูล",
        "st_matched": "✓ จับคู่",
        "st_diff": "⚠ ผลต่าง",
        "st_missing": "❗ ไม่พบใน GL",
        "kpi_total": "ทั้งหมด",
        "kpi_matched": "จับคู่ถูกต้อง",
        "kpi_diff": "มีผลต่าง",
        "kpi_missing": "ไม่พบใน GL",
        "kpi_diff_sum": "ผลต่างรวม",
        "s_label": "รายละเอียด",
        "s_amount": "จำนวนเงิน",
        "s_gl_total": "ยอดรวมตามบัญชีแยกประเภท",
        "s_minus_gl_cr": "หัก : รายการเครดิตที่ไม่มีในรายงานภาษีขาย",
        "s_plus_gl_dr": "บวก : รายการเดบิตที่ไม่มีในรายงานภาษีขาย",
        "s_plus_vat_p": "บวก : รายการยอดขายที่ไม่มีในบัญชีแยกประเภท",
        "s_minus_vat_n": "หัก : รายการลดหนี้ที่ไม่มีในบัญชีแยกประเภท",
        "s_vat_total": "ยอดรวมตามรายงานภาษีขาย",
    },
    "zh": {
        "sheet1": "差异明细",
        "sheet2": "对账汇总",
        "sheet3": "使用说明",
        "h_status": "状态",
        "h_doc_no": "单据号",
        "h_date": "日期",
        "h_customer": "客户名称",
        "h_vat_amt": "VAT报告金额",
        "h_gl_amt": "GL金额",
        "h_diff": "差异",
        "h_acct": "收入科目代码",
        "not_found": "未找到数据",
        "st_matched": "✓ 匹配",
        "st_diff": "⚠ 有差异",
        "st_missing": "❗ GL未找到",
        "kpi_total": "总笔数",
        "kpi_matched": "完全匹配",
        "kpi_diff": "有差异",
        "kpi_missing": "GL未找到",
        "kpi_diff_sum": "差异金额合计",
        "s_label": "项目说明",
        "s_amount": "金额",
        "s_gl_total": "总账金额合计",
        "s_minus_gl_cr": "减：销项税报告中未列的贷方记录",
        "s_plus_gl_dr": "加：销项税报告中未列的借方记录",
        "s_plus_vat_p": "加：总账中未列的销售记录",
        "s_minus_vat_n": "减：总账中未列的贷项凭单(credit note)记录",
        "s_vat_total": "销项税报告金额合计",
    },
    "en": {
        "sheet1": "Detail",
        "sheet2": "Reconciliation",
        "sheet3": "Instructions",
        "h_status": "Status",
        "h_doc_no": "Document No.",
        "h_date": "Date",
        "h_customer": "Customer",
        "h_vat_amt": "VAT Report Amount",
        "h_gl_amt": "GL Amount",
        "h_diff": "Difference",
        "h_acct": "Revenue Account",
        "not_found": "Not found",
        "st_matched": "✓ Matched",
        "st_diff": "⚠ Diff",
        "st_missing": "❗ GL Missing",
        "kpi_total": "Total",
        "kpi_matched": "Matched",
        "kpi_diff": "Diff",
        "kpi_missing": "GL Missing",
        "kpi_diff_sum": "Total Diff Amount",
        "s_label": "Description",
        "s_amount": "Amount",
        "s_gl_total": "Total per General Ledger",
        "s_minus_gl_cr": "Less: GL credits not in VAT Report",
        "s_plus_gl_dr": "Add: GL debits not in VAT Report",
        "s_plus_vat_p": "Add: Sales in VAT Report not in GL",
        "s_minus_vat_n": "Less: Credit notes in VAT Report not in GL",
        "s_vat_total": "Total per VAT Sales Report",
    },
    "ja": {
        "sheet1": "差異明細",
        "sheet2": "照合まとめ",
        "sheet3": "使い方",
        "h_status": "ステータス",
        "h_doc_no": "伝票番号",
        "h_date": "日付",
        "h_customer": "顧客名",
        "h_vat_amt": "VAT報告金額",
        "h_gl_amt": "GL金額",
        "h_diff": "差異",
        "h_acct": "収益科目コード",
        "not_found": "データなし",
        "st_matched": "✓ 一致",
        "st_diff": "⚠ 差異",
        "st_missing": "❗ GL未検出",
        "kpi_total": "合計件数",
        "kpi_matched": "一致",
        "kpi_diff": "差異あり",
        "kpi_missing": "GL未検出",
        "kpi_diff_sum": "差異金額合計",
        "s_label": "項目",
        "s_amount": "金額",
        "s_gl_total": "総勘定元帳合計",
        "s_minus_gl_cr": "減：VAT報告にないGL貸方記録",
        "s_plus_gl_dr": "加：VAT報告にないGL借方記録",
        "s_plus_vat_p": "加：GLにない売上記録",
        "s_minus_vat_n": "減：GLにない赤伝記録",
        "s_vat_total": "VAT売上報告合計",
    },
}


def _t(lang: str, key: str) -> str:
    lang = lang if lang in _I18N else "th"
    return _I18N[lang].get(key, key)


# ─────────────────────────────────────────────────────────────────────
# Excel 导出
# ─────────────────────────────────────────────────────────────────────
def export_gl_vat_excel(
    detail: List[ReconRow],
    summary: GlVatSummary,
    lang: str = "th",
) -> bytes:
    """生成 3 Sheet Excel:
    Sheet1 = 顶部 KPI + 状态列明细
    Sheet2 = 对账汇总
    Sheet3 = 使用说明（4 语言）
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl 未安装")

    if lang not in _I18N:
        lang = "th"

    wb = openpyxl.Workbook()

    # ── 样式 ────────────────────────────────────────────────────────
    H_FILL = PatternFill("solid", fgColor="1A3C5E")
    H_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    NORM = Font(name="Arial", size=10)
    BOLD = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    SECT_FONT = Font(name="Arial", bold=True, size=10, color="111827")  # 分区标题
    KPI_LABEL_FONT = Font(name="Arial", bold=True, size=11, color="1D4ED8")
    KPI_LABEL_FILL = PatternFill("solid", fgColor="EFF6FF")
    KPI_VAL_FONT = Font(name="Arial", bold=True, size=14, color="111827")
    KPI_VAL_OK_FONT = Font(name="Arial", bold=True, size=14, color="16A34A")
    KPI_VAL_MS_FONT = Font(name="Arial", bold=True, size=14, color="D97706")
    OK_FILL = PatternFill("solid", fgColor="D6F5E3")  # 差异=0 绿
    DIFF_FILL = PatternFill("solid", fgColor="FDE8E8")  # 差异≠0 红
    MISS_FILL = PatternFill("solid", fgColor="FFF3CD")  # GL未找到 黄
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")  # 明细子行白底
    KPI_OK_FILL = PatternFill("solid", fgColor="DCFCE7")  # KPI 匹配绿背景
    KPI_MISS_FILL = PatternFill("solid", fgColor="FFEDD5")  # KPI 未找到橙背景
    SECT_FILL = PatternFill("solid", fgColor="EFF6FF")  # 分区标题:淡蓝
    GRPBORDER = Border(top=Side(style="thin", color="BFDBFE"))  # 组间蓝色细分隔线
    CENTER = Alignment(horizontal="center", vertical="center")
    RIGHT = Alignment(horizontal="right", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center")

    # ── 计算 KPI ──────────────────────────────────────────────────
    total = len(detail)
    matched = sum(1 for r in detail if r.gl_amount is not None and (r.diff or 0) == 0)
    diff_cnt = sum(1 for r in detail if r.gl_amount is not None and (r.diff or 0) != 0)
    missing_cnt = sum(1 for r in detail if r.gl_amount is None)
    diff_sum = round(sum((r.diff or 0) for r in detail if r.diff is not None), 2)

    # ── Sheet 1: 明细 + 顶部 KPI ─────────────────────────────────
    ws1 = wb.active
    ws1.title = _t(lang, "sheet1")

    # Row 1: KPI 标签
    kpi_labels = [
        _t(lang, "kpi_total"),
        _t(lang, "kpi_matched"),
        _t(lang, "kpi_diff"),
        _t(lang, "kpi_missing"),
        _t(lang, "kpi_diff_sum"),
    ]
    for i, lbl in enumerate(kpi_labels, 1):
        c = ws1.cell(1, i, lbl)
        c.font, c.fill, c.alignment = KPI_LABEL_FONT, KPI_LABEL_FILL, CENTER
    # Row 2: KPI 数值（背景色 + 字体色双重着色）
    kpi_vals = [total, matched, diff_cnt, missing_cnt, diff_sum]
    kpi_val_fonts = [KPI_VAL_FONT, KPI_VAL_OK_FONT, KPI_VAL_FONT, KPI_VAL_MS_FONT, KPI_VAL_FONT]
    kpi_val_fills = [WHITE_FILL, KPI_OK_FILL, WHITE_FILL, KPI_MISS_FILL, WHITE_FILL]
    for i, (val, vfont, vfill) in enumerate(zip(kpi_vals, kpi_val_fonts, kpi_val_fills), 1):
        c = ws1.cell(2, i, val)
        c.font, c.fill, c.alignment = vfont, vfill, CENTER
        if i == 5:
            c.number_format = "#,##0.00;[Red](#,##0.00);0.00"
    ws1.row_dimensions[2].height = 28

    # Row 3: blank
    ws1.append([])

    # Row 4: 表头（含状态列）
    headers1 = [
        _t(lang, "h_status"),
        _t(lang, "h_doc_no"),
        _t(lang, "h_date"),
        _t(lang, "h_customer"),
        _t(lang, "h_vat_amt"),
        _t(lang, "h_gl_amt"),
        _t(lang, "h_diff"),
        _t(lang, "h_acct"),
    ]
    ws1.append(headers1)
    header_row = ws1.max_row
    ws1.row_dimensions[header_row].height = 22
    for i in range(1, len(headers1) + 1):
        cell = ws1.cell(header_row, i)
        cell.fill, cell.font, cell.alignment = H_FILL, H_FONT, CENTER

    not_found = _t(lang, "not_found")
    st_matched = _t(lang, "st_matched")
    st_diff = _t(lang, "st_diff")
    st_missing = _t(lang, "st_missing")

    for row in detail:
        # 状态判定
        if row.gl_amount is None:
            status = st_missing
            status_fill = MISS_FILL
        elif (row.diff or 0) == 0:
            status = st_matched
            status_fill = OK_FILL
        else:
            status = st_diff
            status_fill = DIFF_FILL

        gl_disp = row.gl_amount if row.gl_amount is not None else not_found
        diff_disp = row.diff if row.diff is not None else not_found
        ws1.append(
            [
                status,
                row.doc_no,
                row.date,
                row.customer_name,
                row.vat_amount,
                gl_disp,
                diff_disp,
                row.account_codes,
            ]
        )
        r = ws1.max_row
        for c in range(1, 9):
            ws1.cell(r, c).font = NORM
        # 状态列：左对齐 + 颜色填充
        ws1.cell(r, 1).fill = status_fill
        ws1.cell(r, 1).alignment = LEFT
        # 数字列右对齐
        for c in (5, 6, 7):
            ws1.cell(r, c).alignment = RIGHT
            ws1.cell(r, c).number_format = "#,##0.00;[Red](#,##0.00);0.00"
        # 差异/GL 列颜色（强化视觉）
        if row.gl_amount is None:
            ws1.cell(r, 6).fill = MISS_FILL
            ws1.cell(r, 7).fill = MISS_FILL
        elif (row.diff or 0) == 0:
            ws1.cell(r, 7).fill = OK_FILL
        else:
            ws1.cell(r, 7).fill = DIFF_FILL

    # 列宽
    for i, w in enumerate([14, 20, 14, 32, 18, 18, 14, 22], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = ws1.cell(header_row + 1, 1).coordinate

    # ── Sheet 2: 对账汇总 ───────────────────────────────────────────
    ws2 = wb.create_sheet(_t(lang, "sheet2"))
    ws2.append([_t(lang, "s_label"), _t(lang, "s_amount")])
    for i in (1, 2):
        c = ws2.cell(1, i)
        c.fill, c.font, c.alignment = H_FILL, H_FONT, CENTER

    # v118.32.5.5.11 · 每个调整项后展开列出单据明细(Korn 反馈)
    # negate=True 表示 items 的 amount 在 Sheet 显示时取负(对应 减 GL 贷方 / 减 VAT 红字)
    summary_rows = [
        {
            "label": _t(lang, "s_gl_total"),
            "amount": summary.gl_total,
            "emphasize": True,
            "items": [],
            "negate": False,
        },
        {
            "label": _t(lang, "s_minus_gl_cr"),
            "amount": -summary.gl_only_credit,
            "emphasize": False,
            "items": summary.gl_only_credit_items,
            "negate": True,
        },
        {
            "label": _t(lang, "s_plus_gl_dr"),
            "amount": summary.gl_only_debit,
            "emphasize": False,
            "items": summary.gl_only_debit_items,
            "negate": False,
        },
        {
            "label": _t(lang, "s_plus_vat_p"),
            "amount": summary.vat_only_positive,
            "emphasize": False,
            "items": summary.vat_only_positive_items,
            "negate": False,
        },
        {
            "label": _t(lang, "s_minus_vat_n"),
            "amount": summary.vat_only_negative,
            "emphasize": False,
            "items": summary.vat_only_negative_items,
            "negate": False,
        },
        {
            "label": _t(lang, "s_vat_total"),
            "amount": summary.vat_total,
            "emphasize": True,
            "items": [],
            "negate": False,
        },
    ]
    ITEM_FONT = Font(name="Arial", size=9, color="475569")  # 明细行 · 灰色小字
    for entry in summary_rows:
        # 分区标题行(emphasize=False) B 列留空 · 金额只显示在下方明细行 · 避免重复混淆
        b_val = round(entry["amount"], 2) if entry["emphasize"] else ""
        ws2.append([entry["label"], b_val])
        r = ws2.max_row
        if entry["emphasize"]:
            ws2.cell(r, 1).fill, ws2.cell(r, 1).font = H_FILL, BOLD
            ws2.cell(r, 2).fill, ws2.cell(r, 2).font = H_FILL, BOLD
        else:
            ws2.cell(r, 1).font = SECT_FONT  # 粗体分区标题
            ws2.cell(r, 2).font = NORM
            ws2.cell(r, 1).fill = SECT_FILL
            ws2.cell(r, 2).fill = SECT_FILL
            if r > 2:  # 非首分区行加蓝色细线分隔组
                ws2.cell(r, 1).border = GRPBORDER
                ws2.cell(r, 2).border = GRPBORDER
        if entry["emphasize"]:
            ws2.cell(r, 2).alignment = RIGHT
            ws2.cell(r, 2).number_format = "#,##0.00;[Red](#,##0.00);0.00"
        # 展开明细行(若有)
        for it in entry.get("items") or []:
            doc_no = it.get("doc_no") or ""
            date = it.get("date") or ""
            name = (it.get("name") or "").strip()
            if len(name) > 55:
                name = name[:53] + "…"
            label = f"    · {doc_no}  ·  {date}  ·  {name}"
            amt = it.get("amount") or 0.0
            if entry.get("negate"):
                amt = -amt
            ws2.append([label, round(amt, 2)])
            rr = ws2.max_row
            ws2.cell(rr, 1).font = ITEM_FONT
            ws2.cell(rr, 2).font = ITEM_FONT
            ws2.cell(rr, 1).fill = WHITE_FILL
            ws2.cell(rr, 2).fill = WHITE_FILL
            ws2.cell(rr, 2).alignment = RIGHT
            ws2.cell(rr, 2).number_format = "#,##0.00;[Red](#,##0.00);0.00"

    ws2.column_dimensions["A"].width = 90
    ws2.column_dimensions["B"].width = 22

    # ── Sheet 3: 使用说明（4 语言并排展示） ─────────────────────
    ws3 = wb.create_sheet(_t(lang, "sheet3"))
    ws3.column_dimensions["A"].width = 100
    TITLE_FONT = Font(name="Arial", bold=True, size=14, color="1A3C5E")
    SECTION_FONT = Font(name="Arial", bold=True, size=11, color="1A3C5E")
    BODY_FONT = Font(name="Arial", size=10)

    instructions_blocks = [
        (
            "中文 · 使用说明",
            [
                "1. 「差异明细」Sheet = VAT 报告与 GL 按单据号一一比对的结果",
                "2. 状态列：✓ 匹配 = 差异 0 · ⚠ 有差异 = 金额不一致 · ❗ GL 未找到 = VAT 有记录但 GL 无对应凭证",
                "3. 「对账汇总」Sheet = 总账金额 + 调整项 = 销项税报告金额（必须相等）",
                "4. 匹配键：VAT 的「เลขที่เอกสารอ้างอิง」(参考单号) ↔ GL 的「ใบสำคัญ」(凭证号)",
                "5. 红字发票（负数）→ 取 GL 借方并转负数；正常销售 → 取 GL 贷方",
                "6. 收入科目筛选：默认 4xxx 开头（可在前端调整前缀）",
            ],
        ),
        (
            "ภาษาไทย · วิธีใช้งาน",
            [
                "1. ชีต «รายละเอียดผลต่าง» = ผลการกระทบ VAT vs GL ตามเลขที่เอกสาร",
                "2. คอลัมน์สถานะ: ✓ จับคู่ · ⚠ ผลต่าง · ❗ ไม่พบใน GL",
                "3. ชีต «กระทบยอด» = ยอด GL + รายการปรับ = ยอดรายงานภาษีขาย (ต้องเท่ากัน)",
                "4. คีย์จับคู่: VAT «เลขที่เอกสารอ้างอิง» ↔ GL «ใบสำคัญ»",
                "5. ใบลดหนี้ (ค่าติดลบ) → ใช้ฝั่งเดบิตของ GL และแปลงเป็นค่าลบ; ขายปกติ → ใช้ฝั่งเครดิตของ GL",
                "6. ตัวกรองบัญชีรายได้: ค่าเริ่มต้น 4xxx (ปรับได้ที่หน้าจอ)",
            ],
        ),
        (
            "English · Instructions",
            [
                '1. "Detail" Sheet = VAT report rows reconciled with GL by document number.',
                "2. Status column: ✓ Matched · ⚠ Diff · ❗ GL Missing",
                '3. "Reconciliation" Sheet = GL Total + adjustments = VAT Report Total (must equal).',
                "4. Matching key: VAT «reference doc no» ↔ GL «voucher no» (ใบสำคัญ)",
                "5. Credit notes (negative) → use GL debit, negated. Normal sales → use GL credit.",
                "6. Revenue account filter: defaults to codes starting with 4 (configurable in UI).",
            ],
        ),
        (
            "日本語 · 使い方",
            [
                "1. 「差異明細」シート = VAT 報告と GL を伝票番号で 1 対 1 照合した結果",
                "2. ステータス列: ✓ 一致 · ⚠ 差異 · ❗ GL 未検出",
                "3. 「照合まとめ」シート = GL 合計 + 調整 = VAT 報告合計 (必ず一致)",
                "4. 照合キー: VAT「参照伝票番号」↔ GL「ใบสำคัญ (証憑番号)」",
                "5. 赤伝 (負数) → GL 借方を負数化; 通常売上 → GL 貸方",
                "6. 収益科目フィルター: デフォルト 4xxx (UI で変更可)",
            ],
        ),
    ]

    ws3.cell(1, 1, "GL vs VAT Reconciliation · Reference Card").font = TITLE_FONT
    cur_row = 3
    for title, lines in instructions_blocks:
        ws3.cell(cur_row, 1, title).font = SECTION_FONT
        cur_row += 1
        for line in lines:
            cell = ws3.cell(cur_row, 1, line)
            cell.font = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cur_row += 1
        cur_row += 1  # 空行隔开

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
