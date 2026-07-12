# -*- coding: utf-8 -*-
"""ใบต่อ ภ.ง.ด.3/53 键入底稿(keying sheet)xlsx 装配(D1-6)。

事务所真实工作流是在 RD e-Filing 网页逐条键入,从不用 RD Prep 程序(T1 勘察实锤)——
本模块产出会计照抄键入用的 xlsx,版式对齐 Express 打印的 ใบต่อ 样张。纯渲染:钱/日期/
分组全由 services/workorder/steps/pnd_prep.py 装配好传入 rows,这里不查库不算钱,只把
行字典摆成表格(与 services/export/excel.py 同一"纯渲染,Decimal 只在写单元格前转 float
以便 Excel 求和"范式,内部合计仍用 Decimal 走 totals())。

键入底稿是辅助件、非官方申报文件,故不受 RD Prep txt 的「一序 ≤3 组收入类型」拼行限制,
也不像官方件那样因地址缺失被剔除——地址无数据源时诚实留空标注(ไม่มีข้อมูล),不编造。
"""

from __future__ import annotations

import io
from decimal import ROUND_HALF_UP, Decimal
from typing import Iterable, Mapping, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.tax import rdprep

_NO_DATA_TH = "ไม่มีข้อมูล"  # 地址等字段无数据源时的诚实占位,禁编造
_TWO_DP = Decimal("0.01")

_HEADER_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")

# (行字典键, 表头文案, 列宽)——照 ใบต่อ ภ.ง.ด.3/53 样张(税表D1-ภงด3-53-方案 附件 1.pdf)列序。
_COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("seq", "ลำดับ (序号)", 6),
    ("tax_id", "เลขประจำตัวผู้เสียภาษี (税号13位)", 18),
    ("name", "คำนำหน้า/ชื่อ (抬头/姓名)", 32),
    ("address", "ที่อยู่ (地址)", 30),
    ("paid_date", "วันที่จ่าย (支付日期)", 14),
    ("income_type", "ประเภทเงินได้ (收入类型)", 22),
    ("rate", "อัตรา % (税率)", 10),
    ("paid_amount", "จำนวนเงินที่จ่าย (支付金额)", 16),
    ("wht_amount", "ภาษีที่หัก (预扣税额)", 16),
    ("condition", "เงื่อนไข (条件)", 18),
)


def build_workbook(form: str, rows: Iterable[Mapping]) -> bytes:
    """键入底稿行 → xlsx 字节流。末行合计 Σจำนวนเงินที่จ่าย/Σภาษีที่หัก,与 rows 的
    Σpaid_amount/Σwht_amount 同源(见 totals()),供守恒断言对齐官方 M 记录总额。"""
    rows = list(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{form} keying"[:31]

    for col, (_, label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=1):
        _write_data_row(ws, idx + 1, idx, row)

    row_totals = totals(rows)
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="รวม (合计)").font = _HEADER_FONT
    _write_amount_cell(ws, total_row, 8, row_totals["paid_amount"], bold=True)
    _write_amount_cell(ws, total_row, 9, row_totals["wht_amount"], bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_data_row(ws, row_no: int, seq: int, row: Mapping) -> None:
    ws.cell(row=row_no, column=1, value=seq)
    ws.cell(row=row_no, column=2, value=row.get("tax_id") or "")
    ws.cell(
        row=row_no,
        column=3,
        value=f"{row.get('title_name') or ''} {row.get('payee_name') or ''}".strip(),
    )
    ws.cell(row=row_no, column=4, value=row.get("address") or _NO_DATA_TH)
    ws.cell(row=row_no, column=5, value=_paid_date_text(row.get("paid_date")))
    ws.cell(row=row_no, column=6, value=row.get("income_type") or "")
    _write_amount_cell(ws, row_no, 7, row.get("rate"))
    _write_amount_cell(ws, row_no, 8, row.get("paid_amount"))
    _write_amount_cell(ws, row_no, 9, row.get("wht_amount"))
    ws.cell(row=row_no, column=10, value=row.get("condition") or "")


def _write_amount_cell(
    ws, row_no: int, col: int, value: Optional[Decimal], *, bold: bool = False
) -> None:
    cell = ws.cell(row=row_no, column=col, value=_money(value))
    cell.alignment = _RIGHT
    cell.number_format = "#,##0.00"
    if bold:
        cell.font = _HEADER_FONT


def totals(rows: Iterable[Mapping]) -> dict[str, Decimal]:
    """Σจำนวนเงินที่จ่าย / Σภาษีที่หัก(Decimal 精确值,供守恒断言与交付物 numbers 快照用,
    不依赖 xlsx 单元格里已转 float 的显示值)。"""
    paid_total = Decimal("0")
    wht_total = Decimal("0")
    for row in rows:
        paid_total += row.get("paid_amount") or Decimal("0")
        wht_total += row.get("wht_amount") or Decimal("0")
    return {"paid_amount": paid_total, "wht_amount": wht_total}


def build_filename(*, form: str, nid: str, tax_year_be: str, tax_month: str) -> str:
    """键入底稿文件名(辅助件,不套官方 RD Prep §3 文件名规约,避免被误当官方申报文件)。"""
    return f"{form}_{nid}_{tax_year_be}_{tax_month}_keying.xlsx"


def _money(value: Optional[Decimal]) -> float:
    """Decimal → 两位小数 float,仅供 xlsx 数值单元格显示(可在 Excel 求和);真值取
    totals()/上游 Decimal,禁止拿这里的 float 回算。"""
    if value is None:
        return 0.0
    q = Decimal(str(value)).quantize(_TWO_DP, rounding=ROUND_HALF_UP)
    return float(q)


def _paid_date_text(d) -> str:
    return rdprep.to_buddhist_date_slashes(d) if d else ""
