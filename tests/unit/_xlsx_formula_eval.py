# -*- coding: utf-8 -*-
"""工作簿公式求值器(测试专用)—— 把 Excel 侧真算一遍,拿去跟权威值对。

存在的理由:openpyxl 读不到公式的**结果**,所以过去账簿工作簿的断言只能扫公式的字符串。
字符串对了不等于口径对了 —— 把销项税从内含 7/107 误写成外加 7%,全套测试照样绿:借方
gross、贷方 (gross−vat)+vat 一起变,借贷仍然相等,试算平衡照样显示 Balanced。会计拿到的
是一份内部自洽、自检绿、销项税错十几铢的表。有了这个求值器,测试才问得出真正该问的问题:
表里算出来的那个数,等不等于我们要推给 ERP 的那个数。

只实现工作簿真用到的那点子集(SUM / SUMIF / ROUND / MAX / IF + 四则 + 比较)。碰到没实现
的函数直接抛 —— 悄悄回个 0 就是再造一个假绿。钱一律 Decimal;ROUND 按 Excel 的 HALF_UP,
不是 Python 默认的 HALF_EVEN。

循环引用不特判:求值时按坐标压栈,重入即抛 CircularReference —— 这正好当循环引用的闸。
"""

from __future__ import annotations

import re
from decimal import ROUND_HALF_UP, Decimal
from typing import Any, List, Optional, Tuple

from openpyxl.utils import column_index_from_string, get_column_letter

ZERO = Decimal("0")


class CircularReference(RuntimeError):
    """某一格在求自己的值时又被引用到 —— Excel 会弹警告并把那一格显示成 0。"""


class UnsupportedFormula(RuntimeError):
    """求值器没实现的语法/函数。宁可抛,也不给一个看起来对的数。"""


_TOKEN = re.compile(
    r"""
      (?P<ws>\s+)
    | (?P<string>"(?:[^"]|"")*")
    | (?P<sheet>'(?:[^']|'')*'!)
    | (?P<ref>\$?[A-Za-z]{1,3}\$?[0-9]+)
    | (?P<number>[0-9]+(?:\.[0-9]+)?)
    | (?P<name>[A-Za-z_][A-Za-z0-9_.]*)
    | (?P<op><>|<=|>=|[-+*/()<>=,:])
    """,
    re.VERBOSE,
)


def _tokenize(formula: str) -> List[Tuple[str, str]]:
    out: List[Tuple[str, str]] = []
    pos = 0
    while pos < len(formula):
        m = _TOKEN.match(formula, pos)
        if not m:
            raise UnsupportedFormula(f"无法解析: {formula[pos:]!r} (整式 {formula!r})")
        pos = m.end()
        kind = m.lastgroup
        if kind != "ws":
            out.append((kind, m.group()))
    return out


def _to_number(value: Any) -> Decimal:
    """单元格值 → Decimal。空格按 Excel 的口径当 0;文本进算术是错的,抛。"""
    if value is None:
        return ZERO
    if isinstance(value, Decimal):
        return value
    if isinstance(value, bool):
        return Decimal(1) if value else ZERO
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    raise UnsupportedFormula(f"文本进了算术: {value!r}")


def _same_text(a: Any, b: Any) -> bool:
    """SUMIF 的文本匹配。Excel 不区分大小写;空 criteria 匹配不到任何文本票号。"""
    left, right = ("" if a is None else str(a)).strip(), ("" if b is None else str(b)).strip()
    if not right:
        return False
    return left.casefold() == right.casefold()


class _Range:
    """一段区间的取值结果 —— 保持顺序,SUMIF 要按下标把两段对齐。"""

    def __init__(self, values: List[Any]):
        self.values = values


class Evaluator:
    """按一份 openpyxl 工作簿求值。构造后可反复问,同一格只算一次。"""

    def __init__(self, workbook):
        self.wb = workbook
        self._cache: dict = {}
        self._stack: List[Tuple[str, str]] = []

    # ---- 对外 ----------------------------------------------------------
    def cell(self, sheet: str, coord: str) -> Any:
        """一格的值:常量原样,公式求值。"""
        key = (sheet, coord.replace("$", "").upper())
        if key in self._cache:
            return self._cache[key]
        if key in self._stack:
            raise CircularReference(
                f"{sheet}!{key[1]} 循环引用: {' → '.join(k[1] for k in self._stack)}"
            )
        raw = self.wb[sheet][key[1]].value
        if not (isinstance(raw, str) and raw.startswith("=")):
            self._cache[key] = raw
            return raw
        self._stack.append(key)
        try:
            value = self.evaluate(raw, sheet)
        finally:
            self._stack.pop()
        self._cache[key] = value
        return value

    def money(self, sheet: str, coord: str) -> Decimal:
        """一格的钱。表里的金额格一律按 Decimal 比,不落 float。"""
        return _to_number(self.cell(sheet, coord))

    def evaluate(self, formula: str, sheet: str) -> Any:
        tokens = _tokenize(formula[1:] if formula.startswith("=") else formula)
        parser = _Parser(self, sheet, tokens)
        value = parser.parse_expression()
        parser.expect_end()
        return value.values if isinstance(value, _Range) else value

    def all_formula_cells(self) -> List[Tuple[str, str, str]]:
        """(表名, 坐标, 公式) 全集 —— 遍历求值就是整份工作簿的循环引用/语法闸。"""
        out = []
        for ws in self.wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        out.append((ws.title, cell.coordinate, cell.value))
        return out

    # ---- 区间 ----------------------------------------------------------
    def range_values(self, sheet: str, start: str, end: str) -> _Range:
        col1, row1 = _split_ref(start)
        col2, row2 = _split_ref(end)
        values = [
            self.cell(sheet, f"{get_column_letter(c)}{r}")
            for r in range(min(row1, row2), max(row1, row2) + 1)
            for c in range(min(col1, col2), max(col1, col2) + 1)
        ]
        return _Range(values)


def _split_ref(ref: str) -> Tuple[int, int]:
    m = re.match(r"^\$?([A-Za-z]{1,3})\$?([0-9]+)$", ref)
    if not m:
        raise UnsupportedFormula(f"不是单元格引用: {ref!r}")
    return column_index_from_string(m.group(1).upper()), int(m.group(2))


def _flatten(args) -> List[Any]:
    out: List[Any] = []
    for a in args:
        out.extend(a.values) if isinstance(a, _Range) else out.append(a)
    return out


def _fn_sum(args) -> Decimal:
    return sum((_to_number(v) for v in _flatten(args) if v is not None), ZERO)


def _fn_sumif(args) -> Decimal:
    if len(args) != 3:
        raise UnsupportedFormula("SUMIF 只支持三参数形式")
    criteria_range, criteria, sum_range = args
    if not isinstance(criteria_range, _Range) or not isinstance(sum_range, _Range):
        raise UnsupportedFormula("SUMIF 的第一、三参数必须是区间")
    total = ZERO
    for key, value in zip(criteria_range.values, sum_range.values):
        if _same_text(key, criteria):
            total += _to_number(value)
    return total


def _fn_round(args) -> Decimal:
    if len(args) != 2:
        raise UnsupportedFormula("ROUND 只支持两参数形式")
    value, digits = _to_number(args[0]), int(_to_number(args[1]))
    return value.quantize(Decimal(1).scaleb(-digits), rounding=ROUND_HALF_UP)


def _fn_max(args) -> Decimal:
    return max((_to_number(v) for v in _flatten(args)), default=ZERO)


def _fn_if(args) -> Any:
    if len(args) != 3:
        raise UnsupportedFormula("IF 只支持三参数形式")
    return args[1] if args[0] else args[2]


_FUNCTIONS = {
    "SUM": _fn_sum,
    "SUMIF": _fn_sumif,
    "ROUND": _fn_round,
    "MAX": _fn_max,
    "IF": _fn_if,
}

_COMPARISONS = {
    "=": lambda a, b: a == b,
    "<>": lambda a, b: a != b,
    ">": lambda a, b: a > b,
    "<": lambda a, b: a < b,
    ">=": lambda a, b: a >= b,
    "<=": lambda a, b: a <= b,
}


class _Parser:
    """递归下降。文法只覆盖本仓工作簿写得出的形状,别拿它当通用 Excel 解析器。"""

    def __init__(self, ev: Evaluator, sheet: str, tokens: List[Tuple[str, str]]):
        self.ev = ev
        self.sheet = sheet
        self.tokens = tokens
        self.pos = 0

    def peek(self) -> Optional[Tuple[str, str]]:
        return self.tokens[self.pos] if self.pos < len(self.tokens) else None

    def take(self) -> Tuple[str, str]:
        tok = self.peek()
        if tok is None:
            raise UnsupportedFormula("公式提前结束")
        self.pos += 1
        return tok

    def accept(self, kind: str, text: Optional[str] = None) -> Optional[Tuple[str, str]]:
        tok = self.peek()
        if tok and tok[0] == kind and (text is None or tok[1].upper() == text):
            self.pos += 1
            return tok
        return None

    def expect_end(self) -> None:
        if self.peek() is not None:
            raise UnsupportedFormula(f"多余的记号: {self.tokens[self.pos:]!r}")

    def parse_expression(self) -> Any:
        left = self.parse_additive()
        tok = self.peek()
        if tok and tok[0] == "op" and tok[1] in _COMPARISONS:
            self.take()
            right = self.parse_additive()
            return _COMPARISONS[tok[1]](_to_number(left), _to_number(right))
        return left

    def parse_additive(self) -> Any:
        value = self.parse_multiplicative()
        while True:
            tok = self.peek()
            if not (tok and tok[0] == "op" and tok[1] in "+-"):
                return value
            self.take()
            right = self.parse_multiplicative()
            value = (
                _to_number(value) + _to_number(right)
                if tok[1] == "+"
                else _to_number(value) - _to_number(right)
            )

    def parse_multiplicative(self) -> Any:
        value = self.parse_unary()
        while True:
            tok = self.peek()
            if not (tok and tok[0] == "op" and tok[1] in "*/"):
                return value
            self.take()
            right = self.parse_unary()
            value = (
                _to_number(value) * _to_number(right)
                if tok[1] == "*"
                else _to_number(value) / _to_number(right)
            )

    def parse_unary(self) -> Any:
        if self.accept("op", "-"):
            return -_to_number(self.parse_unary())
        return self.parse_primary()

    def parse_primary(self) -> Any:
        kind, text = self.take()
        if kind == "number":
            return Decimal(text)
        if kind == "string":
            return text[1:-1].replace('""', '"')
        if kind == "op" and text == "(":
            value = self.parse_expression()
            if not self.accept("op", ")"):
                raise UnsupportedFormula("括号没闭合")
            return value
        if kind == "sheet":
            return self.parse_ref(sheet=text[1:-2].replace("''", "'"))
        if kind == "ref":
            self.pos -= 1
            return self.parse_ref(sheet=self.sheet)
        if kind == "name":
            return self.parse_call(text.upper())
        raise UnsupportedFormula(f"看不懂的记号: {kind}={text!r}")

    def parse_ref(self, *, sheet: str) -> Any:
        _kind, start = self.take()
        if self.accept("op", ":"):
            _kind, end = self.take()
            return self.ev.range_values(sheet, start, end)
        return self.ev.cell(sheet, start)

    def parse_call(self, name: str) -> Any:
        if not self.accept("op", "("):
            raise UnsupportedFormula(f"{name} 后面没有括号")
        args: List[Any] = []
        if not self.accept("op", ")"):
            args.append(self.parse_expression())
            while self.accept("op", ","):
                args.append(self.parse_expression())
            if not self.accept("op", ")"):
                raise UnsupportedFormula(f"{name}( 没闭合")
        fn = _FUNCTIONS.get(name)
        if fn is None:
            raise UnsupportedFormula(f"求值器没实现函数 {name}")
        return fn(args)
