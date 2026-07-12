# -*- coding: utf-8 -*-
"""ใบแนบ ภ.ง.ด.1 键入底稿 xlsx 装配(照 D1-6 pnd_keying_sheet 范式 · 方案 §4.1)。

会计照抄键入 e-Filing 用的人读版式,列=金标 header 语义列,末行 Σ合计(Decimal 守恒,
totals() 同源,不拿 xlsx 里已转 float 的显示值回算)。纯渲染:钱只在写单元格前转 float
以便 Excel 求和,内部合计仍 Decimal。辅助件、非官方申报件,不套官方文件名规约。
"""

from __future__ import annotations

import io
from decimal import ROUND_HALF_UP, Decimal
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

_TWO_DP = Decimal("0.01")
_HEADER_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")

# (PayrollRow 属性名, 表头文案, 列宽)——照金标 header 列序(称谓/名/姓合成一列人读)。
_COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("income_code", "รหัสเงินได้ (收入码)", 12),
    ("seq", "ลำดับ (序号)", 6),
    ("employee_id", "เลข 13 หลัก (身份证)", 18),
    ("_name", "ชื่อ-สกุล (称谓/姓名)", 30),
    ("_paid_date", "วันที่จ่าย (支付日)", 14),
    ("paid_amount", "จำนวนเงินที่จ่าย (支付金额)", 18),
    ("wht_amount", "ภาษีที่หัก (预扣税额)", 16),
    ("condition", "เงื่อนไข (条件)", 10),
)

_AMOUNT_COLS = {"paid_amount", "wht_amount"}


def build_workbook(rows: Iterable) -> bytes:
    """键入底稿行 → xlsx 字节流。末行合计 Σ支付/Σ扣税,与 totals() 同源供守恒断言。"""
    rows = list(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "PND1 keying"

    for col, (_, label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=2):
        _write_row(ws, idx, row)

    row_totals = totals(rows)
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="รวม (合计)").font = _HEADER_FONT
    _write_amount(ws, total_row, 6, row_totals["paid_amount"], bold=True)
    _write_amount(ws, total_row, 7, row_totals["wht_amount"], bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def totals(rows: Iterable) -> dict:
    """Σ支付 / Σ扣税(Decimal 精确值,供守恒断言与交付物快照,禁拿 xlsx float 回算)。"""
    paid = Decimal("0")
    wht = Decimal("0")
    for row in rows:
        paid += row.paid_amount or Decimal("0")
        wht += row.wht_amount or Decimal("0")
    return {"paid_amount": paid, "wht_amount": wht}


def build_filename(*, nid: str, tax_year_be: str, tax_month: str) -> str:
    """键入底稿文件名(辅助件,不套官方 §3 规约,避免误当官方申报文件)。"""
    return f"PND1_{nid}_{tax_year_be}_{tax_month}_keying.xlsx"


def _write_row(ws, row_no: int, row) -> None:
    for col, (attr, _, _) in enumerate(_COLUMNS, start=1):
        if attr in _AMOUNT_COLS:
            _write_amount(ws, row_no, col, getattr(row, attr))
        elif attr == "_name":
            ws.cell(
                row=row_no,
                column=col,
                value=f"{row.title} {row.first_name} {row.last_name}".strip(),
            )
        elif attr == "_paid_date":
            ws.cell(
                row=row_no, column=col, value=row.paid_date.isoformat() if row.paid_date else ""
            )
        else:
            ws.cell(row=row_no, column=col, value=getattr(row, attr))


def _write_amount(
    ws, row_no: int, col: int, value: Optional[Decimal], *, bold: bool = False
) -> None:
    cell = ws.cell(row=row_no, column=col, value=_money(value))
    cell.alignment = _RIGHT
    cell.number_format = "#,##0.00"
    if bold:
        cell.font = _HEADER_FONT


def _money(value: Optional[Decimal]) -> float:
    """Decimal → 两位小数 float,仅供 xlsx 显示求和;真值取 totals()/上游 Decimal。"""
    if value is None:
        return 0.0
    return float(Decimal(str(value)).quantize(_TWO_DP, rounding=ROUND_HALF_UP))
