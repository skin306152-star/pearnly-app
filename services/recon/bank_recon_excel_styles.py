# -*- coding: utf-8 -*-
"""bank_recon_excel 共享调色板 + 通用样式 helper · bank_recon_excel 拆分 leaf.

Sheet 2/5/6 共用的表头/边框/日期格式 helper。Sheet 1 的栏位锚点样式自带局部调色板,
不走这里。0 逻辑改,从 export_bank_recon_excel 原样提出。
"""

from datetime import date
from typing import Optional

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ── Color palette ──────────────────────────────────────────────────
COLOR_HEADER = "2D6A4F"  # dark green
COLOR_SUBHEAD = "52B788"  # medium green
COLOR_MATCHED = "D8F3DC"  # light green
COLOR_L2 = "FFF3CD"  # amber (date tolerance)
COLOR_L3 = "FFE0CC"  # orange (amount only)
COLOR_GL_ONLY = "E8D5F5"  # purple
COLOR_ST_ONLY = "D4E6F1"  # blue
COLOR_DIFF = "FFDAD6"  # red for non-zero diff
COLOR_OK = "D8F3DC"  # green for zero diff
COLOR_ROW_ALT = "F8F9FA"  # alternating row


def _hdr_style(ws, row, col, text, color=COLOR_HEADER, bold=True, size=10):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = Font(bold=bold, color="FFFFFF", size=size)
    cell.fill = PatternFill("solid", fgColor=color)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return cell


def _border_range(ws, min_row, max_row, min_col, max_col):
    thin = Side(style="thin", color="CCCCCC")
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _fmt_date(d: Optional[date]) -> str:
    if d is None:
        return ""
    return d.strftime("%d/%m/%Y")
