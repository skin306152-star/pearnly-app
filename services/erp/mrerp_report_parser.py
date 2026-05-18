# -*- coding: utf-8 -*-
"""
services/erp/mrerp_report_parser.py

Parses the MR.ERP import report xlsx (downloaded from
impartran/component/report.php after a sales_credit bulk upload).

Why this exists:
    importpc.php returns "2" on completion, but "2" does NOT mean every row
    was written to the DB — it means "processing finished, report generated."
    The true per-row outcome lives in the last column ("หมายเหตุ") of every
    sheet in the report xlsx:
        - empty / None  -> row written successfully
        - non-empty     -> rejection reason (Thai text; multiple reasons are
                           joined with '\n' inside a single cell)

    The header sheet ("Worksheet") is the primary truth: one row per invoice.
    Detail sheets ("Worksheet 1" / "Worksheet 2") can also carry per-row notes
    if a child row fails its own validation (e.g. product code missing); the
    parser collects notes from every sheet and aggregates them by invoice_no.

Sample fixture used for tests:
    docs/integrations/samples/report_failure_customer_not_found.xlsx
    (real report.php download from 2026-05-18, customer 99-PEARNLYTEST-001
    that does not exist in master data)
"""

from __future__ import annotations

import io
import logging
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook

logger = logging.getLogger(__name__)


HEADER_INVOICE_NO = "เลขที่"
HEADER_NOTE = "หมายเหตุ"
PRIMARY_SHEET_NAME = "Worksheet"


@dataclass
class ImportReportRow:
    """One invoice's outcome aggregated across all sheets in the report."""
    invoice_no: str
    reasons: List[str] = field(default_factory=list)
    # Per-sheet breakdown so the UI can show "Sheet X said: ...".
    details: List[Dict[str, str]] = field(default_factory=list)

    @property
    def success(self) -> bool:
        return not self.reasons


@dataclass
class ImportReport:
    """Parsed result of a single MR.ERP import report xlsx.

    success: list of invoice_no strings (DB write confirmed)
    failed:  list of ImportReportRow, each with reasons[] + details[]
    """
    success: List[str] = field(default_factory=list)
    failed: List[ImportReportRow] = field(default_factory=list)

    @property
    def total(self) -> int:
        return len(self.success) + len(self.failed)

    @property
    def all_success(self) -> bool:
        return not self.failed and bool(self.success)


def parse_import_report(xlsx_bytes: bytes) -> ImportReport:
    """Parse the xlsx returned by impartran/component/report.php.

    Returns an ImportReport with one entry per invoice_no found in the
    report. The header sheet's invoice_no column is the canonical list;
    notes from detail sheets are merged into the matching invoice.

    Raises ValueError if the xlsx is missing the expected structure.
    """
    if not xlsx_bytes:
        raise ValueError("empty report xlsx bytes")

    try:
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"cannot open report xlsx: {e}") from e

    if PRIMARY_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"report xlsx missing primary sheet {PRIMARY_SHEET_NAME!r}; "
            f"got {wb.sheetnames}"
        )

    invoices: Dict[str, ImportReportRow] = {}

    for sname in wb.sheetnames:
        ws = wb[sname]
        if ws.max_row is None or ws.max_row < 2:
            continue

        invoice_col, note_col = _find_columns(ws)
        if invoice_col is None or note_col is None:
            continue

        for r in range(2, ws.max_row + 1):
            raw_inv = ws.cell(row=r, column=invoice_col).value
            if raw_inv is None:
                continue
            inv = str(raw_inv).strip()
            if not inv:
                continue

            row = invoices.setdefault(inv, ImportReportRow(invoice_no=inv))

            raw_note = ws.cell(row=r, column=note_col).value
            note_text = str(raw_note).strip() if raw_note is not None else ""
            if not note_text:
                continue

            # MR.ERP joins multiple errors with literal \n inside one cell.
            for line in note_text.split("\n"):
                line = line.strip()
                if not line:
                    continue
                row.reasons.append(line)
                row.details.append({"sheet": sname, "reason": line})

    result = ImportReport()
    for inv, row in invoices.items():
        if row.success:
            result.success.append(inv)
        else:
            result.failed.append(row)

    return result


def _find_columns(ws) -> tuple[Optional[int], Optional[int]]:
    """Locate the 'เลขที่' (invoice no) and 'หมายเหตุ' (note) columns by
    exact header-cell text match. Returns (invoice_col, note_col), either
    can be None if absent.

    MR.ERP pads the header row with trailing empty cells (sheet 1's dim is
    A1:Z2 even though headers stop at col 19), so we look up by label
    rather than positional index. Several columns share the prefix
    'หมายเหตุ' (note1/note2/note3 are uploadable, separate from the report's
    final 'หมายเหตุ' verdict column) — exact equality is required.
    """
    invoice_col: Optional[int] = None
    note_col: Optional[int] = None
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=c).value
        if v is None:
            continue
        label = str(v).strip()
        if invoice_col is None and label == HEADER_INVOICE_NO:
            invoice_col = c
        if label == HEADER_NOTE:
            note_col = c
    return invoice_col, note_col
