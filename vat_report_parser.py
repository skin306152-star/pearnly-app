# -*- coding: utf-8 -*-
"""
v118.32.1 · Pearnly · VAT 报告多格式解析器
- .xlsx/.xls               → openpyxl 表头自动识别
- .pdf(电子版)            → pdfplumber 文字+表格提取(零成本)
- .pdf(扫描)/.jpg/.png 等 → Gemini OCR(VAT 报告专用 prompt)
"""

import io
import os
import re
import json
import logging
from typing import List, Dict, Any, Optional

from field_comparator import (
    normalize_tax_id,
    normalize_branch,
    parse_date,
)

# v118.35.0.3 · 包装 pydantic ValidationError 为单行用户友好提示
from services.ocr.error_format import short_error as _short_err

logger = logging.getLogger(__name__)
PARSER_VERSION = "1.1.0"

# ── 列名关键词(中英泰混合) ────────────────────────────
_DATE_H = {"วันที่", "date", "วัน/เดือน/ปี", "วันที่ใบกำกับ", "วันที่เอกสาร"}
_NO_H = {"เลขที่", "invoice no", "เลขที่ใบกำกับ", "เลขที่เอกสาร", "no."}
# v118.32.5 · GL对账·参考单据号列(优先于 invoice_no 检测，专给 GL 匹配用)
_REF_H = {"อ้างอิง", "reference", "ref no", "เอกสารอ้างอิง", "เลขอ้างอิง"}
_NAME_H = {"ชื่อ", "name", "ผู้ซื้อ", "customer"}
_TAX_H = {"เลขประจำตัว", "tax id", "เลขภาษี", "tin"}
_BRANCH_H = {"สาขา", "branch"}
_NET_H = {"ยอดก่อน vat", "มูลค่า", "net", "ก่อน vat", "ยอดก่อน"}
_VAT_H = {"ภาษี", "vat", "7%"}
_TOTAL_H = {"ยอดรวม", "total", "รวม"}
_SKIP_H = {"รวม", "total", "ยอดรวม", "grand", "รวมทั้งสิ้น", "รวมแต่ละหน้า"}


def _hit(header: str, hints: set) -> bool:
    h = header.strip().lower()
    return any(hint in h for hint in hints)


def _to_float(val) -> Optional[float]:
    try:
        s = str(val or "").strip().replace(",", "").replace(" ", "").replace(" ", "")
        if not s or s in {"-", "–"}:
            return None
        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1]
        v = round(float(s), 2)
        return -v if neg else v
    except Exception:
        return None


# v118.32.5.5.3 · 统一过滤 Gemini / pdfplumber 误抓的页眉/分隔/表头行
# 判定：发票号必须含 ≥2 字母数字组合 · 且不全是装饰符 · 不含 VAT 报告表头关键词
_VAT_HEAD_KW = (
    "ใบกำกับภาษี",
    "ใบกํากับภาษี",
    "ลำดับ",
    "ลําดับ",
    "ภาษีมูลค่า",
    "ภาษีมูลคา",
    "เลขที่ออก",
    "เลขประจำตัว",
    "เลขประจําตัว",
    "หมายเหตุ",
    "ผู้ซื้อ",
    "ผูซื้อ",
    "ผู้รับบริการ",
    "ผูรับบริการ",
    "หรือบริการ",
    "ชื่อสถาน",
    "ชื่อผู้",
    "ชื่อผู",
)
_VAT_ORN_RE = re.compile(r"^[\-=<>.\s|*_+]+$")
_VAT_ALNUM_RE = re.compile(r"[A-Za-z0-9]{2,}")


def _filter_garbage_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """剔除 PDF/Gemini 误识别的页眉、分隔线、表头行 · 数据行的 doc_no 应是单 token"""
    out = []
    for r in rows:
        inv = str(r.get("report_invoice_no") or "").strip()
        ref = str(r.get("report_ref_no") or "").strip()
        key = ref or inv
        if not key or len(key) < 3 or len(key) > 30:
            continue
        if _VAT_ORN_RE.match(key):
            continue
        if not _VAT_ALNUM_RE.search(key):
            continue
        if any(kw in key for kw in _VAT_HEAD_KW):
            continue
        # 数据行的 doc_no 是单 token · 不应含多空格(整行误抓会有很多空格)
        if key.count(" ") > 1:
            continue
        out.append(r)
    return out


# ======================================================================
# 1. Excel 解析(零成本)
# ======================================================================


def parse_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl 未安装", "rows": []}

    rows: List[Dict] = []
    meta: Dict[str, Any] = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        header_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            text = " ".join(str(c or "").lower() for c in row)
            if "เลขที่" in text or "invoice no" in text or "วันที่" in text:
                header_idx = i
                break
        if header_idx is None:
            return {"ok": False, "error": "找不到表头行", "rows": []}

        headers = tuple(str(c.value or "") for c in ws[header_idx])
        col_map = _map_columns(headers)

        row_no = 0
        for raw in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            cells = list(raw)
            if not any(c for c in cells if c not in (None, "")):
                continue
            first = str(cells[0] or "").strip().lower()
            if any(k in first for k in _SKIP_H):
                if "amount_pre_vat" in col_map:
                    meta["total_amount_pre_vat"] = _to_float(cells[col_map["amount_pre_vat"]])
                if "vat_amount" in col_map:
                    meta["total_vat"] = _to_float(cells[col_map["vat_amount"]])
                continue
            row_no += 1
            rows.append(_build_row(row_no, cells, col_map))

        return {
            "ok": True,
            "rows": rows,
            "meta": meta,
            "warnings": [],
            "parser_version": PARSER_VERSION,
            "row_count": len(rows),
            "method": "excel",
        }
    except Exception as e:
        logger.error(f"parse_excel failed: {e}")
        return {"ok": False, "error": str(e), "rows": []}


def _map_columns(headers) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for i, h in enumerate(headers):
        h = (str(h) or "").strip()
        if not h:
            continue
        if _hit(h, _DATE_H) and "date" not in col_map:
            col_map["date"] = i
        # v118.32.5 · 参考单据号优先检测（关键词更specific，避免被 _NO_H 抢走）
        elif _hit(h, _REF_H) and "ref_no" not in col_map:
            col_map["ref_no"] = i
        elif _hit(h, _NO_H) and "invoice_no" not in col_map:
            col_map["invoice_no"] = i
        elif _hit(h, _NAME_H) and "buyer_name" not in col_map:
            col_map["buyer_name"] = i
        elif _hit(h, _TAX_H) and "buyer_tax_id" not in col_map:
            col_map["buyer_tax_id"] = i
        elif _hit(h, _BRANCH_H) and "buyer_branch" not in col_map:
            col_map["buyer_branch"] = i
        elif _hit(h, _NET_H) and "amount_pre_vat" not in col_map:
            col_map["amount_pre_vat"] = i
        elif _hit(h, _VAT_H) and "vat_amount" not in col_map:
            col_map["vat_amount"] = i
        elif _hit(h, _TOTAL_H) and "total_amount" not in col_map:
            col_map["total_amount"] = i
    return col_map


def _build_row(row_no: int, cells: list, col_map: Dict[str, int]) -> Dict[str, Any]:
    parsed: Dict[str, Any] = {"row_no": row_no}
    for field, ci in col_map.items():
        raw = cells[ci] if ci < len(cells) else None
        val = str(raw).strip() if raw is not None else ""
        if field == "date":
            d = parse_date(val)
            parsed["report_date"] = d.isoformat() if d else val
        elif field == "invoice_no":
            parsed["report_invoice_no"] = val
        elif field == "ref_no":
            parsed["report_ref_no"] = val  # v118.32.5 · GL对账匹配键
        elif field == "buyer_name":
            parsed["report_buyer_name"] = val
        elif field == "buyer_tax_id":
            parsed["report_buyer_tax_id"] = normalize_tax_id(val)
        elif field == "buyer_branch":
            parsed["report_buyer_branch"] = normalize_branch(val)
        elif field == "amount_pre_vat":
            parsed["report_amount_pre_vat"] = _to_float(val)
        elif field == "vat_amount":
            parsed["report_vat_amount"] = _to_float(val)
        elif field == "total_amount":
            parsed["report_amount"] = _to_float(val)
    parsed["is_individual"] = not bool(parsed.get("report_buyer_tax_id"))
    return parsed


# ======================================================================
# 2. PDF 文本提取(电子版 PDF · 零成本)
# ======================================================================


def parse_pdf_text(file_bytes: bytes) -> Optional[Dict[str, Any]]:
    """成功返回 dict · 失败/扫描件返回 None(回退 Gemini)"""
    try:
        import pdfplumber
    except ImportError:
        logger.info("[vat_pdf_text] pdfplumber 未安装 · 直接回退 Gemini")
        return None

    try:
        rows: List[Dict] = []
        meta: Dict[str, Any] = {}
        total_text_len = 0
        row_no = 0
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = None
                    hi = 0
                    for i, r in enumerate(table[:10]):
                        joined = " ".join(str(c or "").lower() for c in r)
                        if "เลขที่" in joined or "invoice" in joined or "วันที่" in joined:
                            header = r
                            hi = i
                            break
                    if not header:
                        continue
                    col_map = _map_columns(header)
                    for body in table[hi + 1 :]:
                        if not body or not any(c for c in body if c):
                            continue
                        first = str(body[0] or "").strip().lower()
                        if any(k in first for k in _SKIP_H):
                            if "amount_pre_vat" in col_map and col_map["amount_pre_vat"] < len(
                                body
                            ):
                                meta["total_amount_pre_vat"] = _to_float(
                                    body[col_map["amount_pre_vat"]]
                                )
                            if "vat_amount" in col_map and col_map["vat_amount"] < len(body):
                                meta["total_vat"] = _to_float(body[col_map["vat_amount"]])
                            continue
                        row_no += 1
                        rows.append(_build_row(row_no, body, col_map))
                total_text_len += len(page.extract_text() or "")

        # 文本极少 → 扫描件 · 回退 Gemini
        if total_text_len < 200 and not rows:
            logger.info(f"[vat_pdf_text] 文本仅 {total_text_len} 字符 · 判定扫描件 · 回退 Gemini")
            return None
        if not rows:
            # v118.32.5 · 表抽取 0 行但文字层存在 → 试文字行 regex
            # 很多泰文会计软件导出的 VAT 报告没有可见表格线，pdfplumber 抓不到 tables
            logger.info(f"[vat_pdf_text] 表抽取 0 行，文本 {total_text_len} 字符 · 试文字行 regex")
            text_rows = _parse_vat_pdf_text_lines(file_bytes)
            if text_rows:
                logger.info(f"[vat_pdf_text.regex] 文字行模式抽到 {len(text_rows)} 行")
                return {
                    "ok": True,
                    "rows": text_rows,
                    "meta": {},
                    "warnings": [],
                    "parser_version": PARSER_VERSION,
                    "row_count": len(text_rows),
                    "method": "pdf_text_regex",
                }
            return None

        return {
            "ok": True,
            "rows": rows,
            "meta": meta,
            "warnings": [],
            "parser_version": PARSER_VERSION,
            "row_count": len(rows),
            "method": "pdf_text",
        }
    except Exception as e:
        logger.warning(f"[vat_pdf_text] 失败 · 回退 Gemini · {type(e).__name__}: {e}")
        return None


# ── v118.32.5 · VAT 报告文字行 regex 解析（备用通道）──────────────────
# 触发条件：pdfplumber 找不到表格框线（多数泰文会计软件导出的 PDF）
# 行模式：[seq?] DATE INV_NO [REF_NO] BUYER_NAME [TAX_ID] [BRANCH] AMT_PRE_VAT VAT_AMT
# 用启发式：行末必须两个 .2 小数，行内含一个 DD/MM/YYYY
_VAT_DATE_RE = re.compile(r"(\d{1,2}[\-/.]\d{1,2}[\-/.]\d{2,4})")
_VAT_TAIL_2N = re.compile(
    r"((?:\(\s*[\d,]+\.\d{2}\s*\)|-?[\d,]+\.\d{2}))"
    r"\s+"
    r"((?:\(\s*[\d,]+\.\d{2}\s*\)|-?[\d,]+\.\d{2}))"
    r"\s*$"
)
_TAX_ID_RE = re.compile(r"\b(\d{13})\b")
_BRANCH_RE = re.compile(r"\b(0\d{4})\b")
# v118.32.5.5.5 · ref_no token 严格判定:字母开头 + 必须含数字
# 兼容 IV6812001 / OHC06/2025 / DXATM25-010 · 排除纯字母客户名(OHANAHAN/ASAHI/TSUKEMONO/SIM)
_REF_TOKEN_RE = re.compile(r"^(?=.*\d)[A-Za-z][A-Za-z0-9\-/.]+$")
# v118.32.5.5.5 · 切 invoice_no 粘连尾(wnf PDF 字距导致 "DXOHC25-006OHANAHAN" 粘成 1 个 token)
# 主部分 = 字母+数字(可含横杠/斜杠) · 尾巴 = ≥3 个大写字母粘连客户名
_GLUED_TAIL_RE = re.compile(r"^([A-Za-z]+\d+(?:[\-/]\d+)?)([A-Z]{3,}.*)$")


def _parse_vat_pdf_text_lines(file_bytes: bytes) -> List[Dict[str, Any]]:
    """对每页 extract_text() 的行做 regex 抽取
    返回 rows · 失败返回 []
    """
    try:
        import pdfplumber
    except ImportError:
        return []

    rows: List[Dict[str, Any]] = []
    row_no = 0
    try:
        with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
            for page in pdf.pages:
                txt = page.extract_text() or ""
                for raw_line in txt.split("\n"):
                    line = raw_line.strip()
                    if not line:
                        continue
                    # 必须行末两个 .2 金额
                    m_tail = _VAT_TAIL_2N.search(line)
                    if not m_tail:
                        continue
                    # 必须有日期
                    m_date = _VAT_DATE_RE.search(line)
                    if not m_date:
                        continue
                    # 跳过表头/合计：含 "เลขที่" "ยอดรวม" "รวมทั้งสิ้น" 等
                    low = line.lower()
                    if "เลขที่" in line and "ใบกำกับ" in line:
                        continue
                    if any(k in low for k in {"รวมทั้งสิ้น", "รวมแต่ละหน้า", "grand total"}):
                        continue

                    amount_pre_vat = _to_float(m_tail.group(1))
                    vat_amount = _to_float(m_tail.group(2))
                    if amount_pre_vat is None and vat_amount is None:
                        continue

                    # 中间部分：日期之后到行末 2 金额之前
                    mid = line[m_date.end() : m_tail.start()].strip()
                    # 启发：mid 头几个 token 通常是 invoice_no / ref_no
                    tokens = mid.split()
                    invoice_no = ""
                    ref_no = ""
                    if tokens:
                        # v118.32.5.5.4 · 跳过开头的序号(纯数字 1-3 位 · wnf 格式行首是序号)
                        idx = 0
                        if idx < len(tokens) and re.fullmatch(r"\d{1,3}", tokens[idx]):
                            idx += 1
                        if idx < len(tokens):
                            invoice_no = tokens[idx]
                            # v118.32.5.5.5 · 切粘连尾巴(DXOHC25-006OHANAHAN → DXOHC25-006)
                            m_glued = _GLUED_TAIL_RE.match(invoice_no)
                            if m_glued:
                                invoice_no = m_glued.group(1)
                            # ref_no 必须含数字 · 不接受纯字母客户名
                            if (
                                idx + 1 < len(tokens)
                                and _REF_TOKEN_RE.match(tokens[idx + 1])
                                and not _TAX_ID_RE.match(tokens[idx + 1])
                            ):
                                ref_no = tokens[idx + 1]
                    # 提取 13 位税号
                    m_tax = _TAX_ID_RE.search(mid)
                    tax_id = m_tax.group(1) if m_tax else ""
                    # 提取 5 位分支码（0xxxx）· 默认 "00000"
                    m_br = _BRANCH_RE.search(mid)
                    branch = m_br.group(1) if m_br else ("00000" if tax_id else "")
                    # 买方名：去掉这些已识别的 token 后的剩余串
                    name = mid
                    for cut in (tax_id, branch, invoice_no, ref_no):
                        if cut and cut in name:
                            name = name.replace(cut, " ", 1)
                    # v118.32.5.2 · 清洗 PDF 抽文残留：
                    # 1) "00000" 等分支码（pdfplumber 字距问题可能把它从分支号位置漏掉）
                    name = re.sub(r"\b0\d{4}\b", " ", name)
                    # 2) 行首 "SI 1" / "I26 1" 等 — pdfplumber 字距问题把 "SI000001" 抽成 "SI 1"
                    name = re.sub(r"^\s*(?:[A-Za-z]{1,4}\s+\d+\s+){1,3}", " ", name)
                    # 3) 行尾纯数字（残留的金额/seq）
                    name = re.sub(r"\s+\d{1,5}\s*$", "", name)
                    # 4) 合并多空格
                    name = re.sub(r"\s+", " ", name).strip()

                    # 日期归一化（佛历→西历）
                    d = parse_date(m_date.group(1))
                    date_str = d.isoformat() if d else m_date.group(1)

                    row_no += 1
                    rows.append(
                        {
                            "row_no": row_no,
                            "report_date": date_str,
                            "report_invoice_no": invoice_no,
                            "report_ref_no": ref_no,
                            "report_buyer_name": name,
                            "report_buyer_tax_id": normalize_tax_id(tax_id),
                            "report_buyer_branch": normalize_branch(branch),
                            "report_amount_pre_vat": amount_pre_vat,
                            "report_vat_amount": vat_amount,
                            "report_amount": (amount_pre_vat or 0) + (vat_amount or 0),
                            "is_individual": not bool(tax_id),
                        }
                    )
    except Exception as e:
        logger.warning(f"[vat_pdf_text.regex] 失败: {e}")
        return []
    return rows


# ======================================================================
# 3. Gemini OCR(扫描 PDF / 图片)
# ======================================================================

_GEMINI_PROMPT = """You are extracting data from a Thai Sales VAT Report (รายงานภาษีขาย).
This is a monthly report listing all sales tax invoices issued by a VAT-registered business.

⚠ STRICT RULES (must follow):
1. Output EXACTLY what is printed. Do NOT rewrite, normalize, translate, or "correct" any value.
2. For names: preserve the original spelling, spacing, prefixes (คุณ / นาย / นางสาว / บริษัท ... จำกัด), and language (Thai/English). If a name is "คุณสุพัชญ์ สันติวงษ์", output exactly that — do not change to "สุวพัชญ์" or any other variant.
3. For tax IDs: output exactly the 13 digits as printed. If 1-2 digits are unclear, look carefully — DO NOT guess. Output what is most likely printed.
4. For amounts: preserve all decimal places. "7,595.00" → 7595.00 (drop comma, keep decimals). NEVER round or truncate.
5. Read every row left-to-right top-to-bottom. Do not skip rows.

Extract ALL data rows. Skip page headers, signatures, summary/total rows.

For each row, return a JSON object with these fields:
- row_no: integer (sequence number ลำดับ)
- report_date: date "YYYY-MM-DD". If year is Buddhist Era (BE, > 2400), subtract 543. e.g. 01/03/2569 → "2026-03-01"
- report_invoice_no: tax invoice number (เลขที่ใบกำกับภาษี / เลขที่) — copy EXACTLY as printed
- report_ref_no: reference document number (เลขที่เอกสารอ้างอิง / เลขอ้างอิง) — copy EXACTLY as printed; empty string "" if column not present
- report_buyer_name: buyer name (Thai or English) — copy EXACTLY as printed, including all prefixes and spaces
- report_buyer_tax_id: 13-digit Thai tax ID, digits ONLY (drop dashes/spaces). Empty string "" for cash sale / individual buyer (ลูกค้าขายเงินสด/บุคคลธรรมดา)
- report_buyer_branch: "00000" if HQ (สำนักงานใหญ่/สนญ./head office), or 5-digit code, or empty string for individual
- report_amount_pre_vat: pre-VAT amount (มูลค่าสินค้าหรือบริการ) as number
- report_vat_amount: 7% VAT amount (จำนวนเงินภาษีมูลค่าเพิ่ม) as number
- report_amount: total = pre-VAT + VAT (use printed value if shown; compute only if missing)

Self-check before output: For each row, verify report_amount_pre_vat + report_vat_amount ≈ report_amount (±0.01). If not, re-read the row.

Return ONLY:
{"rows": [...], "meta": {"total_amount_pre_vat": number, "total_vat": number}}
"""


# v118.32.4.9.6 · 大 PDF 拆页 + 并行处理(防 Gemini 504)
# 真实国税局 33 行 PDF 单批 2 页 + 45s timeout 仍然 504 · 改单页一批 + 60s + 单次重试
# 单页一批保证 Gemini 处理负担最小 · 总并发 4 路 · 总耗时基本持平
_BATCH_PAGES = 1
_BATCH_WORKERS = 8


def parse_with_gemini_paged(file_bytes: bytes, api_key: Optional[str] = None) -> Dict[str, Any]:
    """PDF 专用:拆页 + 并行调 Gemini · 合并 rows + 重编 row_no
    页数 <= _BATCH_PAGES 时直接调 parse_with_gemini(整文件)"""
    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return parse_with_gemini(file_bytes, "application/pdf", api_key=api_key)

    try:
        reader = PdfReader(io.BytesIO(file_bytes))
        n_pages = len(reader.pages)
    except Exception as e:
        logger.warning(f"[vat_paged] 读页失败 · 回退整文件: {e}")
        return parse_with_gemini(file_bytes, "application/pdf", api_key=api_key)

    if n_pages <= _BATCH_PAGES:
        return parse_with_gemini(file_bytes, "application/pdf", api_key=api_key)

    # 拆批:每 _BATCH_PAGES 页一批
    batches: List[bytes] = []
    try:
        for start in range(0, n_pages, _BATCH_PAGES):
            w = PdfWriter()
            for j in range(start, min(start + _BATCH_PAGES, n_pages)):
                w.add_page(reader.pages[j])
            buf = io.BytesIO()
            w.write(buf)
            batches.append(buf.getvalue())
    except Exception as e:
        logger.warning(f"[vat_paged] 拆页失败 · 回退整文件: {e}")
        return parse_with_gemini(file_bytes, "application/pdf", api_key=api_key)

    logger.info(
        f"[vat_paged] PDF {n_pages} 页 → {len(batches)} 批 × {_BATCH_PAGES}p · 并行 {_BATCH_WORKERS}"
    )

    # 并行
    from concurrent.futures import ThreadPoolExecutor

    rows_all: List[Dict] = []
    meta_all: Dict[str, Any] = {"total_amount_pre_vat": 0.0, "total_vat": 0.0}
    paged_in_tok = 0
    paged_out_tok = 0
    succeeded = 0
    failed_msgs: List[str] = []

    def _one(b: bytes):
        return parse_with_gemini(b, "application/pdf", api_key=api_key)

    with ThreadPoolExecutor(max_workers=min(_BATCH_WORKERS, len(batches))) as ex:
        futures = [ex.submit(_one, b) for b in batches]
        for i, fut in enumerate(futures):
            try:
                r = fut.result()
                if r.get("ok"):
                    succeeded += 1
                    rows_all.extend(r.get("rows", []) or [])
                    m = r.get("meta") or {}
                    if isinstance(m.get("total_amount_pre_vat"), (int, float)):
                        meta_all["total_amount_pre_vat"] += m["total_amount_pre_vat"]
                    if isinstance(m.get("total_vat"), (int, float)):
                        meta_all["total_vat"] += m["total_vat"]
                    paged_in_tok += int(r.get("_input_tokens") or 0)
                    paged_out_tok += int(r.get("_output_tokens") or 0)
                else:
                    failed_msgs.append(f"批{i+1}/{len(batches)}: {(r.get('error') or '?')[:80]}")
            except Exception as e:
                failed_msgs.append(f"批{i+1}/{len(batches)}: {type(e).__name__}: {str(e)[:80]}")

    if succeeded == 0:
        return {
            "ok": False,
            "rows": [],
            "error": f"全部 {len(batches)} 批 OCR 均失败 · {'; '.join(failed_msgs[:3])}",
        }

    # 重编 row_no(每批从 1 开始 · 合并后重新连续编号)
    for i, r in enumerate(rows_all, 1):
        r["row_no"] = i

    warnings = []
    if failed_msgs:
        warnings.append(f"{len(failed_msgs)}/{len(batches)} 批失败 · 部分数据可能缺失")
        for m in failed_msgs[:3]:
            logger.warning(f"[vat_paged] {m}")

    return {
        "ok": True,
        "rows": rows_all,
        "meta": meta_all,
        "warnings": warnings,
        "parser_version": PARSER_VERSION,
        "row_count": len(rows_all),
        "method": f"gemini_paged_{_BATCH_PAGES}p_{len(batches)}b",
        "_input_tokens": paged_in_tok,
        "_output_tokens": paged_out_tok,
    }


# v118.32.4.5 · 大图智能 OCR(预压缩 + 必要时上下分块) · 失败返回 None 让上层回退
_IMG_MAX_LONG_EDGE = 1800  # 长边超过此值 → 等比缩放
_IMG_SPLIT_HEIGHT = 1100  # 高度超此值 → 上下切两块(报告图普遍 1500~2000px · Gemini 单次易 504)
_IMG_OVERLAP_PX = 100  # 切分时上下重叠像素(防止切到表格行中间)


def parse_with_gemini_image_smart(file_bytes: bytes, ext: str, api_key: Optional[str] = None):
    """JPG/PNG 报告 · 大图先缩放 · 高图再切块 · 并行 Gemini · 合并行
    返回 dict(success/fail) 或 None(走老路径)"""
    try:
        from PIL import Image
    except ImportError:
        return None

    try:
        img = Image.open(io.BytesIO(file_bytes))
        if img.mode not in ("RGB", "L"):
            img = img.convert("RGB")
        W, H = img.size
        # 1. 等比缩放
        long_edge = max(W, H)
        if long_edge > _IMG_MAX_LONG_EDGE:
            ratio = _IMG_MAX_LONG_EDGE / long_edge
            new_w, new_h = int(W * ratio), int(H * ratio)
            img = img.resize((new_w, new_h), Image.LANCZOS)
            W, H = new_w, new_h
            logger.info(f"[vat_img] 缩放: {long_edge}→{_IMG_MAX_LONG_EDGE}px")

        # 2. 单图还是切块?
        if H <= _IMG_SPLIT_HEIGHT:
            # 单张直接调
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=85, optimize=True)
            return parse_with_gemini(buf.getvalue(), "image/jpeg", api_key=api_key)

        # 3. 上下切两块 · 中间留重叠区
        mid = H // 2
        top = img.crop((0, 0, W, mid + _IMG_OVERLAP_PX))
        bot = img.crop((0, mid - _IMG_OVERLAP_PX, W, H))
        bufs = []
        for sub in (top, bot):
            b = io.BytesIO()
            sub.save(b, format="JPEG", quality=85, optimize=True)
            bufs.append(b.getvalue())
        logger.info(f"[vat_img] 高图切块: {H}px → 2 块 · 并行")

        # 并行调 Gemini
        from concurrent.futures import ThreadPoolExecutor

        rows_all = []
        meta_all = {"total_amount_pre_vat": 0.0, "total_vat": 0.0}
        img_in_tok = 0
        img_out_tok = 0
        succeeded = 0
        failed_msgs = []

        def _one(b):
            return parse_with_gemini(b, "image/jpeg", api_key=api_key)

        with ThreadPoolExecutor(max_workers=2) as ex:
            futures = [ex.submit(_one, b) for b in bufs]
            for i, fut in enumerate(futures):
                try:
                    r = fut.result()
                    if r.get("ok"):
                        succeeded += 1
                        rows_all.extend(r.get("rows") or [])
                        m = r.get("meta") or {}
                        if isinstance(m.get("total_amount_pre_vat"), (int, float)):
                            meta_all["total_amount_pre_vat"] += m["total_amount_pre_vat"]
                        if isinstance(m.get("total_vat"), (int, float)):
                            meta_all["total_vat"] += m["total_vat"]
                        img_in_tok += int(r.get("_input_tokens") or 0)
                        img_out_tok += int(r.get("_output_tokens") or 0)
                    else:
                        failed_msgs.append(f"块{i+1}: {(r.get('error') or '?')[:80]}")
                except Exception as e:
                    failed_msgs.append(f"块{i+1}: {type(e).__name__}: {str(e)[:80]}")

        if succeeded == 0:
            return {"ok": False, "rows": [], "error": f"两块均失败: {'; '.join(failed_msgs[:2])}"}

        # 重叠区可能重复行 · 按 (date, invoice_no, amount) 去重
        seen = set()
        dedup = []
        for r in rows_all:
            key = (
                r.get("report_date"),
                r.get("report_invoice_no"),
                r.get("report_amount") or r.get("report_amount_pre_vat"),
            )
            if key in seen:
                continue
            seen.add(key)
            dedup.append(r)

        # 重编 row_no
        for i, r in enumerate(dedup, 1):
            r["row_no"] = i

        warnings = []
        if failed_msgs:
            warnings.append(f"{len(failed_msgs)}/2 块失败 · 部分数据可能缺失")

        return {
            "ok": True,
            "rows": dedup,
            "meta": meta_all,
            "warnings": warnings,
            "parser_version": PARSER_VERSION,
            "row_count": len(dedup),
            "method": "gemini_img_split",
            "_input_tokens": img_in_tok,
            "_output_tokens": img_out_tok,
        }
    except Exception as e:
        logger.warning(f"[vat_img] 智能 OCR 异常 · 回退老路径: {type(e).__name__}: {e}")
        return None


def parse_with_gemini(
    file_bytes: bytes, mime_type: str, api_key: Optional[str] = None
) -> Dict[str, Any]:
    try:
        import google.generativeai as genai
    except ImportError:
        return {"ok": False, "error": "google-generativeai 未安装", "rows": []}

    key = api_key or os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not key:
        return {"ok": False, "error": "Gemini API key 未配置", "rows": []}

    text = ""
    try:
        genai.configure(api_key=key)
        model = genai.GenerativeModel(
            "gemini-2.5-flash",
            generation_config={
                "response_mime_type": "application/json",
                "temperature": 0.1,
            },
        )
        # v118.32.4.9.6 · timeout 45→60 + 超时/网络错误单次重试(真实国税局 PDF 504 修)
        response = None
        last_err = None
        for attempt in range(2):
            try:
                response = model.generate_content(
                    [
                        _GEMINI_PROMPT,
                        {"mime_type": mime_type, "data": file_bytes},
                    ],
                    request_options={"timeout": 60},
                )
                break
            except Exception as e:
                last_err = e
                err_name = type(e).__name__
                err_msg = str(e).lower()
                # 仅在超时/网络类错误上重试 · 4xx 业务错直接抛
                if attempt == 0 and (
                    "timeout" in err_msg
                    or "deadline" in err_msg
                    or "503" in err_msg
                    or "504" in err_msg
                    or err_name in ("DeadlineExceeded", "ServiceUnavailable")
                ):
                    logger.warning(f"[vat_gemini] 首次失败({err_name})· 2 秒后重试")
                    import time

                    time.sleep(2)
                    continue
                raise
        if response is None:
            raise last_err or RuntimeError("Gemini 无响应")
        text = (response.text or "").strip()
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
        data = json.loads(text)
        raw_rows = data.get("rows", []) or []

        rows: List[Dict] = []
        for i, r in enumerate(raw_rows, 1):
            parsed = {
                "row_no": int(r.get("row_no") or i),
                "report_date": r.get("report_date") or "",
                "report_invoice_no": str(r.get("report_invoice_no") or "").strip(),
                "report_ref_no": str(
                    r.get("report_ref_no") or ""
                ).strip(),  # v118.32.5 · GL对账匹配键
                "report_buyer_name": str(r.get("report_buyer_name") or "").strip(),
                "report_buyer_tax_id": normalize_tax_id(r.get("report_buyer_tax_id") or ""),
                "report_buyer_branch": normalize_branch(r.get("report_buyer_branch") or ""),
                "report_amount_pre_vat": _to_float(r.get("report_amount_pre_vat")),
                "report_vat_amount": _to_float(r.get("report_vat_amount")),
                "report_amount": _to_float(r.get("report_amount")),
            }
            parsed["is_individual"] = not bool(parsed["report_buyer_tax_id"])
            # v118.32.5.5.2 · 过滤 Gemini 误抓的页眉/分隔行(WNF 报告:列名 / `<----...---->` / `===` 等)
            inv = parsed["report_invoice_no"]
            buyer = parsed["report_buyer_name"]
            amt = parsed["report_amount_pre_vat"] or parsed["report_amount"] or 0
            _is_garbage = (
                not re.search(r"[A-Za-z0-9]{2,}", inv)  # 发票号必须含 ≥2 字母数字
                or "ใบกำกับภาษี" in inv
                or "เลขที่" == inv  # 页眉关键词
                or "ลำดับ" in inv  # 表头"序号"
                or (
                    amt == 0
                    and parsed["report_vat_amount"] == 0
                    and not parsed["report_buyer_tax_id"]
                    and ("---" in inv or "===" in inv or "<--" in inv or "..." in inv)
                )
            )
            if _is_garbage:
                logger.info(f"[vat_gemini] 跳过非数据行 #{i} inv={inv!r} buyer={buyer[:30]!r}")
                continue
            rows.append(parsed)

        _usage = getattr(response, "usage_metadata", None)
        _in_tok = int(getattr(_usage, "prompt_token_count", 0) or 0)
        _out_tok = int(getattr(_usage, "candidates_token_count", 0) or 0)
        return {
            "ok": True,
            "rows": rows,
            "meta": data.get("meta", {}) or {},
            "warnings": [],
            "parser_version": PARSER_VERSION,
            "row_count": len(rows),
            "method": "gemini_ocr",
            "_input_tokens": _in_tok,
            "_output_tokens": _out_tok,
        }
    except json.JSONDecodeError as e:
        logger.error(f"[vat_gemini] JSON 解析失败: {e} · raw: {text[:300]}")
        return {"ok": False, "error": f"Gemini 返回非 JSON: {str(e)[:100]}", "rows": []}
    except Exception as e:
        logger.error(f"[vat_gemini] 失败: {e}")
        return {"ok": False, "error": str(e), "rows": []}


# ======================================================================
# 统一入口 · 按后缀分发 + 自动 fallback
# ======================================================================


def _parse_vat_via_pipeline(
    file_bytes: bytes, filename: str, api_key: Optional[str] = None
) -> Dict[str, Any]:
    """v118.35.0.2 · Route VAT report through services/ocr/pipeline with
    document_type='vat_report'. Used for non-PDF/non-Excel/non-image formats
    (CSV / TSV / DOCX / DOC / TXT / TIFF / BMP / GIF / XLSM).

    Converts pipeline normalized JSON (VatReportDocument) → row dict format
    that downstream reconciliation code expects (report_date / report_invoice_no
    / report_buyer_name / report_buyer_tax_id / report_amount_pre_vat /
    report_vat_amount / report_amount / etc).
    """
    try:
        from services.ocr.pipeline import (
            run_on_image_bytes as _run_image,
            run_on_table_bytes as _run_table,
            IMAGE_EXTENSIONS,
            TABLE_EXTENSIONS,
        )
        from services.ocr.legacy_adapter import pipeline_result_to_legacy_dict
    except ImportError as e:
        return {"ok": False, "rows": [], "row_count": 0, "error": f"pipeline import failed: {e}"}

    ext_dot = "." + (filename or "").lower().rsplit(".", 1)[-1]
    try:
        if ext_dot in IMAGE_EXTENSIONS:
            pr = _run_image(file_bytes, api_key=api_key, document_type="vat_report")
        elif ext_dot in TABLE_EXTENSIONS:
            pr = _run_table(
                file_bytes, filename=filename or "vat", api_key=api_key, document_type="vat_report"
            )
        else:
            return {"ok": False, "rows": [], "row_count": 0, "error": f"pipeline 不支持 {ext_dot}"}
    except Exception as e:
        return {"ok": False, "rows": [], "row_count": 0, "error": _short_err(e)}

    legacy = pipeline_result_to_legacy_dict(pr)
    rows: List[Dict] = []
    row_no = 0
    for page in legacy.get("pages") or []:
        doc = (page or {}).get("document") or {}
        for e in doc.get("entries") or []:
            row_no += 1
            invoice_no = str(e.get("invoice_no") or "").strip()
            if not re.search(r"[A-Za-z0-9]{2,}", invoice_no):
                continue  # skip rows without a real invoice number
            parsed = {
                "row_no": (
                    int(e.get("seq_no") or row_no)
                    if str(e.get("seq_no") or "").isdigit()
                    else row_no
                ),
                "report_date": e.get("transaction_date") or "",
                "report_invoice_no": invoice_no,
                "report_ref_no": invoice_no,
                "report_buyer_name": str(e.get("customer_name") or "").strip(),
                "report_buyer_tax_id": normalize_tax_id(e.get("customer_tax") or ""),
                "report_buyer_branch": normalize_branch(e.get("customer_branch") or ""),
                "report_amount_pre_vat": _to_float(e.get("subtotal")),
                "report_vat_amount": _to_float(e.get("vat")),
                "report_amount": _to_float(e.get("total")),
            }
            parsed["is_individual"] = not bool(parsed["report_buyer_tax_id"])
            rows.append(parsed)

    return {
        "ok": True,
        "rows": rows,
        "row_count": len(rows),
        "meta": {},
        "warnings": [],
        "parser_version": PARSER_VERSION,
        "method": "pipeline_v1",
        "needs_review": legacy.get("_needs_review", False),
    }


def parse_vat_report(
    file_bytes: bytes, filename: str, api_key: Optional[str] = None
) -> Dict[str, Any]:
    ext = (filename or "").lower().rsplit(".", 1)[-1]

    if ext in ("xlsx", "xls"):
        result = parse_excel(file_bytes)
    elif ext == "pdf":
        text_result = parse_pdf_text(file_bytes)
        result = None
        if text_result and text_result.get("rows"):
            # v118.32.5.5.3 · 先过滤再判定 · 过滤后 ≥ 3 行才信
            cleaned = _filter_garbage_rows(text_result["rows"])
            if len(cleaned) >= 3:
                text_result["rows"] = cleaned
                text_result["row_count"] = len(cleaned)
                logger.info(f"[vat] pdf_text 过滤后剩 {len(cleaned)} 行 · 用表抽取结果")
                result = text_result
            else:
                logger.info(f"[vat] pdf_text 表抽取过滤剩 {len(cleaned)} 行 · 试文字行 regex")
        # v118.32.5.5.4 · 表抽取走不通 · 试文字行 regex(不立刻回退 Gemini · 省 40 秒)
        if result is None:
            regex_rows = _parse_vat_pdf_text_lines(file_bytes)
            regex_cleaned = _filter_garbage_rows(regex_rows or [])
            if len(regex_cleaned) >= 3:
                logger.info(f"[vat] pdf_text_regex 抽到 {len(regex_cleaned)} 行 · 跳过 Gemini")
                result = {
                    "ok": True,
                    "rows": regex_cleaned,
                    "meta": {},
                    "warnings": [],
                    "parser_version": PARSER_VERSION,
                    "row_count": len(regex_cleaned),
                    "method": "pdf_text_regex",
                }
            else:
                logger.info(f"[vat] 文字行 regex 也只 {len(regex_cleaned)} 行 · 回退 Gemini")
                result = parse_with_gemini_paged(file_bytes, api_key=api_key)
    elif ext in ("jpg", "jpeg", "png", "webp"):
        # v118.32.4.5 · 大图预压缩 + 必要时上下分块(防 Gemini 504)
        result = parse_with_gemini_image_smart(file_bytes, ext, api_key=api_key)
        if result is None:
            mime = "image/jpeg" if ext == "jpg" else f"image/{ext}"
            result = parse_with_gemini(file_bytes, mime, api_key=api_key)
    elif ext in ("csv", "tsv", "docx", "doc", "txt", "tiff", "tif", "bmp", "gif", "xlsm"):
        # 2026-05-21 v118.35.0.2 · 新增格式走统一 pipeline · document_type=vat_report
        result = _parse_vat_via_pipeline(file_bytes, filename, api_key=api_key)
    else:
        return {
            "ok": False,
            "error": f"暂不支持 .{ext} · 请用 Excel / PDF / 图片 / CSV / Word 等格式",
            "rows": [],
        }

    # v118.32.5.5.3 · 所有路径出口统一过滤一次(Gemini 路径双保险)
    if result.get("ok") and result.get("rows"):
        before = len(result["rows"])
        result["rows"] = _filter_garbage_rows(result["rows"])
        result["row_count"] = len(result["rows"])
        if before != len(result["rows"]):
            logger.info(f"[vat] 最终过滤 {before} → {len(result['rows'])} 行")
    return result
