# -*- coding: utf-8 -*-
"""
Pearnly · 使用明细报告生成器 (PDF + XLSX)
按用户分组 → 日期 → 文件名 → 张数 → 该用户小计 → 总计
四语言表头（zh/en/th/ja）
"""

import io
import os
from typing import Dict, List, Any

_I18N = {
    "th": {
        "title": "รายงานการใช้งาน Pearnly",
        "period": "งวด",
        "company": "บริษัท",
        "date": "วันที่",
        "user": "ผู้ใช้",
        "filename": "ชื่อไฟล์",
        "pages": "จำนวนแผ่น",
        "cost": "ค่าใช้จ่าย",
        "subtotal": "รวม",
        "total": "รวมทั้งหมด",
        "empty": "ไม่มีข้อมูล",
    },
    "zh": {
        "title": "Pearnly 使用明细报告",
        "period": "期间",
        "company": "公司",
        "date": "日期",
        "user": "用户",
        "filename": "文件名",
        "pages": "张数",
        "cost": "费用",
        "subtotal": "小计",
        "total": "总计",
        "empty": "本期无数据",
    },
    "en": {
        "title": "Pearnly Usage Report",
        "period": "Period",
        "company": "Company",
        "date": "Date",
        "user": "User",
        "filename": "File Name",
        "pages": "Pages",
        "cost": "Cost",
        "subtotal": "Subtotal",
        "total": "Total",
        "empty": "No data",
    },
    "ja": {
        "title": "Pearnly 利用レポート",
        "period": "期間",
        "company": "会社",
        "date": "日付",
        "user": "ユーザー",
        "filename": "ファイル名",
        "pages": "枚数",
        "cost": "費用",
        "subtotal": "小計",
        "total": "合計",
        "empty": "データなし",
    },
}


def _t(lang: str, key: str) -> str:
    return _I18N.get(lang, _I18N["en"]).get(key, key)


def _group_rows(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """按 user_id 分组 · 内部保持日期顺序"""
    groups: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for r in rows:
        uid = str(r.get("user_id") or r.get("user_email") or "unknown")
        if uid not in groups:
            groups[uid] = {
                "user_label": r.get("user_name")
                or (r.get("user_email") or "").split("@")[0]
                or "—",
                "user_email": r.get("user_email") or "",
                "rows": [],
                "pages_sum": 0,
                "cost_sum": 0.0,
            }
            order.append(uid)
        groups[uid]["rows"].append(r)
        groups[uid]["pages_sum"] += int(r.get("pages") or 0)
        groups[uid]["cost_sum"] += float(r.get("cost_thb") or 0)
    return [groups[k] for k in order]


_FONTS_REGISTERED = False
_BASE_FONT = "Helvetica"
_BOLD_FONT = "Helvetica-Bold"
_HAS_CJK = False
_HAS_THAI = False


def _try_register(name: str, path: str, subfont_index=None) -> bool:
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    if not os.path.exists(path):
        return False
    try:
        if subfont_index is not None:
            pdfmetrics.registerFont(TTFont(name, path, subfontIndex=subfont_index))
        else:
            pdfmetrics.registerFont(TTFont(name, path))
        return True
    except Exception:
        return False


def _register_fonts():
    """
    Register fonts. Reportlab can NOT consume PostScript-outline (CFF) TTC/OTF
    such as NotoSansCJK; we use TrueType-outline TTC/TTF instead:
      - WQY Zenhei (TTC, TrueType) — CJK
      - Noto Sans Thai (TTF) — Thai (covers ฿ U+0E3F)
      - Helvetica — Latin (built-in)
    """
    global _FONTS_REGISTERED, _BASE_FONT, _BOLD_FONT, _HAS_CJK, _HAS_THAI
    if _FONTS_REGISTERED:
        return _BASE_FONT, _BOLD_FONT

    # CJK
    cjk_paths = [
        "/usr/share/fonts/truetype/wqy/wqy-zenhei.ttc",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "/usr/share/fonts/truetype/arphic/uming.ttc",
    ]
    for p in cjk_paths:
        if _try_register("PR-CJK", p, subfont_index=0):
            _HAS_CJK = True
            break

    # Thai
    thai_paths = [
        "/usr/share/fonts/truetype/noto/NotoSansThai-Regular.ttf",
        "/usr/share/fonts/truetype/tlwg/Norasi.ttf",
        "/usr/share/fonts/truetype/tlwg/Garuda.ttf",
    ]
    thai_bold_paths = [
        "/usr/share/fonts/truetype/noto/NotoSansThai-Bold.ttf",
        "/usr/share/fonts/truetype/tlwg/Norasi-Bold.ttf",
        "/usr/share/fonts/truetype/tlwg/Garuda-Bold.ttf",
    ]
    for p in thai_paths:
        if _try_register("PR-Thai", p):
            _HAS_THAI = True
            break
    for p in thai_bold_paths:
        if _try_register("PR-ThaiBold", p):
            break

    # Use CJK as base since it also covers Latin (WQY Zenhei has full ASCII).
    if _HAS_CJK:
        _BASE_FONT = "PR-CJK"
        _BOLD_FONT = "PR-CJK"  # WQY Zenhei TTC doesn't have a separate bold
    elif _HAS_THAI:
        _BASE_FONT = "PR-Thai"
        _BOLD_FONT = "PR-ThaiBold" if os.path.exists(thai_bold_paths[0]) else "PR-Thai"

    _FONTS_REGISTERED = True
    return _BASE_FONT, _BOLD_FONT


# --- character ranges for script-aware font switching ---
def _is_thai(cp: int) -> bool:
    return 0x0E00 <= cp <= 0x0E7F


def _is_cjk(cp: int) -> bool:
    # CJK Unified, Hiragana, Katakana, Halfwidth/Fullwidth, CJK Symbols
    return (0x3000 <= cp <= 0x9FFF) or (0xFF00 <= cp <= 0xFFEF) or (0x20000 <= cp <= 0x2FFFF)


def _script_runs(s: str):
    """Yield (script, text) runs where script in {'cjk','thai','latin'}."""
    if not s:
        return
    cur_script = None
    cur_text = []
    for ch in s:
        cp = ord(ch)
        if _is_thai(cp):
            sc = "thai"
        elif _is_cjk(cp):
            sc = "cjk"
        else:
            sc = "latin"
        if sc != cur_script and cur_text:
            yield (cur_script, "".join(cur_text))
            cur_text = []
        cur_script = sc
        cur_text.append(ch)
    if cur_text:
        yield (cur_script, "".join(cur_text))


def _xml_escape(s: str) -> str:
    return (s or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


def _build_paragraph_text(s: str, bold: bool = False) -> str:
    """Return reportlab Paragraph markup with per-script <font> tags."""
    if not s:
        return ""
    cjk_font = (
        "PR-CJK"
        if _HAS_CJK
        else ("PR-Thai" if _HAS_THAI else ("Helvetica-Bold" if bold else "Helvetica"))
    )
    thai_font = "PR-ThaiBold" if (bold and _HAS_THAI) else ("PR-Thai" if _HAS_THAI else cjk_font)
    latin_font = "Helvetica-Bold" if bold else "Helvetica"
    parts = []
    for sc, txt in _script_runs(s):
        if sc == "thai":
            fn = thai_font
        elif sc == "cjk":
            fn = cjk_font
        else:
            fn = latin_font
        parts.append(f'<font name="{fn}">{_xml_escape(txt)}</font>')
    return "".join(parts)


def build_pdf(
    *,
    lang: str,
    company: str,
    start_date: str,
    end_date: str,
    rows: List[Dict[str, Any]],
) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        SimpleDocTemplate,
        Paragraph,
        Spacer,
        Table,
        TableStyle,
        KeepTogether,
    )
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

    base_font, bold_font = _register_fonts()
    if lang not in _I18N:
        lang = "en"

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=15 * mm,
        rightMargin=15 * mm,
        topMargin=15 * mm,
        bottomMargin=15 * mm,
        title=_t(lang, "title"),
    )

    # All paragraph text is built via _build_paragraph_text() which embeds
    # per-script <font> tags; the outer ParagraphStyle font is only the
    # fallback for chars no run matched (rare).
    title_style = ParagraphStyle(
        "title", fontName=bold_font, fontSize=16, alignment=TA_CENTER, leading=22, spaceAfter=2
    )
    sub_style = ParagraphStyle(
        "sub",
        fontName=base_font,
        fontSize=10,
        alignment=TA_CENTER,
        leading=14,
        textColor=colors.HexColor("#475569"),
    )
    user_style = ParagraphStyle(
        "user",
        fontName=bold_font,
        fontSize=11.5,
        leading=15,
        spaceBefore=10,
        spaceAfter=3,
        textColor=colors.HexColor("#0f172a"),
    )
    cell_left = ParagraphStyle(
        "cell-left", fontName=base_font, fontSize=9, leading=12, alignment=TA_LEFT
    )
    cell_right = ParagraphStyle(
        "cell-right", fontName=base_font, fontSize=9, leading=12, alignment=TA_RIGHT
    )
    cell_header = ParagraphStyle(
        "cell-hd",
        fontName=bold_font,
        fontSize=9.5,
        leading=12,
        alignment=TA_LEFT,
        textColor=colors.HexColor("#1f2937"),
    )
    cell_header_r = ParagraphStyle(
        "cell-hd-r",
        fontName=bold_font,
        fontSize=9.5,
        leading=12,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#1f2937"),
    )
    cell_subtotal = ParagraphStyle(
        "cell-st",
        fontName=bold_font,
        fontSize=9.5,
        leading=12,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#9a3412"),
    )
    cell_total = ParagraphStyle(
        "cell-tot",
        fontName=bold_font,
        fontSize=11,
        leading=14,
        alignment=TA_RIGHT,
        textColor=colors.HexColor("#7c2d12"),
    )
    empty_style = ParagraphStyle(
        "empty",
        fontName=base_font,
        fontSize=11,
        alignment=TA_CENTER,
        leading=14,
        textColor=colors.HexColor("#94a3b8"),
    )

    def _P(text, style, bold=False):
        return Paragraph(_build_paragraph_text(text, bold=bold), style)

    story = []
    story.append(_P(_t(lang, "title"), title_style, bold=True))
    story.append(_P(f"{_t(lang, 'company')}: {company}", sub_style))
    story.append(_P(f"{_t(lang, 'period')}: {start_date} ~ {end_date}", sub_style))
    story.append(Spacer(1, 8))

    groups = _group_rows(rows)
    grand_pages = 0
    grand_cost = 0.0

    headers_text = [_t(lang, "date"), _t(lang, "filename"), _t(lang, "pages"), _t(lang, "cost")]
    col_widths = [30 * mm, 88 * mm, 22 * mm, 30 * mm]

    if not groups:
        story.append(Spacer(1, 20))
        story.append(_P(_t(lang, "empty"), empty_style))
    else:
        for g in groups:
            label = g["user_label"]
            if g["user_email"] and g["user_email"] != label:
                label = f"{label}  ·  {g['user_email']}"
            block = [_P(label, user_style, bold=True)]

            header_row = [
                _P(headers_text[0], cell_header, bold=True),
                _P(headers_text[1], cell_header, bold=True),
                _P(headers_text[2], cell_header_r, bold=True),
                _P(headers_text[3], cell_header_r, bold=True),
            ]
            data = [header_row]
            for r in g["rows"]:
                date_str = (r.get("date") or "")[:10]
                fname = r.get("filename") or "—"
                if len(fname) > 70:
                    fname = fname[:67] + "…"
                data.append(
                    [
                        _P(date_str, cell_left),
                        _P(fname, cell_left),
                        _P(str(int(r.get("pages") or 0)), cell_right),
                        _P(f"฿{float(r.get('cost_thb') or 0):.2f}", cell_right),
                    ]
                )
            data.append(
                [
                    _P("", cell_left),
                    _P(_t(lang, "subtotal"), cell_subtotal, bold=True),
                    _P(str(g["pages_sum"]), cell_subtotal, bold=True),
                    _P(f"฿{g['cost_sum']:.2f}", cell_subtotal, bold=True),
                ]
            )

            tbl = Table(data, colWidths=col_widths, repeatRows=1)
            tbl.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#f3f4f6")),
                        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#fff7ed")),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e5e7eb")),
                        ("LEFTPADDING", (0, 0), (-1, -1), 4),
                        ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                        ("TOPPADDING", (0, 0), (-1, -1), 4),
                        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                    ]
                )
            )
            block.append(tbl)
            story.append(KeepTogether(block))
            grand_pages += g["pages_sum"]
            grand_cost += g["cost_sum"]

    story.append(Spacer(1, 14))
    total_tbl = Table(
        [
            [
                _P("", cell_total),
                _P(_t(lang, "total"), cell_total, bold=True),
                _P(str(grand_pages), cell_total, bold=True),
                _P(f"฿{grand_cost:.2f}", cell_total, bold=True),
            ]
        ],
        colWidths=col_widths,
    )
    total_tbl.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#fed7aa")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("BOX", (0, 0), (-1, -1), 0.8, colors.HexColor("#fb923c")),
                ("LEFTPADDING", (0, 0), (-1, -1), 4),
                ("RIGHTPADDING", (0, 0), (-1, -1), 4),
                ("TOPPADDING", (0, 0), (-1, -1), 7),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 7),
            ]
        )
    )
    story.append(total_tbl)

    doc.build(story)
    return buf.getvalue()


def build_xlsx(
    *,
    lang: str,
    company: str,
    start_date: str,
    end_date: str,
    rows: List[Dict[str, Any]],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    if lang not in _I18N:
        lang = "en"

    wb = Workbook()
    ws = wb.active
    ws.title = "Usage"

    thin = Side(style="thin", color="E5E7EB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="F3F4F6")
    subtotal_fill = PatternFill("solid", fgColor="FFF7ED")
    total_fill = PatternFill("solid", fgColor="FED7AA")
    title_font = Font(bold=True, size=14)
    meta_font = Font(size=10, color="475569")
    user_font = Font(bold=True, size=11, color="0F172A")
    bold_font = Font(bold=True, size=10)
    subtotal_font = Font(bold=True, size=10, color="9A3412")
    total_font = Font(bold=True, size=11, color="7C2D12")
    right = Alignment(horizontal="right", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")

    row = 1
    c = ws.cell(row=row, column=1, value=_t(lang, "title"))
    c.font = title_font
    c.alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    c = ws.cell(row=row, column=1, value=f"{_t(lang, 'company')}: {company}")
    c.font = meta_font
    c.alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 1
    c = ws.cell(row=row, column=1, value=f"{_t(lang, 'period')}: {start_date} ~ {end_date}")
    c.font = meta_font
    c.alignment = center
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    row += 2

    headers = [_t(lang, "date"), _t(lang, "filename"), _t(lang, "pages"), _t(lang, "cost")]
    groups = _group_rows(rows)
    grand_pages = 0
    grand_cost = 0.0

    if not groups:
        c = ws.cell(row=row, column=1, value=_t(lang, "empty"))
        c.alignment = center
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        row += 2
    else:
        for g in groups:
            label = g["user_label"]
            if g["user_email"] and g["user_email"] != label:
                label = f"{label} · {g['user_email']}"
            c = ws.cell(row=row, column=1, value=label)
            c.font = user_font
            c.alignment = left
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
            row += 1

            for col_idx, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col_idx, value=h)
                c.font = bold_font
                c.fill = header_fill
                c.border = border
                c.alignment = center if col_idx >= 3 else left
            row += 1

            for r in g["rows"]:
                date_str = (r.get("date") or "")[:10]
                c = ws.cell(row=row, column=1, value=date_str)
                c.border = border
                c.alignment = left
                c = ws.cell(row=row, column=2, value=r.get("filename") or "—")
                c.border = border
                c.alignment = left
                c = ws.cell(row=row, column=3, value=int(r.get("pages") or 0))
                c.border = border
                c.alignment = right
                c = ws.cell(row=row, column=4, value=float(r.get("cost_thb") or 0))
                c.number_format = '"฿"#,##0.00'
                c.border = border
                c.alignment = right
                row += 1

            for col_idx in range(1, 5):
                c = ws.cell(row=row, column=col_idx)
                c.fill = subtotal_fill
                c.font = subtotal_font
                c.border = border
            ws.cell(row=row, column=2, value=_t(lang, "subtotal")).alignment = right
            ws.cell(row=row, column=3, value=g["pages_sum"]).alignment = right
            c = ws.cell(row=row, column=4, value=g["cost_sum"])
            c.number_format = '"฿"#,##0.00'
            c.alignment = right
            row += 2

            grand_pages += g["pages_sum"]
            grand_cost += g["cost_sum"]

    for col_idx in range(1, 5):
        c = ws.cell(row=row, column=col_idx)
        c.fill = total_fill
        c.font = total_font
        c.border = border
    ws.cell(row=row, column=2, value=_t(lang, "total")).alignment = right
    ws.cell(row=row, column=3, value=grand_pages).alignment = right
    c = ws.cell(row=row, column=4, value=grand_cost)
    c.number_format = '"฿"#,##0.00'
    c.alignment = right

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
