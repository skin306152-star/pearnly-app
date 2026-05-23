# -*- coding: utf-8 -*-
"""
v118.35.0.61 · 守门测试 · 勾稽匹配诚实化(防 diff=0 恒等式假象误导用户)

真实案例(skin 实测):190 SCB 0 笔匹配、185 AMK 44 行只匹配 1 笔,导出都显示
『差异 0』并染绿『✓』· 用户误以为对平。实际几乎一笔都没对上。

锁定契约:
  1. 0 笔匹配 → 顶部红字横幅(banner_no_match)· 即便 diff≈0 也不染绿
  2. 极低匹配率(<10% 且≥10项)→ 同样不染绿 · 出 low_match 横幅
  3. 正常匹配(diff≈0 且匹配率高)→ 仍染绿 ✓(不误伤正常对账)
  4. 始终输出『已匹配笔数 / 匹配率』指标行
"""

import io
import unittest

import openpyxl

from bank_recon_v2 import BankReconSummary, BankReconRow, export_bank_recon_excel

DIFF_OK_BG = "D1FAE5"  # mint green (对平)
DIFF_BAD_BG = "FEE2E2"  # soft red (未对平/可疑)


def _summary(diff=0.0):
    return BankReconSummary(
        bank_code="scb",
        gl_account_code="1112-01",
        stmt_opening=0.0,
        gl_opening=0.0,
        stmt_closing=1000.0,
        gl_closing=1000.0,
        opening_diff=0.0,
        formula_stmt_closing=1000.0,
        formula_diff=diff,
    )


def _stmt_only_row(i):
    return BankReconRow(
        match_status="stmt_deposit_only",
        match_layer=None,
        stmt_date=None,
        stmt_desc=f"dep{i}",
        stmt_withdrawal=0.0,
        stmt_deposit=100.0,
    )


def _matched_row(i):
    return BankReconRow(
        match_status="matched",
        match_layer=1,
        stmt_date=None,
        stmt_desc=f"m{i}",
        stmt_withdrawal=0.0,
        stmt_deposit=100.0,
        gl_doc_no=f"V{i}",
        gl_desc="gl",
        gl_debit=100.0,
        gl_credit=0.0,
    )


def _texts(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                out.append(v)
    return " ".join(out)


def _diff_cell_fill(ws):
    """找到『差异(应为0)』锚点行 B 列的底色 hex(去 alpha 前缀)。
    注:横幅文案里也含『差异』二字 · 故只认 B 列为数值的锚点行(横幅是合并单元格 · B 为 None)。"""
    for row in ws.iter_rows():
        a = row[0].value
        if (
            isinstance(a, str)
            and a.strip().startswith("差异（应为0")
            and isinstance(row[1].value, (int, float))
        ):
            fill = row[1].fill.fgColor.rgb
            return (fill or "")[-6:].upper()
    return None


class MatchRateHonestyTests(unittest.TestCase):

    def test_zero_match_diff_zero_not_green_and_has_banner(self):
        """0 笔匹配 + diff=0 → 不染绿 + 有 no_match 横幅"""
        rows = [_stmt_only_row(i) for i in range(20)]
        b = export_bank_recon_excel(rows, _summary(0.0), lang="zh")
        ws = openpyxl.load_workbook(io.BytesIO(b)).worksheets[0]
        self.assertEqual(_diff_cell_fill(ws), DIFF_BAD_BG, "0匹配时 diff 不应染绿")
        self.assertIn("0 笔成功匹配", _texts(ws))

    def test_low_match_rate_not_green(self):
        """1/20 匹配(5%) + diff=0 → 不染绿 + low_match 横幅"""
        rows = [_matched_row(0)] + [_stmt_only_row(i) for i in range(19)]
        b = export_bank_recon_excel(rows, _summary(0.0), lang="zh")
        ws = openpyxl.load_workbook(io.BytesIO(b)).worksheets[0]
        self.assertEqual(_diff_cell_fill(ws), DIFF_BAD_BG, "极低匹配率时 diff 不应染绿")
        self.assertIn("匹配率", _texts(ws))

    def test_high_match_diff_zero_stays_green(self):
        """18/20 匹配 + diff=0 → 仍染绿(不误伤正常对账)"""
        rows = [_matched_row(i) for i in range(18)] + [_stmt_only_row(i) for i in range(2)]
        b = export_bank_recon_excel(rows, _summary(0.0), lang="zh")
        ws = openpyxl.load_workbook(io.BytesIO(b)).worksheets[0]
        self.assertEqual(_diff_cell_fill(ws), DIFF_OK_BG, "高匹配率且diff=0应染绿")

    def test_match_rate_row_always_present(self):
        rows = [_matched_row(i) for i in range(5)]
        b = export_bank_recon_excel(rows, _summary(0.0), lang="zh")
        ws = openpyxl.load_workbook(io.BytesIO(b)).worksheets[0]
        joined = _texts(ws)
        self.assertIn("已匹配笔数", joined)
        self.assertIn("匹配率", joined)


if __name__ == "__main__":
    unittest.main()
