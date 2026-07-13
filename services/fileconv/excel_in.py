# -*- coding: utf-8 -*-
"""K2 · Excel/CSV → ConvertResult(会计底稿 → PDF 规范输出的解析入口)。

GL 路复用现成结构识别(services.recon.bank_gl_excel.parse_gl_excel,T4a 已验过),但进
GL 路的门收得很紧(R1 打回收口):仅限【单科目 + 表内真有印刷余额列】的件。原因两条——
① bank_gl_excel 的行余额是 attach_running_balance 从借贷衍生的单条全局链,多科目 GL
(如 MR.ERP 735 行导出)每科目余额各自成链,渲染衍生链 = 悄悄改了用户原表的余额数;
② closing 非印刷值时(表内无余额列),衍生链 vs 衍生 closing 的校验永远自洽,✓ 戳
是自导自演。不满足门槛的一律落 generic 路——忠实转录含原表自己的余额列,零改数,
诚实标「未做数字校验」(同 ocr_bridge._convert_generic 口径)。

GL 路的 ✓ 戳靠真锚立住:衍生链末行余额 vs 表内印刷 closing(bank_gl_excel 读自余额列
最后一笔),对不上进 ISSUE_CLOSING_ANCHOR——同 K1c ocr_bridge._anchor_issues 的独立锚。

老式 .xls:GL 路本身够用(parse_gl_excel 内部已有 openpyxl→xlrd→CSV 嗅探的级联);仅当
GL 认不出、需要落 generic 网格时才单独走 xlrd——环境缺 xlrd 时诚实拒绝,不静默转空表。
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any, Dict, List, Optional

from services.fileconv.ledger import LEDGER_COLUMNS, to_table_rows
from services.fileconv.model import (
    GENERIC_TABLE,
    GL_LEDGER,
    ISSUE_CLOSING_ANCHOR,
    STATUS_OK,
    STATUS_UNSUPPORTED_FORMAT,
    ConvertResult,
    Issue,
    LedgerRow,
    Table,
)
from services.fileconv import validate as validate_mod

_TOL = Decimal("0.01")

_XLSX_EXTS = ("xlsx", "xlsm")
_XLS_EXTS = ("xls",)
_CSV_EXTS = ("csv", "tsv", "txt")
_ALL_EXTS = _XLSX_EXTS + _XLS_EXTS + _CSV_EXTS


def _ext(source_name: str) -> str:
    return (source_name or "").rsplit(".", 1)[-1].lower()


def _reject(source_name: str, reason: str) -> ConvertResult:
    return ConvertResult(
        doc_type="",
        status=STATUS_UNSUPPORTED_FORMAT,
        source_name=source_name,
        stats={"reason": reason},
    )


def _gl_row(row, line_no: int) -> LedgerRow:
    """GlRow(bank_gl_excel 产出,float 展示契约)→ LedgerRow(Decimal)。经 str() 中转,
    禁 float 直转(两位小数金额往返无损,同 gl_upload_adapter 边界口径)。日期只保留
    parse_gl_excel 已归一的公历值——原始文本(含泰历原样)在该解析器内未回传,忠实转录
    只能到归一层,非本模块能力边界。

    account 统一留空(单链 key):GL 路门槛已收紧到单科目件(见 _try_gl),单链即该科目
    自身的连续链;真实科目码仍写回渲染表(_try_gl 的 zip 回填),不丢显示信息。"""
    iso = row.date.isoformat() if row.date else ""
    return LedgerRow(
        line_no=line_no,
        account="",
        date=iso,
        date_ce=iso,
        doc_no=row.doc_no or "",
        description=row.description or "",
        balance=Decimal(str(row.balance)),
        debit=Decimal(str(row.debit)),
        credit=Decimal(str(row.credit)),
    )


def _closing_anchor_issue(rows: List[LedgerRow], printed_closing: Decimal) -> List[Issue]:
    """真锚对照(口径同 ocr_bridge._anchor_issues):表内印刷期末余额 vs 衍生链末行余额。
    衍生链从借贷重算、自身永远自洽——✓ 戳的公信力全靠这条独立锚,对不上必须点名。"""
    if not rows or abs(printed_closing - rows[-1].balance) <= _TOL:
        return []
    return [
        Issue(
            kind=ISSUE_CLOSING_ANCHOR,
            line_no=rows[-1].line_no,
            message="印刷期末余额 ≠ 重算末行余额(疑借贷有误/漏行)",
            expected=f"{printed_closing}",
            actual=f"{rows[-1].balance}",
        )
    ]


def _try_gl(data: bytes, source_name: str) -> Optional[ConvertResult]:
    """走现成 GL 结构识别;不满足 GL 路门槛(见模块顶注)返回 None 落 generic 路。

    门槛(缺一即落 generic):
    · 单科目(accounts ≤ 1)——多科目件的衍生单链余额 ≠ 原表每科目各自成链的印刷余额,
      渲染出来就是改用户的数;
    · closing 读自表内印刷余额列(closing_printed)——否则 closing 也是从借贷衍生的,
      拿它当锚校验自己 = 自导自演,✓ 戳没有公信力。
    """
    from services.recon.bank_gl_excel import parse_gl_excel

    try:
        parsed = parse_gl_excel(data, source_name)
    except Exception:  # noqa: BLE001 · 解析器内部炸(损坏文件等)· 不许 500,落 generic
        return None
    gl_rows = parsed.get("rows") or []
    if not parsed.get("ok") or not gl_rows:
        return None
    if len(parsed.get("accounts") or []) > 1 or not parsed.get("closing_printed"):
        return None

    rows = [_gl_row(r, i) for i, r in enumerate(gl_rows, start=1)]
    opening = {"": Decimal(str(parsed.get("opening") or 0))}
    issues = validate_mod.validate_ledger(rows, opening)
    issues.extend(_closing_anchor_issue(rows, Decimal(str(parsed.get("closing") or 0))))
    stats = validate_mod.ledger_stats(rows, opening)
    stats["engine"] = "excel_gl"
    stats["accounts"] = sorted({r.account_code for r in gl_rows if r.account_code})

    table_rows = to_table_rows(rows)
    for cells, src in zip(table_rows, gl_rows):
        cells[0] = src.account_code or ""  # 换回真实科目码(_gl_row 注:链 key 统一单链)
    table = Table(name="GL Ledger", columns=LEDGER_COLUMNS, rows=table_rows)
    return ConvertResult(
        doc_type=GL_LEDGER,
        status=STATUS_OK,
        source_name=source_name,
        tables=[table],
        issues=issues,
        stats=stats,
    )


def _merged_fill_openpyxl(ws) -> Dict[tuple, Any]:
    """合并格 → {(row,col): 左上值}(1-based)· 渲染非左上格时补这个值,不追求像素级复刻。"""
    fills: Dict[tuple, Any] = {}
    for rng in ws.merged_cells.ranges:
        top = ws.cell(row=rng.min_row, column=rng.min_col).value
        for r in range(rng.min_row, rng.max_row + 1):
            for c in range(rng.min_col, rng.max_col + 1):
                fills[(r, c)] = top
    return fills


def _grid_table(name: str, rows: List[List[Any]]) -> Optional[Table]:
    """裸行列表 → Table(空行已剔除)。列头用占位符(col1..colN)——generic 路不猜表头,
    忠实网格渲染,不假装能识别列语义。"""
    kept = [r for r in rows if any(c not in (None, "") for c in r)]
    if not kept:
        return None
    max_col = max(len(r) for r in kept)
    columns = [f"col{i}" for i in range(1, max_col + 1)]
    return Table(name=name, columns=columns, rows=kept)


def _grid_from_openpyxl(data: bytes, source_name: str) -> Optional[ConvertResult]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception:  # noqa: BLE001 · 损坏 xlsx · 交上层统一诚实拒绝
        return None

    tables: List[Table] = []
    for ws in wb.worksheets:
        fills = _merged_fill_openpyxl(ws)
        rows: List[List[Any]] = []
        for r_idx, row in enumerate(ws.iter_rows(), start=1):
            cells = []
            for c_idx, cell in enumerate(row, start=1):
                val = cell.value
                if val is None:
                    val = fills.get((r_idx, c_idx))
                cells.append("" if val is None else val)
            rows.append(cells)
        table = _grid_table(ws.title or f"Sheet{len(tables) + 1}", rows)
        if table is not None:  # 空 sheet 跳过
            tables.append(table)

    if not tables:
        return None
    return ConvertResult(
        doc_type=GENERIC_TABLE,
        status=STATUS_OK,
        source_name=source_name,
        tables=tables,
        issues=[],
        stats={"engine": "excel_grid", "sheet_count": len(tables)},
    )


def _grid_from_xls(data: bytes, source_name: str) -> Optional[ConvertResult]:
    """老式二进制 .xls 网格路(openpyxl 读不了,xlrd 单独接)。"""
    try:
        import xlrd
    except ImportError:
        return _reject(source_name, "xls_requires_resave")  # 诚实拒绝:请另存为 .xlsx

    try:
        wb = xlrd.open_workbook(file_contents=data)
    except Exception:  # noqa: BLE001 · 损坏文件 · 交上层统一诚实拒绝
        return None

    tables: List[Table] = []
    for sheet in wb.sheets():
        fills: Dict[tuple, Any] = {}
        for rlo, rhi, clo, chi in sheet.merged_cells:
            top = sheet.cell_value(rlo, clo)
            for r in range(rlo, rhi):
                for c in range(clo, chi):
                    fills[(r, c)] = top
        rows: List[List[Any]] = []
        for r in range(sheet.nrows):
            cells = []
            for c in range(sheet.ncols):
                val = sheet.cell_value(r, c)
                if val == "" and (r, c) in fills:
                    val = fills[(r, c)]
                cells.append(val)
            rows.append(cells)
        table = _grid_table(sheet.name or f"Sheet{len(tables) + 1}", rows)
        if table is not None:
            tables.append(table)

    if not tables:
        return None
    return ConvertResult(
        doc_type=GENERIC_TABLE,
        status=STATUS_OK,
        source_name=source_name,
        tables=tables,
        issues=[],
        stats={"engine": "excel_grid_xls", "sheet_count": len(tables)},
    )


def _grid_from_csv(data: bytes, source_name: str) -> Optional[ConvertResult]:
    from services.recon.bank_table_io import _load_csv_sheets

    sheets = _load_csv_sheets(data)
    if not sheets:
        return None
    _, rows_raw = sheets[0]
    table = _grid_table("CSV", [[("" if c is None else c) for c in row] for row in rows_raw])
    if table is None:
        return None
    return ConvertResult(
        doc_type=GENERIC_TABLE,
        status=STATUS_OK,
        source_name=source_name,
        tables=[table],
        issues=[],
        stats={"engine": "excel_grid_csv", "sheet_count": 1},
    )


def _try_generic(data: bytes, source_name: str, ext: str) -> Optional[ConvertResult]:
    if ext in _XLSX_EXTS:
        return _grid_from_openpyxl(data, source_name)
    if ext in _XLS_EXTS:
        return _grid_from_xls(data, source_name)
    return _grid_from_csv(data, source_name)


def convert_excel(data: bytes, source_name: str = "") -> ConvertResult:
    """入口:Excel/CSV 底稿 → ConvertResult。GL 结构识别优先(带守恒校验背书);
    认不出结构落 generic 网格路(忠实渲染,诚实声明未做数字校验);两路都拿不到数据
    → 结构化拒绝,不 500、不出半截表。"""
    if not data:
        return _reject(source_name, "empty_file")

    ext = _ext(source_name)
    if ext not in _ALL_EXTS:
        return _reject(source_name, f"unsupported_ext:{ext or '(none)'}")

    gl = _try_gl(data, source_name)
    if gl is not None:
        return gl

    generic = _try_generic(data, source_name, ext)
    if generic is not None:
        return generic

    return _reject(source_name, "unreadable_or_empty")
