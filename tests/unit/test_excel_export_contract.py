# -*- coding: utf-8 -*-
"""REFACTOR-D2 守门 · 识别结果 Excel 导出 excel_export.py 行为契约

excel_export.py(19 列标准明细 · 4 语列头)此前 0 专属测试。
纯 openpyxl 生成 · 无 DB/网络/凭证(Wave 0 安全网 · 零冲突)。

锁定:19 列 4 语标签完整 · merged_fields→行映射 · 商品明细拼串 ·
列头语言切换 + 未知语言兜底 zh · 数字列格式 · 冻结首行。
"""

import io
import unittest

from openpyxl import load_workbook

import excel_export as ee


class ColumnLabelTests(unittest.TestCase):
    def test_19_columns(self):
        self.assertEqual(len(ee.STANDARD_COLUMNS), 19)

    def test_all_langs_cover_all_keys(self):
        keys = {c[0] for c in ee.STANDARD_COLUMNS}
        for lang in ("zh", "en", "th", "ja"):
            self.assertIn(lang, ee.COLUMN_LABEL_BY_LANG)
            self.assertEqual(set(ee.COLUMN_LABEL_BY_LANG[lang]), keys, f"{lang} 列标签不全")
            for k, v in ee.COLUMN_LABEL_BY_LANG[lang].items():
                self.assertTrue(v, f"{lang}.{k} 空标签")


class FmtItemsTests(unittest.TestCase):
    def test_formats_lines(self):
        out = ee._fmt_items([{"name": "Coffee", "qty": 2, "price": 500, "subtotal": 1000}])
        self.assertEqual(out, "1. Coffee ×2 @500 =1000")

    def test_empty(self):
        self.assertEqual(ee._fmt_items([]), "")

    def test_multiline(self):
        out = ee._fmt_items([{"name": "A"}, {"name": "B"}])
        self.assertEqual(out.split("\n"), ["1. A", "2. B"])


class RowFromRecordTests(unittest.TestCase):
    def test_maps_merged_fields(self):
        rec = {
            "filename": "a.pdf",
            "merged_fields": {
                "invoice_number": "INV1",
                "date": "2026-03-15",
                "total_amount": "1070",
                "items": [{"name": "X"}, {"name": "Y"}],
            },
        }
        row = ee._row_from_record(rec, 3)
        self.assertEqual(row["no"], 3)
        self.assertEqual(row["filename"], "a.pdf")
        self.assertEqual(row["invoice_no"], "INV1")
        self.assertEqual(row["total"], "1070")
        self.assertEqual(row["item_count"], 2)

    def test_fallbacks(self):
        # invoice_no / total 的备用键
        row = ee._row_from_record({"merged_fields": {"invoice_no": "B2", "total": "9"}}, 1)
        self.assertEqual(row["invoice_no"], "B2")
        self.assertEqual(row["total"], "9")


def _load(records, lang="zh"):
    raw = ee.build_xlsx(records, lang=lang)
    assert isinstance(raw, bytes)
    return load_workbook(io.BytesIO(raw)).active


class BuildXlsxTests(unittest.TestCase):
    def _rec(self):
        return [
            {
                "filename": "a.pdf",
                "merged_fields": {
                    "invoice_number": "INV1",
                    "subtotal": "1000",
                    "total_amount": "1070",
                    "items": [{"name": "X", "qty": 2, "price": 500, "subtotal": 1000}],
                },
            }
        ]

    def test_header_english(self):
        ws = _load(self._rec(), lang="en")
        self.assertEqual(ws.title, "Results")
        self.assertEqual(ws.cell(row=1, column=1).value, "No.")
        self.assertEqual(ws.cell(row=1, column=3).value, "Invoice No.")
        self.assertEqual(ws.cell(row=1, column=16).value, "Total")

    def test_unknown_lang_falls_back_to_zh(self):
        ws = _load([], lang="fr")
        self.assertEqual(ws.title, "识别结果")
        self.assertEqual(ws.cell(row=1, column=1).value, "序号")

    def test_data_row_and_items(self):
        ws = _load(self._rec(), lang="en")
        self.assertEqual(ws.cell(row=2, column=1).value, 1)  # no
        self.assertEqual(ws.cell(row=2, column=2).value, "a.pdf")
        self.assertEqual(ws.cell(row=2, column=3).value, "INV1")
        self.assertEqual(ws.cell(row=2, column=16).value, "1070")
        self.assertEqual(ws.cell(row=2, column=17).value, 1)  # item_count
        self.assertIn("X", ws.cell(row=2, column=18).value)  # items

    def test_numeric_columns_get_number_format(self):
        ws = _load(self._rec(), lang="en")
        self.assertEqual(ws.cell(row=2, column=12).number_format, "#,##0.00")  # subtotal

    def test_freeze_first_row(self):
        ws = _load(self._rec())
        self.assertEqual(ws.freeze_panes, "A2")


class FilenameTests(unittest.TestCase):
    def test_default_filename(self):
        fn = ee.default_filename()
        self.assertTrue(fn.startswith("mr-pilot-results-"))
        self.assertTrue(fn.endswith(".xlsx"))


if __name__ == "__main__":
    unittest.main()
