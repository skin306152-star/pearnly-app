# -*- coding: utf-8 -*-
"""对账标准模板生成(按 UI 语言出列头 · 供"下载模板"端点)。

设计:列头跟随用户当前语言(zh/en/th/ja);生成的列正是各对账解析器认得的列,
故"导出即模板、模板即可导入"闭环。每个模板:第 1 行列头 + 一个『说明』sheet;
sheet 标题写 `Pearnly-<TYPE>` 作模板签名(前端"标准模板读取"可据此确定判定)。

statement / gl 是银行对账两侧:
  - 首行示例下有期初『ยอดยกมา』行(余额为字面值);明细行余额走公式自动累计
    (statement = 上一行余额 + 存 − 取;gl = 上一行余额 + 借 − 贷)。
  - gl 借/贷列头带『=存入/=取出』标注,帮用户对齐银行口径(借方=入账、贷方=出账)。

doc_type:
  statement 银行账单 · gl 总账 · vat 税表(VAT 报告)· invoice 销项发票明细
vat 与 invoice 列结构相同(行项 · 无期初/余额)。
"""

import io
from typing import Dict, List, Tuple

# 每列:(canonical_key, {lang: 列头}) · 列头取自各解析器已认得的关键词(4 语均可解析)。
_COLS: Dict[str, List[Tuple[str, Dict[str, str]]]] = {
    "statement": [
        ("date", {"zh": "日期", "en": "Date", "th": "วันที่", "ja": "日付"}),
        ("desc", {"zh": "摘要", "en": "Description", "th": "รายละเอียด", "ja": "摘要"}),
        ("withdrawal", {"zh": "支出", "en": "Withdrawal", "th": "ถอน", "ja": "出金"}),
        ("deposit", {"zh": "存入", "en": "Deposit", "th": "ฝาก", "ja": "入金"}),
        ("balance", {"zh": "余额", "en": "Balance", "th": "ยอดคงเหลือ", "ja": "残高"}),
    ],
    # GL 6 列(无科目码 · 与银行口径对齐):借方=存入(入账)、贷方=取出(出账),末列累计余额。
    "gl": [
        ("date", {"zh": "日期", "en": "Date", "th": "วันที่", "ja": "日付"}),
        ("voucher", {"zh": "凭证号", "en": "Voucher No", "th": "เลขที่ใบสำคัญ", "ja": "伝票番号"}),
        ("desc", {"zh": "摘要", "en": "Description", "th": "รายละเอียด", "ja": "摘要"}),
        (
            "debit",
            {
                "zh": "借方=存入",
                "en": "Debit = Money In",
                "th": "เดบิต=ฝากเงิน",
                "ja": "借方=入金",
            },
        ),
        (
            "credit",
            {
                "zh": "贷方=取出",
                "en": "Credit = Money Out",
                "th": "เครดิต=ถอนเงิน",
                "ja": "貸方=出金",
            },
        ),
        ("balance", {"zh": "余额", "en": "Balance", "th": "คงเหลือ", "ja": "残高"}),
    ],
    "vat": [
        ("date", {"zh": "日期", "en": "Date", "th": "วันที่", "ja": "日付"}),
        (
            "invoice_no",
            {"zh": "发票号", "en": "Invoice No", "th": "เลขที่ใบกำกับ", "ja": "請求書番号"},
        ),
        ("buyer_name", {"zh": "买方名称", "en": "Buyer Name", "th": "ชื่อผู้ซื้อ", "ja": "取引先"}),
        (
            "buyer_tax",
            {
                "zh": "买方税号",
                "en": "Buyer Tax ID",
                "th": "เลขประจำตัวผู้เสียภาษี",
                "ja": "納税者番号",
            },
        ),
        ("branch", {"zh": "分支", "en": "Branch", "th": "สาขา", "ja": "支店"}),
        ("net", {"zh": "税前金额", "en": "Net", "th": "มูลค่า", "ja": "税抜金額"}),
        ("vat", {"zh": "税额", "en": "VAT", "th": "ภาษี", "ja": "消費税"}),
        ("total", {"zh": "合计", "en": "Total", "th": "ยอดรวม", "ja": "合計"}),
    ],
}
_COLS["invoice"] = _COLS["vat"]

# 期初余额行标签(放日期列 · 余额列填字面值)· 4 语跟随。
_OPENING_LABEL = {
    "zh": "期初余额",
    "en": "Opening Balance",
    "th": "ยอดยกมา",
    "ja": "繰越残高",
}
_OPENING_VALUE = 100.0  # 示例期初(删示例行时一并删)

# 明细示例行(不含余额列 · 余额由 generate 按公式补)。
_SAMPLES: Dict[str, List[list]] = {
    "statement": [
        ["02/05/2026", "QR / SCB X1234", "", 1727.46],
        ["02/05/2026", "ค่า SIM", 107.00, ""],
    ],
    "gl": [
        ["02/05/2026", "SVTAX26004029", "QR / SCB", 1727.46, ""],
        ["02/05/2026", "FEE-SIM", "ค่า SIM", "", 107.00],
    ],
    "vat": [
        [
            "02/05/2026",
            "INV-0001",
            "ABC Co., Ltd.",
            "0105556001234",
            "00000",
            1000.00,
            70.00,
            1070.00,
        ],
        ["02/05/2026", "INV-0002", "Mr. Somchai", "", "", 500.00, 35.00, 535.00],
    ],
}
_SAMPLES["invoice"] = _SAMPLES["vat"]

_TITLE = {
    "statement": {
        "zh": "银行账单",
        "en": "Bank Statement",
        "th": "Statement ธนาคาร",
        "ja": "銀行明細",
    },
    "gl": {"zh": "总账 GL", "en": "General Ledger", "th": "บัญชีแยกประเภท GL", "ja": "総勘定元帳"},
    "vat": {"zh": "销项税报告", "en": "VAT Report", "th": "รายงานภาษีขาย", "ja": "消費税レポート"},
    "invoice": {
        "zh": "销项发票明细",
        "en": "Sales Invoices",
        "th": "ใบกำกับภาษีขาย",
        "ja": "売上請求書",
    },
}

# 说明文案(只列要点 · 不出现 OCR/AI 字眼)。
_NOTES = {
    "zh": [
        "Pearnly 标准模板 · {t}",
        "1) 列头勿改,一行一笔。",
        "2) 日期 dd/mm/yyyy(佛历年份自动转)。金额纯数字,勿带逗号或货币符号。",
        "3) 删除下面的示例行后,填入你的数据。请勿保留「合计/小计」等汇总行。",
    ],
    "en": [
        "Pearnly standard template · {t}",
        "1) Keep the header row unchanged. One record per row.",
        "2) Date dd/mm/yyyy. Numbers only (no commas / currency symbols).",
        "3) Delete the sample rows below, then fill in your data. Do not keep total/subtotal rows.",
    ],
    "th": [
        "เทมเพลตมาตรฐาน Pearnly · {t}",
        "1) ห้ามแก้หัวคอลัมน์ · หนึ่งรายการต่อหนึ่งแถว",
        "2) วันที่ dd/mm/yyyy · ตัวเลขล้วน ไม่ใส่จุลภาค/สัญลักษณ์เงิน",
        "3) ลบแถวตัวอย่างด้านล่างแล้วกรอกข้อมูลของคุณ · อย่าทิ้งแถวรวม/ผลรวม",
    ],
    "ja": [
        "Pearnly 標準テンプレート · {t}",
        "1) ヘッダー行は変更しないでください。1行に1件。",
        "2) 日付は dd/mm/yyyy。数字のみ(カンマ・通貨記号なし)。",
        "3) 下のサンプル行を削除してからデータを入力。合計/小計行は残さないでください。",
    ],
}

# 含期初行 + 余额累计列的银行对账两侧;vat/invoice 是纯行项无此结构。
_BALANCE_DOCS = ("statement", "gl")


def template_signature(doc_type: str) -> str:
    return f"Pearnly-{doc_type.upper()}"


def _balance_formula(doc_type: str, row: int) -> str:
    """末列余额 = 上一行余额 + 入账 − 出账。
    statement: 列 C=支出 D=存入 E=余额 → =E{r-1}+D{r}-C{r}
    gl:        列 D=借方 E=贷方 F=余额 → =F{r-1}+D{r}-E{r}
    """
    if doc_type == "statement":
        return f"=E{row - 1}+D{row}-C{row}"
    return f"=F{row - 1}+D{row}-E{row}"


def generate_template(doc_type: str, lang: str = "en") -> bytes:
    """生成对账标准模板 xlsx 字节。doc_type ∈ {statement,gl,vat,invoice};lang ∈ {zh,en,th,ja}。"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    if doc_type not in _COLS:
        raise ValueError(f"unknown doc_type: {doc_type!r}")
    if lang not in ("zh", "en", "th", "ja"):
        lang = "en"

    cols = _COLS[doc_type]
    headers = [c[1].get(lang) or c[1]["en"] for c in cols]
    bal_idx = len(cols)  # 末列即余额列(仅 _BALANCE_DOCS)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template_signature(doc_type)  # 模板签名(前端可据 sheet 名确定判定)
    ws.append(headers)

    if doc_type in _BALANCE_DOCS:
        # 期初行:标签放日期列,期初额放余额列(字面值,识别层据此读期初)。
        open_row = [""] * len(cols)
        open_row[0] = _OPENING_LABEL.get(lang) or _OPENING_LABEL["en"]
        open_row[bal_idx - 1] = _OPENING_VALUE
        ws.append(open_row)
        for s in _SAMPLES[doc_type]:
            r = ws.max_row + 1
            ws.append(list(s) + [_balance_formula(doc_type, r)])
    else:
        for s in _SAMPLES[doc_type]:
            ws.append(s)

    hf = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", fgColor="2563EB")
    for i in range(1, len(headers) + 1):
        c = ws.cell(1, i)
        c.font = hf
        c.fill = fill
        c.alignment = Alignment(horizontal="center")
        ws.column_dimensions[c.column_letter].width = 18

    note = wb.create_sheet(
        {"zh": "说明", "en": "Instructions", "th": "วิธีใช้", "ja": "説明"}[lang]
    )
    title = _TITLE[doc_type].get(lang) or _TITLE[doc_type]["en"]
    for i, line in enumerate(_NOTES[lang], 1):
        note.cell(i, 1, line.replace("{t}", title))
    note.column_dimensions["A"].width = 90

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def template_filename(doc_type: str, lang: str = "en") -> str:
    return f"Pearnly_{doc_type}_{lang}.xlsx"
