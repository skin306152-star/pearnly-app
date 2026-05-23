# -*- coding: utf-8 -*-
"""
P0.2 BUG-B-T2 v118.35.0.38 · 守门测试 · export_bank_recon_excel anchor 标黄 + section

锁定 3 个契约:
  1. anchor_overrides=None / 空 dict → Excel 不含 "手动录入痕迹" section · layout 不变
  2. anchor_overrides 非空 → Sheet 1 含警示行(FEF3C7 浅黄)+ 末尾含『手动录入』section
  3. user_value cell 标黄(FFE082)· cell comment 含 "OCR / User / Diff" 完整对照
"""
import io
import unittest

import openpyxl
from openpyxl.comments import Comment

from bank_recon_v2 import (
    BankReconSummary, BankReconRow,
    export_bank_recon_excel,
)


def _make_minimal_summary():
    return BankReconSummary(
        bank_code="kbank",
        gl_account_code="113001",
        stmt_opening=1000.0, gl_opening=1010.0,
        stmt_closing=2000.0, gl_closing=2010.0,
        opening_diff=-10.0, formula_stmt_closing=2000.0, formula_diff=0.0,
    )


def _make_minimal_rows():
    # 1 matched row · 让 export 不爆 empty
    return [
        BankReconRow(
            match_status="matched", match_layer="L1",
            stmt_date=None, stmt_desc="test", stmt_withdrawal=0.0, stmt_deposit=100.0,
            gl_date=None, gl_doc_no="V001", gl_desc="test gl", gl_debit=100.0, gl_credit=0.0,
        )
    ]


class BankReconAnchorYellowExportTests(unittest.TestCase):

    def test_export_without_overrides_has_no_manual_section(self):
        """P0.2 契约 · 没传 anchor_overrides → Sheet 1 不含『手动录入』section · 老 layout 保持"""
        b = export_bank_recon_excel(
            _make_minimal_rows(), _make_minimal_summary(), lang="en",
            anchor_overrides=None, anchor_ocr=None,
        )
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws1 = wb.worksheets[0]
        all_text = []
        for row in ws1.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    all_text.append(v)
        joined = " ".join(all_text)
        self.assertNotIn("Manual Entry Audit Trail", joined)
        self.assertNotIn("manually entered anchor", joined)

    def test_export_with_overrides_has_top_warning_and_bottom_section(self):
        """P0.2 契约 · 传 anchor_overrides → Sheet 1 含顶部警示 + 末尾『手动录入』section"""
        anchor_overrides = {
            "stmt_opening": {"ocr": 1000.0, "user": 1500.0},
            "gl_closing":   {"ocr": 2010.0, "user": 2500.0},
        }
        b = export_bank_recon_excel(
            _make_minimal_rows(), _make_minimal_summary(), lang="en",
            anchor_overrides=anchor_overrides,
            anchor_ocr={"stmt_opening": 1000.0, "gl_opening": 1010.0, "gl_closing": 2010.0},
        )
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws1 = wb.worksheets[0]
        all_text = []
        for row in ws1.iter_rows(values_only=True):
            for v in row:
                if isinstance(v, str):
                    all_text.append(v)
        joined = " ".join(all_text)
        self.assertIn("Manual Entry Audit Trail", joined)
        self.assertIn("manually entered anchor", joined)

    def test_user_value_cell_is_yellow_with_comment(self):
        """P0.2 契约 · user_value cell 标 FFE082 黄 · cell comment 含 OCR / User / Diff"""
        anchor_overrides = {
            "stmt_opening": {"ocr": 1000.0, "user": 1500.0},
        }
        b = export_bank_recon_excel(
            _make_minimal_rows(), _make_minimal_summary(), lang="en",
            anchor_overrides=anchor_overrides,
        )
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws1 = wb.worksheets[0]
        # 找到含 user_value=1500.0 的 cell · 验证 fill + comment
        found_yellow_user_cell = False
        for row in ws1.iter_rows():
            for cell in row:
                if cell.value == 1500.0 and cell.fill and cell.fill.fgColor:
                    fg = cell.fill.fgColor.rgb or ""
                    # openpyxl 给 rgb 加 alpha 前缀 → 取后 6 位
                    if fg.endswith("FFE082"):
                        found_yellow_user_cell = True
                        # comment 应该含 OCR + User + Diff
                        self.assertIsNotNone(cell.comment, "user_value cell 应该有 comment")
                        comment_text = cell.comment.text or ""
                        self.assertIn("OCR:", comment_text)
                        self.assertIn("User:", comment_text)
                        self.assertIn("Diff:", comment_text)
                        break
            if found_yellow_user_cell:
                break
        self.assertTrue(found_yellow_user_cell,
                        "user_value cell (1500.0) 没找到 FFE082 黄底 · P0.2 标黄失败")


if __name__ == "__main__":
    unittest.main()
