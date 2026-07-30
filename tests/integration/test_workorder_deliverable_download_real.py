# -*- coding: utf-8 -*-
"""交付物下载链:真跑流水线出真文件 → 产品端点真下到字节(交接单 §3.16 的缺口)。

在这份测试之前,这条链只被验到「路由在、鉴权在」:库里的 `artifact_path` 是手工种的、
指的文件根本不存在,端点必然 404,于是「点了到底有没有文件下来」从没被验过一次。
所以本测试的每一条都拒绝在磁盘之外取巧:

  · 文件由真的 `engine.run_work_order(real_handlers())` 跑到 package 写出来,
    不是测试自己造一段 bytes 塞进库(验证目标必须是产品真实产物);
  · 下载走真的 FastAPI 路由 + 真的 Bearer 鉴权,不直接读磁盘、不直接调函数;
  · 断言比对 sha256 与字节数,并把内容真解开一次(json.loads / openpyxl / PDF 尾标),
    专防「响应 200 但正文是一页 HTML 错误」和 0 字节这两种假绿;
  · 反证三条:匿名下载、跨租户下载(文件确实在盘上,别的租户照样拿不到)、
    artifact_path 指到本工单目录之外(取盘防穿越第二道防线)。

夹具只喂两样真实来料:一份走不了税表的附言(sort 按扩展名直接判 non_tax,不烧 OCR)
+ 一笔人工销项(走产品的 `api.record_sales_summary`),数字取语料答案。

跑法(要真库;CI 不跑 tests/integration):
    set PEARNLY_INTEGRATION_DB=1
    set DATABASE_URL=postgresql://pearnly:pearnly_local_dev@127.0.0.1:5432/pearnly
    set PGSSLMODE=disable
    python -m unittest tests.integration.test_workorder_deliverable_download_real -v

账号默认本机开发库那一个。注意 `PEARNLY_E2E_USER/PASS` 在开发机上指的是**生产**账号,
所以这里另起 `PEARNLY_LOCAL_E2E_*` 两个名字,不复用那两个(踩过:拿它当默认值会静默
打到 prod 去)。登录会顶掉该账号在别处的 token(active_jti 单活),别与别的窗口并发跑。
"""

from __future__ import annotations

import hashlib
import io
import json
import os
import shutil
import unittest
from decimal import Decimal
from pathlib import Path

from tests.integration._helpers import get_test_client, require_db

_USER = os.environ.get("PEARNLY_LOCAL_E2E_USER") or "stw_e2e"
_PASS = os.environ.get("PEARNLY_LOCAL_E2E_PASS") or "StwVerify#2026"

# 开单幂等键是 (租户, 账套, 账期, 意图):挑个真实业务不会占用的远古佛历年,
# 就不会撞上本机既有的 2569-xx 工单,重复跑也只会命中同一张夹具单。
_PERIOD = "2400-01"
_PERIOD_TRAVERSAL = "2400-02"
_FOREIGN_TENANT = "00000000-0000-4000-8000-0000000000ff"

# 语料答案(MR.ERP SM · 2569-06 现金销售 609 单合计)。package 只渲染不重算,
# 喂进去多少,底稿里就该出来多少。
_SALES = Decimal("953509.68")
_OUTPUT_VAT = Decimal("66745.59")

_THAI_PP30 = "ภ.พ.30"


def _sha(b: bytes) -> str:
    return hashlib.sha256(b).hexdigest()


class DeliverableDownloadRealTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        require_db()
        cls.client = get_test_client()
        cls._made: list[tuple[str, str]] = []  # (tenant_id, work_order_id) 留给 tearDown 清

        from core import db
        from services.workorder import storage, store

        cls.db, cls.store, cls.storage = db, store, storage

        resp = cls.client.post(
            "/api/login", json={"username": _USER, "password": _PASS, "entry": "ai"}
        )
        if resp.status_code != 200:
            raise unittest.SkipTest(f"本机开发库测试号登不上({resp.status_code}):{resp.text[:200]}")
        cls.auth = {"Authorization": f"Bearer {resp.json()['token']}"}

        with db.get_cursor() as cur:
            cur.execute("SELECT tenant_id FROM users WHERE username = %s", (_USER,))
            row = cur.fetchone()
            cls.tenant_id = str(row["tenant_id"])
            cur.execute(
                "SELECT id FROM workspace_clients WHERE tenant_id = %s ORDER BY id LIMIT 1",
                (cls.tenant_id,),
            )
            ws = cur.fetchone()
        if not ws:
            raise unittest.SkipTest(f"租户 {cls.tenant_id} 名下没有账套,建不了工单夹具")
        cls.ws_id = int(ws["id"])

        cls.order_id, cls.out_dir = cls._packaged_order(cls.tenant_id, _PERIOD)
        # 跨租户反证的对照物:文件真在盘上,只是不属于登录的这个租户。
        cls.foreign_order_id, cls.foreign_dir = cls._packaged_order(_FOREIGN_TENANT, _PERIOD)

    @classmethod
    def tearDownClass(cls):
        # 逐表显式删,不指望 ON DELETE CASCADE:0059 迁移写的是 CASCADE,本机开发库这三条
        # 外键实际是 RESTRICT(建库路径与迁移有漂移),靠级联清理会在别人的机器上炸。
        for tenant_id, work_order_id in getattr(cls, "_made", []):
            shutil.rmtree(cls.storage.order_dir(tenant_id, work_order_id), ignore_errors=True)
            with cls.db.get_cursor(commit=True) as cur:
                for table in (
                    "work_order_deliverables",
                    "work_order_items",
                    "work_order_events",
                ):
                    cur.execute(
                        f"DELETE FROM {table} WHERE tenant_id = %s AND work_order_id = %s",
                        (tenant_id, work_order_id),
                    )
                cur.execute(
                    "DELETE FROM work_orders WHERE tenant_id = %s AND id = %s",
                    (tenant_id, work_order_id),
                )

    @classmethod
    def _packaged_order(cls, tenant_id: str, period: str) -> tuple[str, Path]:
        """开单 → 落一份附言 + 一笔人工销项 → 真跑整条流水线,返回 (工单 id, 交付物版本目录)。"""
        from services.workorder import api, engine
        from services.workorder.steps import real_handlers

        with cls.db.get_cursor(commit=True) as cur:
            wo = cls.store.open_work_order(
                cur, tenant_id=tenant_id, workspace_client_id=cls.ws_id, period=period
            )
            work_order_id = str(wo["id"])
            cls._made.append((tenant_id, work_order_id))
            api.record_sales_summary(
                cur,
                tenant_id=tenant_id,
                work_order_id=work_order_id,
                sales_amount=_SALES,
                output_vat=_OUTPUT_VAT,
                note="integration fixture · 语料答案摘要",
                actor="test:deliverable-download",
            )

        material = cls.storage.save_material(
            tenant_id,
            work_order_id,
            "ใบแนบ (附言,非税件)\n".encode("utf-8"),
            suffix=".txt",
            original_name="fixture-note.txt",
        )
        ctx = engine.StepContext(
            cur=None,
            tenant_id=tenant_id,
            work_order_id=work_order_id,
            data={
                "deliverables_dir": str(cls.storage.deliverables_dir(tenant_id, work_order_id)),
                "intake_files": [str(material)],
            },
            cursor_factory=lambda: cls.db.get_cursor(commit=True),
        )
        out = engine.run_work_order(ctx, handlers=real_handlers())
        if not out.completed:
            raise AssertionError(f"流水线没跑到 package:停在 {out.stopped_at} · {out.result}")
        with cls.db.get_cursor() as cur:
            version = cls.store.current_deliverable_version(
                cur, tenant_id=tenant_id, work_order_id=work_order_id
            )
        return work_order_id, cls.storage.versioned_dir(
            cls.storage.deliverables_dir(tenant_id, work_order_id), version
        )

    def _download(self, kind: str, work_order_id: str | None = None, auth: bool = True):
        return self.client.get(
            f"/api/workorder/orders/{work_order_id or self.order_id}/deliverables/{kind}",
            headers=self.auth if auth else {},
        )

    def _artifact_path(self, work_order_id: str, kind: str, tenant_id: str) -> str:
        with self.db.get_cursor() as cur:
            rows = self.store.list_deliverables(
                cur, tenant_id=tenant_id, work_order_id=work_order_id
            )
        for row in rows:
            if row["kind"] == kind:
                return row["artifact_path"]
        raise AssertionError(f"库里没有 {kind} 这件交付物")

    def test_pipeline_wrote_real_files(self):
        """前置自证:后面每条断言都建立在「盘上真有文件」之上,它得先成立。"""
        names = sorted(p.name for p in self.out_dir.glob("*"))
        self.assertIn("pp30_draft.md", names, f"package 没写出底稿:{names}")
        for name in names:
            self.assertGreater((self.out_dir / name).stat().st_size, 0, f"{name} 是 0 字节")

    def test_download_delivers_the_exact_bytes_on_disk(self):
        listing = self.client.get(
            f"/api/workorder/orders/{self.order_id}/deliverables", headers=self.auth
        )
        self.assertEqual(listing.status_code, 200, listing.text)
        kinds = [d["kind"] for d in listing.json()["deliverables"] if d["has_file"]]
        self.assertGreaterEqual(len(kinds), 5, f"交付物清单太薄,package 可能没出全:{kinds}")

        for kind in kinds:
            with self.subTest(kind=kind):
                got = self._download(kind)
                self.assertEqual(got.status_code, 200, got.text[:300])
                body = got.content
                self.assertGreater(len(body), 0, "下到 0 字节")

                path = self.storage.resolve_within_order(
                    self.tenant_id,
                    self.order_id,
                    self._artifact_path(self.order_id, kind, self.tenant_id),
                )
                self.assertIsNotNone(path, "库里登记的 artifact_path 落不到本工单目录内")
                disk = self.storage.read_bytes(path)
                self.assertEqual(len(body), len(disk))
                self.assertEqual(_sha(body), _sha(disk), "下到的字节与盘上那份不是同一个文件")

                disposition = got.headers["content-disposition"]
                self.assertIn("attachment;", disposition)
                self.assertIn("filename*=UTF-8''", disposition, "缺 RFC 5987 名,泰文文件名会被吃掉")
                self.assertTrue(got.headers.get("content-type"))

                # 内容真解开一次:专防「200 包着一页 HTML 错误」这种假绿。
                text = body.decode("utf-8")
                self.assertFalse(
                    text.lstrip().startswith("<"), f"{kind} 下回来的是标记语言不是交付物"
                )
                if path.suffix == ".json":
                    self.assertIsInstance(json.loads(text), dict)
                else:
                    self.assertTrue(
                        text.startswith("# "), f"{kind} 不是 markdown 底稿:{text[:60]!r}"
                    )

        # 数字要与喂进去的语料答案一致(package 只渲染不重算)。
        pp30 = self._download("pp30_draft").content.decode("utf-8")
        self.assertIn(str(_SALES), pp30)
        self.assertIn(str(_OUTPUT_VAT), pp30)

    def test_financials_xlsx_and_pdf_really_open(self):
        """月度报表走独立端点现场渲染(不落盘),照样要能被 openpyxl / PDF 阅读器打开。"""
        xlsx = self.client.get(
            f"/api/workorder/orders/{self.order_id}/financials/download?format=xlsx&lang=th",
            headers=self.auth,
        )
        self.assertEqual(xlsx.status_code, 200, xlsx.text[:300])
        self.assertIn("spreadsheetml", xlsx.headers["content-type"])
        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(xlsx.content))
        sheet = wb[wb.sheetnames[0]]
        self.assertGreater(sheet.max_row, 1, "xlsx 打得开但一行数据都没有")
        self.assertTrue(sheet.cell(row=1, column=1).value, "首格是空的")

        pdf = self.client.get(
            f"/api/workorder/orders/{self.order_id}/financials/download?format=pdf&lang=th",
            headers=self.auth,
        )
        self.assertEqual(pdf.status_code, 200)
        self.assertEqual(pdf.headers["content-type"], "application/pdf")
        self.assertTrue(pdf.content.startswith(b"%PDF-"), "不是 PDF 文件头")
        self.assertIn(b"%%EOF", pdf.content[-2048:], "PDF 没有结束标记,多半被截断了")

    def test_anonymous_cannot_download(self):
        self.assertEqual(self._download("pp30_draft", auth=False).status_code, 401)

    def test_foreign_tenant_file_exists_on_disk_but_stays_unreachable(self):
        """反证一:文件确确实实在盘上,换一个租户的单子就是拿不到 —— 404 不是「没文件」。"""
        real = self.foreign_dir / "pp30_draft.md"
        self.assertTrue(real.is_file() and real.stat().st_size > 0, "对照组文件没生成,反证无效")

        got = self._download("pp30_draft", work_order_id=self.foreign_order_id)
        self.assertEqual(got.status_code, 404, f"别的租户的交付物被下走了:{got.status_code}")
        self.assertNotIn(_THAI_PP30, got.text)

    def test_artifact_path_outside_the_order_dir_is_refused(self):
        """反证二:登记的路径指到本工单目录之外(误写/被改)时,取盘必须拒。"""
        outside = str(self.foreign_dir / "pp30_draft.md")
        self.assertTrue(Path(outside).is_file(), "对照组文件没生成,反证无效")

        with self.db.get_cursor(commit=True) as cur:
            wo = self.store.open_work_order(
                cur,
                tenant_id=self.tenant_id,
                workspace_client_id=self.ws_id,
                period=_PERIOD_TRAVERSAL,
            )
            traversal_id = str(wo["id"])
            self._made.append((self.tenant_id, traversal_id))
            self.store.upsert_deliverable(
                cur,
                tenant_id=self.tenant_id,
                work_order_id=traversal_id,
                kind="pp30_draft",
                version=1,
                artifact_path=outside,
                numbers={},
            )

        got = self._download("pp30_draft", work_order_id=traversal_id)
        self.assertEqual(got.status_code, 404, "越界的 artifact_path 被放行了")
        self.assertNotIn(_THAI_PP30, got.text)


if __name__ == "__main__":
    unittest.main()
