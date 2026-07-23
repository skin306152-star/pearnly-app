# -*- coding: utf-8 -*-
"""导出工作簿的共用视觉语言。

复核工作簿里销项/进项/待判几张表并排给同一个人看,表头、边框、新建-复用底色必须
完全一致 —— 各写一套迟早会漂。故集中在这里一处定义。
"""

from __future__ import annotations

from typing import Any, Optional, Sequence

try:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:  # 无 openpyxl 环境(纯逻辑测试)不该因此炸
    Alignment = Border = Font = PatternFill = Side = None  # type: ignore
    get_column_letter = None  # type: ignore

# 表头深蓝底白字 · 全站导出统一
HEADER_COLOR = "2C5282"
# 新建=柔和绿 · 复用=淡灰蓝 · 均浅底黑字保可读(黑白打印也分得出深浅)
FILL_NEW_COLOR = "D6F5DE"
FILL_REUSED_COLOR = "EEF2F7"
# 待判=淡琥珀:既不是"成了"也不是"错了",是"还没定"——不能和失败同色
FILL_PENDING_COLOR = "FDF3D8"

MONEY_FMT = "#,##0.00"
QTY_FMT = "#,##0.###"
DATE_FMT = "dd/mm/yyyy"
TEXT_FMT = "@"


def _solid(color: str):
    return PatternFill(start_color=color, end_color=color, fill_type="solid")


def header_fill():
    return _solid(HEADER_COLOR)


def header_font():
    # 不指定字体名 —— 指定了反而破坏泰文字形渲染
    return Font(bold=True, color="FFFFFF")


def fill_new():
    return _solid(FILL_NEW_COLOR)


def fill_reused():
    return _solid(FILL_REUSED_COLOR)


def fill_pending():
    return _solid(FILL_PENDING_COLOR)


def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)


def center():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)


def left():
    return Alignment(horizontal="left", vertical="center", wrap_text=True)


def right():
    return Alignment(horizontal="right", vertical="center")


def write_header_row(ws, headers: Sequence[str], widths: Optional[Sequence[int]] = None) -> None:
    """写表头 + 列宽 + 冻结首行。所有导出表共用,保证视觉一致。"""
    fill, font, border, align = header_fill(), header_font(), thin_border(), center()
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = font
        c.fill = fill
        c.alignment = align
        c.border = border
        if widths and i <= len(widths):
            ws.column_dimensions[get_column_letter(i)].width = widths[i - 1]
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


def style_cell(cell: Any, *, align=None, fmt: Optional[str] = None, fill=None) -> None:
    cell.border = thin_border()
    if align is not None:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    if fill is not None:
        cell.fill = fill
