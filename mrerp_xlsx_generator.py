# -*- coding: utf-8 -*-
"""
v118.27.4.2 · MR.ERP 适配器(方案 C · xlsx 一键导出)
=========================================================
铁律 28-31(2026-05-09 锁定 · 28 于 v118.27.4.2 修订)

格式:
  - .xlsx 不接受 .csv
  - 🔴 铁律 28 修订(v118.27.4.2):**sheet 数随模板动态 1-4**
      · 1 sheet → 名 `Worksheet`(单独 sheet 不带数字 · 客户/商品档案)
      · 2 sheet → `Worksheet` / `Worksheet 1`(会计凭证)
      · 3 sheet → `Worksheet` / `Worksheet 1` / `Worksheet 2`(销售/采购/收款)
      · 4 sheet → `Worksheet` / `Worksheet 1` / `Worksheet 2` / `Worksheet 3`(付款带代扣税)
      · 命名严格 = 空格分隔(不是 Sheet1/Sheet2)
  - 9 种文件类型(销售-赊销/销售-现金/采购-赊购/采购-现金/客户档案/商品档案/财务收款/财务付款/会计凭证)

字段:
  - 日期 = 字符串 YYYY-MM-DD(cell `@` 文本)
  - 单号 = 字符串(cell `@` 文本)
  - 金额 = 数字 cell · 上限 999,999,999.99
  - 客户代码 = 任意字符串(三段式 + 含泰文)· 例 `01-อนุรักษ์-001`
  - 税率 = 字符串枚举 `"7%"` / `"0%"` / `"นอกระบบ"` / `"ยกเว้น"` 等
  - 字段长度 trim 14-50 字符

stub-first 口子(物料到了改这 4 处 · 不动其他代码):
  1. MRERP_SHEET_SCHEMAS['sales_cash']        ← 物料 1(7 列收款方式)
  2. MRERP_SHEET_SCHEMAS['purchase_credit']
     MRERP_SHEET_SCHEMAS['purchase_cash']     ← 物料 2
  3. MRERP_SHEET_SCHEMAS['journal']           ← 物料 3
  4. MRERP_ERROR_FRIENDLY                     ← 物料 4

schema sheet 列字段(按出现顺序映射 sheet):
  - header_columns  → Worksheet     (必填 · 不能空)
  - detail_columns  → Worksheet 1   (可空 · 空则不创建该 sheet)
  - tail_columns    → Worksheet 2   (可空)
  - sheet4_columns  → Worksheet 3   (可空 · 付款代扣税专用)
"""

import io
import logging
import re
from datetime import date, datetime
from decimal import Decimal
from typing import Dict, Any, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
except ImportError:
    Workbook = None  # type: ignore

logger = logging.getLogger(__name__)


# ============================================================
# v27.8.1.17 · 商品名归一化 + 映射查表(给 detail 行写真 product_code 用)
#   - 归一化规则跟 db.py 的 _product_name_norm_for_db 完全一致(必须一致才能命中映射表的 item_name_norm 列)
#   - mappings['products'] 是 db.get_mrerp_mappings_bundle 返的 list · 每项含 item_name_norm / erp_code
#   - 找不到返 None · 调用方 fallback '123' 占位
# ============================================================
_PRODUCT_NAME_NORM_RE = re.compile(r"[\s\.,\-_/\\()&\"'`*]+")

def _norm_product_name(s: Any) -> str:
    if not s:
        return ""
    out = _PRODUCT_NAME_NORM_RE.sub('', str(s))
    return out.lower().strip()[:256]


def _build_product_lookup(mappings: Optional[Dict[str, Any]]) -> Dict[str, str]:
    """从 mappings['products'] 建 item_name_norm → erp_code 查表
    map 的 key 用归一化形式 · 跟 erp_product_mappings.item_name_norm 列一致
    """
    out: Dict[str, str] = {}
    if not mappings:
        return out
    products = mappings.get('products') if isinstance(mappings, dict) else None
    if not isinstance(products, list):
        return out
    for p in products:
        if not isinstance(p, dict):
            continue
        erp_code = p.get('erp_code')
        if not erp_code:
            continue
        # 优先用 DB 算好的 norm · 退到自己算
        norm = p.get('item_name_norm') or _norm_product_name(p.get('item_name') or '')
        if norm:
            out[norm] = str(erp_code).strip()
    return out


def _resolve_product_code(item_name: Any, lookup: Dict[str, str]) -> Optional[str]:
    """OCR item_name → ERP product code · 找不到返 None
    需配合 _build_product_lookup(mappings) 一起用 · 在 detail rows 循环外建 lookup 一次 · 内循环 O(1) 查表
    """
    if not item_name or not lookup:
        return None
    norm = _norm_product_name(item_name)
    return lookup.get(norm) if norm else None

# ============================================================
# 铁律 29 · 字段格式工具
# ============================================================
MAX_AMOUNT = Decimal("999999999.99")

# 服务端业务校验上限(2026-05-18 集成测试发现 · 见 mrerp-known-facts.md §7)
# xlsx schema 字段标 str(30) / str(50) 但 MR.ERP 服务端实际更严
# 校验顺序:服务端先做长度校验 → 长度过 → 才查主数据存在
# adapter 必须 client-side 拦截,避免 "ไม่พบ" 类错误被长度错遮盖
MRERP_INVOICE_NO_MAX = 18
MRERP_BILL_NO_MAX = 20
MRERP_CUSTOMER_CODE_MAX = 20
MRERP_CUSTOMER_BILL_MAX = 20

# sales_credit 允许的 Pearnly tax_kind 枚举(derive_tax_kind 的返回值之一)
# wht_* 是预扣税 · 销项发票不适用 · 仅 vat_* 或 non_vat 合法
MRERP_VALID_TAX_KINDS_SC = ("vat_7", "vat_0", "vat_exempt", "non_vat")

# 日期检查阈值(单位:天)
MRERP_DATE_FUTURE_HARD_REJECT_DAYS = 30   # > today + 30d → ERR_DATE_FUTURE
MRERP_DATE_FUTURE_WARN_DAYS = 7           # > today + 7d → 警告(不拒)
MRERP_DATE_PAST_WARN_DAYS = 730           # < today - 730d → 警告(2 年前)


def fmt_date(value: Any) -> str:
    """日期 → YYYY-MM-DD 字符串(cell 用 @ 文本格式)"""
    if not value:
        return ""
    if isinstance(value, str):
        s = value.strip()
        if not s:
            return ""
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            return s[:10]
        try:
            dt = datetime.fromisoformat(s.replace("Z", "+00:00"))
            return dt.strftime("%Y-%m-%d")
        except Exception:
            return s[:10]
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value)[:10]


def fmt_str(value: Any, max_len: int = 50) -> str:
    """字符串 trim 到 max_len · None / NaN 返回空串"""
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() in ("none", "null", "nan"):
        return ""
    if len(s) > max_len:
        return s[:max_len]
    return s


def fmt_number(value: Any) -> Optional[float]:
    """金额 → float · 超上限或非法返回 None · 负数静默 clamp 到 -MAX_AMOUNT
    (保留原行为:用于摘要/展示场景 · 不抛错)"""
    if value is None or value == "":
        return None
    try:
        n = Decimal(str(value))
    except Exception:
        return None
    if n > MAX_AMOUNT:
        return float(MAX_AMOUNT)
    if n < -MAX_AMOUNT:
        return float(-MAX_AMOUNT)
    return float(n)


def fmt_number_strict(value: Any) -> float:
    """金额严格模式 → float · 负数 / 超 MAX_AMOUNT / 非法都 raise ValueError
    用于 sales_credit 上传前 preflight · 销项发票净额必须 > 0
    (P1-A §3.3 · 2026-05-18)"""
    if value is None or value == "":
        raise ValueError("amount is missing")
    try:
        n = Decimal(str(value))
    except Exception as e:
        raise ValueError(f"amount not parseable: {value!r}") from e
    if n < 0:
        raise ValueError(
            f"negative amount {n} not allowed for sales_credit upload "
            "(use a credit-note flow instead)"
        )
    if n > MAX_AMOUNT:
        raise ValueError(
            f"amount {n} exceeds MR.ERP ceiling {MAX_AMOUNT}; "
            "split the invoice or escalate"
        )
    return float(n)


def derive_mrerp_invoice_no(history: Dict[str, Any]) -> str:
    """v27.8.1.5 · 根据 OCR invoice_date 生成 MR.ERP 期望格式 YYMMDD-NNN
    Korn 真样本:'690507-001' = BE 年(西历+543 末 2 位)+ 月 + 日 + 序号

    v118.34.28 (Zihao 2026-05-19 拍板) · 修死链:之前 seq 写死 "001" ·
    所有 push 都 derive 出 YYMMDD-001 · MR.ERP 服务端不会自动 increment ·
    第一笔成功后第二笔重复就 "เลขที่ดังกล่าวมีอยู่ในระบบแล้ว".
    序号现在按 docstring 设计取 history.id 末 3 位 hex 转 dec mod 1000 ·
    同一 history 重传幂等 · 不同 history 序号不同."""
    inv_date_str = fmt_date(history.get('invoice_date'))
    if inv_date_str and len(inv_date_str) >= 10:
        try:
            ce_year = int(inv_date_str[:4])
            be_year = (ce_year + 543) % 100
            mm = inv_date_str[5:7]
            dd = inv_date_str[8:10]
            date_part = f"{be_year:02d}{mm}{dd}"
        except Exception:
            date_part = "690101"
    else:
        from datetime import datetime as _dt
        d = _dt.utcnow()
        date_part = f"{(d.year + 543) % 100:02d}{d.month:02d}{d.day:02d}"

    # 取 history.id 末 6 位 hex 转 dec mod 99999 + 1 (1-99999 · 5 位 seq) ·
    # MR.ERP test 数据库 cumulative · 3 位 seq (1-999) 跟历史 push 撞率太高 ·
    # 5 位 seq 几乎不撞. 同 history 仍幂等. invoice_no 总长 6+1+5 = 12 char ·
    # 远小于 18 char 上限.
    hist_id = str(history.get('id') or '').replace('-', '')
    if hist_id:
        try:
            seq_int = int(hist_id[-6:], 16) % 99999 + 1   # 1-99999
            seq = f"{seq_int:05d}"
        except ValueError:
            seq = "00001"
    else:
        seq = "00001"
    return f"{date_part}-{seq}"


# ============================================================
# 9 种 sheet schema(stub-first · 物料到了改这里)
# ============================================================
# schema 格式:
# {
#     'header_columns':  [{'key': 'invoice_no', 'label': '...', 'type': 'str', 'max': 30}, ...],
#     'detail_columns':  [...] | None,   # 空 / None → 不创建 Worksheet 1
#     'tail_columns':    [...] | None,   # 空 / None → 不创建 Worksheet 2
#     'sheet4_columns':  [...] | None,   # 空 / None → 不创建 Worksheet 3(付款代扣税专用)
#     'stub': True | False,              # True = 物料未到 · 字段是猜测
# }

MRERP_SHEET_SCHEMAS: Dict[str, Dict[str, Any]] = {

    # ✅ 销售-赊销(物料 1 真实样本对齐 v27.8.1.4 · Korn 实际导出 SC 单)
    # Worksheet: 18 列 · Worksheet 1: 8 列 · Worksheet 2: 3 列
    'sales_credit': {
        'stub': False,
        'sheet_count_expected': 3,
        'label_zh': '销售-赊销',
        'label_th': 'ขายเชื่อ',
        'label_en': 'Sales (credit)',
        'label_ja': '売上(掛)',
        # Worksheet · 单据头(严格按 Korn 真样本列顺序)
        'header_columns': [
            {'key': 'invoice_no',     'label': 'เลขที่',           'type': 'str',  'max': 30},
            {'key': 'invoice_date',   'label': 'วันที่',            'type': 'date'},
            {'key': 'tax_rate_str',   'label': 'อัตราภาษี',         'type': 'str',  'max': 14, 'default': '7 (แยก)'},
            {'key': 'branch_code',    'label': 'สาขา',             'type': 'str',  'max': 14, 'default': '00000'},
            {'key': 'department',     'label': 'แผนก',             'type': 'str',  'max': 14, 'default': 'BOI1'},
            {'key': 'job',            'label': 'งาน',              'type': 'str',  'max': 14, 'default': '00002'},
            {'key': 'salesman',       'label': 'พนักงานขาย',        'type': 'str',  'max': 30, 'default': 'กร ทดสอบ'},
            {'key': 'delivery_date',  'label': 'กำหนดส่งสินค้า',     'type': 'date'},
            {'key': 'customer_code',  'label': 'รหัสลูกค้า',        'type': 'str',  'max': 50},
            {'key': 'customer_bill',  'label': 'รหัสลูกค้า (บิล)',  'type': 'str',  'max': 50},
            {'key': 'bill_no',        'label': 'เลขที่บิล',         'type': 'str',  'max': 30},
            {'key': 'bill_date',      'label': 'วันที่',            'type': 'date'},
            {'key': 'sales_area',     'label': 'พื้นที่การขาย',     'type': 'str',  'max': 30, 'default': 'สุพรรณบุรี'},
            {'key': 'shipping_type',  'label': 'ประเภทขนส่ง',       'type': 'str',  'max': 30, 'default': 'ขนส่งโดยบริษัท'},
            {'key': 'discount',       'label': 'หักส่วนลด',          'type': 'num', 'default': 0},
            {'key': 'note1',          'label': 'หมายเหตุ 1',        'type': 'str',  'max': 50},
            {'key': 'note2',          'label': 'หมายเหตุ 2',        'type': 'str',  'max': 50},
            {'key': 'note3',          'label': 'หมายเหตุ 3',        'type': 'str',  'max': 50},
        ],
        # Worksheet 1 · 商品明细(每行 1 件商品 · 关联 invoice_no)
        'detail_columns': [
            {'key': 'invoice_no',   'label': 'เลขที่',         'type': 'str', 'max': 30},
            {'key': 'product_code', 'label': 'รหัสสินค้า',     'type': 'str', 'max': 30, 'default': '123'},
            {'key': 'department',   'label': 'แผนก',          'type': 'str', 'max': 14, 'default': 'BOI1'},
            {'key': 'job',          'label': 'งาน',           'type': 'str', 'max': 14, 'default': '00002'},
            {'key': 'warehouse',    'label': 'คลัง',          'type': 'str', 'max': 14, 'default': '0000'},
            {'key': 'qty',          'label': 'จำนวน',         'type': 'num'},
            {'key': 'unit_price',   'label': 'ราคา/หน่วย',     'type': 'num'},
            {'key': 'amount',       'label': 'จำนวนเงิน',     'type': 'num'},
        ],
        # Worksheet 2 · 尾(押金 / 是否开销售单)
        'tail_columns': [
            {'key': 'invoice_no',     'label': 'เลขที่',           'type': 'str', 'max': 30},
            {'key': 'deposit_no',     'label': 'เลขที่เงินมัดจำ',   'type': 'str', 'max': 30},
            {'key': 'is_sales_issued','label': 'ออกใบขาย',         'type': 'str', 'max': 14},
        ],
    },

    # 🟡 销售-现金(stub · TODO 物料 1 · 7 列收款方式待定 · 预期 3 sheet)
    'sales_cash': {
        'stub': True,
        'sheet_count_expected': 3,
        'label_zh': '销售-现金 · 待物料',
        'label_th': 'ขายสด · รอตัวอย่าง',
        'label_en': 'Sales (cash) · WAIT MATERIAL',
        'label_ja': '売上(現金) · 物料待ち',
        # TODO 物料 1 · 现金销售真实样本到达后填:头 18 列 + 7 列收款方式 + 明细 8 + 尾 3
        'header_columns': [],
        'detail_columns': [],
        'tail_columns': [],
    },

    # 🟡 采购-赊购(stub · TODO 物料 2 · 头 16 + 明细 8 + 尾 3 · 预期 3 sheet)
    'purchase_credit': {
        'stub': True,
        'sheet_count_expected': 3,
        'label_zh': '采购-赊购 · 待物料',
        'label_th': 'ซื้อเชื่อ · รอตัวอย่าง',
        'label_en': 'Purchase (credit) · WAIT MATERIAL',
        'label_ja': '仕入(掛) · 物料待ち',
        # TODO 物料 2 · 采购赊购样本到达后填字段名
        'header_columns': [],
        'detail_columns': [],
        'tail_columns': [],
    },

    # 🟡 采购-现金(stub · TODO 物料 2 · 预期 3 sheet)
    'purchase_cash': {
        'stub': True,
        'sheet_count_expected': 3,
        'label_zh': '采购-现金 · 待物料',
        'label_th': 'ซื้อสด · รอตัวอย่าง',
        'label_en': 'Purchase (cash) · WAIT MATERIAL',
        'label_ja': '仕入(現金) · 物料待ち',
        # TODO 物料 2 · 采购现金样本到达后填字段名
        'header_columns': [],
        'detail_columns': [],
        'tail_columns': [],
    },

    # ✅ 客户档案(已有完整 25 列样本)· 1 sheet
    'customer': {
        'stub': False,
        'sheet_count_expected': 1,
        'label_zh': '客户档案',
        'label_th': 'แฟ้มลูกค้า',
        'label_en': 'Customer master',
        'label_ja': '取引先マスタ',
        'header_columns': [
            {'key': 'customer_code',  'label': 'รหัสลูกค้า',     'type': 'str', 'max': 50},
            {'key': 'customer_name',  'label': 'ชื่อลูกค้า',      'type': 'str', 'max': 50},
            {'key': 'tax_id',         'label': 'เลขผู้เสียภาษี',  'type': 'str', 'max': 20},
            {'key': 'branch_code',    'label': 'สาขา',           'type': 'str', 'max': 14, 'default': '00000'},
            {'key': 'address1',       'label': 'ที่อยู่ 1',      'type': 'str', 'max': 50},
            {'key': 'address2',       'label': 'ที่อยู่ 2',      'type': 'str', 'max': 50},
            {'key': 'address3',       'label': 'ที่อยู่ 3',      'type': 'str', 'max': 50},
            {'key': 'phone',          'label': 'โทรศัพท์',       'type': 'str', 'max': 30},
            {'key': 'fax',            'label': 'แฟกซ์',          'type': 'str', 'max': 30},
            {'key': 'email',          'label': 'อีเมล',          'type': 'str', 'max': 50},
            {'key': 'contact_person', 'label': 'ผู้ติดต่อ',      'type': 'str', 'max': 50},
            {'key': 'credit_limit',   'label': 'วงเงินเครดิต',   'type': 'num'},
            {'key': 'credit_days',    'label': 'เครดิต(วัน)',    'type': 'num'},
            {'key': 'salesman_code',  'label': 'พนักงานขาย',     'type': 'str', 'max': 14},
            {'key': 'group_code',     'label': 'กลุ่มลูกค้า',    'type': 'str', 'max': 14},
            {'key': 'zone_code',      'label': 'เขตการขาย',      'type': 'str', 'max': 14},
            {'key': 'price_level',    'label': 'ระดับราคา',      'type': 'str', 'max': 14},
            {'key': 'tax_rate_str',   'label': 'อัตราภาษี',      'type': 'str', 'max': 14},
            {'key': 'is_wht',         'label': 'หัก ณ ที่จ่าย',  'type': 'str', 'max': 14},
            {'key': 'wht_rate_str',   'label': 'อัตราหัก',       'type': 'str', 'max': 14},
            {'key': 'bank_name',      'label': 'ธนาคาร',         'type': 'str', 'max': 30},
            {'key': 'bank_account',   'label': 'เลขบัญชี',       'type': 'str', 'max': 30},
            {'key': 'note',           'label': 'หมายเหตุ',       'type': 'str', 'max': 50},
            {'key': 'is_active',      'label': 'สถานะ',          'type': 'str', 'max': 14, 'default': 'A'},
            {'key': 'created_date',   'label': 'วันที่สร้าง',    'type': 'date'},
        ],
        'detail_columns': None,
        'tail_columns': None,
    },

    # ✅ 商品档案(已知 29 列结构)· 1 sheet
    'product': {
        'stub': False,
        'sheet_count_expected': 1,
        'label_zh': '商品档案',
        'label_th': 'แฟ้มสินค้า',
        'label_en': 'Product master',
        'label_ja': '商品マスタ',
        'header_columns': [
            {'key': f'col_{i}', 'label': f'Col{i}', 'type': 'str', 'max': 50}
            for i in range(1, 30)
        ],
        'detail_columns': None,
        'tail_columns': None,
    },

    # ✅ 财务收款(已知 头20 + 核销3 + 支票5)· 3 sheet
    'receipt': {
        'stub': False,
        'sheet_count_expected': 3,
        'label_zh': '财务收款',
        'label_th': 'รับชำระ',
        'label_en': 'Receipt',
        'label_ja': '入金',
        'header_columns': [
            {'key': f'h_{i}', 'label': f'H{i}', 'type': 'str', 'max': 50}
            for i in range(1, 21)
        ],
        'detail_columns': [
            {'key': f'd_{i}', 'label': f'D{i}', 'type': 'str', 'max': 50}
            for i in range(1, 4)
        ],
        'tail_columns': [
            {'key': f't_{i}', 'label': f'T{i}', 'type': 'str', 'max': 50}
            for i in range(1, 6)
        ],
    },

    # ✅ 财务付款(已知 多 Sheet4 = 15 列代扣税)· 4 sheet
    # 当前 stub-ish:Sheet1-3 列名待物料 · Sheet4 用 P1-P15 占位
    'payment': {
        'stub': False,
        'sheet_count_expected': 4,
        'label_zh': '财务付款',
        'label_th': 'จ่ายชำระ',
        'label_en': 'Payment',
        'label_ja': '出金',
        'header_columns': [
            {'key': f'h_{i}', 'label': f'H{i}', 'type': 'str', 'max': 50}
            for i in range(1, 21)
        ],
        'detail_columns': [
            {'key': f'd_{i}', 'label': f'D{i}', 'type': 'str', 'max': 50}
            for i in range(1, 4)
        ],
        'tail_columns': [
            {'key': f't_{i}', 'label': f'T{i}', 'type': 'str', 'max': 50}
            for i in range(1, 6)
        ],
        'sheet4_columns': [
            {'key': f'wht_{i}', 'label': f'WHT{i}', 'type': 'str', 'max': 50}
            for i in range(1, 16)
        ],
    },

    # 🟡 会计凭证(stub · TODO 物料 3 · 头 6 + 分录 7 · 预期 2 sheet)
    'journal': {
        'stub': True,
        'sheet_count_expected': 2,
        'label_zh': '会计凭证 · 待物料',
        'label_th': 'สมุดรายวัน · รอตัวอย่าง',
        'label_en': 'Journal · WAIT MATERIAL',
        'label_ja': '仕訳 · 物料待ち',
        # TODO 物料 3 · 会计凭证样本到达后填字段名
        'header_columns': [],
        'detail_columns': [],
    },
}


# ============================================================
# 错误码 → 4 语友好提示(stub · TODO 物料 4)
# ============================================================
MRERP_ERROR_FRIENDLY: Dict[str, Dict[str, str]] = {}


def get_error_friendly(error_code: str, lang: str = 'zh') -> str:
    """物料 4 未到时 fallback · 直接显示原始错误码"""
    entry = MRERP_ERROR_FRIENDLY.get(error_code or "", {})
    return entry.get(lang) or entry.get('en') or (error_code or "")


# ============================================================
# Mapping 拼装(读 v118.27.0 的 3 张表)
# ============================================================
def lookup_customer_code(client_id: int, mappings: Dict[str, Any]) -> str:
    """从 erp_client_mappings 拿当前 client 在 mrerp 下的代码 · 没配返回空"""
    cli = (mappings or {}).get('clients') or []
    for m in cli:
        if m.get('erp_type') == 'mrerp' and int(m.get('client_id') or 0) == int(client_id or 0):
            return fmt_str(m.get('erp_code'), 50)
    return ""


def lookup_account_code(category: str, mappings: Dict[str, Any]) -> str:
    acc = (mappings or {}).get('accounts') or []
    for m in acc:
        if m.get('erp_type') == 'mrerp' and m.get('pearnly_category') == category:
            return fmt_str(m.get('erp_code'), 50)
    return ""


def lookup_tax_code(tax_kind: str, mappings: Dict[str, Any]) -> str:
    """从 erp_tax_mappings 拿税种字符串(铁律 29 · 税率字符串枚举)"""
    tax = (mappings or {}).get('taxes') or []
    for m in tax:
        if m.get('erp_type') == 'mrerp' and m.get('pearnly_tax_kind') == tax_kind:
            return fmt_str(m.get('erp_code'), 14)
    return ""


def derive_tax_kind(history: Dict[str, Any]) -> str:
    """从 OCR 结果推断 Pearnly tax_kind 枚举"""
    rate = history.get('tax_rate_pct') or history.get('vat_rate')
    wht = history.get('wht_rate_pct') or history.get('wht_rate')
    try:
        if wht:
            w = int(float(wht))
            return f'wht_{w}' if w in (1, 3, 5) else 'wht_3'
    except Exception:
        pass
    try:
        if rate is not None:
            r = float(rate)
            if r == 7:
                return 'vat_7'
            if r == 0:
                return 'vat_0'
            if r < 0:
                return 'vat_exempt'
    except Exception:
        pass
    return 'vat_7'


# ============================================================
# 数据装配:从 OCR history 装出 sales_credit 的 row dict
# ============================================================
def build_sales_credit_row(history: Dict[str, Any], mappings: Dict[str, Any]) -> Dict[str, Any]:
    """v27.8.1.4 · 严格对齐 Korn 真实样本 18 列 header
    v27.8.1.5 · invoice_no 转 MR.ERP YYMMDD-NNN 标准格式"""
    cid = history.get('client_id') or 0
    customer_code = lookup_customer_code(cid, mappings)

    # v27.8.1.5 · invoice_no 转 MR.ERP 标准格式(原 OCR 号 'INV-...' 不被认)
    mrerp_invoice_no = derive_mrerp_invoice_no(history)
    inv_date = fmt_date(history.get('invoice_date'))

    return {
        'invoice_no':     mrerp_invoice_no,
        'invoice_date':   inv_date,
        # 'tax_rate_str' / 'branch_code' / 'department' / 'job' / 'salesman'
        # / 'sales_area' / 'shipping_type' 都用 schema default(TEST2019 已知存在的值)
        'delivery_date':  inv_date,  # 通常 = 开票日(没运单时)
        'customer_code':  customer_code,
        'customer_bill':  customer_code,
        'bill_no':        ('SI' + mrerp_invoice_no)[:30],   # 'SI690415-501'
        'bill_date':      inv_date,
        # 'discount' default = 0
        # v27.8.1.6 · 严格对齐 Korn 真样本 · 三个 note 都留空(MR.ERP 可能对备注字段格式有限制)
        'note1':          '',
        'note2':          '',
        'note3':          '',
    }


def build_sales_credit_detail_rows(history: Dict[str, Any], mappings: Dict[str, Any]) -> List[Dict[str, Any]]:
    """v27.8.1.4 · OCR items → detail rows · v27.8.1.5 · invoice_no 同步用 MR.ERP 标准格式
    v27.8.1.17 · 加 product_code · 从 mappings['products'] 查 item_name → erp_code · 找不到 fallback '123'(下游 korn_clone 已 fallback)
    """
    mrerp_invoice_no = derive_mrerp_invoice_no(history)

    # v27.8.1.17 · 建商品 lookup 一次（O(N)）· 内循环 O(1) 查
    product_lookup = _build_product_lookup(mappings)

    items = []
    for src_field in (
        history.get('items'),
        (history.get('fields') or {}).get('items') if isinstance(history.get('fields'), dict) else None,
    ):
        if isinstance(src_field, list) and src_field:
            items = src_field
            break
    if not items:
        pages = history.get('pages')
        if isinstance(pages, list):
            for p in pages:
                if not isinstance(p, dict):
                    continue
                pf = p.get('fields') or {}
                if isinstance(pf, dict):
                    pi = pf.get('items')
                    if isinstance(pi, list) and pi:
                        items = pi
                        break

    rows = []
    if items:
        for it in items:
            if not isinstance(it, dict):
                continue
            qty = fmt_number(it.get('qty') or it.get('quantity'))
            price = fmt_number(it.get('unit_price') or it.get('price'))
            amt = fmt_number(it.get('amount') or it.get('total'))
            if qty is None and price is not None and amt is not None and price != 0:
                qty = round(amt / price, 4)
            if price is None and qty is not None and amt is not None and qty != 0:
                price = round(amt / qty, 4)
            if amt is None and qty is not None and price is not None:
                amt = round(qty * price, 2)
            # v27.8.1.17 · 查商品映射 · 找不到返 None · 下游 korn_clone 兜底 '123'
            item_name = it.get('name') or it.get('description') or ''
            erp_code = _resolve_product_code(item_name, product_lookup)
            rows.append({
                'invoice_no':   mrerp_invoice_no,
                'qty':          qty if qty is not None else 1,
                'unit_price':   price if price is not None else 0,
                'amount':       amt if amt is not None else 0,
                'product_code': erp_code,
                'item_name':    str(item_name or ''),
            })
    if not rows:
        sub = fmt_number(history.get('subtotal') or history.get('amount_before_tax'))
        tot = fmt_number(history.get('total_amount'))
        if sub is None and tot is not None:
            sub = round(tot / 1.07, 2)
        unit = sub or tot or 0
        rows.append({
            'invoice_no':   mrerp_invoice_no,
            'qty':          1,
            'unit_price':   unit,
            'amount':       unit,
            'product_code': None,
            'item_name':    '',
        })
    return rows


def build_sales_credit_tail_row(history: Dict[str, Any]) -> Dict[str, Any]:
    """v27.8.1.4 · tail · v27.8.1.5 同步 invoice_no 格式"""
    mrerp_invoice_no = derive_mrerp_invoice_no(history)
    return {
        'invoice_no':     mrerp_invoice_no,
        'deposit_no':     '',
        'is_sales_issued': '',
    }


# ============================================================
# 验证:发票装得出有效 sales_credit 行吗?
# ============================================================
def validate_history_for_sales_credit(
    history: Dict[str, Any],
    mappings: Dict[str, Any],
) -> Tuple[bool, Optional[str], List[str]]:
    """Preflight check before sending a history through the MR.ERP sales_credit
    upload. Returns (ok, error_code, warnings).

    error_code (when ok=False) — one of:
        ERR_NO_HISTORY            history dict empty
        ERR_NO_CLIENT             client_id missing
        ERR_NO_CUSTOMER_MAPPING   erp_client_mappings has no mrerp row
        ERR_NO_INVOICE_NO         neither invoice_no nor invoice_number set
        ERR_NO_INVOICE_DATE       invoice_date missing
        ERR_NO_TOTAL_AMOUNT       total_amount missing or 0
        ERR_NEGATIVE_AMOUNT       total_amount < 0  (P1-A §3.3)
        ERR_INVOICE_NO_TOO_LONG   derived invoice_no > 18 chars  (P1-A §3.1)
        ERR_BILL_NO_TOO_LONG      SI+invoice_no > 20 chars       (P1-A §3.1)
        ERR_CUSTOMER_CODE_TOO_LONG       customer_code > 20 chars  (P1-A §3.1)
        ERR_CUSTOMER_BILL_TOO_LONG       customer_bill > 20 chars  (P1-A §3.1)
        ERR_TAX_RATE_INVALID      derive_tax_kind ∉ MRERP_VALID_TAX_KINDS_SC  (P1-A §3.2)
        ERR_DATE_FUTURE           invoice_date > today + 30 days  (P1-A §3.5)

    warnings (non-blocking; caller can push these into the exceptions table):
        WARN_DATE_NEAR_FUTURE     invoice_date in (today + 7d, today + 30d]
        WARN_DATE_TOO_OLD         invoice_date < today - 730 days

    The check order intentionally matches MR.ERP's own server-side ordering
    (length → existence → enum → date) so error messages line up.
    """
    if not history:
        return False, 'ERR_NO_HISTORY', []
    cid = history.get('client_id') or 0
    if not cid:
        return False, 'ERR_NO_CLIENT', []
    customer_code = lookup_customer_code(cid, mappings)
    if not customer_code:
        return False, 'ERR_NO_CUSTOMER_MAPPING', []
    if not (history.get('invoice_no') or history.get('invoice_number')):
        return False, 'ERR_NO_INVOICE_NO', []
    if not history.get('invoice_date'):
        return False, 'ERR_NO_INVOICE_DATE', []

    # Amount: strict mode rejects negative + overflow.
    try:
        total = fmt_number_strict(history.get('total_amount'))
    except ValueError as e:
        msg = str(e).lower()
        if 'negative' in msg:
            return False, 'ERR_NEGATIVE_AMOUNT', []
        if 'exceeds' in msg or 'missing' in msg:
            return False, 'ERR_NO_TOTAL_AMOUNT', []
        return False, 'ERR_NO_TOTAL_AMOUNT', []
    if total <= 0:
        return False, 'ERR_NO_TOTAL_AMOUNT', []

    # Length pre-flight (P1-A §3.1).
    invoice_no = derive_mrerp_invoice_no(history)
    if len(invoice_no) > MRERP_INVOICE_NO_MAX:
        return False, 'ERR_INVOICE_NO_TOO_LONG', []
    bill_no = 'SI' + invoice_no
    if len(bill_no) > MRERP_BILL_NO_MAX:
        return False, 'ERR_BILL_NO_TOO_LONG', []
    if len(customer_code) > MRERP_CUSTOMER_CODE_MAX:
        return False, 'ERR_CUSTOMER_CODE_TOO_LONG', []
    # customer_bill defaults to customer_code; allow override via mapping
    # if a future schema introduces it.
    customer_bill = customer_code
    if len(customer_bill) > MRERP_CUSTOMER_BILL_MAX:
        return False, 'ERR_CUSTOMER_BILL_TOO_LONG', []

    # Tax rate enum gate (P1-A §3.2).
    tax_kind = derive_tax_kind(history)
    if tax_kind not in MRERP_VALID_TAX_KINDS_SC:
        return False, 'ERR_TAX_RATE_INVALID', []

    # Date sanity (P1-A §3.5).
    warnings: List[str] = []
    inv_date_str = fmt_date(history.get('invoice_date'))
    parsed_inv: Optional[date] = None
    if inv_date_str and len(inv_date_str) >= 10:
        try:
            parsed_inv = datetime.strptime(inv_date_str[:10], "%Y-%m-%d").date()
        except Exception:
            parsed_inv = None
    if parsed_inv is not None:
        today = date.today()
        delta = (parsed_inv - today).days
        if delta > MRERP_DATE_FUTURE_HARD_REJECT_DAYS:
            return False, 'ERR_DATE_FUTURE', []
        if delta > MRERP_DATE_FUTURE_WARN_DAYS:
            warnings.append('WARN_DATE_NEAR_FUTURE')
        if delta < -MRERP_DATE_PAST_WARN_DAYS:
            warnings.append('WARN_DATE_TOO_OLD')

    return True, None, warnings


# ============================================================
# 内部工具 · 收集 schema 的所有 sheet 列(动态 1-4 sheet)
# ============================================================
def _collect_sheet_groups(schema: Dict[str, Any]) -> List[Tuple[str, List[Dict[str, Any]]]]:
    """
    根据 schema 收集要建的 sheet · 按固定顺序:
      header → Worksheet
      detail → Worksheet 1
      tail   → Worksheet 2
      sheet4 → Worksheet 3
    
    某段为空(None / 空 list)→ 不创建该 sheet。
    返回:[(group_name, columns), ...] · 至少 1 个(否则报错)
    """
    groups: List[Tuple[str, List[Dict[str, Any]]]] = []
    for key in ('header_columns', 'detail_columns', 'tail_columns', 'sheet4_columns'):
        cols = schema.get(key)
        if cols:  # 非空 list / None / [] 都跳过
            group_name = key.replace('_columns', '')
            groups.append((group_name, cols))
    return groups


def _sheet_title(idx: int) -> str:
    """idx=0 → 'Worksheet' · idx>=1 → 'Worksheet N'(空格分隔 · 铁律 28)"""
    if idx == 0:
        return "Worksheet"
    return f"Worksheet {idx}"


# ============================================================
# 主入口:生成 .xlsx bytes(sheet 数动态 · 铁律 28 修订)
# ============================================================
def generate_xlsx(
    histories: List[Dict[str, Any]],
    mappings: Dict[str, Any],
    sheet_kind: str = 'sales_credit',
) -> bytes:
    """
    返回 .xlsx 二进制字节 · 调用方负责发给前端
    
    histories: List of OCR history dicts(1 个文件可多行 · 同 sheet 类型)
    mappings:  {'clients': [...], 'accounts': [...], 'taxes': [...]}(从 db 拿)
    sheet_kind: 9 选 1 · 默认销售-赊销
    
    🔴 铁律 28 修订(v118.27.4.2):
      根据 schema 实际有的 columns 段动态创建 1-4 sheet · 不再硬编码 3 sheet
      命名严格 = `Worksheet` / `Worksheet 1` / `Worksheet 2` / `Worksheet 3`(空格分隔)
    """
    if Workbook is None:
        raise RuntimeError("openpyxl not installed · pip install openpyxl")

    schema = MRERP_SHEET_SCHEMAS.get(sheet_kind)
    if not schema:
        raise ValueError(f"unknown sheet_kind: {sheet_kind}")
    if schema.get('stub') and not schema.get('header_columns'):
        raise RuntimeError(f"sheet_kind={sheet_kind} 物料未到 · 字段未填")

    # sales_credit 走 Korn 真样本克隆路径(100% PhpSpreadsheet 兼容)
    # 详见 docs/integrations/mrerp-known-facts.md §6 xlsx 字节级冷知识
    if sheet_kind == 'sales_credit':
        try:
            return _generate_xlsx_sales_credit_korn_clone(histories, mappings)
        except FileNotFoundError as e:
            logger.warning(f"Korn template missing (fallback to openpyxl): {e}")
        except Exception as e:
            logger.warning(f"Korn clone failed (fallback to openpyxl): {e}")

    groups = _collect_sheet_groups(schema)
    if not groups:
        raise RuntimeError(f"sheet_kind={sheet_kind} 无任何 columns 配置")

    wb = Workbook()
    bold = Font(bold=True)
    sheets: List[Tuple[str, Any, List[Dict[str, Any]]]] = []

    # 动态建 sheet · 第 1 个用 active 改名 · 后续用 create_sheet
    for idx, (group_name, cols) in enumerate(groups):
        title = _sheet_title(idx)
        if idx == 0:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        # 标题行
        for i, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=i, value=col['label'])
            c.font = bold
        sheets.append((group_name, ws, cols))

    # v27.8.1.11 · 空 cell 直接返 None · 让 openpyxl skip cell
    # 后处理时再把缺的 cell 补成完全空 cell `<c r="X#"/>`(对齐 Korn 风格)
    def _safe_val(val, col):
        if val is None or val == '':
            val = col.get('default', None)
        if val is None or val == '':
            return None  # ← skip cell · 后处理插完全空 cell
        return val

    for group_name, ws, cols in sheets:
        if sheet_kind == 'sales_credit' and group_name == 'header':
            for row_idx, history in enumerate(histories, start=2):
                row_data = build_sales_credit_row(history, mappings)
                for col_idx, col in enumerate(cols, start=1):
                    val = _safe_val(row_data.get(col['key']), col)
                    if val is None:
                        continue   # v27.8.1.11 · 完全 skip · 后处理补完全空 cell
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if col['type'] in ('str', 'date'):
                        cell.number_format = '@'
        elif sheet_kind == 'sales_credit' and group_name == 'detail':
            cur_row = 2
            for history in histories:
                detail_rows = build_sales_credit_detail_rows(history, mappings)
                for row_data in detail_rows:
                    for col_idx, col in enumerate(cols, start=1):
                        val = _safe_val(row_data.get(col['key']), col)
                        if val is None:
                            continue
                        cell = ws.cell(row=cur_row, column=col_idx, value=val)
                        if col['type'] in ('str', 'date'):
                            cell.number_format = '@'
                    cur_row += 1
        elif sheet_kind == 'sales_credit' and group_name == 'tail':
            # v27.8.1.11 · Korn 真样本 tail sheet 只 header · 没 data row
            # MR.ERP 期望 tail 是「条件可选」(押金/已开销售单)· 没数据就别写
            pass
        # 其他 sheet_kind:首版只写表头(等 v27.8.2)

    # v27.8.1.11 · spacer 占位改在后处理阶段插「完全空 cell」<c r="S2"/>
    # 这里只生成 18 列 · 后处理时 expand 到 19 列(跟 Korn 真样本完全一致风格)

    buf = io.BytesIO()
    wb.save(buf)
    out = buf.getvalue()

    # v27.8.1.9 + v27.8.1.11 · 关键!后处理:
    # 1. inlineStr → sharedStrings(PhpSpreadsheet 不识别 inline)
    # 2. 缺失 cell 补「完全空 cell <c r='X#'/>」(让 row 显式声明每列存在)
    # 3. row 加 spans 属性(跟 Korn 真样本对齐)
    # 4. sheet1 末尾 col 19 加完全空 cell(让 dim=A1:S2)
    if sheet_kind == 'sales_credit':
        try:
            sheet_col_map = {}
            # idx 对应 sheets 列表 · sheet1=header sheet2=detail sheet3=tail
            for idx, (group_name, _ws, cols) in enumerate(sheets):
                sheet_col_map[f"sheet{idx+1}"] = len(cols)
            out = _convert_inline_to_shared_strings(out, sheet_col_map=sheet_col_map)
        except Exception as e:
            logger.warning(f"_convert_inline_to_shared_strings failed (使用原 xlsx): {e}")

    return out


def _convert_inline_to_shared_strings(xlsx_bytes: bytes, sheet_col_map=None) -> bytes:
    """v27.8.1.9 + v27.8.1.11 · 把 openpyxl 输出转成 MR.ERP / PhpSpreadsheet 兼容格式

    1. inlineStr → sharedStrings(PhpSpreadsheet 不识别 inline)
    2. 缺失的 cell 补「完全空 cell <c r='X#'/>」(跟 Korn 真样本风格一致)
    3. row 加 spans="1:N" 属性(让 PhpSpreadsheet 正确数列)
    4. sheet1 在 row 1/2 加 col 19 完全空 cell(让 dim=A1:S2 跟 Korn 一致)
    5. 去掉 t="n" 属性(Korn 真样本数值 cell 不带 t · default 就是 numeric)

    sheet_col_map: {'sheet1': 18, 'sheet2': 8, 'sheet3': 3} · 各 sheet schema 列数
    """
    import zipfile, re as _re
    from collections import OrderedDict

    src_buf = io.BytesIO(xlsx_bytes)
    files = {}
    with zipfile.ZipFile(src_buf, 'r') as src_zip:
        for name in src_zip.namelist():
            files[name] = src_zip.read(name)

    shared = OrderedDict()

    def _get_idx(text):
        if text not in shared:
            shared[text] = len(shared)
        return shared[text]

    def _decode_xml_entities(s):
        s = (s.replace('&amp;', '&')
              .replace('&lt;', '<')
              .replace('&gt;', '>')
              .replace('&quot;', '"')
              .replace('&apos;', "'"))
        s = _re.sub(r'&#(\d+);', lambda m: chr(int(m.group(1))), s)
        s = _re.sub(r'&#x([0-9a-fA-F]+);', lambda m: chr(int(m.group(1), 16)), s)
        return s

    inline_full_re = _re.compile(
        r'<c([^>]*?)\st="inlineStr"([^>]*?)>\s*<is>\s*<t(?:\s[^>]*)?>([^<]*)</t>\s*</is>\s*</c>',
        _re.DOTALL,
    )

    def _replace_full(m):
        pre, post, text = m.group(1), m.group(2), m.group(3)
        text = _decode_xml_entities(text)
        idx = _get_idx(text)
        return f'<c{pre} t="s"{post}><v>{idx}</v></c>'

    inline_empty_re = _re.compile(
        r'<c([^>]*?)\st="inlineStr"([^>]*?)\s*(?:/>|>\s*</c>|>\s*<is\s*/>\s*</c>|>\s*<is>\s*</is>\s*</c>)',
        _re.DOTALL,
    )

    def _replace_empty(m):
        pre, post = m.group(1), m.group(2)
        # v27.8.1.11 · 直接输出完全空 cell(无 t 无 v · 跟 Korn 风格一致)
        return f'<c{pre}{post}/>'

    # 去掉数值 cell 的 t="n" 属性(Korn 不带 t)
    numeric_t_re = _re.compile(r'<c([^>]*?)\st="n"([^>]*?)>(<v>[^<]*</v>)</c>')
    def _strip_numeric_t(m):
        pre, post, vtag = m.group(1), m.group(2), m.group(3)
        return f'<c{pre}{post}>{vtag}</c>'

    def _col_letter(n):
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    def _process_sheet_xml(xml, sheet_name):
        # 阶段 1:cell 类型转换
        xml = inline_full_re.sub(_replace_full, xml)
        xml = inline_empty_re.sub(_replace_empty, xml)
        xml = numeric_t_re.sub(_strip_numeric_t, xml)

        col_target = (sheet_col_map or {}).get(sheet_name)
        if not col_target:
            return xml

        # sheet1 末尾 +1 占位 cell(对齐 Korn dim=A1:S2)
        if sheet_name == 'sheet1':
            col_target_total = col_target + 1
        else:
            col_target_total = col_target

        # 阶段 2:每个 row 补缺失 cell + 加 spans
        def _process_row(rm):
            row_attr = rm.group(1)
            row_inner = rm.group(2)
            rn_match = _re.search(r'r="(\d+)"', row_attr)
            if not rn_match:
                return rm.group(0)
            row_num = int(rn_match.group(1))
            existing_cols = set()
            for cm in _re.finditer(r'<c\s+r="([A-Z]+)' + str(row_num) + r'"', row_inner):
                existing_cols.add(cm.group(1))
            missing = []
            for c in range(1, col_target_total + 1):
                letter = _col_letter(c)
                if letter not in existing_cols:
                    missing.append(f'<c r="{letter}{row_num}"/>')
            if missing:
                row_inner = row_inner + ''.join(missing)
            if 'spans=' not in row_attr:
                row_attr = row_attr + f' spans="1:{col_target_total}"'
            return f'<row{row_attr}>{row_inner}</row>'

        xml = _re.sub(r'<row([^>]*)>(.*?)</row>', _process_row, xml, flags=_re.DOTALL)

        # 阶段 3:重算 dimension
        new_dim_letter = _col_letter(col_target_total)
        max_row = 1
        for rm in _re.finditer(r'<row r="(\d+)"', xml):
            rn = int(rm.group(1))
            if rn > max_row:
                max_row = rn
        new_dim = f'A1:{new_dim_letter}{max_row}'
        xml = _re.sub(r'<dimension ref="[^"]+"', f'<dimension ref="{new_dim}"', xml)
        return xml

    for name in list(files.keys()):
        if name.startswith('xl/worksheets/sheet') and name.endswith('.xml'):
            xml = files[name].decode('utf-8')
            sheet_name = name.split('/')[-1].replace('.xml', '')
            xml = _process_sheet_xml(xml, sheet_name)
            files[name] = xml.encode('utf-8')

    if shared:
        def _xml_escape(s):
            return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
        parts = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
                 f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
                 f'count="{len(shared)}" uniqueCount="{len(shared)}">']
        for text in shared:
            parts.append(f'<si><t xml:space="preserve">{_xml_escape(text)}</t></si>')
        parts.append('</sst>')
        files['xl/sharedStrings.xml'] = ''.join(parts).encode('utf-8')

        ct = files['[Content_Types].xml'].decode('utf-8')
        if 'sharedStrings.xml' not in ct:
            ct = ct.replace(
                '</Types>',
                '<Override PartName="/xl/sharedStrings.xml" '
                'ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/>'
                '</Types>'
            )
            files['[Content_Types].xml'] = ct.encode('utf-8')

        rels_path = 'xl/_rels/workbook.xml.rels'
        if rels_path in files:
            rels = files[rels_path].decode('utf-8')
            if 'sharedStrings.xml' not in rels:
                rid_nums = [int(m.group(1)) for m in _re.finditer(r'Id="rId(\d+)"', rels)]
                new_rid = (max(rid_nums) if rid_nums else 0) + 1
                rels = rels.replace(
                    '</Relationships>',
                    f'<Relationship Id="rId{new_rid}" '
                    f'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                    f'Target="sharedStrings.xml"/></Relationships>'
                )
                files[rels_path] = rels.encode('utf-8')

    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as out_zip:
        for name, data in files.items():
            out_zip.writestr(name, data)
    return out_buf.getvalue()


def _generate_xlsx_sales_credit_korn_clone(histories: List[Dict[str, Any]],
                                            mappings: Dict[str, Any]) -> bytes:
    """v27.8.1.12 · 用 Korn 真样本作模板 · 克隆方式生成 sales_credit xlsx
    
    根因:openpyxl 输出 workbook.xml / [Content_Types].xml 跟 PhpSpreadsheet 期望差异大
    解法:克隆 Korn 真样本(已验证可 import) · 只重写 sharedStrings + sheetData
    保留 metadata:workbook.xml / styles.xml / theme.xml / [Content_Types].xml 等全部不动
    """
    import zipfile, os, re as _re
    from collections import OrderedDict

    template_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        'test_data_mrerp_sample_SC.xlsx'
    )
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Korn template missing: {template_path}")

    with open(template_path, 'rb') as f:
        template_bytes = f.read()

    files: Dict[str, bytes] = {}
    with zipfile.ZipFile(io.BytesIO(template_bytes), 'r') as zf:
        for name in zf.namelist():
            files[name] = zf.read(name)

    shared: "OrderedDict[str, int]" = OrderedDict()

    def _get_idx(text: str) -> int:
        text = str(text) if text is not None else ''
        if text not in shared:
            shared[text] = len(shared)
        return shared[text]

    def _col_letter(n: int) -> str:
        s = ""
        while n > 0:
            n, r = divmod(n - 1, 26)
            s = chr(65 + r) + s
        return s

    def _xml_escape(s: str) -> str:
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # ─── Sheet 1 (header) ──────────────────────────────────────────
    HEADERS_1 = ['เลขที่', 'วันที่', 'อัตราภาษี', 'สาขา', 'แผนก', 'งาน',
                 'พนักงานขาย', 'กำหนดส่งสินค้า', 'รหัสลูกค้า', 'รหัสลูกค้า (บิล)',
                 'เลขที่บิล', 'วันที่', 'พื้นที่การขาย', 'ประเภทขนส่ง',
                 'หักส่วนลด', 'หมายเหตุ 1', 'หมายเหตุ 2', 'หมายเหตุ 3']
    row1_cells = []
    for i, lbl in enumerate(HEADERS_1, 1):
        row1_cells.append(
            f'<c r="{_col_letter(i)}1" s="2" t="s"><v>{_get_idx(lbl)}</v></c>'
        )
    rows_xml = [
        f'<row r="1" spans="1:19" ht="23.1" customHeight="1" '
        f'x14ac:dyDescent="0.2">{"".join(row1_cells)}</row>'
    ]

    for ridx, history in enumerate(histories, start=2):
        row_data = build_sales_credit_row(history, mappings)
        invoice_no = row_data.get('invoice_no', '')
        invoice_date = row_data.get('invoice_date', '')
        bill_no = row_data.get('bill_no', '')
        cust_code = row_data.get('customer_code', '')
        STR_VALUES = [
            invoice_no, invoice_date, '7 (แยก)', '00000', 'BOI1', '00002',
            'กร ทดสอบ', invoice_date, cust_code, cust_code, bill_no,
            invoice_date, 'สุพรรณบุรี', 'ขนส่งโดยบริษัท',
        ]
        cells = []
        for i, val in enumerate(STR_VALUES, 1):
            cells.append(
                f'<c r="{_col_letter(i)}{ridx}" s="3" t="s">'
                f'<v>{_get_idx(val)}</v></c>'
            )
        # 第 15 列(O)= 折扣 0(数值 · 不带 t)
        cells.append(f'<c r="O{ridx}" s="5"><v>0</v></c>')
        # 第 16-18 列(P/Q/R)= 备注完全空 cell
        for i in (16, 17, 18):
            cells.append(f'<c r="{_col_letter(i)}{ridx}" s="3"/>')
        # 第 19 列(S)= 末尾 spacer 完全空 cell(对齐 Korn dim=A1:S2)
        cells.append(f'<c r="S{ridx}" s="6"/>')
        rows_xml.append(
            f'<row r="{ridx}" spans="1:19" ht="23.1" customHeight="1" '
            f'x14ac:dyDescent="0.2">{"".join(cells)}</row>'
        )

    new_sheet_data = '<sheetData>' + ''.join(rows_xml) + '</sheetData>'
    s1 = files['xl/worksheets/sheet1.xml'].decode('utf-8')
    s1 = _re.sub(r'<sheetData>.+?</sheetData>', new_sheet_data, s1, flags=_re.DOTALL)
    s1 = _re.sub(r'<dimension ref="[^"]+"',
                 f'<dimension ref="A1:S{1 + len(histories)}"', s1)
    files['xl/worksheets/sheet1.xml'] = s1.encode('utf-8')

    # ─── Sheet 2 (detail) ──────────────────────────────────────────
    HEADERS_2 = ['เลขที่', 'รหัสสินค้า', 'แผนก', 'งาน', 'คลัง',
                 'จำนวน', 'ราคา/หน่วย', 'จำนวนเงิน']
    rows2 = []
    h_cells = []
    for i, lbl in enumerate(HEADERS_2, 1):
        h_cells.append(
            f'<c r="{_col_letter(i)}1" s="2" t="s"><v>{_get_idx(lbl)}</v></c>'
        )
    rows2.append(
        f'<row r="1" spans="1:8" ht="23.1" customHeight="1" '
        f'x14ac:dyDescent="0.2">{"".join(h_cells)}</row>'
    )

    cur_row = 2
    total_detail_rows = 0
    for history in histories:
        invoice_no = derive_mrerp_invoice_no(history)
        detail_rows = build_sales_credit_detail_rows(history, mappings)
        for row_data in detail_rows:
            qty = row_data.get('qty', 0) or 0
            unit_price = row_data.get('unit_price', 0) or 0
            amount = row_data.get('amount', 0) or 0
            product_code = row_data.get('product_code') or '123'
            cells = [
                f'<c r="A{cur_row}" s="3" t="s"><v>{_get_idx(invoice_no)}</v></c>',
                f'<c r="B{cur_row}" s="3" t="s"><v>{_get_idx(product_code)}</v></c>',
                f'<c r="C{cur_row}" s="3" t="s"><v>{_get_idx("BOI1")}</v></c>',
                f'<c r="D{cur_row}" s="3" t="s"><v>{_get_idx("00002")}</v></c>',
                f'<c r="E{cur_row}" s="3" t="s"><v>{_get_idx("0000")}</v></c>',
                f'<c r="F{cur_row}" s="5"><v>{_format_num(qty)}</v></c>',
                f'<c r="G{cur_row}" s="5"><v>{_format_num(unit_price)}</v></c>',
                f'<c r="H{cur_row}" s="5"><v>{_format_num(amount)}</v></c>',
            ]
            rows2.append(
                f'<row r="{cur_row}" spans="1:8" ht="23.1" customHeight="1" '
                f'x14ac:dyDescent="0.2">{"".join(cells)}</row>'
            )
            cur_row += 1
            total_detail_rows += 1

    if total_detail_rows == 0:
        # 至少留 header · 否则 MR.ERP 可能拒
        total_detail_rows = 0  # row 1 only

    new_sheet2_data = '<sheetData>' + ''.join(rows2) + '</sheetData>'
    s2 = files['xl/worksheets/sheet2.xml'].decode('utf-8')
    s2 = _re.sub(r'<sheetData>.+?</sheetData>', new_sheet2_data, s2, flags=_re.DOTALL)
    s2_max_row = 1 + total_detail_rows
    s2 = _re.sub(r'<dimension ref="[^"]+"',
                 f'<dimension ref="A1:H{max(s2_max_row, 1)}"', s2)
    files['xl/worksheets/sheet2.xml'] = s2.encode('utf-8')

    # ─── Sheet 3 (tail) · 只 header(跟 Korn 真样本完全一致) ──────
    HEADERS_3 = ['เลขที่', 'เลขที่เงินมัดจำ', 'ออกใบขาย']
    h3_cells = []
    for i, lbl in enumerate(HEADERS_3, 1):
        h3_cells.append(
            f'<c r="{_col_letter(i)}1" s="2" t="s"><v>{_get_idx(lbl)}</v></c>'
        )
    new_sheet3_data = (
        f'<sheetData><row r="1" spans="1:3" ht="23.1" customHeight="1" '
        f'x14ac:dyDescent="0.2">{"".join(h3_cells)}</row></sheetData>'
    )
    s3 = files['xl/worksheets/sheet3.xml'].decode('utf-8')
    s3 = _re.sub(r'<sheetData>.+?</sheetData>', new_sheet3_data, s3, flags=_re.DOTALL)
    s3 = _re.sub(r'<dimension ref="[^"]+"', '<dimension ref="A1:C1"', s3)
    files['xl/worksheets/sheet3.xml'] = s3.encode('utf-8')

    # ─── 重写 sharedStrings.xml ────────────────────────────────────
    parts = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n',
             f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
             f'count="{len(shared)}" uniqueCount="{len(shared)}">']
    for text in shared:
        parts.append(f'<si><t xml:space="preserve">{_xml_escape(text)}</t></si>')
    parts.append('</sst>')
    files['xl/sharedStrings.xml'] = ''.join(parts).encode('utf-8')

    # ─── 重新打包(保留 Korn 所有其他文件不动)──────────────────
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        for name, data in files.items():
            zf.writestr(name, data)
    return out_buf.getvalue()


def _format_num(n):
    """数值格式化 · 整数不要小数点(跟 Korn 风格一致)"""
    try:
        f = float(n)
        if f == int(f):
            return str(int(f))
        return repr(f)
    except (ValueError, TypeError):
        return '0'


def make_filename(sheet_kind: str, history_id: str) -> str:
    """
    .xlsx 文件名 · stub sheet 加 -DRAFT 提醒用户字段是猜测
    例:Pearnly_MRERP_sales_credit_20260510_a1b2c3.xlsx
    """
    schema = MRERP_SHEET_SCHEMAS.get(sheet_kind, {})
    is_stub = bool(schema.get('stub'))
    today = datetime.utcnow().strftime("%Y%m%d")
    sid = (history_id or "")[:8]
    suffix = "-DRAFT" if is_stub else ""
    return f"Pearnly_MRERP_{sheet_kind}_{today}_{sid}{suffix}.xlsx"
