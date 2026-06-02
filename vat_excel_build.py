# -*- coding: utf-8 -*-
"""VAT 对账 4-sheet Excel 生成 · 编排器 + Sheet1/2/4 builder · vat_excel_export 拆分。

样式 → vat_excel_styles · Sheet3(一对一对账)→ vat_excel_sheet3 · sheet body 0 逻辑改。"""

import io
import logging
from typing import List, Dict, Any, Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from vat_recon_core import _derive_period
from vat_ocr_extract import _ocr_validate_invoice
from vat_excel_i18n import _I18N
from vat_excel_sheet3 import _build_sheet3
from vat_excel_styles import (
    FONT_NAME,
    F_NORM,
    F_BOLD,
    F_TITLE,
    FILL_SUM,
    FILL_OK,
    FILL_MISS,
    BORDER_TH,
    AL_R,
    AL_L,
    ROW_HEIGHT,
    _style_header,
    _zebra,
)

logger = logging.getLogger(__name__)


def _build_sheet1(wb, invoices, L):
    # ════════════ Sheet 1 · 发票明细 ════════════
    ws1 = wb.active
    ws1.title = L["sh1"]
    ws1.sheet_properties.tabColor = "2563EB"  # Tab 蓝
    headers1 = L["h_inv"]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1))

    # v4.10.22 · OCR 校验色
    FILL_OCR_WARN = FILL_MISS  # 有问题 → 红底(复用)
    FILL_OCR_OK = FILL_OK  # 全通过 → 绿底(复用)
    F_OCR_WARN = Font(name=FONT_NAME, size=10, color="DC2626")
    F_OCR_OK = Font(name=FONT_NAME, size=10, color="16A34A", bold=True)

    # 校验问题 key → 发票明细列号对应关系
    _WARN_COL = {
        "w_invoice_no_empty": 4,
        "w_buyer_name_empty": 3,
        "w_tax_id_bad_length": 2,
        "w_date_parse_fail": 5,
        "w_total_zero": 9,
        "w_vat_rate_mismatch": 8,
        "w_amount_sum_mismatch": 9,
    }

    # v4.10.22 · 先收集每行 OCR 校验结果(稍后在样式循环后应用)
    _inv_ocr_warns: List[List[str]] = []
    for i, inv in enumerate(invoices, 1):
        # Bug 1 · 期间降级
        period_val = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")
        warn_keys = _ocr_validate_invoice(inv)
        _inv_ocr_warns.append(warn_keys)
        warn_text = " · ".join(L.get(k, k) for k in warn_keys)
        ws1.append(
            [
                i,
                inv.get("buyer_tax_id") or "",
                inv.get("buyer_name") or "",
                inv.get("invoice_no") or "",
                inv.get("invoice_date") or "",
                period_val,
                inv.get("amount_pre_vat") or 0,
                inv.get("vat_amount") or 0,
                inv.get("total_amount") or 0,
                inv.get("filename") or "",
                warn_text if warn_text else L.get("ocr_ok", "✓ OK"),
            ]
        )

    # 合计行
    if invoices:
        last1 = len(invoices) + 1
        sum_row = last1 + 1
        ws1.cell(row=sum_row, column=1, value=L["sum"]).font = F_BOLD
        for col in (7, 8, 9):
            cell = ws1.cell(
                row=sum_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last1})",
            )
            cell.font = F_BOLD
            cell.fill = FILL_SUM
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=sum_row, column=c).fill = FILL_SUM
            ws1.cell(row=sum_row, column=c).border = BORDER_TH
            ws1.cell(row=sum_row, column=c).font = F_BOLD

    # 列宽 + 数字格式 + 斑马 + 行高(先统一设 F_NORM · 后面 OCR pass 再覆盖)
    widths1 = [5, 18, 28, 18, 13, 10, 14, 14, 14, 28, 30]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, len(invoices) + 2):
        ws1.row_dimensions[r].height = ROW_HEIGHT
        for c in range(1, len(headers1) + 1):
            cell = ws1.cell(row=r, column=c)
            cell.font = F_NORM
            cell.border = BORDER_TH
        for col in (7, 8, 9):
            ws1.cell(row=r, column=col).alignment = AL_R
            ws1.cell(row=r, column=col).number_format = "#,##0.00"
        ws1.cell(row=r, column=11).alignment = AL_L
    _zebra(ws1, 2, len(invoices) + 1, len(headers1))

    # v4.10.22 · OCR 高亮 pass(在通用样式之后应用 · 确保覆盖 F_NORM)
    for i, warn_keys in enumerate(_inv_ocr_warns, 1):
        data_row = i + 1
        ocr_cell = ws1.cell(row=data_row, column=11)
        if warn_keys:
            ocr_cell.fill = FILL_OCR_WARN
            ocr_cell.font = F_OCR_WARN
            for wk in warn_keys:
                col = _WARN_COL.get(wk)
                if col:
                    ws1.cell(row=data_row, column=col).fill = FILL_OCR_WARN
        else:
            ocr_cell.fill = FILL_OCR_OK
            ocr_cell.font = F_OCR_OK

    ws1.freeze_panes = "A2"


def _build_sheet2(wb, report_rows, period_year, period_month, L):
    # ════════════ Sheet 2 · VAT 报告明细 ════════════
    ws2 = wb.create_sheet(L["sh2"])
    ws2.sheet_properties.tabColor = "16A34A"  # Tab 绿
    headers2 = L["h_rep"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    for i, row in enumerate(report_rows, 1):
        # Bug 1 · 期间降级 · 优先用 report_date 算 · fallback period_year/month
        date_str = row.get("report_date") or ""
        period_val = _derive_period(date_str, "")
        if not period_val and period_year and period_month:
            period_val = f"{period_month:02d}/{period_year}"
        pre = row.get("report_amount_pre_vat") or row.get("report_amount") or 0
        vat = row.get("report_vat_amount") or 0
        try:
            total = round(float(pre or 0) + float(vat or 0), 2)
        except Exception:
            total = 0
        ws2.append(
            [
                i,
                row.get("report_buyer_tax_id") or "",
                row.get("report_buyer_name") or "",
                date_str,
                period_val,
                pre,
                vat,
                total,
            ]
        )

    if report_rows:
        last2 = len(report_rows) + 1
        sum_row2 = last2 + 1
        ws2.cell(row=sum_row2, column=1, value=L["sum"]).font = F_BOLD
        for col in (6, 7, 8):  # Sheet 2 数字列 偏移 1(加了 日期 列)
            cell = ws2.cell(
                row=sum_row2,
                column=col,
                value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last2})",
            )
            cell.font = F_BOLD
            cell.fill = FILL_SUM
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=sum_row2, column=c).fill = FILL_SUM
            ws2.cell(row=sum_row2, column=c).border = BORDER_TH
            ws2.cell(row=sum_row2, column=c).font = F_BOLD

    widths2 = [5, 18, 28, 13, 12, 14, 14, 14]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, len(report_rows) + 2):
        ws2.row_dimensions[r].height = ROW_HEIGHT
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.font = F_NORM
            cell.border = BORDER_TH
        for col in (6, 7, 8):
            ws2.cell(row=r, column=col).alignment = AL_R
            ws2.cell(row=r, column=col).number_format = "#,##0.00"
    _zebra(ws2, 2, len(report_rows) + 1, len(headers2))
    ws2.freeze_panes = "A2"


def _build_sheet4(wb, lang, L):
    # ════════════ Sheet 4 · 使用说明 ════════════
    ws4 = wb.create_sheet(L["sh4"])
    ws4.sheet_properties.tabColor = "6B7280"  # Tab 灰
    ws4.column_dimensions["A"].width = 100
    ws4.cell(row=1, column=1, value=L["title"]).font = F_TITLE
    ws4.row_dimensions[1].height = 28
    for i, line in enumerate(L["help_lines"], start=3):
        ws4.cell(row=i, column=1, value=line).font = F_NORM
        ws4.cell(row=i, column=1).alignment = AL_L
        ws4.row_dimensions[i].height = 22
    other_langs = [k for k in _I18N.keys() if k != lang]
    base_row = 3 + len(L["help_lines"]) + 2
    for lk in other_langs:
        for line in _I18N[lk]["help_lines"]:
            c = ws4.cell(row=base_row, column=1, value=line)
            c.font = Font(name=FONT_NAME, size=9, color="6B7280")
            c.alignment = AL_L
            base_row += 1
        base_row += 1


def build_excel(
    invoices: List[Dict[str, Any]],
    report_rows: List[Dict[str, Any]],
    client_name: str = "",
    client_tax_id: str = "",
    period_year: Optional[int] = None,
    period_month: Optional[int] = None,
    lang: str = "th",
):
    """v118.32.4.9.6 · 生成 4-sheet Excel · 返回 (xlsx bytes, task_summary)。"""
    if lang not in _I18N:
        lang = "th"
    L = _I18N[lang]
    wb = openpyxl.Workbook()
    _build_sheet1(wb, invoices, L)
    _build_sheet2(wb, report_rows, period_year, period_month, L)
    n_total, n_ok, n_diff, diff_amount_total, task_rows = _build_sheet3(
        wb, invoices, report_rows, client_name, period_year, period_month, L
    )
    _build_sheet4(wb, lang, L)
    bio = io.BytesIO()
    wb.save(bio)
    task_summary = {
        "n_total": n_total,
        "n_ok": n_ok,
        "n_diff": n_diff,
        "diff_amount_total": round(diff_amount_total, 2),
        "rows": task_rows,
    }
    return bio.getvalue(), task_summary
