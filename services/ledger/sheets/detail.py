# -*- coding: utf-8 -*-
"""明细表:逐行商品 + 按票汇总(含税小计 / 内含 VAT / 净额)。

这是全链公式的**源头**:下游三张表的每一个金额都按票号 SUMIF 回这里的「จำนวนเงิน」列。
所以这一张表是唯一可编辑的数据表 —— 会计改一个数,VAT/净额/分录/分类账/试算平衡当场
全动,这才是会计用 Excel 的方式。

同类工作簿的金标在这里把 VAT 写成了死值,链条断在中间:改了行金额,小计动了、VAT 不动,
试算平衡照样显示「Balanced」。那种自检比没有自检更坏。这里 VAT 一律写成公式。
"""

from __future__ import annotations

from dataclasses import dataclass
from typing import Sequence, Tuple

from openpyxl.utils import get_column_letter

from services.excel import xlsx_style as sty
from services.ledger import xlsx_common as xc
from services.ledger.models import SETTLE_BANK, SETTLE_CASH, SalesDoc

SHEET_DETAIL = "รายละเอียดใบเสร็จ"

COL_INVOICE = 1
COL_DATE = 2
COL_ITEM = 3
COL_QTY = 4
COL_PRICE = 5
COL_AMOUNT = 6
COL_PAYMENT = 7
COL_CUSTOMER = 8
COL_ROW_KEY = 9

_HEADERS: Tuple[str, ...] = (
    "เลขที่ใบเสร็จ",
    "วันที่ (ค.ศ.)",
    "รายการสินค้า",
    "จำนวน",
    "ราคาต่อหน่วย",
    "จำนวนเงิน (บาท)",
    "วิธีชำระเงิน",
    "ลูกค้า",
    "รหัสอ้างอิงระบบ",
)
_WIDTHS: Tuple[int, ...] = (16, 14, 34, 10, 14, 16, 24, 24, 26)

_SUMMARY_TITLE = "สรุปยอดตามใบเสร็จ"
_SUMMARY_HEADERS: Tuple[str, ...] = (
    "เลขที่ใบเสร็จ",
    "ยอดรวมสินค้า (Subtotal)",
    "ภาษีมูลค่าเพิ่ม 7% (รวมใน)",
    "รายได้สุทธิ (ไม่รวม VAT)",
)
_TOTAL_LABEL = "รวมทั้งสิ้น"

# 付款方式印在**每一行**而不是只印首行:会计核对时会删掉核对无误的行,只印首行的话
# 删掉第一行整张票就看不出是现金还是转账了。
_PAYMENT_LABEL = {
    SETTLE_CASH: "ชำระเงินสด",
    SETTLE_BANK: "ชำระโอนเงินผ่านธนาคาร",
}


@dataclass(frozen=True)
class DetailRefs:
    """下游表向明细表取数的唯一入口 —— 谁也不许自己拼明细表的坐标。

    sheet 为空时给的是同表内引用(明细表自己的汇总块用),非空时带表名前缀。
    """

    sheet: str
    invoice_range: str
    amount_range: str

    def gross(self, criteria: str) -> str:
        return xc.sumif(self.invoice_range, criteria, self.amount_range)

    def vat(self, criteria: str) -> str:
        return xc.vat_included(self.gross(criteria))

    def net(self, criteria: str) -> str:
        return f"{self.gross(criteria)}-{self.vat(criteria)}"


def _refs(sheet, last_data: int) -> DetailRefs:
    return DetailRefs(
        sheet=sheet or "",
        invoice_range=xc.abs_range(sheet, COL_INVOICE, xc.FIRST_DATA_ROW, last_data),
        amount_range=xc.abs_range(sheet, COL_AMOUNT, xc.FIRST_DATA_ROW, last_data),
    )


def _write_line(ws, row: int, doc: SalesDoc, line, payment: str) -> None:
    xc.text_cell(ws, row, COL_INVOICE, doc.invoice_number, align=sty.center())
    xc.date_cell(ws, row, COL_DATE, doc.doc_date)
    xc.text_cell(ws, row, COL_ITEM, line.description)
    qty = ws.cell(row=row, column=COL_QTY, value=line.qty)
    sty.style_cell(qty, align=sty.right(), fmt=sty.QTY_FMT)
    xc.money_cell(ws, row, COL_PRICE, line.unit_price)
    # 有单价就写公式,改数量当场生效;没单价时行金额本身就是源数据 —— 写公式等于凭空
    # 造一个我们没从票面读到的单价出来。
    amount = line.amount
    if line.unit_price is not None:
        amount = f"=ROUND({get_column_letter(COL_QTY)}{row}*{get_column_letter(COL_PRICE)}{row},2)"
    xc.money_cell(ws, row, COL_AMOUNT, amount)
    xc.text_cell(ws, row, COL_PAYMENT, payment, align=sty.center())
    xc.text_cell(ws, row, COL_CUSTOMER, doc.customer_name)
    xc.text_cell(ws, row, COL_ROW_KEY, doc.row_key)


def write_detail_sheet(ws, docs: Sequence[SalesDoc], *, title: str, banner: str) -> DetailRefs:
    """写明细表,返回下游取数用的跨表区间引用。"""
    xc.write_title(ws, f"{title} · {SHEET_DETAIL} / Tax Invoice (ABB)", banner=banner)
    xc.write_header(ws, _HEADERS, _WIDTHS)
    xc.hide_column(ws, COL_ROW_KEY)

    row = xc.FIRST_DATA_ROW
    for doc in docs:
        payment = _PAYMENT_LABEL.get(doc.settlement, doc.payment_method_raw or "-")
        for line in doc.lines:
            _write_line(ws, row, doc, line, payment)
            row += 1

    last_data = max(row - 1, xc.FIRST_DATA_ROW)
    _write_summary_block(ws, docs, _refs(None, last_data), start_row=last_data + 2)
    return _refs(SHEET_DETAIL, last_data)


def _write_summary_block(ws, docs: Sequence[SalesDoc], refs: DetailRefs, *, start_row: int) -> None:
    """按票汇总:小计 SUMIF 回上面的明细行,VAT/净额再从小计派生 —— 一条链,不断在中间。"""
    title = ws.cell(row=start_row, column=COL_INVOICE, value=_SUMMARY_TITLE)
    title.font = sty.Font(bold=True)
    header_row = start_row + 1
    xc.write_header(ws, _SUMMARY_HEADERS, (), row=header_row, freeze=False)

    row = header_row + 1
    for doc in docs:
        inv = xc.text_cell(ws, row, 1, doc.invoice_number, align=sty.center())
        xc.money_cell(ws, row, 2, f"={refs.gross(inv.coordinate)}")
        xc.money_cell(ws, row, 3, f"=ROUND(B{row}*{xc.VAT_NUMERATOR}/{xc.VAT_DENOMINATOR},2)")
        xc.money_cell(ws, row, 4, f"=B{row}-C{row}")
        row += 1

    first, last = header_row + 1, row - 1
    xc.text_cell(ws, row, 1, _TOTAL_LABEL, align=sty.center())
    for letter, col in (("B", 2), ("C", 3), ("D", 4)):
        xc.money_cell(ws, row, col, f"=SUM({letter}{first}:{letter}{last})" if docs else 0)
    xc.bold_row(ws, row, (1, 2, 3, 4))
