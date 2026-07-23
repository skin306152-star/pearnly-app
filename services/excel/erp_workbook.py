# -*- coding: utf-8 -*-
"""ERP 复核工作簿:一批料推完 ERP 后,给会计核对并回导的那个文件。

四张表,各守各的合同:
  ขาย       销项 · 每行一个商品行 · 前 12 列是 MR.ERP 公式合同(见 excel_template_th)
  ซื้อ       进项 · 每行一张票    · 沿用 ภ.พ.30 进项税明细的法定列
  รอจำแนก   待判 · 方向判不出/被闸拦下的票 + 原因 · 会计裁决后把行剪到上面两张表里
  สรุป       汇总 · 本批新建的商品/客户/供应商清单 · 只读不回导

粒度差异是真实的,不该硬捏平:销项要拆到商品行(Express 靠它动库存、算成本),
进项是整张票进一个费用科目。

「สรุป」不是装饰:在 Express 里删掉单据【不会】删掉我们为它新建的商品/客户档。
若某行之所以要重做,恰恰是因为品名读错、照错名建了档,那么删单之后垃圾档还在,
改名回导又会再建一个 —— 主档只增不减。这张表就是那份清理清单。
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional, Sequence

try:
    from openpyxl import Workbook
except ImportError:
    Workbook = None  # type: ignore

from services.excel import erp_roundtrip as rt
from services.excel import xlsx_style as sty
from services.excel.excel_template_th import sales_detail_filename, write_sales_sheet

logger = logging.getLogger(__name__)


def _s(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "null", "nan") else s


def _f(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("฿", "").strip())
    except ValueError:
        return None


def _write_grid(
    ws,
    headers: Sequence[str],
    widths: Sequence[int],
    rows: List[Sequence[Any]],
    money_cols: Sequence[int] = (),
    row_fill=None,
) -> int:
    """通用网格写入。money_cols 是 1-based 列号。"""
    sty.write_header_row(ws, headers, widths)
    money = set(money_cols)
    for r_off, values in enumerate(rows):
        row = 2 + r_off
        for c_off, val in enumerate(values, start=1):
            cell = ws.cell(row=row, column=c_off, value=val)
            if c_off in money:
                sty.style_cell(cell, align=sty.right(), fmt=sty.MONEY_FMT, fill=row_fill)
            elif c_off == 1:
                sty.style_cell(cell, align=sty.center(), fill=row_fill)
            else:
                sty.style_cell(cell, align=sty.left(), fill=row_fill)
    return len(rows)


def _first(f: Dict[str, Any], *keys: str) -> Optional[float]:
    """按顺序取第一个能解成数的键。

    OCR 侧(ThaiInvoice)用 subtotal/vat,回导解析侧用 amount_before_vat/vat_amount ——
    两条来源都要喂进同一张表,只认一套键就会漏一半。
    """
    for k in keys:
        v = _f(f.get(k))
        if v is not None:
            return v
    return None


def _purchase_rows(records: List[Dict[str, Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for i, rec in enumerate(records or [], start=1):
        f = (rec or {}).get("merged_fields") or {}
        base = _first(f, "amount_before_vat", "subtotal")
        vat = _first(f, "vat_amount", "vat")
        total = _f(f.get("total_amount"))
        out.append(
            [
                i,
                _s(f.get("date")),
                _s(f.get("invoice_number")),
                _s(f.get("seller_name")),
                _s(f.get("seller_tax")),
                _s(f.get("seller_branch")),
                base,
                vat,
                total,
                _s(f.get("category")),
                *rt.roundtrip_values(
                    party_tax=f.get("seller_tax") or f.get("seller_tax_id"),
                    doc_vat=vat,
                    docnum=f.get("erp_docnum"),
                    party_code=f.get("erp_party_code"),
                    push_status=f.get("push_status"),
                    push_reason=f.get("push_reason"),
                    row_key=rt.encode_row_key(f.get("history_id"), 0),
                ),
            ]
        )
    return out


def _pending_rows(records: List[Dict[str, Any]]) -> List[List[Any]]:
    out: List[List[Any]] = []
    for rec in records or []:
        f = (rec or {}).get("merged_fields") or {}
        out.append(
            [
                _s(f.get("date")),
                _s(f.get("invoice_number")),
                _s(f.get("seller_name")) or _s(f.get("buyer_name")),
                _s(f.get("seller_tax")) or _s(f.get("buyer_tax")),
                _f(f.get("total_amount")),
                _s(rec.get("reason") or f.get("pending_reason")),
                *rt.roundtrip_values(
                    party_tax=f.get("seller_tax") or f.get("buyer_tax"),
                    push_status=f.get("push_status"),
                    push_reason=f.get("push_reason"),
                    row_key=rt.encode_row_key(f.get("history_id"), 0),
                ),
            ]
        )
    return out


_SUMMARY_HEADERS = ("ประเภท", "รหัส", "ชื่อ", "เอกสารที่สร้าง")
_SUMMARY_WIDTHS = (16, 18, 40, 20)
_KIND_LABEL = {"item": "สินค้า/บริการ", "customer": "ลูกค้า", "supplier": "ผู้ขาย"}


def _summary_rows(created: List[Dict[str, Any]]) -> List[List[Any]]:
    return [
        [
            _KIND_LABEL.get(_s(c.get("kind")), _s(c.get("kind")) or "-"),
            _s(c.get("code")),
            _s(c.get("name")),
            _s(c.get("docnum")),
        ]
        for c in (created or [])
    ]


def build_review_workbook(
    *,
    sales: Optional[List[Dict[str, Any]]] = None,
    purchase: Optional[List[Dict[str, Any]]] = None,
    pending: Optional[List[Dict[str, Any]]] = None,
    created_masters: Optional[List[Dict[str, Any]]] = None,
) -> bytes:
    """生成复核工作簿。四张表恒在 —— 空表也写表头。

    空表不省:会计打开发现没有「รอจำแนก」这张表,分不清是"没有待判的"还是"这版没这功能"。
    表在、行数为 0,才是诚实的空态。
    """
    if Workbook is None:
        raise RuntimeError("openpyxl not installed")

    wb = Workbook()
    ws_sales = wb.active
    ws_sales.title = rt.SHEET_SALES
    write_sales_sheet(ws_sales, sales or [])

    ws_pur = wb.create_sheet(rt.SHEET_PURCHASE)
    _write_grid(
        ws_pur,
        [*rt.PURCHASE_HEADERS, *rt.ROUNDTRIP_HEADERS],
        [*rt.PURCHASE_WIDTHS, *rt.ROUNDTRIP_WIDTHS],
        _purchase_rows(purchase or []),
        money_cols=(7, 8, 9),
    )

    ws_pend = wb.create_sheet(rt.SHEET_PENDING)
    _write_grid(
        ws_pend,
        [*rt.PENDING_HEADERS, *rt.ROUNDTRIP_HEADERS],
        [*rt.PENDING_WIDTHS, *rt.ROUNDTRIP_WIDTHS],
        _pending_rows(pending or []),
        money_cols=(5,),
        row_fill=sty.fill_pending(),
    )

    ws_sum = wb.create_sheet(rt.SHEET_SUMMARY)
    _write_grid(ws_sum, _SUMMARY_HEADERS, _SUMMARY_WIDTHS, _summary_rows(created_masters or []))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def review_workbook_filename(prefix: str = "Pearnly_ERP_Review") -> str:
    return sales_detail_filename(prefix)
