# -*- coding: utf-8 -*-
"""表里算出来的数 = 我们要推的数 —— 拿求值器把整份工作簿的公式真算一遍,逐格对权威值。

这一层测试补的是老窟窿:过去所有断言都只扫公式的**字符串**,于是把销项税从内含 7/107
误写成外加 7%,29 个测试全绿(借方 gross、贷方 (gross−vat)+vat 一起变,借贷仍然相等,
试算平衡照样显示 Balanced),会计拿到的是一份内部自洽、自检绿、销项税错 16.71 บาท 的表。

顺带把两个会计一定会做的动作也在这里验:删核对无误的行、在表尾补一件漏识别的商品。
"""

import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.ledger import entries
from services.ledger.accounts import resolve_chart
from services.ledger.entries import AMOUNT_GROSS, AMOUNT_NET, SIDE_DEBIT
from services.ledger.models import parse_sales_docs
from services.ledger.recipes import sales_books
from services.ledger.sheets import detail as detail_sheet
from services.ledger.sheets import journal as journal_sheet
from services.ledger.sheets import ledger as ledger_sheet
from services.ledger.sheets import trial_balance as tb_sheet
from tests.unit._ledger_golden import GOLDEN_GROSS, build_golden, load, record
from tests.unit._xlsx_formula_eval import CircularReference, Evaluator

D = Decimal


def _reload(wb):
    """存盘再打开 —— 会计手上的是文件,不是内存里的对象。"""
    buf = io.BytesIO()
    wb.save(buf)
    return load_workbook(io.BytesIO(buf.getvalue()))


def _data_rows(ws, col: int, first: int = 5):
    """某列从第一数据行起的非空行号。"""
    return [
        r
        for r in range(first, ws.max_row + 1)
        if ws.cell(row=r, column=col).value not in (None, "")
    ]


def _first_spare_row(ws) -> int:
    """明细区第一个空行 —— 会计补漏就补在这儿。"""
    return next(
        r
        for r in range(5, ws.max_row + 2)
        if not ws.cell(row=r, column=detail_sheet.COL_INVOICE).value
    )


def _summary_rows(ws, invoices):
    """按票汇总块的行 —— 汇总行的第二列是公式,明细行的第二列是日期文本,以此分开。"""
    return [
        r
        for r in _data_rows(ws, detail_sheet.COL_INVOICE)
        if str(ws.cell(row=r, column=1).value or "") in invoices
        and str(ws.cell(row=r, column=2).value or "").startswith("=")
    ]


class WorkbookMathTests(unittest.TestCase):
    """四表逐格求值 —— 每一格都要等于分录里那个 Decimal。"""

    def setUp(self):
        self.docs, self.result = build_golden()
        self.wb = load(self.result)
        self.ev = Evaluator(self.wb)
        self.legs = entries.build_legs([d for d in self.docs if d.bookable], resolve_chart())

    def test_detail_summary_block_matches_authoritative_numbers(self):
        ws = self.wb[detail_sheet.SHEET_DETAIL]
        by_invoice = {d.invoice_number: d for d in self.docs}
        rows = _summary_rows(ws, by_invoice)
        self.assertEqual(len(rows), 3)
        for row in rows:
            doc = by_invoice[str(ws.cell(row=row, column=1).value)]
            invoice = doc.invoice_number
            self.assertEqual(self.ev.money(ws.title, f"B{row}"), doc.gross, invoice)
            self.assertEqual(self.ev.money(ws.title, f"C{row}"), doc.vat, invoice)
            self.assertEqual(self.ev.money(ws.title, f"D{row}"), doc.net, invoice)

    def test_every_journal_leg_evaluates_to_its_authoritative_amount(self):
        """一格一格对:VAT 腿算出来的必须就是 result.documents 里那个销项税。"""
        ws = self.wb[journal_sheet.SHEET_JOURNAL]
        rows = _data_rows(ws, journal_sheet.COL_INVOICE)
        self.assertEqual(len(rows), len(self.legs))
        for row, leg in zip(rows, self.legs):
            letter = "E" if leg.side == SIDE_DEBIT else "F"
            other = "F" if letter == "E" else "E"
            got = self.ev.money(ws.title, f"{letter}{row}")
            self.assertEqual(got, leg.amount, f"{leg.doc.invoice_number} {leg.amount_kind} 行{row}")
            self.assertIsNone(ws[f"{other}{row}"].value, "一腿只能占一侧")

    def test_journal_totals_equal_authoritative_totals(self):
        ws = self.wb[journal_sheet.SHEET_JOURNAL]
        total_row = max(_data_rows(ws, journal_sheet.COL_MEMO))
        self.assertEqual(
            self.ev.money(ws.title, f"E{total_row}"), self.result.numbers["debit_total"]
        )
        self.assertEqual(
            self.ev.money(ws.title, f"F{total_row}"), self.result.numbers["credit_total"]
        )
        self.assertEqual(self.ev.money(ws.title, f"E{total_row}"), GOLDEN_GROSS)

    def test_every_ledger_cell_evaluates_to_its_authoritative_amount(self):
        ws = self.wb[ledger_sheet.SHEET_LEDGER]
        expected = [leg.amount for leg in self.legs]
        got = []
        for row in _data_rows(ws, ledger_sheet.COL_INVOICE):
            for col in (ledger_sheet.COL_DEBIT, ledger_sheet.COL_CREDIT):
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    got.append(self.ev.money(ws.title, cell.coordinate))
        self.assertEqual(sorted(got), sorted(expected))

    def test_trial_balance_totals_and_check_cell(self):
        ws = self.wb[tb_sheet.SHEET_TRIAL_BALANCE]
        account_rows = [
            r
            for r in _data_rows(ws, tb_sheet.COL_ACCOUNT)
            if str(ws.cell(row=r, column=2).value or "").startswith("=MAX(")
        ]
        debit = sum((self.ev.money(ws.title, f"B{r}") for r in account_rows), D(0))
        credit = sum((self.ev.money(ws.title, f"C{r}") for r in account_rows), D(0))
        self.assertEqual(debit, self.result.numbers["debit_total"])
        self.assertEqual(credit, self.result.numbers["credit_total"])
        check = [c for c in ws["B"] if isinstance(c.value, str) and c.value.startswith("=IF(")][0]
        self.assertEqual(self.ev.cell(ws.title, check.coordinate), "ถูกต้อง (Balanced)")

    def test_vat_across_the_workbook_is_seven_over_one_o_seven(self):
        """把每一处 VAT 格的求值结果加起来 —— 误写成外加 7% 时这里会高报,金标 238.72。"""
        ws = self.wb[detail_sheet.SHEET_DETAIL]
        rows = _summary_rows(ws, {d.invoice_number for d in self.docs})
        total = sum((self.ev.money(ws.title, f"C{r}") for r in rows), D(0))
        self.assertEqual(total, self.result.numbers["vat_total"])
        journal = self.wb[journal_sheet.SHEET_JOURNAL]
        vat_legs = [
            self.ev.money(journal.title, f"F{row}")
            for row, leg in zip(_data_rows(journal, journal_sheet.COL_INVOICE), self.legs)
            if leg.amount_kind not in (AMOUNT_GROSS, AMOUNT_NET)
        ]
        self.assertEqual(sum(vat_legs, D(0)), self.result.numbers["vat_total"])

    def test_no_circular_reference_anywhere(self):
        for sheet, coord, formula in self.ev.all_formula_cells():
            try:
                self.ev.cell(sheet, coord)
            except CircularReference as exc:
                self.fail(f"{sheet}!{coord} {formula}: {exc}")


class EmptyBatchWorkbookTests(unittest.TestCase):
    """全批进待判 —— 一张能入账的票都没有时,文件也得能打开。"""

    def setUp(self):
        docs = parse_sales_docs(
            [
                record(
                    "02000139",
                    dates={"date_raw": "Time In 30/04/2569 22:40 Time Out 01/05/2569 00:20"},
                ),
                record("02000138", dates={"date": "27/05/2569"}, payment="card"),
            ]
        )
        self.result = sales_books.build(docs, title="SM", period_label="2569-05")
        self.ev = Evaluator(load(self.result))

    def test_journal_total_does_not_sum_itself(self):
        """=SUM(E5:E5) 写在 E5 上 —— Excel/WPS 弹循环引用警告、两格显示 0,会计判「文件坏了」。"""
        self.assertEqual(self.result.numbers["doc_count"], 0)
        for sheet, coord, formula in self.ev.all_formula_cells():
            try:
                self.ev.cell(sheet, coord)
            except CircularReference as exc:
                self.fail(f"{sheet}!{coord} {formula}: {exc}")

    def test_totals_read_zero(self):
        ws = self.ev.wb[journal_sheet.SHEET_JOURNAL]
        total_row = max(r for r in range(5, ws.max_row + 1) if ws.cell(row=r, column=3).value)
        self.assertEqual(self.ev.money(ws.title, f"E{total_row}"), D(0))
        self.assertEqual(self.ev.money(ws.title, f"F{total_row}"), D(0))


class DeleteRowsTests(unittest.TestCase):
    """会计核对的标准动作之一:把核对无误的行删掉。删完不能炸,剩下的仍要算得出来。"""

    def setUp(self):
        self.docs, self.result = build_golden()
        wb = load(self.result)
        wb[detail_sheet.SHEET_DETAIL].delete_rows(5, 3)  # 删掉 02000138 的三行
        self.wb = _reload(wb)
        self.ev = Evaluator(self.wb)

    def test_no_ref_error_anywhere(self):
        for _sheet, coord, formula in self.ev.all_formula_cells():
            self.assertNotIn("#REF", formula, coord)

    def test_deleted_document_falls_to_zero_and_the_rest_keep_their_numbers(self):
        ws = self.wb[journal_sheet.SHEET_JOURNAL]
        legs = entries.build_legs([d for d in self.docs if d.bookable], resolve_chart())
        rows = [r for r in range(5, ws.max_row + 1) if ws.cell(row=r, column=2).value]
        for row, leg in zip(rows, legs):
            letter = "E" if leg.side == SIDE_DEBIT else "F"
            got = self.ev.money(ws.title, f"{letter}{row}")
            expected = D(0) if leg.doc.invoice_number == "02000138" else leg.amount
            self.assertEqual(got, expected, f"{leg.doc.invoice_number} {leg.amount_kind}")


class AddRowsTests(unittest.TestCase):
    """另一个标准动作:补一件漏识别的商品。加在表尾的行必须被三张派生表算进去。"""

    def setUp(self):
        self.docs, self.result = build_golden()
        wb = load(self.result)
        ws = wb[detail_sheet.SHEET_DETAIL]
        self.appended_row = _first_spare_row(ws)
        ws.cell(row=self.appended_row, column=detail_sheet.COL_INVOICE, value="02000138")
        ws.cell(row=self.appended_row, column=detail_sheet.COL_ITEM, value="สินค้าที่ตกหล่น")
        ws.cell(row=self.appended_row, column=detail_sheet.COL_QTY, value=1)
        ws.cell(row=self.appended_row, column=detail_sheet.COL_AMOUNT, value=300)
        self.wb = _reload(wb)
        self.ev = Evaluator(self.wb)

    def test_appended_line_reaches_the_derived_sheets(self):
        """加在 SUMIF 区间之外 = 明细里看得见那 300,派生表当它不存在,而自检格仍然绿。"""
        ws = self.wb[detail_sheet.SHEET_DETAIL]
        summary_row = _summary_rows(ws, {"02000138"})[0]
        self.assertEqual(self.ev.money(ws.title, f"B{summary_row}"), D("789.00"))

        journal = self.wb[journal_sheet.SHEET_JOURNAL]
        gross_row = [
            r
            for r in range(5, journal.max_row + 1)
            if journal.cell(row=r, column=2).value == "02000138"
            and journal.cell(row=r, column=5).value is not None
        ][0]
        self.assertEqual(self.ev.money(journal.title, f"E{gross_row}"), D("789.00"))

    def test_journal_total_picks_up_the_appended_amount(self):
        ws = self.wb[journal_sheet.SHEET_JOURNAL]
        total_row = max(r for r in range(5, ws.max_row + 1) if ws.cell(row=r, column=3).value)
        self.assertEqual(self.ev.money(ws.title, f"E{total_row}"), GOLDEN_GROSS + D("300"))


if __name__ == "__main__":
    unittest.main()
