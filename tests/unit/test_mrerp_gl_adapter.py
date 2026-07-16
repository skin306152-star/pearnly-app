# -*- coding: utf-8 -*-
"""MR.ERP 分类账适配器单测(services/accounting/mrerp_gl_adapter.py · T4b-SM)。

守恒金标:合成 SM 式逐科目分组分类账(标题行→明细行→รวม 小计→空行)经 adapter → GlRow →
Σ借=Σ贷 差 0.00。另覆盖:逐科目分组分流(标题/明细/小计/空行)、佛历日期解析(str 与
datetime 两路)、空科目不产幽灵行、明细无归属如实丢弃、is_mrerp_gl 识别(分类账 xlsx 命中,
试算平衡表 xlsx / csv 不命中)、parse_gl_bytes 路由(MR.ERP xlsx 走本适配器,非 MR.ERP 原路)。
"""

import io
import unittest
from datetime import date, datetime
from decimal import Decimal

import openpyxl

from services.accounting import gl_upload_adapter
from services.accounting import mrerp_gl_adapter as adapter

_HEADER = ["วันที่", "สมุด", "ใบสำคัญ", "คำอธิบาย", "เดบิต", "เครดิต", "ยอดคงเหลือ"]

# 合成 SM 式分类账:6 明细行 · 名对 preset · Σ借=Σ贷=2070.50 · 含一个空科目(仅标题无明细)
_ACCOUNTS = [
    ("1111-01", "เงินสด", [("2569-05-28", "รับ", "FR001", "ขายสด", "1000.50", "0")]),
    ("1161-10", "ภาษีซื้อ", [("2569-05-28", "ซื้อ", "PV001", "ภาษีซื้อ", "70.00", "0")]),
    ("5010-01", "ต้นทุนขาย", [("2569-05-28", "ซื้อ", "PV001", "ต้นทุน", "1000.00", "0")]),
    # datetime 单元格(openpyxl 也会遇到)· 佛历 2569 → 公历 2026
    ("4110-01", "รายได้จากการขาย", [(datetime(2569, 1, 29), "ขาย", "SE001", "ขาย", "0", "934.58")]),
    ("2160-11", "ภาษีขาย", [("2569-05-28", "ขาย", "SE001", "ภาษีขาย", "0", "65.92")]),
    ("1112-01", "เงินฝากธนาคาร", [("2569-05-28", "จ่าย", "PV002", "ธนาคาร", "0", "1070.00")]),
    ("1113-01", "เงินฝาก-ว่าง", []),  # 空科目:标题在,无明细 → 不应产行
]


def _build_ledger(accounts, *, with_header=True) -> bytes:
    """合成 MR.ERP 分类账 xlsx 字节(逐科目分组 · 抬头 3 行 + 表头 + 每科目标题/明细/รวม/空行)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["บริษัท ทดสอบ จำกัด", "", "", "", "", "", ""])
    ws.append(["รายงานสมุดแยกประเภท", "", "", "", "", "", ""])
    ws.append(["วันที่  01/01/2569  ถึง", "", "", "", "", "วันที่  16/07/2569", ""])
    ws.append(["รหัสบัญชี  1111-01  ถึง", "", "", "", "", "", ""])
    if with_header:
        ws.append(list(_HEADER))
    for code, name, details in accounts:
        ws.append([f"{code}  {name}", "", "", "", "", "", "0"])
        td = tc = Decimal("0")
        for dt, book, vch, desc, dr, cr in details:
            ws.append([dt, book, vch, desc, dr, cr, "0"])
            td += Decimal(str(dr))
            tc += Decimal(str(cr))
        if details:
            ws.append(["รวม", "", "", "", str(td), str(tc), ""])
        ws.append(["", "", "", "", "", "", ""])
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _dec_sum(rows, field):
    return sum((Decimal(str(getattr(r, field))) for r in rows), Decimal("0"))


class ConservationGoldenTests(unittest.TestCase):
    def test_debit_equals_credit_zero_diff(self):
        rows, issues = adapter.parse_mrerp_gl_xlsx(_build_ledger(_ACCOUNTS), "gl.xlsx")
        self.assertEqual(issues, [])
        self.assertEqual(len(rows), 6)  # 空科目 1113-01 不产行
        d, c = _dec_sum(rows, "debit"), _dec_sum(rows, "credit")
        self.assertEqual(d, Decimal("2070.50"))
        self.assertEqual(c, Decimal("2070.50"))
        self.assertEqual((d - c).copy_abs(), Decimal("0.00"))


class GroupingAndParseTests(unittest.TestCase):
    def setUp(self):
        self.rows, _ = adapter.parse_mrerp_gl_xlsx(_build_ledger(_ACCOUNTS), "gl.xlsx")

    def test_grouped_account_codes(self):
        cash = self.rows[0]
        self.assertEqual(cash.account_code, "1111-01")
        self.assertEqual((cash.debit, cash.credit), (1000.50, 0.0))
        self.assertEqual(cash.doc_no, "FR001")
        self.assertEqual(cash.description, "ขายสด")
        # 每明细行归到其所属标题科目,不串组。
        codes = [r.account_code for r in self.rows]
        self.assertEqual(codes, ["1111-01", "1161-10", "5010-01", "4110-01", "2160-11", "1112-01"])

    def test_buddhist_date_from_string_and_datetime(self):
        self.assertEqual(self.rows[0].date, date(2026, 5, 28))  # str "2569-05-28"
        sales = next(r for r in self.rows if r.account_code == "4110-01")
        self.assertEqual(sales.date, date(2026, 1, 29))  # datetime(2569,1,29)
        self.assertEqual((sales.debit, sales.credit), (0.0, 934.58))

    def test_empty_account_yields_no_ghost_row(self):
        self.assertNotIn("1113-01", [r.account_code for r in self.rows])


class HonestDropTests(unittest.TestCase):
    def test_detail_before_any_title_dropped(self):
        # 明细日期行出现在任何科目标题之前 → 无归属,如实丢一条 issue,不臆造。
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(list(_HEADER))
        ws.append(["2569-05-28", "รับ", "X001", "โผล่มาก่อนหัวบัญชี", "10.00", "0"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        rows, issues = adapter.parse_mrerp_gl_xlsx(buf.getvalue(), "bad.xlsx")
        self.assertEqual(rows, [])
        self.assertEqual(len(issues), 1)
        self.assertIn("无所属科目", issues[0])


class DetectionAndRoutingTests(unittest.TestCase):
    def test_is_mrerp_gl_true_for_ledger(self):
        self.assertTrue(adapter.is_mrerp_gl(_build_ledger(_ACCOUNTS), "gl.xlsx"))

    def test_is_mrerp_gl_false_for_trial_balance(self):
        # 试算平衡表(รหัสบัญชี 抬头的纯表)无 วันที่/สมุด → 不误吞。
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["รายงานงบทดลอง", "", "", "", "", ""])
        ws.append(["รหัสบัญชี", "ชื่อบัญชี", "ยอดยกมา", "เดบิต", "เครดิต", "ยอดคงเหลือ"])
        ws.append(["1111-01", "เงินสด", "0", "12900", "0", "12900"])
        buf = io.BytesIO()
        wb.save(buf)
        wb.close()
        self.assertFalse(adapter.is_mrerp_gl(buf.getvalue(), "tb.xlsx"))

    def test_is_mrerp_gl_false_for_csv_ext(self):
        self.assertFalse(adapter.is_mrerp_gl(_build_ledger(_ACCOUNTS), "gl.csv"))

    def test_parse_gl_bytes_routes_mrerp_xlsx(self):
        out = gl_upload_adapter.parse_gl_bytes(_build_ledger(_ACCOUNTS), "gl.xlsx")
        self.assertEqual(len(out["rows"]), 6)
        self.assertEqual(out["rows"][0].account_code, "1111-01")
        self.assertEqual(out["rows"][0].source_file, "gl.xlsx")


class AccountTitleExtractionTests(unittest.TestCase):
    def test_titles_extracted_including_empty_account(self):
        titles = adapter.iter_account_titles(_build_ledger(_ACCOUNTS))
        codes = [t["code"] for t in titles]
        self.assertEqual(len(titles), 7)  # 含空科目 1113-01
        self.assertIn("1113-01", codes)
        cash = next(t for t in titles if t["code"] == "1111-01")
        self.assertEqual(cash["name_th"], "เงินสด")


if __name__ == "__main__":
    unittest.main()
