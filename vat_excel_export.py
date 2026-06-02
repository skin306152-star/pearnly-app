# -*- coding: utf-8 -*-
"""
v118.32.4.10.1 · Pearnly · Excel 公式对账模块(全网开放)
完全独立于 vat_report_parser / reconciliation_matcher

设计哲学(Zihao 2026-05-13 拍板):
  AI 只抽字段 · 不做对账判定 · 让 Excel 公式让用户/Excel 自己核对
  A/B 测试是否比"AI 全程对账"更稳更便宜

调用方式:
  from vat_excel_export import extract_invoice_fields, merge_vat_reports, build_excel
"""

import io
import os
import json
import logging
from typing import List, Dict, Any, Optional, Tuple
from concurrent.futures import ThreadPoolExecutor, TimeoutError as FuturesTimeout

from vat_report_parser import parse_vat_report  # 复用但不修改
from field_comparator import (  # noqa: F401  parse_date 直用 · 余 re-export (safety-net 测试)
    parse_date,
    normalize_invoice_no,
    normalize_tax_id,
    normalize_str,
    tax_id_fuzzy_distance,
)

# 对账核心 · moved to vat_recon_core.py
from vat_recon_core import (  # noqa: F401  re-export (tests) + facade-internal(build_excel/OCR/merge)
    _to_float,
    _derive_period,
    _dominant_report_period,
    _eq_amount,
    _get_inv_total,
    _get_rep_total,
    _build_recon_pairs,
    _diff_dims,
)

logger = logging.getLogger(__name__)
MODULE_VERSION = "1.5.0"  # v4.10.23 · Sheet3 KPI 区升级为 Korn 模板格式(EFF6FF 行 · 标签+彩色大值)


# 单张发票 OCR · moved to vat_ocr_extract.py
from vat_ocr_extract import (  # noqa: F401  re-export (routes/handlers/tests) + facade-internal
    extract_invoice_fields,
    _ocr_validate_invoice,
    extract_invoices_parallel,
    _ocr_with_hard_timeout,
    _VEX_OCR_PER_FILE_TIMEOUT,
)

# ════════════════════════════════════════════════════════════════════════
# v118.32.5 · 性能优化 B · 多发票批量 OCR（800+ 张场景，减少 5x API 调用）
# ════════════════════════════════════════════════════════════════════════
_INVOICE_BATCH_PROMPT = """You are reading {n} Thai tax invoices (ใบกำกับภาษี) attached in order: invoice_1, invoice_2, ..., invoice_{n}.

For EACH invoice, extract the SAME 8 fields. Do NOT interpret · do NOT match · do NOT clean.

Output JSON ONLY in this exact shape:
{{
  "invoices": [
    {{
      "index": 1,
      "buyer_tax_id": "...",
      "buyer_name":   "...",
      "buyer_branch": "...",
      "invoice_no":   "...",
      "invoice_date": "...",
      "period":       "...",
      "amount_pre_vat": 0.00,
      "vat_amount":     0.00,
      "total_amount":   0.00
    }},
    ... (one object per attached invoice, in same order)
  ]
}}

Field rules (identical to single-invoice mode):
- buyer_tax_id: 13-digit Thai tax ID of the BUYER · digits only · "" if missing
- buyer_name:   Buyer name EXACTLY as printed (keep prefixes like บริษัท ... จำกัด)
- buyer_branch: "สำนักงานใหญ่" or 5-digit code · "" if cash customer
- invoice_no:   Invoice number EXACTLY as printed · keep prefixes (INV/IV/TAX) · do NOT strip leading zeros
- invoice_date: Date EXACTLY as printed · format DD/MM/YYYY · keep BE year as printed
- period:       MM/YYYY (Gregorian only · BE-543 if Buddhist Era)
- amount_pre_vat / vat_amount / total_amount: number · 2 decimals · no commas

STRICT RULES:
1. Output exactly {n} objects in the same order as the attached files
2. If a field is partially visible or unreadable · output "" · do NOT guess
3. Cash customer → buyer_tax_id="" buyer_branch=""
4. Numbers: digits and dot only · no commas · no currency
"""


def _mime_for(filename: str) -> Optional[str]:
    ext = (filename or "").lower().rsplit(".", 1)[-1]
    return {
        "pdf": "application/pdf",
        "jpg": "image/jpeg",
        "jpeg": "image/jpeg",
        "png": "image/png",
        "webp": "image/webp",
    }.get(ext)


def extract_invoice_fields_batch(
    invoice_files: List[Dict[str, Any]],
    api_key: Optional[str] = None,
    timeout: int = 90,
) -> List[Dict[str, Any]]:
    """一次 Gemini 调用抽取多张发票字段
    invoice_files: [{filename, bytes}]，建议 ≤ 5 张/批
    返回顺序与输入一致；任何一张失败/缺失 → 该 index 标记 ok=False
    """
    n = len(invoice_files)
    if n == 0:
        return []
    if n == 1:
        # 单张直接走原 single 流程，少一层包装
        f = invoice_files[0]
        return [extract_invoice_fields(f["bytes"], f["filename"], api_key=api_key)]

    # 校验 mime
    parts: List[Any] = [_INVOICE_BATCH_PROMPT.format(n=n)]
    for f in invoice_files:
        mime = _mime_for(f.get("filename") or "")
        if not mime:
            return [
                {
                    "ok": False,
                    "filename": x.get("filename"),
                    "error": "batch contains unsupported format",
                }
                for x in invoice_files
            ]
        parts.append({"mime_type": mime, "data": f["bytes"]})

    try:
        import google.generativeai as genai
    except ImportError:
        return [
            {"ok": False, "filename": f.get("filename"), "error": "google-generativeai 未安装"}
            for f in invoice_files
        ]

    key = api_key or os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
    if not key:
        return [
            {"ok": False, "filename": f.get("filename"), "error": "Gemini key 未配置"}
            for f in invoice_files
        ]

    text = ""
    try:
        genai.configure(api_key=key)
        model = genai.GenerativeModel(
            "gemini-2.5-flash",
            generation_config={
                "response_mime_type": "application/json",
                "temperature": 0.0,
            },
        )
        response = model.generate_content(
            parts,
            request_options={"timeout": timeout},
        )
        text = (response.text or "").strip()
        data = json.loads(text)
        items = data.get("invoices") or []
        _usage = getattr(response, "usage_metadata", None)
        _in_tok = int(getattr(_usage, "prompt_token_count", 0) or 0)
        _out_tok = int(getattr(_usage, "candidates_token_count", 0) or 0)
        # token 均摊到每张
        _in_per = _in_tok // max(n, 1)
        _out_per = _out_tok // max(n, 1)

        out: List[Dict[str, Any]] = [None] * n  # type: ignore
        for r in items:
            try:
                idx = int(r.get("index", 0)) - 1
            except Exception:
                continue
            if not (0 <= idx < n):
                continue
            out[idx] = {
                "ok": True,
                "filename": invoice_files[idx].get("filename"),
                "buyer_tax_id": str(r.get("buyer_tax_id") or "").strip(),
                "buyer_name": str(r.get("buyer_name") or "").strip(),
                "buyer_branch": str(r.get("buyer_branch") or "").strip(),
                "invoice_no": str(r.get("invoice_no") or "").strip(),
                "invoice_date": str(r.get("invoice_date") or "").strip(),
                "period": str(r.get("period") or "").strip(),
                "amount_pre_vat": _to_float(r.get("amount_pre_vat")),
                "vat_amount": _to_float(r.get("vat_amount")),
                "total_amount": _to_float(r.get("total_amount")),
                "_input_tokens": _in_per,
                "_output_tokens": _out_per,
                "_batch_size": n,
            }
        # 漏掉的 index → 标 fail，调用方会 fallback 单张
        for i in range(n):
            if out[i] is None:
                out[i] = {
                    "ok": False,
                    "filename": invoice_files[i].get("filename"),
                    "error": "batch_missing_index",
                }
        return out  # type: ignore
    except json.JSONDecodeError as e:
        logger.warning(f"[vex.batch] JSON 解析失败 n={n}: {e} · raw={text[:200]}")
        return [
            {"ok": False, "filename": f.get("filename"), "error": f"AI 返回格式异常: {str(e)[:60]}"}
            for f in invoice_files
        ]
    except Exception as e:
        logger.error(f"[vex.batch] n={n} 失败: {type(e).__name__}: {e}")
        return [
            {"ok": False, "filename": f.get("filename"), "error": str(e)[:120]}
            for f in invoice_files
        ]


def extract_invoices_batched_parallel(
    invoice_files: List[Dict[str, Any]],
    api_key: Optional[str] = None,
    batch_size: int = 5,
    max_workers: int = 4,
    auto_fallback_single: bool = True,
) -> List[Dict[str, Any]]:
    """v118.32.5 · 批量 + 并行：
    - 每批 batch_size 张走一次 Gemini
    - 多批并行 max_workers 路
    - 批失败时 auto_fallback_single 自动回退单张重试（保证不丢数据）
    """
    n = len(invoice_files)
    if n == 0:
        return []
    results: List[Optional[Dict]] = [None] * n
    batches: List[Tuple[int, List[Dict[str, Any]]]] = []
    for start in range(0, n, batch_size):
        chunk = invoice_files[start : start + batch_size]
        batches.append((start, chunk))

    def _run_batch(start: int, chunk: List[Dict[str, Any]]):
        out = extract_invoice_fields_batch(chunk, api_key=api_key)
        # fallback：批失败 / 批内部分失败 → 单张重试
        if auto_fallback_single:
            for j, r in enumerate(out):
                if not (r and r.get("ok")):
                    try:
                        f = chunk[j]
                        out[j] = extract_invoice_fields(f["bytes"], f["filename"], api_key=api_key)
                    except Exception as e:
                        logger.error(
                            f"[vex.batch] fallback 单张失败 {chunk[j].get('filename')}: {e}"
                        )
        for j, r in enumerate(out):
            results[start + j] = r

    # P0-2:每批加硬超时(批含 batch_size 张 + 可能的单张 fallback)· 挂起的批不阻塞整体 · job 能完成
    _batch_timeout = _VEX_OCR_PER_FILE_TIMEOUT * (batch_size + 1)
    pool = ThreadPoolExecutor(max_workers=max_workers)
    fut_to_batch = {pool.submit(_run_batch, s, c): (s, c) for (s, c) in batches}
    for fut, (s, c) in fut_to_batch.items():
        try:
            fut.result(timeout=_batch_timeout)
        except FuturesTimeout:
            logger.error(
                f"[vex.batch.parallel] 批硬超时({_batch_timeout}s) · start={s} · 落 ok=False"
            )
            for j, f in enumerate(c):
                if results[s + j] is None:
                    results[s + j] = {
                        "ok": False,
                        "filename": f.get("filename") or f"invoice_{s + j}",
                        "error": f"OCR 超时(>{_batch_timeout}s)",
                        "error_code": "ocr_timeout",
                    }
        except Exception as e:
            logger.error(f"[vex.batch.parallel] 批失败: {e}")
    pool.shutdown(wait=False)

    return [r or {"ok": False, "error": "worker_returned_none"} for r in results]


# ════════════════════════════════════════════════════════════════════════
# VAT 报告多份拼接 · 校验同卖方同期间
# ════════════════════════════════════════════════════════════════════════
def merge_vat_reports(
    report_files: List[Dict[str, Any]], api_key: Optional[str] = None
) -> Dict[str, Any]:
    """多份 VAT 报告 PDF 串行解析 · 拼接 rows · 校验同卖方同期间
    report_files: [{filename, bytes}]
    返回:{ok, rows, seller_name, seller_tax_id, period_year, period_month,
           sources(各文件解析摘要), error}
    """
    if not report_files:
        return {"ok": False, "error": "未上传 VAT 报告", "rows": []}

    parsed_list: List[Dict] = []
    sources: List[Dict] = []
    _rep_in_tok = 0
    _rep_out_tok = 0
    for f in report_files:
        # P0-2:报告 OCR 加线程层硬超时 · 偶发挂起的 Gemini 调用不再让 job 无限 running
        _fb, _fn = f["bytes"], f["filename"]
        r = _ocr_with_hard_timeout(
            lambda fb=_fb, fn=_fn: parse_vat_report(fb, fn, api_key=api_key),
            timeout_sec=int(os.environ.get("VEX_REPORT_OCR_TIMEOUT_SEC", "120")),
            on_timeout=lambda: {
                "ok": False,
                "rows": [],
                "error": "OCR 超时",
                "error_code": "ocr_timeout",
            },
        )
        if not r.get("ok"):
            return {
                "ok": False,
                "rows": [],
                "error": f"「{f['filename']}」解析失败: {r.get('error', '未知')}",
            }
        parsed_list.append(r)
        _rep_in_tok += int(r.get("_input_tokens") or 0)
        _rep_out_tok += int(r.get("_output_tokens") or 0)
        sources.append(
            {
                "filename": f["filename"],
                "row_count": r.get("row_count", len(r.get("rows", []))),
                "method": r.get("method", ""),
            }
        )

    # 校验:多份必须同卖方 + 同期间(从第 1 份起对照)
    # P1-2 修(2026-05-25 销项税回归):此前 period 只在 meta 有时才比 · 但 pdf_text_regex /
    #   pipeline 路径 meta={} → 永远拿不到 period → 混月份不拦(TC04 两月报告被合并成 6 行继续对账)。
    #   改:每份报告用 _dominant_report_period(meta 优先 · 否则行日期众数)· 跨文件比对期间。
    metas = [p.get("meta", {}) for p in parsed_list]
    sellers = set()
    for m in metas:
        sid = m.get("seller_tax_id") or m.get("issuer_tax_id") or ""
        if sid:
            sellers.add(sid)
    periods = set()
    for p in parsed_list:
        dp = _dominant_report_period(p)
        if dp:
            periods.add(dp)

    if len(sellers) > 1:
        return {
            "ok": False,
            "rows": [],
            "error": f"多份报告卖方不一致(税号: {' / '.join(sorted(sellers))})· 请检查",
        }
    if len(periods) > 1:
        _ps = " / ".join(f"{y}-{mn:02d}" for (y, mn) in sorted(periods))
        return {"ok": False, "rows": [], "error": f"多份报告期间不一致({_ps})· 请检查"}

    # 合并 rows · 重排 row_no 连续递增
    # v118.35.0.27 · 多页 PDF 溯源:每行加 source_filename + source_page
    # 让对账差异行能定位回原 PDF 第几页
    all_rows: List[Dict] = []
    seq = 0
    for p, src in zip(parsed_list, report_files):
        src_filename = src.get("filename", "")
        for row in p.get("rows", []):
            seq += 1
            row = dict(row)
            row["row_no"] = seq
            row["source_filename"] = src_filename
            # 保留原 row 里的 page_no(如果 parse_vat_report 返回)· 否则空
            row.setdefault("source_page", row.get("page_no") or row.get("page") or "")
            all_rows.append(row)

    seller_tax_id = next(iter(sellers)) if sellers else ""
    seller_name = ""
    for m in metas:
        if m.get("seller_name"):
            seller_name = m["seller_name"]
            break
    period_year = next(iter(periods))[0] if periods else None
    period_month = next(iter(periods))[1] if periods else None

    return {
        "ok": True,
        "rows": all_rows,
        "seller_tax_id": seller_tax_id,
        "seller_name": seller_name,
        "period_year": period_year,
        "period_month": period_month,
        "sources": sources,
        "row_count": len(all_rows),
        "_input_tokens": _rep_in_tok,
        "_output_tokens": _rep_out_tok,
    }


# ════════════════════════════════════════════════════════════════════════
# 4 语文案 · moved to vat_excel_i18n.py
from vat_excel_i18n import _I18N  # noqa: F401  facade-internal(build_excel)


def build_excel(
    invoices: List[Dict[str, Any]],
    report_rows: List[Dict[str, Any]],
    client_name: str = "",
    client_tax_id: str = "",
    period_year: Optional[int] = None,
    period_month: Optional[int] = None,
    lang: str = "th",
) -> bytes:
    """v118.32.4.9.6 · 生成 4-sheet Excel · 返回 xlsx bytes
    Bug 1-5 修复 + Sheet 3 一对一重做 + UI 美化(Sarabun/斑马/Tab 色/4 KPI 大卡)"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if lang not in _I18N:
        lang = "th"
    L = _I18N[lang]

    wb = openpyxl.Workbook()

    # ── 通用样式(Sarabun 优先 · 泰文友好) ──
    FONT_NAME = "Sarabun"
    F_HEAD = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    F_NORM = Font(name=FONT_NAME, size=10)
    F_BOLD = Font(name=FONT_NAME, size=10, bold=True)
    F_TITLE = Font(name=FONT_NAME, size=18, bold=True, color="111827")
    # F_KPI_LBL / F_KPI_VAL removed in v4.10.23 (replaced by Korn-style inline KPI)
    F_DIFF_RED = Font(name=FONT_NAME, size=10, color="DC2626")

    FILL_HEAD = PatternFill("solid", fgColor="2563EB")  # 蓝表头
    FILL_SUM = PatternFill("solid", fgColor="EFF6FF")
    FILL_ZEBRA = PatternFill("solid", fgColor="F9FAFB")  # 斑马偶数行
    FILL_OK = PatternFill("solid", fgColor="DCFCE7")  # 绿底匹配
    FILL_DIFF = PatternFill("solid", fgColor="FEF3C7")  # 黄底差异
    FILL_MISS = PatternFill("solid", fgColor="FEE2E2")  # 红底缺一边
    FILL_FUZZY = PatternFill("solid", fgColor="DBEAFE")  # 蓝底疑似
    FILL_OCRMSG = PatternFill("solid", fgColor="FED7AA")  # 橙底 OCR 漏抽
    # FILL_KPI_B/G/R/O removed in v4.10.23 (KPI row now uses EFF6FF bg + colored text)

    BORDER_TH = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right", vertical="center")
    AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

    HEAD_HEIGHT = 32  # 表头行高 32
    ROW_HEIGHT = 22

    def _style_header(ws, row, num_cols):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.font = F_HEAD
            cell.fill = FILL_HEAD
            cell.alignment = AL_C
            cell.border = BORDER_TH
        ws.row_dimensions[row].height = HEAD_HEIGHT

    def _zebra(ws, start_row, end_row, num_cols):
        """偶数行(相对 start_row)斑马条纹"""
        for r in range(start_row, end_row + 1):
            if (r - start_row) % 2 == 1:
                for c in range(1, num_cols + 1):
                    cell = ws.cell(row=r, column=c)
                    if cell.fill.fgColor and cell.fill.fgColor.rgb in (None, "00000000"):
                        cell.fill = FILL_ZEBRA

    # ════════════ Sheet 1 · 发票明细 ════════════
    ws1 = wb.active
    ws1.title = L["sh1"]
    ws1.sheet_properties.tabColor = "2563EB"  # Tab 蓝
    headers1 = L["h_inv"]
    ws1.append(headers1)
    _style_header(ws1, 1, len(headers1))

    # v4.10.22 · OCR 校验色
    FILL_OCR_WARN = FILL_MISS  # 有问题 → 红底(复用)
    FILL_OCR_OK = FILL_OK  # 全通过 → 绿底(复用)
    F_OCR_WARN = Font(name=FONT_NAME, size=10, color="DC2626")
    F_OCR_OK = Font(name=FONT_NAME, size=10, color="16A34A", bold=True)

    # 校验问题 key → 发票明细列号对应关系
    _WARN_COL = {
        "w_invoice_no_empty": 4,
        "w_buyer_name_empty": 3,
        "w_tax_id_bad_length": 2,
        "w_date_parse_fail": 5,
        "w_total_zero": 9,
        "w_vat_rate_mismatch": 8,
        "w_amount_sum_mismatch": 9,
    }

    # v4.10.22 · 先收集每行 OCR 校验结果(稍后在样式循环后应用)
    _inv_ocr_warns: List[List[str]] = []
    for i, inv in enumerate(invoices, 1):
        # Bug 1 · 期间降级
        period_val = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")
        warn_keys = _ocr_validate_invoice(inv)
        _inv_ocr_warns.append(warn_keys)
        warn_text = " · ".join(L.get(k, k) for k in warn_keys)
        ws1.append(
            [
                i,
                inv.get("buyer_tax_id") or "",
                inv.get("buyer_name") or "",
                inv.get("invoice_no") or "",
                inv.get("invoice_date") or "",
                period_val,
                inv.get("amount_pre_vat") or 0,
                inv.get("vat_amount") or 0,
                inv.get("total_amount") or 0,
                inv.get("filename") or "",
                warn_text if warn_text else L.get("ocr_ok", "✓ OK"),
            ]
        )

    # 合计行
    if invoices:
        last1 = len(invoices) + 1
        sum_row = last1 + 1
        ws1.cell(row=sum_row, column=1, value=L["sum"]).font = F_BOLD
        for col in (7, 8, 9):
            cell = ws1.cell(
                row=sum_row,
                column=col,
                value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last1})",
            )
            cell.font = F_BOLD
            cell.fill = FILL_SUM
        for c in range(1, len(headers1) + 1):
            ws1.cell(row=sum_row, column=c).fill = FILL_SUM
            ws1.cell(row=sum_row, column=c).border = BORDER_TH
            ws1.cell(row=sum_row, column=c).font = F_BOLD

    # 列宽 + 数字格式 + 斑马 + 行高(先统一设 F_NORM · 后面 OCR pass 再覆盖)
    widths1 = [5, 18, 28, 18, 13, 10, 14, 14, 14, 28, 30]
    for i, w in enumerate(widths1, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, len(invoices) + 2):
        ws1.row_dimensions[r].height = ROW_HEIGHT
        for c in range(1, len(headers1) + 1):
            cell = ws1.cell(row=r, column=c)
            cell.font = F_NORM
            cell.border = BORDER_TH
        for col in (7, 8, 9):
            ws1.cell(row=r, column=col).alignment = AL_R
            ws1.cell(row=r, column=col).number_format = "#,##0.00"
        ws1.cell(row=r, column=11).alignment = AL_L
    _zebra(ws1, 2, len(invoices) + 1, len(headers1))

    # v4.10.22 · OCR 高亮 pass(在通用样式之后应用 · 确保覆盖 F_NORM)
    for i, warn_keys in enumerate(_inv_ocr_warns, 1):
        data_row = i + 1
        ocr_cell = ws1.cell(row=data_row, column=11)
        if warn_keys:
            ocr_cell.fill = FILL_OCR_WARN
            ocr_cell.font = F_OCR_WARN
            for wk in warn_keys:
                col = _WARN_COL.get(wk)
                if col:
                    ws1.cell(row=data_row, column=col).fill = FILL_OCR_WARN
        else:
            ocr_cell.fill = FILL_OCR_OK
            ocr_cell.font = F_OCR_OK

    ws1.freeze_panes = "A2"

    # ════════════ Sheet 2 · VAT 报告明细 ════════════
    ws2 = wb.create_sheet(L["sh2"])
    ws2.sheet_properties.tabColor = "16A34A"  # Tab 绿
    headers2 = L["h_rep"]
    ws2.append(headers2)
    _style_header(ws2, 1, len(headers2))

    for i, row in enumerate(report_rows, 1):
        # Bug 1 · 期间降级 · 优先用 report_date 算 · fallback period_year/month
        date_str = row.get("report_date") or ""
        period_val = _derive_period(date_str, "")
        if not period_val and period_year and period_month:
            period_val = f"{period_month:02d}/{period_year}"
        pre = row.get("report_amount_pre_vat") or row.get("report_amount") or 0
        vat = row.get("report_vat_amount") or 0
        try:
            total = round(float(pre or 0) + float(vat or 0), 2)
        except Exception:
            total = 0
        ws2.append(
            [
                i,
                row.get("report_buyer_tax_id") or "",
                row.get("report_buyer_name") or "",
                date_str,
                period_val,
                pre,
                vat,
                total,
            ]
        )

    if report_rows:
        last2 = len(report_rows) + 1
        sum_row2 = last2 + 1
        ws2.cell(row=sum_row2, column=1, value=L["sum"]).font = F_BOLD
        for col in (6, 7, 8):  # Sheet 2 数字列 偏移 1(加了 日期 列)
            cell = ws2.cell(
                row=sum_row2,
                column=col,
                value=f"=SUM({get_column_letter(col)}2:{get_column_letter(col)}{last2})",
            )
            cell.font = F_BOLD
            cell.fill = FILL_SUM
        for c in range(1, len(headers2) + 1):
            ws2.cell(row=sum_row2, column=c).fill = FILL_SUM
            ws2.cell(row=sum_row2, column=c).border = BORDER_TH
            ws2.cell(row=sum_row2, column=c).font = F_BOLD

    widths2 = [5, 18, 28, 13, 12, 14, 14, 14]
    for i, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r in range(2, len(report_rows) + 2):
        ws2.row_dimensions[r].height = ROW_HEIGHT
        for c in range(1, len(headers2) + 1):
            cell = ws2.cell(row=r, column=c)
            cell.font = F_NORM
            cell.border = BORDER_TH
        for col in (6, 7, 8):
            ws2.cell(row=r, column=col).alignment = AL_R
            ws2.cell(row=r, column=col).number_format = "#,##0.00"
    _zebra(ws2, 2, len(report_rows) + 1, len(headers2))
    ws2.freeze_panes = "A2"

    # ════════════ Sheet 3 · 对账结果(一对一 15 列 · 后端预算结果)════════════
    ws3 = wb.create_sheet(L["sh3"])
    ws3.sheet_properties.tabColor = "D97706"  # Tab 橙

    # 跑配对(Bug 2/3/4/5 全在这里)
    match_result = _build_recon_pairs(invoices, report_rows)
    pairs = match_result["pairs"]
    unmatched_inv = match_result["unmatched_inv"]
    unmatched_rep = match_result["unmatched_rep"]

    # 计算 KPI
    n_total = len(pairs) + len(unmatched_inv) + len(unmatched_rep)
    n_ok = 0
    for _p in pairs:
        _inv = invoices[_p["inv_idx"]]
        _rep = report_rows[_p["rep_idx"]]
        if not _eq_amount(_get_inv_total(_inv), _get_rep_total(_rep)):
            continue
        if _p["kind"] == "matched_cash":
            n_ok += 1
        elif _p["kind"] == "matched":
            _d = _diff_dims(_inv, _rep)
            if not any(_d.values()):
                n_ok += 1
    n_diff = n_total - n_ok
    diff_amount_total = 0.0
    for p in pairs:
        if p["kind"] in ("matched", "matched_cash"):
            a = _get_inv_total(invoices[p["inv_idx"]])
            b = _get_rep_total(report_rows[p["rep_idx"]])
            if not _eq_amount(a, b):
                diff_amount_total += abs(a - b)
    for ii in unmatched_inv:
        diff_amount_total += _get_inv_total(invoices[ii])
    for ri in unmatched_rep:
        diff_amount_total += _get_rep_total(report_rows[ri])

    # R1 · 标题行(Korn 样式 · 合并全行 · sz=18 · 高 36)
    _n_cols3 = len(L["h_recon"])
    ws3.merge_cells(f"A1:{get_column_letter(_n_cols3)}1")
    _c_title = ws3.cell(row=1, column=1, value=L["title"])
    _c_title.font = F_TITLE
    _c_title.alignment = AL_C
    ws3.row_dimensions[1].height = 36

    # R2 · 客户+期间 meta
    meta_parts = []
    if client_name:
        meta_parts.append(f"{L['client']}: {client_name}")
    if period_year and period_month:
        meta_parts.append(f"{L['period']}: {period_month:02d}/{period_year}")
    if meta_parts:
        ws3.cell(row=2, column=1, value=" · ".join(meta_parts)).font = F_NORM

    # R3 · 空行

    # R4-R5 · KPI 4 大卡(每卡 4 列宽 · 共 16 列 · 彩色底色)
    F_KPI_LBL2 = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    F_KPI_VAL2 = Font(name=FONT_NAME, size=22, bold=True, color="FFFFFF")
    FILL_KPI_B = PatternFill("solid", fgColor="2563EB")
    FILL_KPI_G = PatternFill("solid", fgColor="16A34A")
    FILL_KPI_R = PatternFill("solid", fgColor="DC2626")
    FILL_KPI_O = PatternFill("solid", fgColor="D97706")
    KPI_ROW_LBL = 4
    KPI_ROW_VAL = 5
    ws3.row_dimensions[KPI_ROW_LBL].height = 22
    ws3.row_dimensions[KPI_ROW_VAL].height = 44

    def _kpi(col_start, label, value, fill):
        ws3.merge_cells(
            start_row=KPI_ROW_LBL,
            start_column=col_start,
            end_row=KPI_ROW_LBL,
            end_column=col_start + 3,
        )
        c1 = ws3.cell(row=KPI_ROW_LBL, column=col_start, value=label)
        c1.font = F_KPI_LBL2
        c1.fill = fill
        c1.alignment = AL_C
        ws3.merge_cells(
            start_row=KPI_ROW_VAL,
            start_column=col_start,
            end_row=KPI_ROW_VAL,
            end_column=col_start + 3,
        )
        c2 = ws3.cell(row=KPI_ROW_VAL, column=col_start, value=value)
        c2.font = F_KPI_VAL2
        c2.fill = fill
        c2.alignment = AL_C

    _kpi(1, L["kpi_total"], n_total, FILL_KPI_B)
    _kpi(5, L["kpi_ok"], n_ok, FILL_KPI_G)
    _kpi(9, L["kpi_diff"], n_diff, FILL_KPI_R)
    _kpi(13, L["kpi_amt"], f"฿ {diff_amount_total:,.2f}", FILL_KPI_O)

    # R6 · 表头(15 列)
    HEADER_ROW = 6
    DATA_START = 7
    headers3 = L["h_recon"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=HEADER_ROW, column=c, value=h)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = AL_C
        cell.border = BORDER_TH
    ws3.row_dimensions[HEADER_ROW].height = HEAD_HEIGHT

    # ── 写数据行 · 配对 + 孤儿
    def _status_for(pair, dims) -> tuple:
        """v118.32.4.10.0 · 维度感知状态判定 · 返回 (status_key, text, fill)"""
        kind = pair["kind"]
        if kind == "fuzzy":
            return "st_fuzzy", L["st_fuzzy"], FILL_FUZZY
        if kind == "ocr_missing":
            return "st_ocr_missing", L["st_ocr_missing"], FILL_OCRMSG
        if dims.get("tax_id", "").startswith("~"):
            return "st_fuzzy", L["st_fuzzy"], FILL_FUZZY
        if kind == "matched_cash":
            a = _get_inv_total(invoices[pair["inv_idx"]])
            b = _get_rep_total(report_rows[pair["rep_idx"]])
            if _eq_amount(a, b):
                return "st_cash", L["st_cash"], FILL_OK
            return "st_diff", L["st_diff"], FILL_DIFF
        diff_keys = [k for k, v in dims.items() if v and not (k == "tax_id" and v.startswith("~"))]
        a = _get_inv_total(invoices[pair["inv_idx"]])
        b = _get_rep_total(report_rows[pair["rep_idx"]])
        if not _eq_amount(a, b):
            diff_keys.append("_amt")
        n = len(diff_keys)
        if n == 0:
            return "st_ok", L["st_ok"], FILL_OK
        if n >= 2:
            return "st_multi_diff", L["st_multi_diff"], FILL_DIFF
        key = diff_keys[0]
        if key == "_amt":
            return "st_diff", L["st_diff"], FILL_DIFF
        if key in ("date", "period"):
            return "st_date_diff", L["st_date_diff"], FILL_DIFF
        if key == "branch":
            return "st_branch_diff", L["st_branch_diff"], FILL_DIFF
        if key == "name":
            return "st_name_diff", L["st_name_diff"], FILL_DIFF
        if key == "inv_no":
            return "st_inv_no_diff", L["st_inv_no_diff"], FILL_DIFF
        if key == "tax_id":
            return "st_tax_id_diff", L["st_tax_id_diff"], FILL_DIFF
        return "st_diff", L["st_diff"], FILL_DIFF

    row_cursor = DATA_START
    seq_no = 0
    diff_col_indices = (9, 10, 11, 12, 13, 14)  # 6 维度差异列(含税号差)
    F_FUZZY_BLUE = Font(name=FONT_NAME, size=10, color="2563EB")
    task_rows: List[Dict] = []  # v118.32.4.10.0 · 任务摘要行

    for pair in pairs:
        seq_no += 1
        inv = invoices[pair["inv_idx"]]
        rep = report_rows[pair["rep_idx"]]
        dims = _diff_dims(inv, rep)
        # Bug 2 · 散客:不比较/不显示分公司和税号差
        if pair["kind"] == "matched_cash":
            dims["branch"] = ""
            dims["tax_id"] = ""
        # tax_id fuzzy ~ 前缀处理(显示时去掉 · 用蓝色替换红色)
        tax_id_raw = dims.get("tax_id", "")
        tax_id_is_fuzzy = tax_id_raw.startswith("~")
        tax_id_display = tax_id_raw[1:] if tax_id_is_fuzzy else tax_id_raw
        status_key, status_text, row_fill = _status_for(pair, dims)
        amt_inv = _get_inv_total(inv)
        amt_rep = _get_rep_total(rep)
        amt_diff = round(amt_inv - amt_rep, 2)
        period_inv = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")

        # v4.10.13 · 净额/VAT 分字段差异备注
        amt_pre_inv = float(inv.get("amount_pre_vat") or 0)
        amt_vat_inv = float(inv.get("vat_amount") or 0)
        amt_pre_rep = float((rep.get("report_amount_pre_vat") or rep.get("report_amount")) or 0)
        amt_vat_rep = float(rep.get("report_vat_amount") or 0)
        pre_diff = round(amt_pre_inv - amt_pre_rep, 2)
        vat_diff_ = round(amt_vat_inv - amt_vat_rep, 2)
        if not _eq_amount(amt_inv, amt_rep):
            if not _eq_amount(pre_diff, 0) and not _eq_amount(vat_diff_, 0):
                _amt_note = f"净额差 {pre_diff:+,.2f} · VAT 差 {vat_diff_:+,.2f}"
            elif not _eq_amount(pre_diff, 0):
                _amt_note = f"净额差 {pre_diff:+,.2f} · VAT 一致"
            else:
                _amt_note = f"VAT 差 {vat_diff_:+,.2f} · 净额一致"
        else:
            _amt_note = ""

        # v4.10.13 · ocr_missing · 备注追加提醒(dims 原样保留)
        _base_note = pair.get("note") or ""
        if pair["kind"] == "ocr_missing":
            _base_note = (_base_note + " · OCR 抽取可能不完整 · 请核对原 PDF").lstrip(" · ")
        _note_val = " · ".join(filter(None, [_base_note, _amt_note]))

        values = [
            seq_no,  # 1 #
            status_text,  # 2 status
            inv.get("buyer_name") or rep.get("report_buyer_name") or "",  # 3 客户
            inv.get("invoice_no") or "",  # 4 发票号
            period_inv,  # 5 期间
            amt_inv,  # 6 金额(发)
            amt_rep,  # 7 金额(报)
            amt_diff,  # 8 差异金额
            dims["inv_no"],  # 9 发票号差
            dims["date"],  # 10 日期差
            dims["period"],  # 11 期间差
            tax_id_display,  # 12 税号差(新)
            dims["branch"],  # 13 分公司差
            dims["name"],  # 14 客户名差
            _note_val,  # 15 备注
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = row_fill
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
            # 差异维度列若有内容 · 税号 fuzzy 蓝 · 其余红
            if c in diff_col_indices and v:
                if c == 12 and tax_id_is_fuzzy:
                    cell.font = F_FUZZY_BLUE
                else:
                    cell.font = F_DIFF_RED
        # v118.32.4.10.0 · 收集摘要行
        task_rows.append(
            {
                "status_key": status_key,
                "status": status_text,
                "customer": inv.get("buyer_name") or rep.get("report_buyer_name") or "",
                "invoice_no": inv.get("invoice_no") or "",
                "amount_inv": amt_inv,
                "amount_rep": amt_rep,
                "dims": {
                    k: (v[1:] if k == "tax_id" and v.startswith("~") else v)
                    for k, v in dims.items()
                },
                "kind": pair["kind"],
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 发票孤儿(报告无)
    for ii in unmatched_inv:
        seq_no += 1
        inv = invoices[ii]
        amt_inv = _get_inv_total(inv)
        period_inv = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")
        values = [
            seq_no,
            L["st_no_rep"],
            inv.get("buyer_name") or "",
            inv.get("invoice_no") or "",
            period_inv,
            amt_inv,
            0,
            amt_inv,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = FILL_MISS
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
        task_rows.append(
            {
                "status_key": "st_no_rep",
                "status": L["st_no_rep"],
                "customer": inv.get("buyer_name") or "",
                "invoice_no": inv.get("invoice_no") or "",
                "amount_inv": amt_inv,
                "amount_rep": 0,
                "dims": {},
                "kind": "invoice_orphan",
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 报告孤儿(发票无)
    for ri in unmatched_rep:
        seq_no += 1
        rep = report_rows[ri]
        amt_rep = _get_rep_total(rep)
        rep_date = rep.get("report_date") or ""
        d = parse_date(rep_date)
        period_rep = f"{d.month:02d}/{d.year}" if d else ""
        values = [
            seq_no,
            L["st_no_inv"],
            rep.get("report_buyer_name") or "",
            rep.get("report_invoice_no") or "",
            period_rep,
            0,
            amt_rep,
            -amt_rep,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = FILL_MISS
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
        task_rows.append(
            {
                "status_key": "st_no_inv",
                "status": L["st_no_inv"],
                "customer": rep.get("report_buyer_name") or "",
                "invoice_no": rep.get("report_invoice_no") or "",
                "amount_inv": 0,
                "amount_rep": amt_rep,
                "dims": {},
                "kind": "report_orphan",
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 列宽
    widths3 = [5, 22, 26, 16, 10, 14, 14, 13, 22, 14, 16, 22, 18, 26, 22]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = f"A{DATA_START}"

    # ════════════ Sheet 4 · 使用说明 ════════════
    ws4 = wb.create_sheet(L["sh4"])
    ws4.sheet_properties.tabColor = "6B7280"  # Tab 灰
    ws4.column_dimensions["A"].width = 100
    ws4.cell(row=1, column=1, value=L["title"]).font = F_TITLE
    ws4.row_dimensions[1].height = 28
    for i, line in enumerate(L["help_lines"], start=3):
        ws4.cell(row=i, column=1, value=line).font = F_NORM
        ws4.cell(row=i, column=1).alignment = AL_L
        ws4.row_dimensions[i].height = 22
    other_langs = [k for k in _I18N.keys() if k != lang]
    base_row = 3 + len(L["help_lines"]) + 2
    for lk in other_langs:
        for line in _I18N[lk]["help_lines"]:
            c = ws4.cell(row=base_row, column=1, value=line)
            c.font = Font(name=FONT_NAME, size=9, color="6B7280")
            c.alignment = AL_L
            base_row += 1
        base_row += 1

    # ── 保存为 bytes ──
    bio = io.BytesIO()
    wb.save(bio)
    task_summary = {
        "n_total": n_total,
        "n_ok": n_ok,
        "n_diff": n_diff,
        "diff_amount_total": round(diff_amount_total, 2),
        "rows": task_rows,
    }
    return bio.getvalue(), task_summary
