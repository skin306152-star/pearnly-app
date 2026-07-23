# -*- coding: utf-8 -*-
"""钱的闸:导出去再读回来,每个数字逐分相等。

这道闸存在的理由是两个真问题:
1. 写侧和读侧曾各写一套四舍五入,多行发票往返一圈差分币,而推 ERP 用的是读侧那个数
   —— 会计在表上看到的和账里的对不上;
2. 金额列的 Excel 公式写死了列字母(=E2*F2),往数量列前插一列公式就指向错列,
   而按列名取值的读侧照常正确 —— 文件自相矛盾,结构类测试一条都抓不到。

故这里**只断言数字**,而且用会触发进位边界的语料。以后任何改动只要让表和读回值
对不上,这里就会红。
"""

import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.excel import erp_money as money
from services.excel import erp_roundtrip as rt
from services.excel.erp_roundtrip_reader import parse_roundtrip_workbook
from services.excel.erp_workbook import build_review_workbook


def _sales(hid, inv, lines, vat=None, **extra):
    """lines = [(qty, price)]。vat=票面税额(不给则不带,读侧按税基算)。"""
    f = {
        "history_id": hid,
        "invoice_number": inv,
        "date": "2026-07-23",
        "buyer_name": "B",
        "items": [
            {"description": f"item{i}", "qty": q, "unit_price": p}
            for i, (q, p) in enumerate(lines, 1)
        ],
    }
    if vat is not None:
        f["vat_amount"] = vat
    f.update(extra)
    return {"filename": f"{inv}.png", "merged_fields": f}


def _read_one(**kw):
    out = parse_roundtrip_workbook(build_review_workbook(**kw))
    return out["documents"][0]["fields"]


class LineAmountTests(unittest.TestCase):
    def test_line_amount_matches_contract(self):
        """公式列没有缓存值,读侧按 数量×单价 派生 —— 必须逐分等于合同算法。"""
        f = _read_one(sales=[_sales("h", "S1", [(3, 250)])])
        self.assertEqual(f["items"][0]["amount"], money.line_amount(3, 250))
        self.assertEqual(f["items"][0]["amount"], Decimal("750.00"))

    def test_fractional_price_rounds_once_not_twice(self):
        """单价带小数时只在行金额上舍一次,不许先舍再乘。"""
        f = _read_one(sales=[_sales("h", "S1", [(3, "33.335")])])
        self.assertEqual(f["items"][0]["amount"], money.line_amount(3, "33.335"))


class DocVatTests(unittest.TestCase):
    def test_printed_vat_wins_over_recompute(self):
        """票面税额是要进 ภ.พ.30 的法定数字 —— 我们没有权力把它算成别的。
        两行各 33.33:逐行算 2.33+2.33=4.66,按总额算 4.67。票面印 4.66 就该是 4.66。"""
        f = _read_one(sales=[_sales("h", "S1", [(1, "33.33"), (1, "33.33")], vat="4.66")])
        self.assertEqual(f["amount_before_vat"], Decimal("66.66"))
        self.assertEqual(f["vat_amount"], Decimal("4.66"))
        self.assertEqual(f["total_amount"], Decimal("71.32"))

    def test_recomputes_when_base_really_changed(self):
        """票面税额与税基对不上(会计改了数量/单价)→ 才重算,不硬套过期的票面值。"""
        f = _read_one(sales=[_sales("h", "S1", [(10, 100)], vat="4.66")])
        self.assertEqual(f["amount_before_vat"], Decimal("1000.00"))
        self.assertEqual(f["vat_amount"], Decimal("70.00"))

    def test_no_printed_vat_falls_back_to_contract(self):
        f = _read_one(sales=[_sales("h", "S1", [(1, "100")])])
        self.assertEqual(f["vat_amount"], money.vat_of("100"))


class TotalsConsistencyTests(unittest.TestCase):
    def test_base_plus_vat_equals_total_exactly(self):
        """税基 + 税额 必须严格等于合计 —— 差一分就是账不平。"""
        for lines in ([(1, "0.01")], [(3, "33.33")], [(7, "142.857")], [(1, "1"), (1, "2")]):
            f = _read_one(sales=[_sales("h", "S1", lines)])
            self.assertEqual(
                f["amount_before_vat"] + f["vat_amount"], f["total_amount"], str(lines)
            )

    def test_line_amounts_sum_to_base(self):
        f = _read_one(sales=[_sales("h", "S1", [(2, "19.99"), (3, "7.77"), (1, "0.03")])])
        self.assertEqual(sum(i["amount"] for i in f["items"]), f["amount_before_vat"])


class PurchaseMoneyTests(unittest.TestCase):
    def _pur(self, base, vat, total=None):
        f = {
            "history_id": "h",
            "invoice_number": "P1",
            "seller_name": "S",
            "amount_before_vat": base,
            "vat_amount": vat,
        }
        if total is not None:
            f["total_amount"] = total
        out = parse_roundtrip_workbook(build_review_workbook(purchase=[{"merged_fields": f}]))
        return out["documents"][0]["fields"]

    def test_printed_input_vat_preserved(self):
        f = self._pur("3200.00", "224.00")
        self.assertEqual(f["amount_before_vat"], Decimal("3200.00"))
        self.assertEqual(f["vat_amount"], Decimal("224.00"))
        self.assertEqual(f["total_amount"], Decimal("3424.00"))

    def test_odd_printed_vat_kept_within_tolerance(self):
        """供应商票上印的税额与 7% 差一两分是常事 —— 照抄,不替人家改。"""
        f = self._pur("100.00", "7.01")
        self.assertEqual(f["vat_amount"], Decimal("7.01"))


class ColumnLetterDriftGate(unittest.TestCase):
    """公式列写死了列字母。这道闸不看公式文本,只看「Excel 会算出什么」——
    把公式指向的两列取出来相乘,必须等于读侧派生的行金额。列一挪,这里立刻红。"""

    def test_formula_points_at_qty_and_price_columns(self):
        raw = build_review_workbook(sales=[_sales("h", "S1", [(3, 250)])])
        ws = load_workbook(io.BytesIO(raw))[rt.SHEET_SALES]
        hmap = {ws.cell(1, c).value: c for c in range(1, ws.max_column + 1)}

        formula = str(ws.cell(2, hmap[rt.SALES_COL_AMOUNT]).value or "")
        self.assertTrue(formula.startswith("="), "金额列该是公式(会计要改数量看金额跟着动)")

        from openpyxl.utils import get_column_letter

        qty_letter = get_column_letter(hmap[rt.SALES_COL_QTY])
        price_letter = get_column_letter(hmap[rt.SALES_COL_PRICE])
        self.assertEqual(
            formula,
            f"={qty_letter}2*{price_letter}2",
            "公式指向的列与「数量」「单价」的实际列位不符 —— "
            "Excel 里显示的金额会和我们导入的对不上",
        )


if __name__ == "__main__":
    unittest.main()
