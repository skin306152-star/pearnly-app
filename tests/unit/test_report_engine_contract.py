# -*- coding: utf-8 -*-
"""REFACTOR-D2 守门 · 统一报表生成器 report_engine.py 行为契约

report_engine.py(4 模板 input_vat/standard/erp/print · 4 语)此前 0 专属测试。
纯 openpyxl + i18n_reports · 无 DB/网络/凭证(Wave 0 安全网 · 零冲突)。

锁定:行归一化(多源字段 + 7% VAT 逆算)· list_templates 契约 ·
build_report 4 模板均能出可加载的 xlsx + 未知模板报错 + 未知语言兜底 ·
default_filename 非法字符清洗。
"""

import io
import unittest

from openpyxl import load_workbook

from services.report import report_engine as re_eng


class NormalizeRowTests(unittest.TestCase):
    def test_maps_top_level(self):
        row = re_eng._normalize_row(
            {"invoice_no": "INV1", "seller_name": "ACME", "total_amount": "1070"}, 1
        )
        self.assertEqual(row["no"], 1)
        self.assertEqual(row["invoice_no"], "INV1")
        self.assertEqual(row["seller_name"], "ACME")
        self.assertEqual(row["total_amount"], 1070.0)

    def test_merged_fields_fallback(self):
        row = re_eng._normalize_row({"merged_fields": {"invoice_number": "M9"}}, 2)
        self.assertEqual(row["invoice_no"], "M9")

    def test_vat_reverse_calc_from_total(self):
        # 有 total 没 subtotal/vat → 按 7% 还原
        row = re_eng._normalize_row({"total_amount": "107"}, 1)
        self.assertEqual(row["amount_before_vat"], 100.0)
        self.assertEqual(row["vat_amount"], 7.0)

    def test_rec_takes_precedence_over_merged(self):
        row = re_eng._normalize_row(
            {"invoice_no": "TOP", "merged_fields": {"invoice_no": "NESTED"}}, 1
        )
        self.assertEqual(row["invoice_no"], "TOP")


class ListTemplatesTests(unittest.TestCase):
    def test_returns_all_templates_with_contract(self):
        out = re_eng.list_templates(lang="en")
        codes = {t["code"] for t in out}
        self.assertEqual(codes, {"input_vat", "standard", "erp", "print"})
        for t in out:
            for key in ("code", "name", "desc", "category", "category_code", "recommended"):
                self.assertIn(key, t)
            self.assertIsInstance(t["recommended"], bool)
            self.assertTrue(t["name"])

    def test_lang_switch_changes_name(self):
        zh = {t["code"]: t["name"] for t in re_eng.list_templates("zh")}
        en = {t["code"]: t["name"] for t in re_eng.list_templates("en")}
        self.assertNotEqual(zh["standard"], en["standard"])


class BuildReportTests(unittest.TestCase):
    def _rows(self):
        return [
            {
                "invoice_no": "INV1",
                "invoice_date": "2026-03-15",
                "seller_name": "ACME",
                "total_amount": "107",
            }
        ]

    def test_each_template_produces_loadable_xlsx(self):
        meta = {"client_name": "X Co", "period_label": "2026-05"}
        for code in ("input_vat", "standard", "erp", "print"):
            raw = re_eng.build_report(code, self._rows(), meta=meta, lang="en")
            self.assertIsInstance(raw, bytes)
            ws = load_workbook(io.BytesIO(raw)).active
            self.assertGreaterEqual(ws.max_row, 1, f"{code} 报表无内容")

    def test_unknown_template_raises(self):
        with self.assertRaises(ValueError):
            re_eng.build_report("no_such", self._rows())

    def test_unknown_lang_falls_back_no_crash(self):
        raw = re_eng.build_report("standard", self._rows(), lang="fr")
        self.assertIsInstance(raw, bytes)

    def test_empty_rows_still_builds(self):
        raw = re_eng.build_report("standard", [], lang="zh")
        self.assertIsInstance(raw, bytes)
        load_workbook(io.BytesIO(raw))  # 不抛即可


class DefaultFilenameTests(unittest.TestCase):
    def test_basic(self):
        fn = re_eng.default_filename("standard")
        self.assertTrue(fn.startswith("mrpilot-"))
        self.assertTrue(fn.endswith(".xlsx"))
        self.assertIn("standard", fn)

    def test_illegal_client_chars_cleaned(self):
        fn = re_eng.default_filename("erp", client="A/B:C*", period="03/2026")
        self.assertNotIn("/", fn.replace(".xlsx", ""))  # period 的 / 也转 -
        self.assertNotIn(":", fn)
        self.assertNotIn("*", fn)
        self.assertIn("03-2026", fn)


if __name__ == "__main__":
    unittest.main()
