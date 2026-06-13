# -*- coding: utf-8 -*-
"""VAT 报告解析 · Excel(.xlsx/.xls)路径(零成本 openpyxl 表头识别)。"""

import io
import logging
from typing import List, Dict, Any

from services.vat.vat_parser_common import (
    _to_float,
    PARSER_VERSION,
    _SKIP_H,
    _map_columns,
    _build_row,
)

logger = logging.getLogger(__name__)


# ======================================================================
# 1. Excel 解析(零成本)
# ======================================================================


def parse_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl 未安装", "rows": []}

    rows: List[Dict] = []
    meta: Dict[str, Any] = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        # 语言无关表头检测(泰/英/中/日):用列头关键词映射判定,而非硬编码几个泰/英词。
        # 真表头 = 命中日期列 + 至少一个(票号/参考号/任一金额列)· 数据行(值)不会命中列头关键词。
        header_idx = None
        col_map: Dict[str, int] = {}
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            cm = _map_columns([str(c or "") for c in row])
            if "date" in cm and (
                "invoice_no" in cm
                or "ref_no" in cm
                or "amount_pre_vat" in cm
                or "vat_amount" in cm
                or "total_amount" in cm
            ):
                header_idx = i
                col_map = cm
                break
        if header_idx is None:
            return {"ok": False, "error": "找不到表头行", "rows": []}

        row_no = 0
        for raw in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            cells = list(raw)
            if not any(c for c in cells if c not in (None, "")):
                continue
            first = str(cells[0] or "").strip().lower()
            if any(k in first for k in _SKIP_H):
                if "amount_pre_vat" in col_map:
                    meta["total_amount_pre_vat"] = _to_float(cells[col_map["amount_pre_vat"]])
                if "vat_amount" in col_map:
                    meta["total_vat"] = _to_float(cells[col_map["vat_amount"]])
                continue
            row_no += 1
            rows.append(_build_row(row_no, cells, col_map))

        return {
            "ok": True,
            "rows": rows,
            "meta": meta,
            "warnings": [],
            "parser_version": PARSER_VERSION,
            "row_count": len(rows),
            "method": "excel",
        }
    except Exception as e:
        logger.error(f"parse_excel failed: {e}")
        return {"ok": False, "error": str(e), "rows": []}
