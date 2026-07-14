# -*- coding: utf-8 -*-
"""影子分录 → Express 导出(services/workorder/entries_export.py · M1-3KEY 键二)。

锁定:①coa 码经桥翻 Express 码,桥缺码留空 + 标 unmapped(禁臆造)②桥整个未配置→首行提示 +
Express 列全空 ③金额按借/贷分列,Decimal 保精度,总计 Σ借=Σ贷 与影子试算平衡一致 ④产出是
可被 openpyxl 读回的合法 xlsx。金标影子取自 workorder_shadow_adapter.build_shadow(与
test_workorder_financials_routes_contract 同源),不手搓假数据。
"""

from __future__ import annotations

import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.accounting import workorder_shadow_adapter
from services.workorder import entries_export


def _golden_shadow() -> dict:
    return workorder_shadow_adapter.build_shadow(
        purchase_entries=[{"net": Decimal("1000"), "vat": Decimal("70"), "grand": Decimal("1070")}],
        sales_amount=Decimal("5000"),
        output_vat=Decimal("350"),
        period="2569-05",
    ).as_gate_payload()


def _sheet_rows(content: bytes, sheet: str) -> list[tuple]:
    wb = load_workbook(io.BytesIO(content))
    return list(wb[sheet].iter_rows(values_only=True))


class EntryRowsTests(unittest.TestCase):
    def test_maps_present_code_and_flags_missing_as_unmapped(self):
        shadow = _golden_shadow()
        first_code = shadow["entries"][0]["account_code"]
        bridge = {first_code: "1234-56"}  # 只配第一码,其余应 unmapped
        rows = entries_export.build_entry_rows(shadow, bridge)

        self.assertEqual(len(rows), len(shadow["entries"]))
        mapped = [r for r in rows if r["coa_code"] == first_code]
        self.assertTrue(mapped)
        self.assertTrue(all(r["express_code"] == "1234-56" and not r["unmapped"] for r in mapped))
        others = [r for r in rows if r["coa_code"] != first_code]
        self.assertTrue(others, "金标应有不止一个科目")
        self.assertTrue(all(r["express_code"] == "" and r["unmapped"] for r in others))

    def test_debit_credit_split_and_decimal(self):
        rows = entries_export.build_entry_rows(_golden_shadow(), {})
        for r in rows:
            self.assertTrue((r["debit"] is None) != (r["credit"] is None), "一行只落借或贷")
            amount = r["debit"] if r["debit"] is not None else r["credit"]
            self.assertIsInstance(amount, Decimal)


class XlsxBuildTests(unittest.TestCase):
    def test_valid_xlsx_totals_balanced(self):
        shadow = _golden_shadow()
        tb = shadow["trial_balance"]
        self.assertEqual(str(tb["debit"]), str(tb["credit"]))  # 金标影子本就配平
        content = entries_export.build_entries_xlsx(
            shadow, {}, bridge_configured=True, period="2569-05", client_name="Sister Makeup"
        )
        total_row = next(
            r for r in _sheet_rows(content, "Express Entries") if r[0] and "合计" in str(r[0])
        )
        self.assertAlmostEqual(float(total_row[4]), float(total_row[5]), places=2)
        self.assertAlmostEqual(float(total_row[4]), float(Decimal(tb["debit"])), places=2)

    def test_unmapped_note_when_bridge_lacks_code(self):
        content = entries_export.build_entries_xlsx(
            _golden_shadow(), {}, bridge_configured=True, period="2569-05", client_name="X"
        )
        notes = [str(r[7] or "") for r in _sheet_rows(content, "Express Entries")]
        self.assertTrue(any("unmapped" in v for v in notes))

    def test_bridge_missing_hint_row_and_empty_express_col(self):
        content = entries_export.build_entries_xlsx(
            _golden_shadow(), {}, bridge_configured=False, period="2569-05", client_name="X"
        )
        rows = _sheet_rows(content, "Express Entries")
        self.assertIn("科目桥未配置", str(rows[0][0] or ""))
        # 数据行从第 3 行起(首行提示 + 表头);Express 码列(索引 1)全空。
        data_express = [r[1] for r in rows[2:] if r[0] and "合计" not in str(r[0])]
        self.assertTrue(all((v is None or v == "") for v in data_express))

    def test_meta_sheet_carries_client_and_period(self):
        content = entries_export.build_entries_xlsx(
            _golden_shadow(),
            {},
            bridge_configured=True,
            period="2569-05",
            client_name="Sister Makeup",
        )
        joined = " ".join(
            str(v) for r in _sheet_rows(content, "ข้อมูล (信息)") for v in r if v is not None
        )
        self.assertIn("Sister Makeup", joined)
        self.assertIn("2569-05", joined)


if __name__ == "__main__":
    unittest.main()
