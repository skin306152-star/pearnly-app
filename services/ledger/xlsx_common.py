# -*- coding: utf-8 -*-
"""账簿工作簿的写表原语 —— 视觉沿用 services/excel/xlsx_style,公式策略住这里。

公式策略只有一条,但它是整个设计的地基:**跨表引用一律按票号 SUMIF,绝不用绝对单元格
引用。** 会计核对的标准动作是「把核对无误的行删掉」,`='สมุดรายวันทั่วไป'!$E$13` 这种
引用一删行就 #REF! 整表废;SUMIF 删行只让结果变小,不炸。

泰文表名跨表引用必须加单引号(`='สมุดรายวันทั่วไป'!…`)—— 不加的话部分 Excel/WPS 拒解析。
"""

from __future__ import annotations

from decimal import Decimal
from typing import Any, Optional, Sequence

from openpyxl.utils import get_column_letter

from services.excel import xlsx_style as sty
from services.sales_agg import vat

# 税率取 sales_agg.vat:公式算出来的必须与拆税算出来的同一个比,否则底稿与申报表对不上。
VAT_NUMERATOR, VAT_DENOMINATOR = vat.VAT_PART, vat.GROSS_PART

TITLE_ROW = 1
BANNER_ROW = 2
HINT_ROW = 3
HEADER_ROW = 4
FIRST_DATA_ROW = HEADER_ROW + 1


def quote_sheet(name: str) -> str:
    """泰文/含空格的表名进公式必须带单引号。表名里的单引号按 Excel 规则翻倍转义。"""
    return "'" + str(name).replace("'", "''") + "'"


def abs_range(sheet: Optional[str], col: int, row_start: int, row_end: int) -> str:
    """绝对**区间**引用(不是单格)。删区间内的行,Excel 自动缩范围,不产生 #REF!。"""
    letter = get_column_letter(col)
    body = f"${letter}${row_start}:${letter}${row_end}"
    return f"{quote_sheet(sheet)}!{body}" if sheet else body


def sumif(criteria_range: str, criteria: str, sum_range: str) -> str:
    return f"SUMIF({criteria_range},{criteria},{sum_range})"


def vat_included(gross_expr: str) -> str:
    """内含 VAT:gross × 7/107 四舍五入到分。与 services/sales_agg/vat.split_gross 同口径。"""
    return f"ROUND({gross_expr}*{VAT_NUMERATOR}/{VAT_DENOMINATOR},2)"


def write_title(ws, text: str, *, banner: str = "") -> None:
    """标题 + 科目码来源横幅。横幅不是装饰:它回答「这张表上的码是不是我账套里的码」。"""
    title = ws.cell(row=TITLE_ROW, column=1, value=text)
    title.font = sty.Font(bold=True, size=13)
    if banner:
        cell = ws.cell(row=BANNER_ROW, column=1, value=banner)
        cell.font = sty.Font(italic=True, color="9C4221")
        cell.alignment = sty.left()


def write_hint(ws, text: str) -> None:
    """表头之上的操作提示。放在冻结区之上,滚到表尾也还看得见 —— 提示要在她动手的地方。"""
    cell = ws.cell(row=HINT_ROW, column=1, value=text)
    cell.font = sty.Font(bold=True, color="2C5282")
    cell.alignment = sty.left()


def write_header(
    ws, headers: Sequence[str], widths: Sequence[int], *, row: int = HEADER_ROW, freeze: bool = True
) -> None:
    sty.write_header_row(ws, headers, widths, row=row, freeze=freeze)


def money_cell(ws, row: int, col: int, value: Any):
    cell = ws.cell(row=row, column=col, value=value)
    sty.style_cell(cell, align=sty.right(), fmt=sty.MONEY_FMT)
    return cell


def text_cell(ws, row: int, col: int, value: Any, *, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    sty.style_cell(cell, align=align or sty.left())
    return cell


def date_cell(ws, row: int, col: int, value):
    """日期按票面原样显示(ISO 文本)—— 我们不推算日期,只定读哪一个。"""
    return text_cell(ws, row, col, value.isoformat() if value else "", align=sty.center())


def bold_row(ws, row: int, cols: Sequence[int]) -> None:
    for col in cols:
        ws.cell(row=row, column=col).font = sty.Font(bold=True)


def hide_column(ws, col: int) -> None:
    ws.column_dimensions[get_column_letter(col)].hidden = True


def dec(value: Any) -> Decimal:
    return value if isinstance(value, Decimal) else Decimal(str(value or "0"))
