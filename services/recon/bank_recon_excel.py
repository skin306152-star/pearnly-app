# -*- coding: utf-8 -*-
"""bank_recon_excel.py · Pearnly · 4-sheet openpyxl reconciliation export.

Split verbatim from bank_recon_v2.py. Pure presentation: i18n labels +
openpyxl workbook build, no matching/judgement logic.
"""

import io
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

from services.recon.bank_recon_types import BankReconRow, BankReconSummary

# 4 语标签 + 翻译 helper · moved to bank_recon_excel_i18n.py
from services.recon.bank_recon_excel_i18n import (  # noqa: F401  facade-internal
    _t,
    _layer_label,
    _status_label,
)

# 使用说明文案 · moved to bank_recon_excel_usage.py
from services.recon.bank_recon_excel_usage import _USAGE_BLOCKS  # noqa: F401  facade-internal


def export_bank_recon_excel(
    recon_rows: List[BankReconRow],
    summary: BankReconSummary,
    lang: str = "th",
    task_info: Optional[Dict[str, Any]] = None,
    parse_info: Optional[Dict[str, Any]] = None,
    anchor_overrides: Optional[Dict[str, Dict[str, float]]] = None,
    anchor_ocr: Optional[Dict[str, float]] = None,
    warnings: Optional[List[str]] = None,
) -> bytes:
    """Generate Excel report with File Info + 4 data sheets, all headers i18n.

    P0.2 BUG-B-T2 v118.35.0.38 · anchor_overrides + anchor_ocr 来自 summary_json
    · anchor_overrides 非空时 · sheet 1 顶部加警示行 + 末尾加 "手动录入痕迹" section
    · 标黄(FFE082)被用户覆盖的 cell · 灰字显示 OCR 原值参考
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError("openpyxl not installed")

    lang = lang if lang in ("th", "en", "zh", "ja") else "th"

    wb = openpyxl.Workbook()

    # ── Color palette ──────────────────────────────────────────────────
    COLOR_HEADER = "2D6A4F"  # dark green
    COLOR_SUBHEAD = "52B788"  # medium green
    COLOR_MATCHED = "D8F3DC"  # light green
    COLOR_L2 = "FFF3CD"  # amber (date tolerance)
    COLOR_L3 = "FFE0CC"  # orange (amount only)
    COLOR_GL_ONLY = "E8D5F5"  # purple
    COLOR_ST_ONLY = "D4E6F1"  # blue
    COLOR_DIFF = "FFDAD6"  # red for non-zero diff
    COLOR_OK = "D8F3DC"  # green for zero diff
    COLOR_ROW_ALT = "F8F9FA"  # alternating row

    def _hdr_style(ws, row, col, text, color=COLOR_HEADER, bold=True, size=10):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=bold, color="FFFFFF", size=size)
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        return cell

    def _label_style(ws, row, col, text, bold=False):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(bold=bold, size=9)
        cell.alignment = Alignment(vertical="center")
        return cell

    def _num_style(ws, row, col, val, fmt="#,##0.00", fill_color=None):
        cell = ws.cell(row=row, column=col, value=val)
        cell.number_format = fmt
        cell.alignment = Alignment(horizontal="right", vertical="center")
        if fill_color:
            cell.fill = PatternFill("solid", fgColor=fill_color)
        return cell

    def _border_range(ws, min_row, max_row, min_col, max_col):
        thin = Side(style="thin", color="CCCCCC")
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                ws.cell(r, c).border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def _fmt_date(d: Optional[date]) -> str:
        if d is None:
            return ""
        return d.strftime("%d/%m/%Y")

    # ══════════════════════════════════════════════════════════════════
    # SHEET 1: สรุป (Consolidated Summary · v118.34)
    # 3 sections in one sheet (was 3 separate sheets pre-v118.34):
    #   A. Reconciliation summary (vertical itemized layout, 2 cols)
    #   B. File Info (parse diagnostics, folded to 2 cols)
    #   C. How to Use (usage instructions, merged A:B)
    # Style: 2-col (label | amount/status), clear color tiers:
    #   - Dark navy: title + main anchor rows (GL期末/账单期末)
    #   - Light gray: section headers
    #   - White: detail rows (each unmatched item itemized)
    #   - Blue: subtotal (计算期末余额)
    #   - Red/green: final diff
    # ══════════════════════════════════════════════════════════════════
    ws1 = wb.active  # reuse auto-created first sheet (was File Info pre-v118.34)
    ws1.title = _t("sh_summary", lang)
    ws1.sheet_view.showGridLines = False
    ws1.column_dimensions["A"].width = 78
    ws1.column_dimensions["B"].width = 22  # v118.33.13.6 · fit (7-digit) amounts with parens

    # Color palette
    NAVY = "1F2937"  # dark slate - main anchor rows
    NAVY_LIGHT = "374151"  # slightly lighter for sub-anchor
    SECTION_BG = "EEF2F6"  # very light blue-gray for section headers
    DETAIL_BG = "FFFFFF"
    DETAIL_ALT = "FAFBFC"
    SUBTOTAL_BG = "DBEAFE"  # soft blue for calc-close subtotal
    DIFF_OK_BG = "D1FAE5"  # mint green for zero diff
    DIFF_BAD_BG = "FEE2E2"  # soft red for non-zero diff
    INFO_BG = "F9FAFB"  # very subtle gray for bank/acct info

    NUM_FORMAT = "#,##0.00;[Red](#,##0.00)"

    def _fmt_d(d):
        if not d:
            return ""
        try:
            return d.strftime("%d/%m/%Y")
        except Exception:
            return ""

    def _title_row(row, text):
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws1.cell(row, 1, text)
        c.font = Font(bold=True, size=14, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=NAVY)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[row].height = 32

    def _header_row(row, label_text, amount_text):
        for col, txt in ((1, label_text), (2, amount_text)):
            c = ws1.cell(row, col, txt)
            c.font = Font(bold=True, size=11, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=NAVY)
            c.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[row].height = 26

    def _anchor_row(row, label, value, *, bg=NAVY, fg="FFFFFF", size=12, bold=True):
        a = ws1.cell(row, 1, label)
        a.font = Font(bold=bold, size=size, color=fg)
        a.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b = ws1.cell(row, 2, value)
        b.font = Font(bold=bold, size=size, color=fg)
        b.fill = PatternFill("solid", fgColor=bg)
        b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        b.number_format = NUM_FORMAT
        ws1.row_dimensions[row].height = 24

    def _section_row(row, label):
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws1.cell(row, 1, label)
        c.font = Font(bold=True, size=10, color="111827")
        c.fill = PatternFill("solid", fgColor=SECTION_BG)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.row_dimensions[row].height = 22

    def _detail_row(row, label, value, alt=False, italic=False, color="333333"):
        bg = DETAIL_ALT if alt else DETAIL_BG
        a = ws1.cell(row, 1, "  · " + (label if label else ""))
        a.font = Font(size=9, color=color, italic=italic)
        a.fill = PatternFill("solid", fgColor=bg)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=2, wrap_text=False)
        b = ws1.cell(row, 2, value)
        b.font = Font(size=10, color=color, italic=italic)
        b.fill = PatternFill("solid", fgColor=bg)
        b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        if isinstance(value, (int, float)):
            b.number_format = NUM_FORMAT
        ws1.row_dimensions[row].height = 18

    def _info_row(row, label, value):
        a = ws1.cell(row, 1, label)
        a.font = Font(size=10, color="6B7280")
        a.fill = PatternFill("solid", fgColor=INFO_BG)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        b = ws1.cell(row, 2, value)
        b.font = Font(size=10, color="111827", bold=True)
        b.fill = PatternFill("solid", fgColor=INFO_BG)
        b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
        ws1.row_dimensions[row].height = 20

    # ── 1. Title ──
    _RECON_TITLE = {
        "en": "Bank Reconciliation",
        "zh": "银行对账",
        "th": "กระทบยอด GL กับบัญชีธนาคาร",
        "ja": "銀行照合",
    }
    r = 1
    _title_row(r, f"{_RECON_TITLE.get(lang, 'Bank Reconciliation')} · {summary.bank_code.upper()}")
    r += 1

    # ── 2. Info: bank + GL account ──
    _info_row(r, _t("lbl_bank", lang), summary.bank_code.upper())
    r += 1
    _info_row(r, _t("lbl_gl_acct", lang), summary.gl_account_code or "—")
    r += 1

    # ── 2b. v118.35.0.61 · 勾稽匹配诚实化:0/极低匹配时顶部红字横幅 ──
    # 防『差异=0 是会计恒等式假象』误导用户以为对平。matched/总项 自算 · 历史任务也生效。
    _matched_n = sum(1 for rr in recon_rows if rr.match_status == "matched")
    _total_items = len(recon_rows) or 1
    _match_rate = _matched_n / _total_items
    _low_match = (_matched_n == 0) or (_match_rate < 0.10 and _total_items >= 10)

    def _banner(row, text, *, bg="FEE2E2", fg="991B1B"):
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        c = ws1.cell(row=row, column=1, value=text)
        c.font = Font(bold=True, size=10, color=fg)
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
        ws1.row_dimensions[row].height = 46

    _banner_msgs = []
    if _matched_n == 0:
        _banner_msgs.append(_t("banner_no_match", lang))
    elif _low_match:
        _banner_msgs.append(
            _t("banner_low_match", lang).format(n=_matched_n, r=round(_match_rate * 100, 1))
        )
    # 调用方传入的输入不匹配警告(期间/科目/规模)· 与前端提示条同源
    for _w in warnings or []:
        if _w:
            _banner_msgs.append(str(_w))
    if _banner_msgs:
        r += 1  # spacer
        for _bm in _banner_msgs:
            _banner(r, _bm)
            r += 1
    # P0.2 BUG-B-T2 v118.35.0.38 · 有 anchor 被覆盖 → 顶部一行警示『含手动录入 · 看末尾对照』
    if anchor_overrides:
        r += 1  # 警示前空 1 行 · 视觉舒服
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        warn_cell = ws1.cell(row=r, column=1, value=_t("lbl_manual_warn", lang))
        warn_cell.font = Font(bold=True, size=10, color="92400E")
        warn_cell.fill = PatternFill("solid", fgColor="FEF3C7")  # 浅黄底
        warn_cell.alignment = Alignment(
            horizontal="left", vertical="center", indent=1, wrap_text=True
        )
        ws1.row_dimensions[r].height = 24
        r += 1
    r += 1  # blank-row spacer

    # ── 3. Column headers: 项目说明 | 金额 ──
    _header_row(r, _t("col_summary_item", lang), _t("col_summary_amount", lang))
    r += 1

    # ── 4. Start anchor: GL 期末余额 ──
    _anchor_row(r, _t("lbl_gl_close", lang), summary.gl_closing, bg=NAVY, size=12)
    r += 1

    # ── 5. + 期初差异 (signed) ──
    sign_char = "+" if summary.opening_diff >= 0 else "−"
    _section_row(r, f"{sign_char} {_t('sec_open_diff_expand', lang)}")
    r += 1
    open_diff_label = f"{summary.stmt_opening:,.2f} − {summary.gl_opening:,.2f}"
    _detail_row(r, open_diff_label, summary.opening_diff)
    r += 1

    # ── 6. Itemized unmatched sections (4 categories) ──
    def _add_itemized(sign_int, section_key, status_filter, get_fields):
        """sign_int ∈ {-1, +1}.  get_fields(row) → (date_str, doc, desc, amt)."""
        nonlocal r
        ch = "+" if sign_int > 0 else "−"
        rows_match = [rr for rr in recon_rows if rr.match_status == status_filter]
        _section_row(r, f"{ch} {_t(section_key, lang)}")
        r += 1
        if not rows_match:
            _detail_row(r, _t("detail_no_items", lang), 0.0, italic=True, color="9CA3AF")
            r += 1
            return
        for i, rr in enumerate(rows_match):
            date_str, doc, desc, amt = get_fields(rr)
            parts = [p for p in (date_str, doc, desc) if p]
            label = " · ".join(parts) if parts else ""
            _detail_row(r, label, sign_int * (amt or 0), alt=(i % 2 == 1))
            r += 1

    def _gl_fields(rr):
        amt = rr.gl_debit if rr.match_status == "gl_debit_only" else rr.gl_credit
        return _fmt_d(rr.gl_date), rr.gl_doc_no or "", rr.gl_desc or "", amt

    def _stmt_fields(rr):
        amt = rr.stmt_withdrawal if rr.match_status == "stmt_withdrawal_only" else rr.stmt_deposit
        return _fmt_d(rr.stmt_date), "", rr.stmt_desc or "", amt

    _add_itemized(-1, "sec_gl_debit_only_full", "gl_debit_only", _gl_fields)
    _add_itemized(+1, "sec_gl_credit_only_full", "gl_credit_only", _gl_fields)
    _add_itemized(-1, "sec_stmt_wd_only_full", "stmt_withdrawal_only", _stmt_fields)
    _add_itemized(+1, "sec_stmt_dep_only_full", "stmt_deposit_only", _stmt_fields)

    # ── 7. Subtotal: 计算期末余额 (light blue) ──
    r += 1  # spacer
    _anchor_row(
        r,
        _t("lbl_formula_calc", lang),
        summary.formula_stmt_closing,
        bg=SUBTOTAL_BG,
        fg="1E3A8A",
        size=12,
    )
    r += 1

    # ── 8. Target: 账单期末余额 (dark anchor, same style as GL_close) ──
    _anchor_row(r, _t("lbl_stmt_close", lang), summary.stmt_closing, bg=NAVY, size=12)
    r += 1

    # ── 8b. v118.35.0.61 · 真实勾稽指标:已匹配笔数 + 匹配率(诚实化核心) ──
    _info_row(r, _t("lbl_matched_n", lang), f"{_matched_n} / {len(recon_rows)}")
    r += 1
    _info_row(r, _t("lbl_match_rate", lang), f"{round(_match_rate * 100, 1)}%")
    r += 1

    # ── 9. Final: 差异 ──
    # v118.35.0.61 · 诚实化:diff≈0 但匹配率极低时 → 不染绿(那只是会计恒等式 · 不是真对平)
    diff_zero = abs(summary.formula_diff) < 0.05
    diff_ok = diff_zero and not _low_match
    diff_bg = DIFF_OK_BG if diff_ok else DIFF_BAD_BG
    diff_fg = "065F46" if diff_ok else "991B1B"
    _anchor_row(
        r, _t("lbl_formula_diff", lang), summary.formula_diff, bg=diff_bg, fg=diff_fg, size=13
    )
    r += 1
    if diff_zero and _low_match:
        # diff 为 0 却几乎没匹配 → 明确告知这是恒等式 · 不代表勾稽成功
        _detail_row(
            r, _t("diff_identity_note", lang).format(n=_matched_n), "", italic=True, color="991B1B"
        )
        r += 1

    # ── 10. OCR accuracy check (only if any warnings) ──
    warn_balance = sum(1 for rr in recon_rows if rr.stmt_balance_ok is False)
    warn_lowconf = sum(1 for rr in recon_rows if rr.stmt_confidence == "low")
    fixed_n = sum(1 for rr in recon_rows if getattr(rr, "stmt_autocorrected", False))
    if warn_balance or warn_lowconf or fixed_n:
        r += 1  # spacer
        _section_row(r, _t("lbl_ocr_check", lang))
        r += 1
        if fixed_n:
            # v118.35.0.62 · 系统按余额自动修正的行 · 黄字 · 建议复核
            _detail_row(r, _t("lbl_ocr_autofixed", lang), fixed_n, color="B45309")
            r += 1
        if warn_balance:
            _detail_row(r, _t("lbl_ocr_bal_warn", lang), warn_balance, color="DC2626")
            r += 1
        if warn_lowconf:
            _detail_row(r, _t("lbl_ocr_lowconf", lang), warn_lowconf, color="EA580C", alt=True)
            r += 1

    # ── 11. File Info sub-section ──
    r += 2  # spacer between summary and file info
    _section_row(r, _t("sh_fileinfo", lang))
    r += 1

    fi_pairs: List[Tuple[str, str]] = []
    if parse_info:
        for f in parse_info.get("stmt_files") or []:
            ok_status = (
                _t("fi_ok", lang)
                if (f.get("ok") and f.get("rows", 0) > 0)
                else (
                    _t("fi_warn", lang)
                    if (f.get("ok") and f.get("rows", 0) == 0)
                    else _t("fi_fail", lang)
                )
            )
            bank_part = f" · {f.get('bank_code')}" if f.get("bank_code") else ""
            err_part = f" · {f.get('error')}" if f.get("error") else ""
            label = f"{_t('fi_stmt_type', lang)}: {f.get('file', '')} · {f.get('rows', 0)} {_t('fi_rows', lang).lower()}{bank_part}{err_part}"
            fi_pairs.append((label, ok_status))
        for f in parse_info.get("gl_files") or []:
            ok_status = (
                _t("fi_ok", lang)
                if (f.get("ok") and f.get("rows", 0) > 0)
                else (
                    _t("fi_warn", lang)
                    if (f.get("ok") and f.get("rows", 0) == 0)
                    else _t("fi_fail", lang)
                )
            )
            accts = ", ".join(f.get("accounts") or [])
            acct_part = f" · {accts}" if accts else ""
            err_part = f" · {f.get('error')}" if f.get("error") else ""
            label = f"{_t('fi_gl_type', lang)}: {f.get('file', '')} · {f.get('rows', 0)} {_t('fi_rows', lang).lower()}{acct_part}{err_part}"
            fi_pairs.append((label, ok_status))
    elif task_info:
        for fname in (task_info.get("stmt_files") or "").split(";"):
            fname = fname.strip()
            if not fname:
                continue
            rc = task_info.get("stmt_row_count", 0)
            ok_status = _t("fi_ok", lang) if rc > 0 else _t("fi_warn", lang)
            bank_code = task_info.get("bank_code", "")
            bank_part = f" · {bank_code}" if bank_code else ""
            label = f"{_t('fi_stmt_type', lang)}: {fname} · {rc} {_t('fi_rows', lang).lower()}{bank_part}"
            fi_pairs.append((label, ok_status))
        for fname in (task_info.get("gl_files") or "").split(";"):
            fname = fname.strip()
            if not fname:
                continue
            rc = task_info.get("gl_row_count", 0)
            ok_status = _t("fi_ok", lang) if rc > 0 else _t("fi_warn", lang)
            gl_acct = task_info.get("gl_account", "")
            acct_part = f" · {gl_acct}" if gl_acct else ""
            label = (
                f"{_t('fi_gl_type', lang)}: {fname} · {rc} {_t('fi_rows', lang).lower()}{acct_part}"
            )
            fi_pairs.append((label, ok_status))

    _fi_status_colors = {
        _t("fi_ok", lang): "D8F3DC",
        _t("fi_warn", lang): "FFF3CD",
        _t("fi_fail", lang): "FFDAD6",
    }
    if fi_pairs:
        for label, status_text in fi_pairs:
            a = ws1.cell(r, 1, label)
            a.font = Font(size=9, color="111827")
            a.fill = PatternFill("solid", fgColor=INFO_BG)
            a.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
            b = ws1.cell(r, 2, status_text)
            b.font = Font(size=9, bold=True, color="111827")
            b.fill = PatternFill("solid", fgColor=_fi_status_colors.get(status_text, INFO_BG))
            b.alignment = Alignment(horizontal="center", vertical="center")
            ws1.row_dimensions[r].height = 22
            r += 1
    else:
        a = ws1.cell(r, 1, _t("detail_no_items", lang))
        a.font = Font(size=9, italic=True, color="9CA3AF")
        a.fill = PatternFill("solid", fgColor=INFO_BG)
        a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        r += 1

    # ── 11b. P0.2 BUG-B-T2 v118.35.0.38 · 手动录入痕迹 section
    # 只在 anchor_overrides 非空时显示 · 列 3 个 anchor 的 OCR 值 vs 用户填值 vs 差额
    # cell user_value 标黄 (FFE082) · cell ocr_value 灰字 · 给 P0.3 历史详情对照同款数据
    if anchor_overrides:
        r += 2  # spacer
        _section_row(r, _t("sec_manual_entry", lang))
        r += 1
        # 4 列 mini-header(label | OCR | user | diff)· 暂借 2 列 layout · 一行 4 segment 用单元格 + 文本
        # 因 Sheet 1 是 2 列 layout · 这里特殊用 cell A 写 anchor label · cell B 写 'OCR XXX → 用户 YYY (差 ZZZ)' 标黄
        _ANCHOR_LABEL_KEYS = [
            ("stmt_opening", "lbl_anchor_stmt_open"),
            ("gl_opening", "lbl_anchor_gl_open"),
            ("gl_closing", "lbl_anchor_gl_close"),
            ("stmt_closing", "lbl_anchor_stmt_close"),  # BUG-FIX-T3 v118.35.0.44 · 4th anchor
        ]
        YELLOW_FILL = PatternFill("solid", fgColor="FFE082")
        for idx, (anchor_key, lbl_key) in enumerate(_ANCHOR_LABEL_KEYS):
            ov = (anchor_overrides or {}).get(anchor_key)
            if not ov:
                continue
            ocr_val = float(ov.get("ocr") or 0.0)
            user_val = float(ov.get("user") or 0.0)
            diff = user_val - ocr_val
            # cell A · anchor label
            a = ws1.cell(r, 1, "  · " + _t(lbl_key, lang))
            a.font = Font(size=10, color="111827")
            a.fill = PatternFill("solid", fgColor=DETAIL_BG)
            a.alignment = Alignment(horizontal="left", vertical="center", indent=2)
            # cell B · 黄底 · user 值粗体 + 后跟灰色 OCR 原值 + diff
            b = ws1.cell(r, 2, user_val)
            b.font = Font(size=11, bold=True, color="92400E")  # 暖棕色文字
            b.fill = YELLOW_FILL
            b.alignment = Alignment(horizontal="right", vertical="center", indent=1)
            b.number_format = NUM_FORMAT
            # cell comment 写完整对照(hover Excel 看)
            try:
                from openpyxl.comments import Comment as _XLComment

                b.comment = _XLComment(
                    f"OCR: {ocr_val:,.2f}\nUser: {user_val:,.2f}\nDiff: {diff:+,.2f}", "Pearnly"
                )
            except Exception:
                pass  # comment 失败不阻塞导出 · 数据本身已经在 cell value 里
            ws1.row_dimensions[r].height = 22
            r += 1
        # 3 行 "OCR 原值" 灰字小字提示行(每 anchor 1 行)· cell A 留空 · cell B 显示 OCR 值
        # 简化:不加 OCR 原值小字行 · 已有 comment + 后续历史详情(P0.3)可看
        # 列头提示(标在 section 末尾)· 4 语
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        ref_cell = ws1.cell(
            r,
            1,
            "ℹ "
            + _t("col_manual_ocr", lang)
            + " / "
            + _t("col_manual_user", lang)
            + " · "
            + _t("col_manual_diff", lang),
        )
        ref_cell.font = Font(size=9, italic=True, color="6B7280")
        ref_cell.fill = PatternFill("solid", fgColor=INFO_BG)
        ref_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.row_dimensions[r].height = 18
        r += 1

    # ── 12. How to Use sub-section ──
    r += 2  # spacer
    _section_row(r, _t("sh_usage", lang))
    r += 1

    usage_block = _USAGE_BLOCKS.get(lang, _USAGE_BLOCKS["en"])
    for text, bold in usage_block:
        ws1.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        cell = ws1.cell(row=r, column=1, value=text)
        cell.font = Font(bold=bold, size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        if bold and text:
            cell.fill = PatternFill("solid", fgColor="E5E7EB")
        ws1.row_dimensions[r].height = 22 if bold else 18
        r += 1

    # ── 13. Final border around the whole Summary sheet ──
    _border_range(ws1, 1, r - 1, 1, 2)
    # Freeze header so it stays visible while scrolling
    ws1.freeze_panes = "A6"

    # ══════════════════════════════════════════════════════════════════
    # SHEET 2: ผลการจับคู่ (Consolidated Match Results · v118.34)
    # Combines what were previously 3 sheets (matched + unmatched_gl + unmatched_stmt).
    # First column "Status" distinguishes:
    #   - "✓ Matched"  (matched rows, color by match layer L1/L2/L3)
    #   - "GL Debit/Credit Only"  (purple tint)
    #   - "Stmt Withdrawal/Deposit Only"  (blue tint)
    # Match Layer column shows L1/L2/L3 for matched rows, "—" for unmatched.
    # ══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(_t("sh_match_results", lang))
    ws2.sheet_view.showGridLines = False

    match_cols = [
        (_t("col_status", lang), 18),
        (_t("col_match_layer", lang), 12),
        (_t("col_date", lang), 12),
        (_t("col_desc", lang), 26),
        (_t("col_withdrawal", lang), 12),
        (_t("col_deposit", lang), 12),
        (_t("col_balance", lang), 12),
        (_t("col_gl_date", lang), 12),
        (_t("col_gl_doc", lang), 14),
        (_t("col_gl_acct", lang), 11),
        (_t("col_gl_desc", lang), 26),
        (_t("col_gl_debit", lang), 12),
        (_t("col_gl_credit", lang), 12),
        (_t("col_date_diff", lang), 10),
        (_t("col_source_stmt", lang), 18),
        (_t("col_source_gl", lang), 18),
    ]
    for ci, (hdr, width) in enumerate(match_cols, 1):
        _hdr_style(ws2, 1, ci, hdr)
        ws2.column_dimensions[get_column_letter(ci)].width = width

    # Group + sort the recon_rows by category
    matched_rows_for_export = sorted(
        [rr for rr in recon_rows if rr.match_status == "matched"],
        key=lambda x: (x.stmt_date or date.min, x.gl_date or date.min),
    )
    gl_only_rows_for_export = sorted(
        [rr for rr in recon_rows if rr.match_status in ("gl_debit_only", "gl_credit_only")],
        key=lambda x: (x.gl_date or date.min, x.gl_doc_no or ""),
    )
    stmt_only_rows_for_export = sorted(
        [
            rr
            for rr in recon_rows
            if rr.match_status in ("stmt_withdrawal_only", "stmt_deposit_only")
        ],
        key=lambda x: (x.stmt_date or date.min, x.stmt_desc or ""),
    )

    _DASH = "—"

    ri = 2
    # Matched block (tinted by match layer)
    for row in matched_rows_for_export:
        layer_fill_color = (
            COLOR_MATCHED
            if row.match_layer == 1
            else COLOR_L2 if row.match_layer == 2 else COLOR_L3
        )
        fill = PatternFill("solid", fgColor=layer_fill_color)
        vals = [
            _t("status_matched", lang),
            _layer_label(row.match_layer, lang),
            _fmt_date(row.stmt_date),
            row.stmt_desc,
            row.stmt_withdrawal or "",
            row.stmt_deposit or "",
            row.stmt_balance or "",
            _fmt_date(row.gl_date),
            row.gl_doc_no,
            row.gl_account_code,
            row.gl_desc,
            row.gl_debit or "",
            row.gl_credit or "",
            row.date_diff_days if row.date_diff_days is not None else "",
            row.source_stmt_file,
            row.source_gl_file,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = fill
            if isinstance(val, float) and val != "":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=9)
        ri += 1

    # GL-only block (purple tint)
    for row in gl_only_rows_for_export:
        fill = PatternFill("solid", fgColor=COLOR_GL_ONLY if ri % 2 == 0 else "F3E8FF")
        vals = [
            _status_label(row.match_status, lang),
            _DASH,
            "",  # stmt date
            "",  # stmt desc
            "",  # stmt withdrawal
            "",  # stmt deposit
            "",  # stmt balance
            _fmt_date(row.gl_date),
            row.gl_doc_no,
            row.gl_account_code,
            row.gl_desc,
            row.gl_debit or "",
            row.gl_credit or "",
            "",  # date diff
            "",  # source stmt
            row.source_gl_file,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = fill
            if isinstance(val, float) and val != "":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=9)
        ri += 1

    # Stmt-only block (blue tint)
    for row in stmt_only_rows_for_export:
        fill = PatternFill("solid", fgColor=COLOR_ST_ONLY if ri % 2 == 0 else "EBF5FB")
        vals = [
            _status_label(row.match_status, lang),
            _DASH,
            _fmt_date(row.stmt_date),
            row.stmt_desc,
            row.stmt_withdrawal or "",
            row.stmt_deposit or "",
            row.stmt_balance or "",
            "",  # gl date
            "",  # gl doc
            "",  # gl acct
            "",  # gl desc
            "",  # gl debit
            "",  # gl credit
            "",  # date diff
            row.source_stmt_file,
            "",  # source gl
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws2.cell(ri, ci, val)
            cell.fill = fill
            if isinstance(val, float) and val != "":
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
            cell.font = Font(size=9)
        ri += 1

    _border_range(ws2, 1, max(1, ri - 1), 1, len(match_cols))
    ws2.freeze_panes = "C2"  # freeze status + match layer cols

    # ══════════════════════════════════════════════════════════════════
    # SHEET 5: Statement Detail (all parsed statement rows + OCR check)
    # v118.33.13.0
    # ══════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet(_t("sh_stmt_detail", lang))
    ws5.sheet_view.showGridLines = False

    sd_cols = [
        (_t("col_date", lang), 12),
        (_t("col_desc", lang), 38),
        (_t("col_withdrawal", lang), 14),
        (_t("col_deposit", lang), 14),
        (_t("col_balance", lang), 14),
        (_t("col_confidence", lang), 12),
        (_t("col_balance_ok", lang), 12),
        (_t("col_source_file", lang), 22),
    ]
    for ci, (hdr, width) in enumerate(sd_cols, 1):
        _hdr_style(ws5, 1, ci, hdr)
        ws5.column_dimensions[get_column_letter(ci)].width = width

    CONF_LBL = {
        "high": _t("conf_high", lang),
        "medium": _t("conf_medium", lang),
        "low": _t("conf_low", lang),
    }
    CONF_FILL = {"high": "D8F3DC", "medium": "FFF3CD", "low": "FFDAD6"}

    # Source: stmt-side rows (all of them — matched + stmt-only)
    stmt_side_rows = [
        r
        for r in recon_rows
        if r.stmt_date is not None
        or r.stmt_balance != 0
        or r.stmt_withdrawal != 0
        or r.stmt_deposit != 0
    ]
    # Sort by stmt_date
    stmt_side_rows.sort(key=lambda x: (x.stmt_date or date.min, x.stmt_desc))

    for ri, row in enumerate(stmt_side_rows, 2):
        conf = (row.stmt_confidence or "high").lower()
        if getattr(row, "stmt_autocorrected", False):
            # v118.35.0.62 · 系统按余额自动修正过 · 显式标黄『已修正』· 透明 · 提示可复核
            bal_str = _t("bal_fixed", lang)
            bal_fill = "FFE082"
        elif row.stmt_balance_ok is True:
            bal_str = _t("bal_ok", lang)
            bal_fill = "D8F3DC"
        elif row.stmt_balance_ok is False:
            bal_str = _t("bal_warn", lang)
            bal_fill = "FFDAD6"
        else:
            bal_str = _t("bal_na", lang)
            bal_fill = None
        vals = [
            _fmt_date(row.stmt_date),
            row.stmt_desc,
            row.stmt_withdrawal or "",
            row.stmt_deposit or "",
            row.stmt_balance or "",
            CONF_LBL.get(conf, conf),
            bal_str,
            row.source_stmt_file,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws5.cell(ri, ci, val)
            cell.font = Font(size=9)
            if isinstance(val, float) and val:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            # Highlight confidence/balance columns
            if ci == 6:
                cell.fill = PatternFill("solid", fgColor=CONF_FILL.get(conf, "FFFFFF"))
                cell.alignment = Alignment(horizontal="center")
            if ci == 7 and bal_fill:
                cell.fill = PatternFill("solid", fgColor=bal_fill)
                cell.alignment = Alignment(horizontal="center")
        # Tint the whole row red if balance check failed
        if row.stmt_balance_ok is False:
            for ci in range(1, len(vals) + 1):
                if ws5.cell(ri, ci).fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    ws5.cell(ri, ci).fill = PatternFill("solid", fgColor="FEF2F2")

    _border_range(ws5, 1, max(1, len(stmt_side_rows) + 1), 1, len(sd_cols))
    ws5.freeze_panes = "A2"

    # ══════════════════════════════════════════════════════════════════
    # SHEET 6: GL Detail (all GL rows reconstructed from recon_rows)
    # v118.34 · Mirrors Sheet 5 (Statement Detail) — same visual idiom
    # ══════════════════════════════════════════════════════════════════
    ws_gl = wb.create_sheet(_t("sh_gl_detail", lang))
    ws_gl.sheet_view.showGridLines = False

    gld_cols = [
        (_t("col_date", lang), 12),
        (_t("col_doc_no", lang), 16),
        (_t("col_account_code", lang), 14),
        (_t("col_desc", lang), 38),
        (_t("col_debit", lang), 14),
        (_t("col_credit", lang), 14),
        (_t("col_source_file", lang), 22),
    ]
    for ci, (hdr, width) in enumerate(gld_cols, 1):
        _hdr_style(ws_gl, 1, ci, hdr)
        ws_gl.column_dimensions[get_column_letter(ci)].width = width

    # Source: every recon_row that carries GL data
    # (matched rows + gl_debit_only + gl_credit_only).
    # Stmt-only rows have no GL data → excluded.
    gl_data_rows = [
        r
        for r in recon_rows
        if r.match_status == "matched" or r.match_status in ("gl_debit_only", "gl_credit_only")
    ]
    gl_data_rows.sort(
        key=lambda x: (x.gl_date or date.min, x.gl_doc_no or "", x.gl_account_code or "")
    )

    for ri, row in enumerate(gl_data_rows, 2):
        alt_fill = "F8F9FA" if ri % 2 == 0 else None
        vals = [
            _fmt_date(row.gl_date),
            row.gl_doc_no,
            row.gl_account_code,
            row.gl_desc,
            row.gl_debit if row.gl_debit else "",
            row.gl_credit if row.gl_credit else "",
            row.source_gl_file,
        ]
        for ci, val in enumerate(vals, 1):
            cell = ws_gl.cell(ri, ci, val)
            cell.font = Font(size=9)
            if isinstance(val, float) and val:
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center")
            if alt_fill:
                cell.fill = PatternFill("solid", fgColor=alt_fill)

    _border_range(ws_gl, 1, max(1, len(gl_data_rows) + 1), 1, len(gld_cols))
    ws_gl.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
