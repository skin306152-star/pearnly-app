# -*- coding: utf-8 -*-
"""MR.ERP 分类账(สมุดแยกประเภท xlsx)→ GlRow 适配器(T4b-SM · F2 对平 erp 侧 · Decimal 全程)。

MR.ERP 分类账不是纯分录表,而是逐科目分组:每科目一个 `[扁平码 科目名]` 标题行 → 若干明细行
(วันที่日期 | สมุด账簿 | ใบสำคัญ凭证 | คำอธิบาย摘要 | เดบิต借 | เครดิต贷 | ยอดคงเหลือ余额)→ 一个
`รวม` 小计行。本层按行判据分流:明细行第0列匹配佛历日期 `^\\d{4}-\\d\\d-\\d\\d`,标题行匹配
`^\\d{4}-\\d\\d\\s+名`,`รวม`/表头/公司抬头/空行一律跳过——把明细行整形成 GlRow(account_code=
当前科目扁平码如 1111-01,debit/credit 按借贷两列直取)。

金额以字符串直进 Decimal(不经 float 中转),仅在 GlRow 边界转 float(GlRow 是展示型 float 契约,
下游 shadow_gl_recon 经 str() 回 Decimal 清算,两位小数往返无损——与 express_gl_adapter 同口径)。
日期 YYYY-MM-DD 佛历(年减 543)如实解析但不参与对平(F2 只聚合科目发生额);日期单元格可能是
str 或 datetime(openpyxl 两种都遇到),两路都归一。识别 = xlsx 且表头含 วันที่/สมุด/เดบิต/เครดิต
(逐科目分组特征),与试算平衡表(รหัสบัญชี 抬头的纯表)及银行对账 xlsx 区分开。

openpyxl read_only + data_only,句柄用完即 close(Windows 文件锁)。
"""

from __future__ import annotations

import io
import re
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

import openpyxl

from services.recon.bank_recon_types import GlRow

_ZERO = Decimal("0")
_BE_OFFSET = 543  # 佛历年 → 公历年

_XLSX_EXTS = ("xlsx", "xlsm")
_TOTAL_MARK = "รวม"  # 小计行首列
# 明细行:第0列是佛历日期(2569-05-28...);标题行:扁平码 + 空格 + 科目名(1111-01 เงินสด)
_DETAIL_DATE = re.compile(r"^\d{4}-\d\d-\d\d")
_TITLE = re.compile(r"^(\d{4}-\d\d)\s+(.+)$")
# 表头识别 token:逐科目分组分类账表头四列俱全(วันที่/สมุด/เดบิต/เครดิต)。สมุด(账簿列)是
# 分组分类账独有——试算平衡表首列是 รหัสบัญชี 且无 วันที่;科目码列成表(รหัสบัญชี 当列)的扁平
# GL 有 วันที่/เดบิต/เครดิต 却无 สมุด,靠 สมุด + 结构上的科目标题行两条一起把它们排除。
_HEADER_TOKENS = ("วันที่", "สมุด", "เดบิต", "เครดิต")


def _dec(value) -> Decimal:
    """金额单元格 → Decimal(去千分位,空/解不出 → 0)。不经 float 中转。"""
    if value is None:
        return _ZERO
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    try:
        s = str(value).replace(",", "").strip()
        return Decimal(s) if s else _ZERO
    except InvalidOperation:
        return _ZERO


def _be_date(cell) -> date | None:
    """佛历日期单元格(datetime 或 `YYYY-MM-DD ...` 字符串)→ 公历 date(年减 543);解不出 → None。"""
    if isinstance(cell, datetime):
        y, m, d = cell.year, cell.month, cell.day
    else:
        s = str(cell or "").strip()[:10]
        if not re.fullmatch(r"\d{4}-\d\d-\d\d", s):
            return None
        y, m, d = (int(p) for p in s.split("-"))
    try:
        return date(y - _BE_OFFSET, m, d)
    except ValueError:
        return None


def _cell(cells, i: int):
    return cells[i] if i < len(cells) else None


def _s(cell) -> str:
    return "" if cell is None else str(cell).strip()


def is_mrerp_gl(data: bytes, filename: str) -> bool:
    """xlsx 且前若干行同时具备分类账表头(วันที่/สมุด/เดบิต/เครดิต)与逐科目分组标题行 → MR.ERP
    分类账(分流到本适配器)。两条一起判:光有 date/debit/credit 列的扁平 GL 或试算平衡表都不命中。"""
    if (filename or "").rsplit(".", 1)[-1].lower() not in _XLSX_EXTS:
        return False
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception:
        return False
    try:
        has_header = has_title = False
        for cells in wb.active.iter_rows(min_row=1, max_row=15, values_only=True):
            tokens = {_s(c) for c in cells}
            if all(t in tokens for t in _HEADER_TOKENS):
                has_header = True
            head = _s(_cell(cells, 0))
            if head != _TOTAL_MARK and not _DETAIL_DATE.match(head) and _TITLE.match(head):
                has_title = True
            if has_header and has_title:
                return True
        return False
    finally:
        wb.close()


def parse_mrerp_gl_xlsx(data: bytes, filename: str = ""):
    """MR.ERP 分类账 xlsx 字节 → (GlRow 列表, 行级丢弃清单)。

    逐行分流:`รวม` 小计/表头/抬头/空行跳过;标题行切换当前科目;明细行按当前科目扁平码建 GlRow
    (debit=เดบิต列, credit=เครดิต列)。明细行出现在任何标题行之前(结构异常)→ 如实丢一条 issue
    不臆造归属。金额全程 Decimal,仅 GlRow 边界转 float。
    """
    out: list[GlRow] = []
    issues: list[str] = []
    current: str | None = None

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for line_no, cells in enumerate(wb.active.iter_rows(values_only=True), start=1):
            head = _s(_cell(cells, 0))
            if not head or head == _TOTAL_MARK:
                continue
            if _DETAIL_DATE.match(head):
                if current is None:
                    issues.append(f"line {line_no}: 明细行无所属科目标题,如实丢弃")
                    continue
                debit = _dec(_cell(cells, 4))
                credit = _dec(_cell(cells, 5))
                out.append(
                    GlRow(
                        date=_be_date(_cell(cells, 0)),
                        doc_no=_s(_cell(cells, 2)),
                        account_code=current,
                        description=_s(_cell(cells, 3)),
                        debit=float(debit),
                        credit=float(credit),
                        source_file=filename,
                        balance=float(_dec(_cell(cells, 6))),
                    )
                )
                continue
            m = _TITLE.match(head)
            if m:
                current = m.group(1)
    finally:
        wb.close()
    return out, issues


def iter_account_titles(data: bytes) -> list[dict]:
    """扫科目标题行 → [{code, name_th}](科目桥的输入;MR.ERP 无独立科目表,从 GL 标题行提取)。"""
    accounts: list[dict] = []
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        for cells in wb.active.iter_rows(values_only=True):
            head = _s(_cell(cells, 0))
            if not head or head == _TOTAL_MARK or _DETAIL_DATE.match(head):
                continue
            m = _TITLE.match(head)
            if m:
                accounts.append({"code": m.group(1), "name_th": m.group(2).strip()})
    finally:
        wb.close()
    return accounts
