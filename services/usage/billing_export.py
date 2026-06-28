# -*- coding: utf-8 -*-
"""Pearnly · 账单明细导出(扣费 / 充值 / 识别 三 sheet 工作簿)。

build_billing_xlsx 把三类记录写进一个 xlsx,每类一个 sheet:
  扣费记录(credit_transactions · usage + subscription)
  充值记录(topup_requests · 含审核状态)
  识别记录(ocr_history · 派生三态 成功/待复核/失败)
表头、sheet 名、状态/计费词随系统语言(zh/en/th/ja)。金额列用会计格式;
计费如实区分「实扣 / 套餐内免费 / 月费」、识别如实标「成功 / 失败」、充值如实标
「到账 / 待审核 / 驳回」——不混淆已扣未扣、成功失败。
"""

from __future__ import annotations

import io
from typing import Any, Dict, List

# 四语言词典 · 仅本导出用(与 usage_report._I18N 平行 · 列更全)。
_I18N: Dict[str, Dict[str, str]] = {
    "zh": {
        "title": "Pearnly 账单明细",
        "company": "公司",
        "sheet_usage": "扣费记录",
        "sheet_topup": "充值记录",
        "sheet_ocr": "识别记录",
        "c_date": "时间",
        "c_type": "类型",
        "c_billing": "计费",
        "c_desc": "说明 / 文件",
        "c_pages": "张数",
        "c_amount": "金额",
        "c_balance": "余额",
        "c_payer": "付款人",
        "c_status": "状态",
        "c_review_time": "审核时间",
        "c_note": "备注",
        "c_file": "文件名",
        "c_invoice": "发票号",
        "c_seller": "卖方",
        "c_source": "来源",
        "t_scan": "扫描扣费",
        "t_sub": "订阅 / 续订",
        "b_charged": "实扣",
        "b_free": "套餐内免费",
        "b_monthly": "月费",
        "tp_pending": "待审核",
        "tp_approved": "已到账",
        "tp_rejected": "已驳回",
        "o_confirmed": "识别成功",
        "o_pending": "待复核",
        "o_failed": "识别失败",
        "s_upload": "上传",
        "s_line": "LINE",
        "s_email": "邮件",
        "empty": "暂无记录",
    },
    "en": {
        "title": "Pearnly Billing Detail",
        "company": "Company",
        "sheet_usage": "Charges",
        "sheet_topup": "Top-ups",
        "sheet_ocr": "Recognition",
        "c_date": "Time",
        "c_type": "Type",
        "c_billing": "Billing",
        "c_desc": "Detail / File",
        "c_pages": "Pages",
        "c_amount": "Amount",
        "c_balance": "Balance",
        "c_payer": "Payer",
        "c_status": "Status",
        "c_review_time": "Reviewed",
        "c_note": "Note",
        "c_file": "File Name",
        "c_invoice": "Invoice No.",
        "c_seller": "Seller",
        "c_source": "Source",
        "t_scan": "Scan charge",
        "t_sub": "Subscription",
        "b_charged": "Charged",
        "b_free": "In quota (free)",
        "b_monthly": "Monthly fee",
        "tp_pending": "Pending",
        "tp_approved": "Credited",
        "tp_rejected": "Rejected",
        "o_confirmed": "Success",
        "o_pending": "Needs review",
        "o_failed": "Failed",
        "s_upload": "Upload",
        "s_line": "LINE",
        "s_email": "Email",
        "empty": "No records",
    },
    "th": {
        "title": "รายละเอียดค่าใช้จ่าย Pearnly",
        "company": "บริษัท",
        "sheet_usage": "การหักเงิน",
        "sheet_topup": "การเติมเงิน",
        "sheet_ocr": "การอ่านเอกสาร",
        "c_date": "เวลา",
        "c_type": "ประเภท",
        "c_billing": "การคิดเงิน",
        "c_desc": "รายละเอียด / ไฟล์",
        "c_pages": "จำนวนแผ่น",
        "c_amount": "จำนวนเงิน",
        "c_balance": "ยอดคงเหลือ",
        "c_payer": "ผู้ชำระ",
        "c_status": "สถานะ",
        "c_review_time": "เวลาตรวจสอบ",
        "c_note": "หมายเหตุ",
        "c_file": "ชื่อไฟล์",
        "c_invoice": "เลขที่ใบกำกับ",
        "c_seller": "ผู้ขาย",
        "c_source": "แหล่งที่มา",
        "t_scan": "หักค่าสแกน",
        "t_sub": "สมาชิก / ต่ออายุ",
        "b_charged": "หักจริง",
        "b_free": "ในแพ็กเกจ (ฟรี)",
        "b_monthly": "ค่ารายเดือน",
        "tp_pending": "รอตรวจสอบ",
        "tp_approved": "เข้าบัญชีแล้ว",
        "tp_rejected": "ถูกปฏิเสธ",
        "o_confirmed": "สำเร็จ",
        "o_pending": "รอตรวจทาน",
        "o_failed": "ล้มเหลว",
        "s_upload": "อัปโหลด",
        "s_line": "LINE",
        "s_email": "อีเมล",
        "empty": "ไม่มีข้อมูล",
    },
    "ja": {
        "title": "Pearnly 請求明細",
        "company": "会社",
        "sheet_usage": "課金履歴",
        "sheet_topup": "チャージ履歴",
        "sheet_ocr": "認識履歴",
        "c_date": "日時",
        "c_type": "種類",
        "c_billing": "課金",
        "c_desc": "内容 / ファイル",
        "c_pages": "枚数",
        "c_amount": "金額",
        "c_balance": "残高",
        "c_payer": "支払者",
        "c_status": "状態",
        "c_review_time": "審査日時",
        "c_note": "備考",
        "c_file": "ファイル名",
        "c_invoice": "請求書番号",
        "c_seller": "売り手",
        "c_source": "経路",
        "t_scan": "スキャン課金",
        "t_sub": "サブスク / 更新",
        "b_charged": "実課金",
        "b_free": "プラン内(無料)",
        "b_monthly": "月額",
        "tp_pending": "審査待ち",
        "tp_approved": "入金済み",
        "tp_rejected": "却下",
        "o_confirmed": "成功",
        "o_pending": "要確認",
        "o_failed": "失敗",
        "s_upload": "アップロード",
        "s_line": "LINE",
        "s_email": "メール",
        "empty": "記録なし",
    },
}

_MONEY_FMT = '"฿"#,##0.00'


def _t(lang: str, key: str) -> str:
    return _I18N.get(lang, _I18N["en"]).get(key, _I18N["en"].get(key, key))


def _src_label(lang: str, source: str) -> str:
    s = (source or "").strip().lower()
    if s == "line":
        return _t(lang, "s_line")
    if s == "email":
        return _t(lang, "s_email")
    return _t(lang, "s_upload")  # manual / upload / 其余


def _ocr_status_label(lang: str, status: str) -> str:
    s = (status or "").strip().lower()
    if s == "confirmed":
        return _t(lang, "o_confirmed")
    if s == "failed":
        return _t(lang, "o_failed")
    return _t(lang, "o_pending")


def _topup_status_label(lang: str, status: str) -> str:
    s = (status or "").strip().lower()
    if s == "approved":
        return _t(lang, "tp_approved")
    if s == "rejected":
        return _t(lang, "tp_rejected")
    return _t(lang, "tp_pending")


def build_billing_xlsx(
    *,
    lang: str,
    company: str,
    usage_rows: List[Dict[str, Any]],
    topup_rows: List[Dict[str, Any]],
    ocr_rows: List[Dict[str, Any]],
) -> bytes:
    """三 sheet 工作簿 → bytes。每个 sheet:标题行 + 公司行 + 表头 + 数据(空则一行 empty)。"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    if lang not in _I18N:
        lang = "en"

    thin = Side(style="thin", color="DDDDDD")
    st = {
        "border": Border(left=thin, right=thin, top=thin, bottom=thin),
        "head_font": Font(bold=True, color="FFFFFF"),
        "head_fill": PatternFill("solid", fgColor="7C4DFF"),
        "title_font": Font(bold=True, size=13),
        "center": Alignment(horizontal="center", vertical="center"),
        "left": Alignment(horizontal="left", vertical="center"),
        "right": Alignment(horizontal="right", vertical="center"),
    }

    def write_sheet(ws, sheet_title, headers, data_rows, money_cols, widths):
        ncol = len(headers)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        c = ws.cell(row=1, column=1, value=f"{_t(lang, 'title')} · {sheet_title}")
        c.font = st["title_font"]
        c.alignment = st["left"]
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncol)
        c = ws.cell(row=2, column=1, value=f"{_t(lang, 'company')}: {company or '—'}")
        c.alignment = st["left"]
        hr = 4
        for i, h in enumerate(headers, start=1):
            c = ws.cell(row=hr, column=i, value=h)
            c.font = st["head_font"]
            c.fill = st["head_fill"]
            c.alignment = st["center"]
            c.border = st["border"]
        r = hr + 1
        if not data_rows:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=ncol)
            c = ws.cell(row=r, column=1, value=_t(lang, "empty"))
            c.alignment = st["center"]
        for drow in data_rows:
            for i, val in enumerate(drow, start=1):
                c = ws.cell(row=r, column=i, value=val)
                c.border = st["border"]
                if i in money_cols:
                    c.number_format = _MONEY_FMT
                    c.alignment = st["right"]
                else:
                    c.alignment = st["left"]
            r += 1
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.freeze_panes = ws.cell(row=hr + 1, column=1)

    wb = Workbook()

    # ── Sheet 1 · 扣费记录 ───────────────────────────────────────────────
    ws1 = wb.active
    ws1.title = _t(lang, "sheet_usage")[:31]
    u_headers = [
        _t(lang, "c_date"),
        _t(lang, "c_type"),
        _t(lang, "c_billing"),
        _t(lang, "c_desc"),
        _t(lang, "c_pages"),
        _t(lang, "c_amount"),
        _t(lang, "c_balance"),
    ]
    u_data = []
    for r in usage_rows:
        is_sub = (r.get("type") or "") == "subscription"
        cost = float(r.get("cost_thb") or 0)
        if is_sub:
            typ = _t(lang, "t_sub")
            billing = _t(lang, "b_monthly")
        else:
            typ = _t(lang, "t_scan")
            billing = _t(lang, "b_free") if cost == 0 else _t(lang, "b_charged")
        desc = r.get("description") or r.get("filename") or ""
        bal = r.get("balance_after")
        u_data.append(
            [
                r.get("date") or "",
                typ,
                billing,
                desc,
                int(r.get("pages") or 0),
                round(abs(cost), 2),
                round(float(bal), 2) if bal is not None else "",
            ]
        )
    write_sheet(
        ws1, ws1.title, u_headers, u_data, money_cols={6, 7}, widths=[18, 12, 13, 40, 8, 12, 13]
    )

    # ── Sheet 2 · 充值记录 ───────────────────────────────────────────────
    ws2 = wb.create_sheet(title=_t(lang, "sheet_topup")[:31])
    t_headers = [
        _t(lang, "c_date"),
        _t(lang, "c_amount"),
        _t(lang, "c_payer"),
        _t(lang, "c_status"),
        _t(lang, "c_review_time"),
        _t(lang, "c_note"),
    ]
    t_data = []
    for r in topup_rows:
        t_data.append(
            [
                r.get("created_at") or "",
                round(float(r.get("amount_thb") or 0), 2),
                r.get("payer_name") or "",
                _topup_status_label(lang, r.get("status") or ""),
                r.get("reviewed_at") or "",
                r.get("note") or "",
            ]
        )
    write_sheet(ws2, ws2.title, t_headers, t_data, money_cols={2}, widths=[18, 12, 18, 12, 18, 30])

    # ── Sheet 3 · 识别记录 ───────────────────────────────────────────────
    ws3 = wb.create_sheet(title=_t(lang, "sheet_ocr")[:31])
    o_headers = [
        _t(lang, "c_date"),
        _t(lang, "c_file"),
        _t(lang, "c_status"),
        _t(lang, "c_invoice"),
        _t(lang, "c_seller"),
        _t(lang, "c_amount"),
        _t(lang, "c_pages"),
        _t(lang, "c_source"),
    ]
    o_data = []
    for r in ocr_rows:
        total = r.get("total_amount")
        o_data.append(
            [
                r.get("created_at") or "",
                r.get("filename") or "",
                _ocr_status_label(lang, r.get("status") or ""),
                r.get("invoice_no") or "",
                r.get("seller_name") or "",
                round(float(total), 2) if total is not None else "",
                int(r.get("page_count") or 0),
                _src_label(lang, r.get("source") or ""),
            ]
        )
    write_sheet(
        ws3, ws3.title, o_headers, o_data, money_cols={6}, widths=[18, 32, 12, 16, 24, 12, 8, 10]
    )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
