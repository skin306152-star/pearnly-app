#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
tests/unit/test_mrerp_report_parser.py

Offline test of services.erp.mrerp_report_parser against the real failure
sample saved in docs/integrations/samples/. No network or browser needed.

Usage:
    python -m unittest tests.unit.test_mrerp_report_parser -v
    or
    python tests/unit/test_mrerp_report_parser.py
"""

import io
import sys
import unittest
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(PROJECT_ROOT))

from openpyxl import Workbook   # noqa: E402

from services.erp.mrerp_report_parser import (   # noqa: E402
    HEADER_INVOICE_NO,
    HEADER_NOTE,
    ImportReport,
    parse_import_report,
)


FAILURE_SAMPLE = (
    PROJECT_ROOT / "docs" / "integrations" / "samples"
    / "report_failure_customer_not_found.xlsx"
)


class RealFailureSampleTests(unittest.TestCase):
    """The real 2026-05-18 report.php download where customer
    99-PEARNLYTEST-001 does not exist."""

    def test_one_invoice_two_reasons(self):
        self.assertTrue(FAILURE_SAMPLE.exists(),
                        f"fixture missing: {FAILURE_SAMPLE}")
        xlsx_bytes = FAILURE_SAMPLE.read_bytes()

        report = parse_import_report(xlsx_bytes)

        self.assertIsInstance(report, ImportReport)
        self.assertEqual(report.total, 1)
        self.assertEqual(report.success, [])
        self.assertEqual(len(report.failed), 1)

        row = report.failed[0]
        self.assertEqual(row.invoice_no, "PEARNLY-TEST-001")
        self.assertFalse(row.success)
        self.assertEqual(len(row.reasons), 2)
        self.assertIn("ไม่พบข้อมูลรหัสลูกค้า", row.reasons)
        self.assertIn("ไม่พบข้อมูลรหัสลูกค้า (บิล)", row.reasons)
        for detail in row.details:
            self.assertEqual(detail["sheet"], "Worksheet")


class SyntheticReportTests(unittest.TestCase):
    """Synthesize reports to exercise edge cases the real sample does not
    cover (success-only, mixed, multi-line, detail-sheet aggregation)."""

    @staticmethod
    def _build(sheets_data):
        """sheets_data = [(sheet_name, [(label, label, ...), (val, val, ...), ...]), ...]"""
        wb = Workbook()
        first = True
        for sname, rows in sheets_data:
            if first:
                ws = wb.active
                ws.title = sname
                first = False
            else:
                ws = wb.create_sheet(sname)
            for r_idx, row_vals in enumerate(rows, start=1):
                for c_idx, val in enumerate(row_vals, start=1):
                    if val is not None:
                        ws.cell(row=r_idx, column=c_idx, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_all_success(self):
        xlsx = self._build([
            ("Worksheet", [
                (HEADER_INVOICE_NO, HEADER_NOTE),
                ("OK-INV-001", None),
                ("OK-INV-002", None),
            ]),
        ])
        report = parse_import_report(xlsx)
        self.assertEqual(report.total, 2)
        self.assertEqual(sorted(report.success), ["OK-INV-001", "OK-INV-002"])
        self.assertEqual(report.failed, [])
        self.assertTrue(report.all_success)

    def test_mixed_with_multiline_error(self):
        xlsx = self._build([
            ("Worksheet", [
                (HEADER_INVOICE_NO, HEADER_NOTE),
                ("GOOD-001", None),
                ("BAD-001", "reason A\nreason B\nreason C"),
            ]),
        ])
        report = parse_import_report(xlsx)
        self.assertEqual(report.success, ["GOOD-001"])
        self.assertEqual(len(report.failed), 1)
        bad = report.failed[0]
        self.assertEqual(bad.invoice_no, "BAD-001")
        self.assertEqual(bad.reasons, ["reason A", "reason B", "reason C"])
        self.assertFalse(report.all_success)

    def test_detail_sheet_note_aggregates_to_same_invoice(self):
        xlsx = self._build([
            ("Worksheet", [
                (HEADER_INVOICE_NO, HEADER_NOTE),
                ("INV-X", None),
            ]),
            ("Worksheet 1", [
                (HEADER_INVOICE_NO, "other", HEADER_NOTE),
                ("INV-X", "ignored", "product missing"),
            ]),
        ])
        report = parse_import_report(xlsx)
        self.assertEqual(report.success, [])
        self.assertEqual(len(report.failed), 1)
        self.assertEqual(report.failed[0].invoice_no, "INV-X")
        self.assertEqual(report.failed[0].reasons, ["product missing"])
        self.assertEqual(report.failed[0].details[0]["sheet"], "Worksheet 1")

    def test_empty_bytes_raises(self):
        with self.assertRaises(ValueError):
            parse_import_report(b"")

    def test_missing_primary_sheet_raises(self):
        xlsx = self._build([("NotWorksheet", [(HEADER_INVOICE_NO, HEADER_NOTE)])])
        with self.assertRaises(ValueError):
            parse_import_report(xlsx)


if __name__ == "__main__":
    unittest.main(verbosity=2)
