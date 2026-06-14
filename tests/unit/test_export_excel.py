# -*- coding: utf-8 -*-
"""导出行 → xlsx 内存流(export.excel.build_workbook)· openpyxl round-trip(阶段二)。

写出字节流再读回:表头 == COLUMNS · 行数对 · Decimal 写成数值(可求和)· evidence URL 渲超链。
"""

import io
import unittest

from openpyxl import load_workbook

from services.export.excel import build_workbook
from services.export.rows import COLUMNS, build_export_rows
from services.export.entries import summarize_voucher


def _rows():
    item = {
        "doc": {
            "id": "D1",
            "doc_date": "2026-06-01",
            "doc_no": "INV1",
            "doc_kind": "expense",
            "payment_status": "paid",
            "has_vat": False,
            "discount_total": "0",
            "rounding": "0",
        },
        "lines": [
            {
                "description": "Lunch",
                "qty": "1",
                "unit_price": "120",
                "line_total": "120",
                "discount": "0",
                "vat_rate": "0",
                "vat_applicable": False,
                "wht_rate": "0",
                "item_type": "service",
            }
        ],
        "supplier": {"name": "Cafe"},
        "posting": summarize_voucher(None),
        "evidence_url": "https://drive.google.com/folder/abc",
    }
    return build_export_rows([item], category_names={})


class BuildWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.wb = load_workbook(io.BytesIO(build_workbook(_rows(), sheet_title="进项明细")))
        self.ws = self.wb.active

    def test_header_matches_columns(self):
        headers = [c.value for c in self.ws[1]]
        self.assertEqual(headers, [h for _, h in COLUMNS])

    def test_one_data_row(self):
        self.assertEqual(self.ws.max_row, 2)  # 1 表头 + 1 数据

    def test_numeric_cell_is_number(self):
        keys = [k for k, _ in COLUMNS]
        net_col = keys.index("line_net") + 1
        self.assertEqual(self.ws.cell(row=2, column=net_col).value, 120.0)

    def test_evidence_url_rendered_as_hyperlink(self):
        keys = [k for k, _ in COLUMNS]
        ev_col = keys.index("evidence") + 1
        cell = self.ws.cell(row=2, column=ev_col)
        self.assertEqual(cell.value, "查看")
        self.assertIsNotNone(cell.hyperlink)

    def test_sheet_title_truncated_safe(self):
        wb = load_workbook(io.BytesIO(build_workbook([], sheet_title="x" * 50)))
        self.assertLessEqual(len(wb.active.title), 31)

    def test_empty_rows_only_header(self):
        wb = load_workbook(io.BytesIO(build_workbook([])))
        self.assertEqual(wb.active.max_row, 1)


if __name__ == "__main__":
    unittest.main()
