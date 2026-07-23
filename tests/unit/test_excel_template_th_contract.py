# -*- coding: utf-8 -*-
"""REFACTOR-D2 守门 · 泰国销售明细导出模板 excel_template_th.py 行为契约

excel_template_th.py(12 列泰文表头 · 按 items 拆行 · =E*F 公式)此前 0 专属测试。
纯生成函数 · 无 DB/网络/凭证(Wave 0 安全网 · 零冲突)。

锁定 Korn 反馈修过的真实合同(v27.8.1.2):G/H/I 每行 =E*F · J 每行 =I*0.07,
+ 单号列纯文本格式(防 Excel 把 IV69/00271 当日期)+ 12 列泰文表头不漂移。
"""

import io
import unittest

from openpyxl import load_workbook

from services.excel import excel_template_th as et
from services.excel.erp_roundtrip import ROUNDTRIP_HEADERS


class NormDateTests(unittest.TestCase):
    def test_parses_western_formats(self):
        from datetime import datetime

        self.assertEqual(et._norm_date("2026-03-15"), datetime(2026, 3, 15))
        self.assertEqual(et._norm_date("15/03/2026"), datetime(2026, 3, 15))
        self.assertEqual(et._norm_date("15.03.2026"), datetime(2026, 3, 15))

    def test_empty_and_bad(self):
        self.assertIsNone(et._norm_date(None))
        self.assertIsNone(et._norm_date(""))
        self.assertIsNone(et._norm_date("not-a-date"))

    def test_buddhist_era_converted(self):
        """已修(REFACTOR-D2):_norm_date 解析后对 year>2400 减 543 · 佛历转西历。"""
        from datetime import datetime

        self.assertEqual(et._norm_date("2569-03-15"), datetime(2026, 3, 15))
        self.assertEqual(et._norm_date("15/03/2569"), datetime(2026, 3, 15))


class ToFloatStrTests(unittest.TestCase):
    def test_to_float_strips_symbols(self):
        self.assertEqual(et._to_float("1,070.00"), 1070.0)
        self.assertEqual(et._to_float("฿500"), 500.0)
        self.assertEqual(et._to_float("100 THB"), 100.0)

    def test_to_float_none_and_bad(self):
        self.assertIsNone(et._to_float(None))
        self.assertIsNone(et._to_float(""))
        self.assertIsNone(et._to_float("abc"))

    def test_str_cleans_junk(self):
        self.assertEqual(et._str(None), "")
        self.assertEqual(et._str("  X "), "X")
        self.assertEqual(et._str("none"), "")
        self.assertEqual(et._str("NULL"), "")
        self.assertEqual(et._str("nan"), "")


def _load(records):
    raw = et.build_sales_detail_xlsx(records)
    assert isinstance(raw, (bytes, bytearray))
    return load_workbook(io.BytesIO(raw)).active


class HeaderTests(unittest.TestCase):
    def test_header_row_matches_declared_headers(self):
        ws = _load([])
        n = len(et.HEADERS_TH)
        headers = [ws.cell(row=1, column=i).value for i in range(1, n + 1)]
        self.assertEqual(headers, et.HEADERS_TH)

    def test_accounting_contract_columns_1_to_12_frozen(self):
        """1-12 列是与 MR.ERP 约定的公式合同 · 列位/列名一格都不许动。
        后续加列(回导列等)只能追加在其后。"""
        self.assertEqual(
            et.HEADERS_TH[:12],
            [
                "วันที่",
                "เลขที่",
                "รหัสลูกค้า",
                "รหัสสินค้า",
                "จำนวน",
                "ราคาต่อหน่วย",
                "จำนวนเงิน",
                "รวมจำนวนเงิน",
                "รวมจำนวนเงินก่อนVAT",
                "VAT",
                "อ้างอิง",
                "Status",
            ],
        )

    def test_status_columns_present(self):
        self.assertEqual(et.HEADERS_TH[12], "สถานะสินค้า")  # 13 商品状态
        self.assertEqual(et.HEADERS_TH[13], "สถานะลูกค้า")  # 14 客户状态

    def test_roundtrip_columns_appended_after_contract(self):
        """回导列必须在合同列之后 · 且顺序由 erp_roundtrip 单点定义(写读共用)。"""
        self.assertEqual(tuple(et.HEADERS_TH[14:]), ROUNDTRIP_HEADERS)
        self.assertEqual(len(et.HEADERS_TH), len(et.COLUMN_WIDTHS))

    def test_header_has_fill_and_sheet_frozen(self):
        ws = _load([])
        self.assertEqual(ws.cell(row=1, column=1).fill.fgColor.rgb[-6:], "2C5282")
        self.assertEqual(ws.freeze_panes, "A2")


class FormulaContractTests(unittest.TestCase):
    def _rec_two_items(self):
        return [
            {
                "filename": "a.pdf",
                "engine": "x",
                "merged_fields": {
                    "date": "2026-03-15",
                    "invoice_number": "IV69/00271",
                    "buyer_name": "ACME",
                    "total_amount": "1070.00",
                    "items": [
                        {"description": "Coffee", "qty": "2", "unit_price": "500"},
                        {"description": "Tea", "amount": "100"},  # 缺 qty/price
                    ],
                },
            }
        ]

    def test_qty_price_row_uses_formulas(self):
        ws = _load(self._rec_two_items())
        # 第 1 个 item(row2):有 qty/price → G/H/I = =E2*F2 · J = =I2*0.07
        self.assertEqual(ws.cell(row=2, column=7).value, "=E2*F2")
        self.assertEqual(ws.cell(row=2, column=8).value, "=E2*F2")
        self.assertEqual(ws.cell(row=2, column=9).value, "=E2*F2")
        self.assertEqual(ws.cell(row=2, column=10).value, "=I2*0.07")
        self.assertEqual(ws.cell(row=2, column=5).value, 2.0)
        self.assertEqual(ws.cell(row=2, column=6).value, 500.0)

    def test_invoice_no_is_text_format(self):
        ws = _load(self._rec_two_items())
        c = ws.cell(row=2, column=2)
        self.assertEqual(c.value, "IV69/00271")
        self.assertEqual(c.number_format, "@")  # 纯文本 · 防被当日期

    def test_customer_and_description(self):
        ws = _load(self._rec_two_items())
        self.assertEqual(ws.cell(row=2, column=3).value, "ACME")
        self.assertEqual(ws.cell(row=2, column=4).value, "Coffee")

    def test_missing_qty_price_uses_line_amount_literal(self):
        ws = _load(self._rec_two_items())
        # 第 2 个 item(row3):缺 qty/price · 用 line_amount=100 死值 · J=round(100*0.07,2)
        self.assertIsNone(ws.cell(row=3, column=5).value)
        self.assertIsNone(ws.cell(row=3, column=6).value)
        self.assertEqual(ws.cell(row=3, column=7).value, 100.0)
        self.assertEqual(ws.cell(row=3, column=8).value, 100.0)
        self.assertEqual(ws.cell(row=3, column=9).value, 100.0)
        self.assertEqual(ws.cell(row=3, column=10).value, 7.0)


class EdgeCaseTests(unittest.TestCase):
    def test_no_items_single_fallback_row(self):
        ws = _load(
            [
                {
                    "merged_fields": {
                        "date": "2026-03-15",
                        "invoice_number": "X",
                        "total_amount": "500",
                    }
                }
            ]
        )
        # 无 items → 单行兜底(空商品 + total 死值)· openpyxl 空串读回为 None
        self.assertIn(ws.cell(row=2, column=4).value, (None, ""))
        self.assertEqual(ws.cell(row=2, column=7).value, 500.0)
        self.assertIsNone(ws.cell(row=3, column=1).value)  # 只有 1 行数据

    def test_non_dict_records_skipped(self):
        ws = _load([None, "junk", {"merged_fields": {"invoice_number": "Y", "total_amount": "10"}}])
        # 前两个非法被跳过 · 只 Y 落第 2 行
        self.assertEqual(ws.cell(row=2, column=2).value, "Y")
        self.assertIsNone(ws.cell(row=3, column=2).value)

    def test_customer_fallback_to_customer_name(self):
        ws = _load([{"merged_fields": {"customer_name": "FALLBACK", "items": [{"amount": "1"}]}}])
        self.assertEqual(ws.cell(row=2, column=3).value, "FALLBACK")


class NormActionTests(unittest.TestCase):
    def test_string_aliases(self):
        self.assertEqual(et._norm_action("new"), et.ACTION_NEW)
        self.assertEqual(et._norm_action("erp_auto_created"), et.ACTION_NEW)
        self.assertEqual(et._norm_action("reused"), et.ACTION_REUSED)
        self.assertEqual(et._norm_action("matched"), et.ACTION_REUSED)
        self.assertEqual(et._norm_action("cache_hit"), et.ACTION_REUSED)

    def test_bool_is_new_slots_in(self):
        # ERP 层直接回 is_new 布尔 · 无缝落位
        self.assertEqual(et._norm_action(True), et.ACTION_NEW)
        self.assertEqual(et._norm_action(False), et.ACTION_REUSED)

    def test_empty_and_unknown(self):
        self.assertEqual(et._norm_action(None), et.ACTION_UNKNOWN)
        self.assertEqual(et._norm_action(""), et.ACTION_UNKNOWN)
        self.assertEqual(et._norm_action("garbage"), et.ACTION_UNKNOWN)


class ErpActionColumnTests(unittest.TestCase):
    def _rec(self):
        return [
            {
                "merged_fields": {
                    "invoice_number": "IV/001",
                    "buyer_name": "ACME",
                    "customer_erp_action": "new",
                    "items": [
                        {
                            "description": "Coffee",
                            "qty": "2",
                            "unit_price": "500",
                            "erp_action": "new",
                        },
                        {
                            "description": "Tea",
                            "qty": "1",
                            "unit_price": "100",
                            "erp_action": "reused",
                        },
                        {"description": "Water", "qty": "3", "unit_price": "10"},  # 无 erp_action
                    ],
                }
            }
        ]

    def test_product_status_labels(self):
        ws = _load(self._rec())
        self.assertEqual(ws.cell(row=2, column=13).value, "ใหม่")  # new
        self.assertEqual(ws.cell(row=3, column=13).value, "เดิม")  # reused
        self.assertEqual(ws.cell(row=4, column=13).value, "-")  # 未知 → 占位

    def test_customer_status_repeats_per_row(self):
        ws = _load(self._rec())
        # 客户动作单据级 · 每行都标 ใหม่
        for r in (2, 3, 4):
            self.assertEqual(ws.cell(row=r, column=14).value, "ใหม่")

    def test_new_gets_green_reused_gets_subtle_unknown_none(self):
        ws = _load(self._rec())
        # 商品单元格(col4)按动作填色
        self.assertEqual(ws.cell(row=2, column=4).fill.fgColor.rgb[-6:], "D6F5DE")  # new 绿
        self.assertEqual(ws.cell(row=3, column=4).fill.fgColor.rgb[-6:], "EEF2F7")  # reused 淡灰蓝
        self.assertEqual(ws.cell(row=4, column=4).fill.patternType, None)  # 未知 无填色
        # 客户单元格(col3)= new → 绿
        self.assertEqual(ws.cell(row=2, column=3).fill.fgColor.rgb[-6:], "D6F5DE")

    def test_is_new_bool_fallback(self):
        ws = _load(
            [
                {
                    "merged_fields": {
                        "invoice_number": "IV/002",
                        "customer_is_new": False,
                        "items": [
                            {"description": "X", "qty": "1", "unit_price": "5", "is_new": True}
                        ],
                    }
                }
            ]
        )
        self.assertEqual(ws.cell(row=2, column=13).value, "ใหม่")  # item is_new=True
        self.assertEqual(ws.cell(row=2, column=14).value, "เดิม")  # customer is_new=False


class MoneyFormatTests(unittest.TestCase):
    def test_money_cols_right_aligned_thousands(self):
        ws = _load(
            [{"merged_fields": {"items": [{"description": "A", "qty": "2", "unit_price": "1500"}]}}]
        )
        for col in (6, 7, 8, 9, 10):
            c = ws.cell(row=2, column=col)
            self.assertEqual(c.number_format, "#,##0.00")
            self.assertEqual(c.alignment.horizontal, "right")


class FilenameTests(unittest.TestCase):
    def test_filename_format(self):
        fn = et.sales_detail_filename()
        self.assertTrue(fn.startswith("Pearnly_SalesDetail_"))
        self.assertTrue(fn.endswith(".xlsx"))

    def test_filename_custom_prefix(self):
        self.assertTrue(et.sales_detail_filename(prefix="X").startswith("X_"))


if __name__ == "__main__":
    unittest.main()
