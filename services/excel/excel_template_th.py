# -*- coding: utf-8 -*-
"""
导出模板「泰国销售明细」· sales_detail_th
=========================================================
按泰国本地销售清单习惯生成 · 用户处理完一批发票后导出复盘「都发生了什么」。

表结构(19 列):
  1-12  会计核算区(与 MR.ERP Korn 反馈的公式合同一致 · 不动列位/公式):
        วันที่ / เลขที่ / ชื่อลูกค้า / ชื่อสินค้า / จำนวน(E) / ราคาต่อหน่วย(F) /
        จำนวนเงิน(G==E*F) / รวมจำนวนเงิน(H==E*F) / รวมก่อนVAT(I==E*F) /
        VAT(J==I*0.07) / อ้างอิง / Status(留空给会计)
  13    สถานะสินค้า  · 每商品行:ERP 里是新建(ใหม่)还是复用(เดิม)
  14    สถานะลูกค้า  · 每张发票:客户是新建还是复用
  15-19 回导列(见 erp_roundtrip)· 让「导出→会计改→回导重推」闭得上环

复盘信号靠两处呈现:客户/商品单元格底色(绿=新建·淡灰蓝=复用)+ 末两列显式文本
(可筛选、黑白打印也读得出)。表头深蓝底白字、冻结首行、金额千分位右对齐。

数据来源:
  records: list of {filename, engine, merged_fields}
  merged_fields.items[].erp_action = 'new' | 'reused' | ''   (商品行动作)
  merged_fields.customer_erp_action = 'new' | 'reused' | ''  (客户动作 · 单据级)
  两字段由 ERP 推送环节回填(is_new 布尔亦可)· 未回填时该列留 '-' 不填色。
"""

import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None  # type: ignore

from services.excel import erp_roundtrip as rt
from services.excel.erp_roundtrip import (
    ROUNDTRIP_HEADERS,
    ROUNDTRIP_WIDTHS,
    SHEET_SALES,
    encode_row_key,
    roundtrip_values,
)

logger = logging.getLogger(__name__)


# 列名取自 erp_roundtrip —— 读侧按同一批常量取值。两边各写一份泰文串的话,改任一边
# 读侧都不会报错,只是查不到列、静默降级成空值。
HEADERS_TH = [
    rt.SALES_COL_DATE,  # 1  date
    rt.SALES_COL_INVOICE,  # 2  invoice_no
    rt.SALES_COL_PARTY,  # 3  customer
    rt.SALES_COL_ITEM,  # 4  product / description
    rt.SALES_COL_QTY,  # 5  qty (E)
    rt.SALES_COL_PRICE,  # 6  unit_price (F)
    rt.SALES_COL_AMOUNT,  # 7  amount = E*F
    "รวมจำนวนเงิน",  # 8  line total = E*F
    rt.SALES_COL_PRE_VAT,  # 9  pre-VAT = E*F
    "VAT",  # 10 = I*0.07
    "อ้างอิง",  # 11 ref / note(留空)
    "Status",  # 12 会计手填过账状态(留空)
    "สถานะสินค้า",  # 13 商品:ใหม่ / เดิม / -
    "สถานะลูกค้า",  # 14 客户:ใหม่ / เดิม / -
    *ROUNDTRIP_HEADERS,  # 15-19 回导列
]

COLUMN_WIDTHS = [12, 14, 30, 24, 8, 12, 14, 14, 18, 10, 12, 10, 13, 13, *ROUNDTRIP_WIDTHS]

MONEY_COLUMNS = (6, 7, 8, 9, 10)
QTY_COLUMN = 5
PRODUCT_STATUS_COLUMN = 13
CUSTOMER_STATUS_COLUMN = 14

_MONEY_FMT = "#,##0.00"
_QTY_FMT = "#,##0.###"
_DATE_FMT = "dd/mm/yyyy"

# 动作归一后的三态
ACTION_NEW = "new"
ACTION_REUSED = "reused"
ACTION_UNKNOWN = ""

# 泰文状态标签 · ใหม่=新 · เดิม=原有
_STATUS_LABEL = {ACTION_NEW: "ใหม่", ACTION_REUSED: "เดิม", ACTION_UNKNOWN: "-"}

_thin = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
# 表头深蓝(与 standard 模板同色 · 统一设计语言)· 不指定字体名 · 保泰文字形渲染
_HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
# 新建=柔和绿 · 复用=淡灰蓝 · 均浅底黑字保可读
_FILL_NEW = PatternFill(start_color="D6F5DE", end_color="D6F5DE", fill_type="solid")
_FILL_REUSED = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")

_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center")


def _norm_date(raw: Any) -> Optional[datetime]:
    """日期字符串 → datetime · 让 cell.number_format 自动渲染"""
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    # 容错多种格式 · 含泰国佛历(2569 → 公历 2026)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            d = datetime.strptime(s[:10], fmt)
            if d.year > 2400:  # 佛历纪年 · 减 543 转西历
                d = d.replace(year=d.year - 543)
            return d
        except Exception:
            pass
    return None


def _to_float(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        s = str(raw).replace(",", "").replace("฿", "").replace("THB", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _str(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    return "" if s.lower() in ("none", "null", "nan") else s


def _norm_action(raw: Any) -> str:
    """把 ERP 动作归一到 new / reused / ''(未知)。

    接受 erp_action 字符串,或直接接 ERP 层的 is_new 布尔
    (CustomerSyncResult / ProductSyncResult.is_new)· 让真实数据无缝落位。
    """
    if isinstance(raw, bool):
        return ACTION_NEW if raw else ACTION_REUSED
    if raw is None:
        return ACTION_UNKNOWN
    s = str(raw).strip().lower()
    if s in ("new", "created", "auto_created", "erp_auto_created"):
        return ACTION_NEW
    if s in ("reused", "reuse", "existing", "matched", "match", "cache_hit", "db_mapping"):
        return ACTION_REUSED
    return ACTION_UNKNOWN


def _fill_for(action: str) -> Optional[PatternFill]:
    if action == ACTION_NEW:
        return _FILL_NEW
    if action == ACTION_REUSED:
        return _FILL_REUSED
    return None


def _customer_action(f: Dict[str, Any]) -> str:
    """客户动作 · 单据级 · 主字段 customer_erp_action · 兼容 buyer_erp_action / is_new"""
    raw = f.get("customer_erp_action")
    if raw in (None, ""):
        raw = f.get("buyer_erp_action")
    if raw in (None, ""):
        raw = f.get("customer_is_new")
    return _norm_action(raw)


def _item_action(it: Dict[str, Any]) -> str:
    raw = it.get("erp_action")
    if raw in (None, ""):
        raw = it.get("is_new")
    return _norm_action(raw)


def _write_header(ws) -> None:
    for i, h in enumerate(HEADERS_TH, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
        ws.column_dimensions[get_column_letter(i)].width = COLUMN_WIDTHS[i - 1]
    ws.row_dimensions[1].height = 30


def _write_status_cell(ws, row: int, column: int, action: str) -> None:
    c = ws.cell(row=row, column=column, value=_STATUS_LABEL[action])
    c.alignment = _CENTER
    fill = _fill_for(action)
    if fill is not None:
        c.font = Font(bold=True)
        c.fill = fill


def _style_data_row(ws, row: int) -> None:
    """整行 14 格补边框 + 对齐 + 金额/数量格式(空的备注/状态格也补边框成完整网格)"""
    for col in range(1, len(HEADERS_TH) + 1):
        c = ws.cell(row=row, column=col)
        c.border = _BORDER
        if col == 1:
            c.alignment = _CENTER
        elif col == QTY_COLUMN:
            c.alignment = _RIGHT
            c.number_format = _QTY_FMT
        elif col in MONEY_COLUMNS:
            c.alignment = _RIGHT
            c.number_format = _MONEY_FMT
        elif col in (PRODUCT_STATUS_COLUMN, CUSTOMER_STATUS_COLUMN):
            c.alignment = _CENTER
        else:
            c.alignment = _LEFT


def write_sales_sheet(ws, records: List[Dict[str, Any]]) -> int:
    """把销项记录写进给定工作表,返回写了多少数据行。

    抽出来是为了让它既能单独成表(build_sales_detail_xlsx),也能当复核工作簿里的
    「ขาย」那一张(erp_workbook)—— 两处必须是同一段代码,格式才不会漂。
    """
    _write_header(ws)

    row = 2
    for rec in records or []:
        if not isinstance(rec, dict):
            continue
        f = rec.get("merged_fields") or {}
        if not isinstance(f, dict):
            continue

        date_v = _norm_date(f.get("date"))
        invoice_no = _str(f.get("invoice_number"))
        customer = _str(f.get("buyer_name")) or _str(f.get("customer_name"))
        cust_action = _customer_action(f)
        cust_fill = _fill_for(cust_action)
        total = _to_float(f.get("total_amount"))
        # 回导用的单据级事实(由 export_actions 从 erp_push_logs 回填 · 缺则留空不假装)
        erp_docnum = _str(f.get("erp_docnum"))
        erp_party = _str(f.get("erp_party_code"))
        # 对手方税号:销项看买方 · ภ.พ.30 销项附表要它,合同 12 列里没有,故进回导列
        party_tax = _str(f.get("buyer_tax")) or _str(f.get("buyer_tax_id"))
        # 票面税额原样带走:合同里的 VAT 列是每行的,多行票逐行相加与它可能差分币,
        # 而它才是要进 ภ.พ.30 的法定数字
        doc_vat = _to_float(f.get("vat_amount")) or _to_float(f.get("vat"))
        push_status = f.get("push_status")
        push_reason = f.get("push_reason")
        history_id = _str(f.get("history_id")) or _str(rec.get("history_id"))

        items = f.get("items")
        if not isinstance(items, list) or len(items) == 0:
            # 没明细 · 单行兜底(空商品 + 总额)
            items = [{"description": "", "qty": None, "unit_price": None, "amount": total}]

        for line_idx, it in enumerate(items):
            if not isinstance(it, dict):
                it = {}
            qty = _to_float(it.get("qty") or it.get("quantity"))
            unit_price = _to_float(it.get("unit_price") or it.get("price"))
            description = (
                _str(it.get("description")) or _str(it.get("product_name")) or _str(it.get("name"))
            )
            line_amount = _to_float(it.get("amount") or it.get("total"))
            item_action = _item_action(it)

            cell_date = ws.cell(row=row, column=1, value=date_v)
            if date_v:
                cell_date.number_format = _DATE_FMT
            # 单号纯文本 · 防 Excel 把 IV69/00271 当日期
            ws.cell(row=row, column=2, value=invoice_no).number_format = "@"
            cust_cell = ws.cell(row=row, column=3, value=customer)
            if cust_fill is not None:
                cust_cell.fill = cust_fill
            prod_cell = ws.cell(row=row, column=4, value=description)
            item_fill = _fill_for(item_action)
            if item_fill is not None:
                prod_cell.fill = item_fill
            ws.cell(row=row, column=5, value=qty if qty is not None else None)
            ws.cell(row=row, column=6, value=unit_price if unit_price is not None else None)

            # 会计公式合同(Korn 反馈 · v27.8.1.2):G/H/I 每行 =E*F · J 每行 =I*0.07。
            # qty/price 缺(OCR 没拆出来)时用预算好的 line_amount 死值兜底。
            if qty is not None and unit_price is not None:
                ws.cell(row=row, column=7, value=f"=E{row}*F{row}")
                ws.cell(row=row, column=8, value=f"=E{row}*F{row}")
                ws.cell(row=row, column=9, value=f"=E{row}*F{row}")
                ws.cell(row=row, column=10, value=f"=I{row}*0.07")
            elif line_amount is not None:
                ws.cell(row=row, column=7, value=line_amount)
                ws.cell(row=row, column=8, value=line_amount)
                ws.cell(row=row, column=9, value=line_amount)
                ws.cell(row=row, column=10, value=round(line_amount * 0.07, 2))

            _style_data_row(ws, row)
            _write_status_cell(ws, row, PRODUCT_STATUS_COLUMN, item_action)
            _write_status_cell(ws, row, CUSTOMER_STATUS_COLUMN, cust_action)

            # 回导列 · 第 15 列起(合同列位之后追加)
            rt = roundtrip_values(
                party_tax=party_tax,
                doc_vat=doc_vat,
                docnum=erp_docnum,
                item_code=it.get("erp_item_code"),
                party_code=erp_party,
                push_status=push_status,
                push_reason=push_reason,
                row_key=encode_row_key(history_id, line_idx),
            )
            for off, val in enumerate(rt):
                ws.cell(row=row, column=CUSTOMER_STATUS_COLUMN + 1 + off, value=val)
            row += 1

    ws.freeze_panes = "A2"
    return row - 2


def build_sales_detail_xlsx(records: List[Dict[str, Any]], lang: str = "zh") -> bytes:
    """返回单表 .xlsx 二进制。records: list of {filename, engine, merged_fields}。
    每张发票按 items 拆 N 行 · 共享单据级字段(日期/单号/客户)。

    工作表名用 SHEET_SALES —— 表名即方向,单表导出也因此可直接回导重推。
    """
    if Workbook is None:
        raise RuntimeError("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_SALES
    write_sales_sheet(ws, records)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def sales_detail_filename(prefix: str = "Pearnly_SalesDetail") -> str:
    return f'{prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
