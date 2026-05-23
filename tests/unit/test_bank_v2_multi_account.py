# -*- coding: utf-8 -*-
"""
v118.35.0.61 · 守门测试 · 多账户 .xls 分账户解析 + 逐账户独立余额校验

真实案例(skin 实测 KTB 8258 และ 8606):一个文件 2 个账户,8258 期初 -39、
8606 期初 -740万。此前合并成一条余额链 → 从几万跳到 -737万整链作废、误标。
现在每账户用『自己的期初』独立 verify · 表头上方的 ยกมา 期初也要捕获。

锁定契约:
  1. 多账户文件 → multi_account=True + account_codes 含两个账户号
  2. 每账户独立校验:不会因跨账户跳变而误标 balance_ok=False
  3. 表头上方带标签的期初(ยกมา/opening)被捕获
  4. 末期取『最后一个有余额的行』(末行常是无余额的 Sweep 行)
  5. 单账户文件 → multi_account=False(不误伤)
"""

import io
import unittest

import openpyxl

from bank_recon_v2 import parse_bank_stmt_xlsx_direct


def _build_multi():
    """两个 sheet = 两个账户,各自带表头上方期初(ยกมา)+ 一条自洽余额链。"""
    wb = openpyxl.Workbook()
    specs = [
        (
            "984-2-99825-8",
            -39.15,
            [("01/01/2026", "dep1", 2000.0, 1960.85), ("02/01/2026", "dep2", 3000.0, 4960.85)],
        ),
        (
            "984-2-99860-6",
            -7409714.58,
            [
                ("01/01/2026", "tr1", 29989.15, -7379725.43),
                ("02/01/2026", "tr2", 10995.0, -7368730.43),
            ],
        ),
    ]
    for si, (acct, opening, txns) in enumerate(specs):
        ws = wb.active if si == 0 else wb.create_sheet(acct)
        if si == 0:
            ws.title = acct
        ws.append(["Account No.", acct])
        ws.append(["", "", "", "", "ยอดยกมา", opening])  # 表头上方期初
        ws.append([])
        ws.append(["Date", "Description", "Amount", "Balance"])  # 单一带符号金额
        for d, desc, amt, bal in txns:
            ws.append([d, desc, amt, bal])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_single():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Acct"
    ws.append(["Account No.", "123-4-56789-0"])
    ws.append(["Date", "Description", "Amount", "Balance"])
    ws.append(["01/01/2026", "dep", 1000.0, 1000.0])
    ws.append(["02/01/2026", "dep", 500.0, 1500.0])
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class MultiAccountParseTests(unittest.TestCase):

    def test_multi_account_flag_and_codes(self):
        res = parse_bank_stmt_xlsx_direct(_build_multi(), "ktb.xlsx")
        self.assertTrue(res["ok"])
        self.assertTrue(res.get("multi_account"))
        self.assertEqual(set(res.get("account_codes") or []), {"984-2-99825-8", "984-2-99860-6"})
        self.assertEqual(len(res.get("accounts") or []), 2)

    def test_per_account_opening_captured_and_balance_clean(self):
        res = parse_bank_stmt_xlsx_direct(_build_multi(), "ktb.xlsx")
        by = {a["account_no"]: a for a in res["accounts"]}
        # 表头上方 ยกมา 期初被捕获
        self.assertAlmostEqual(by["984-2-99860-6"]["opening"], -7409714.58, places=2)
        # 逐账户独立校验 → 不应因 -740万 这条独立链被误标(全 sheet 自洽)
        for a in res["accounts"]:
            warn = sum(1 for r in a["rows"] if r.balance_ok is False)
            self.assertEqual(warn, 0, f"账户 {a['account_no']} 不应有余额误标")
        # 每行打了 account_no 标签
        for a in res["accounts"]:
            self.assertTrue(all(r.account_no == a["account_no"] for r in a["rows"]))

    def test_single_account_not_flagged_multi(self):
        res = parse_bank_stmt_xlsx_direct(_build_single(), "single.xlsx")
        self.assertTrue(res["ok"])
        self.assertFalse(res.get("multi_account"), "单账户不应被标多账户")


def _build_zero_close_blank_opening():
    """v118.35.0.66 真实案例(KTB 8258):期初汇总区被银行清空(无 ยกมา 标签),
    末笔 Sweep 把账户余额归零(期末真值 = 0.0)。锁定:
      - 期末 0.0 不被当成『空单元格』丢掉(此前 str(c or '') 把 0.0 变 '')
      - 期初被首笔『余额 − 净额』数学反推回来(39.15)"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "984-2-99825-8"
    ws.append(["Account No.", "984-2-99825-8"])
    ws.append([])  # 期初汇总区留白(无标签)
    ws.append(["Date", "Description", "Amount", "Balance"])  # 单一带符号金额
    ws.append(["01/01/2026", "dep", 2000.0, 2039.15])  # 首笔 → 反推期初 39.15
    ws.append(["31/01/2026", "Sweep MINBAL to xxx", -2039.15, 0.0])  # 归零 → 期末 0.0
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_known_opening_with_gap():
    """期初有印刷标签(ยกมา 1000)但中间一笔余额跳变与金额对不上(模拟漏读/读错),
    锁定完整性交叉校验会产出 closing_mismatch(不再静默)。"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "111-1-11111-1"
    ws.append(["Account No.", "111-1-11111-1"])
    ws.append(["", "", "", "", "ยอดยกมา", 1000.0])
    ws.append(["Date", "Description", "Amount", "Balance"])
    ws.append(["01/01/2026", "ok", 500.0, 1500.0])  # 1000+500=1500 ✓
    ws.append(["02/01/2026", "bad", 500.0, 3000.0])  # 余额跳 1500 但金额只 500 → 缺口
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


class XlsxClosingAndCompletenessTests(unittest.TestCase):
    """v118.35.0.66 · .xls 期末 0 归零 + 空期初反推 + 完整性交叉校验。"""

    def test_zero_closing_preserved_and_opening_derived(self):
        res = parse_bank_stmt_xlsx_direct(_build_zero_close_blank_opening(), "ktb8258.xlsx")
        self.assertTrue(res["ok"])
        a = res["accounts"][0]
        self.assertAlmostEqual(a["closing"], 0.0, places=2)  # 期末真值 0(此前误报 2039.15)
        self.assertTrue(a["opening_known"])  # 期初被反推
        self.assertAlmostEqual(a["opening"], 39.15, places=2)
        self.assertTrue(res["completeness"]["ok"])  # 链对平 · 无误报

    def test_completeness_catches_mismatch(self):
        res = parse_bank_stmt_xlsx_direct(_build_known_opening_with_gap(), "gap.xlsx")
        self.assertTrue(res["ok"])
        self.assertFalse(res["completeness"]["ok"], "余额链对不上必须被抓到 · 不能静默")
        types = [it["type"] for it in res["completeness"]["issues"]]
        self.assertIn("closing_mismatch", types)


if __name__ == "__main__":
    unittest.main()
