# -*- coding: utf-8 -*-
"""
Pearnly · Excel 导出模块
"普通模板":17 列标准发票结构,系统默认,所有套餐可用
"""

import io
from typing import List, Dict, Any
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# 普通模板列定义
# ============================================================
STANDARD_COLUMNS = [
    # (key, 中文, 英文, 泰文, 日文, width)
    ("no", "序号", "No.", "ลำดับ", "番号", 6),
    ("filename", "文件名", "Filename", "ชื่อไฟล์", "ファイル名", 28),
    ("invoice_no", "发票号", "Invoice No.", "เลขที่", "請求書番号", 18),
    ("invoice_date", "日期", "Date", "วันที่", "日付", 14),
    ("invoice_date_raw", "原始日期", "Date (raw)", "วันที่ต้นฉบับ", "原文日付", 14),
    ("seller_name", "卖方名称", "Seller Name", "ผู้ขาย", "売主名称", 30),
    ("seller_tax", "卖方税号", "Seller Tax ID", "เลขภาษีผู้ขาย", "売主税番号", 16),
    ("seller_addr", "卖方地址", "Seller Addr.", "ที่อยู่ผู้ขาย", "売主住所", 30),
    ("buyer_name", "买方名称", "Buyer Name", "ผู้ซื้อ", "買主名称", 30),
    ("buyer_tax", "买方税号", "Buyer Tax ID", "เลขภาษีผู้ซื้อ", "買主税番号", 16),
    ("buyer_addr", "买方地址", "Buyer Addr.", "ที่อยู่ผู้ซื้อ", "買主住所", 30),
    ("subtotal", "未税金额", "Subtotal", "ยอดก่อน VAT", "税抜金額", 14),
    ("vat", "税额", "VAT", "ภาษี", "税額", 12),
    ("wht_rate", "预扣税率", "WHT %", "อัตรา หัก ณ ที่จ่าย", "源泉%", 10),
    ("wht_amount", "预扣税额", "WHT Amount", "ภาษีหัก ณ ที่จ่าย", "源泉額", 14),
    ("total", "总金额", "Total", "ยอดรวม", "合計", 14),
    ("item_count", "明细数", "Items", "จำนวนรายการ", "明細数", 10),
    ("items", "商品明细", "Line Items", "รายการสินค้า", "明細", 40),
    ("notes", "备注", "Notes", "หมายเหตุ", "備考", 20),
]

COLUMN_LABEL_BY_LANG = {
    "zh": {c[0]: c[1] for c in STANDARD_COLUMNS},
    "en": {c[0]: c[2] for c in STANDARD_COLUMNS},
    "th": {c[0]: c[3] for c in STANDARD_COLUMNS},
    "ja": {c[0]: c[4] for c in STANDARD_COLUMNS},
}

# 样式
_THIN_SIDE = Side(style="thin", color="E2E8F0")
_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="Calibri", size=10)
_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


def _fmt_items(items: List[Dict[str, Any]]) -> str:
    """商品明细 → 多行字符串"""
    if not items:
        return ""
    lines = []
    for i, it in enumerate(items, start=1):
        name = it.get("name") or ""
        qty = it.get("qty") or it.get("quantity") or ""
        price = it.get("price") or it.get("unit_price") or ""
        sub = it.get("subtotal") or it.get("total") or ""
        parts = [f"{i}."]
        if name:
            parts.append(str(name))
        if qty:
            parts.append(f"×{qty}")
        if price:
            parts.append(f"@{price}")
        if sub:
            parts.append(f"={sub}")
        lines.append(" ".join(parts))
    return "\n".join(lines)


def _row_from_record(rec: Dict[str, Any], idx: int) -> Dict[str, Any]:
    """把识别结果转成一行"""
    # rec 来自前端,已经过合并:{filename, pages, merged_fields, edits}
    # 前端会把最终字段(含编辑)传到 merged_fields
    mf = rec.get("merged_fields") or {}

    return {
        "no": idx,
        "filename": rec.get("filename") or "",
        "invoice_no": mf.get("invoice_number") or mf.get("invoice_no") or "",
        "invoice_date": mf.get("date") or mf.get("invoice_date") or "",
        "invoice_date_raw": mf.get("date_raw") or "",
        "seller_name": mf.get("seller_name") or "",
        "seller_tax": mf.get("seller_tax") or "",
        "seller_addr": mf.get("seller_addr") or "",
        "buyer_name": mf.get("buyer_name") or "",
        "buyer_tax": mf.get("buyer_tax") or "",
        "buyer_addr": mf.get("buyer_addr") or "",
        "subtotal": mf.get("subtotal") or "",
        "vat": mf.get("vat") or "",
        "wht_rate": mf.get("wht_rate") or "",
        "wht_amount": mf.get("wht_amount") or "",
        "total": mf.get("total_amount") or mf.get("total") or "",
        "item_count": len(mf.get("items") or []),
        "items": _fmt_items(mf.get("items") or []),
        "notes": mf.get("notes") or "",
    }


def build_xlsx(records: List[Dict[str, Any]], lang: str = "zh") -> bytes:
    """
    生成 Excel 字节流
    records: 前端传来的识别结果数组
    lang:    列头语言
    """
    if lang not in COLUMN_LABEL_BY_LANG:
        lang = "zh"

    wb = Workbook()
    ws = wb.active
    ws.title = {"zh": "识别结果", "en": "Results", "th": "ผลลัพธ์", "ja": "認識結果"}[lang]

    # 列头
    labels = COLUMN_LABEL_BY_LANG[lang]
    for col_idx, (key, *_, width) in enumerate(STANDARD_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=labels[key])
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 28

    # 数据行
    for row_idx, rec in enumerate(records, start=1):
        row_data = _row_from_record(rec, row_idx)
        for col_idx, (key, *_) in enumerate(STANDARD_COLUMNS, start=1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx + 1, column=col_idx, value=value)
            cell.font = _CELL_FONT
            cell.border = _BORDER
            # 对齐
            if key in ("no", "invoice_date", "invoice_date_raw", "item_count", "wht_rate"):
                cell.alignment = _CENTER
            elif key in ("subtotal", "vat", "total", "wht_amount"):
                cell.alignment = _RIGHT
                # 数字格式
                try:
                    if value != "":
                        float(str(value).replace(",", ""))
                        cell.number_format = "#,##0.00"
                except (ValueError, TypeError):
                    pass
            else:
                cell.alignment = _LEFT

    # 冻结首行
    ws.freeze_panes = "A2"

    # 写入字节
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def default_filename() -> str:
    return f"mr-pilot-results-{datetime.now().strftime('%Y%m%d-%H%M%S')}.xlsx"
