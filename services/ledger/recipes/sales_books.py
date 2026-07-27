# -*- coding: utf-8 -*-
"""配方 · 销售票 → 四表账簿工作簿(明细 / 日记账 / 分类账 / 试算平衡)。

会计把几张销售小票丢进来,拿走一个她能核对、改完能塞回来的文件。第五张「รายการรอจำแนก」
不是第五个功能,是这四张表的诚实边界:日期跨月、付款方式读不出的票不入账,但要看得见。

内含 VAT 全程 Decimal(gross × 7/107 四舍五入到分,走 services/sales_agg/vat.split_gross),
表里对应的格子写公式;两边同口径,会计不改数时表里算出来的就等于这里报的数。
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any, Dict, Optional, Sequence, Tuple

from openpyxl import Workbook

from services.ledger import entries
from services.ledger.accounts import ChartResolution, resolve_chart
from services.ledger.models import SalesDoc
from services.ledger.registry import RecipeResult, RecipeSpec, register
from services.ledger.sheets.detail import SHEET_DETAIL, write_detail_sheet
from services.ledger.sheets.journal import SHEET_JOURNAL, write_journal_sheet
from services.ledger.sheets.ledger import SHEET_LEDGER, write_ledger_sheet
from services.ledger.sheets.pending import SHEET_PENDING, write_pending_sheet
from services.ledger.sheets.trial_balance import SHEET_TRIAL_BALANCE, write_trial_balance_sheet

RECIPE_CODE = "sales_books"
SHEETS: Tuple[str, ...] = (
    SHEET_DETAIL,
    SHEET_JOURNAL,
    SHEET_LEDGER,
    SHEET_TRIAL_BALANCE,
    SHEET_PENDING,
)

ZERO = Decimal("0")

# 单批上限。超了如实说「这批太大,按期分批导」,不给会计转三分钟的圈再交一个打不开的文件。
MAX_LINES = 2000


def _safe_stem(text: str) -> str:
    keep = [c if c.isalnum() or c in "-_" else "_" for c in (text or "").strip()]
    return ("".join(keep).strip("_") or "ledger")[:48]


def build_filename(title: str, period_label: str) -> str:
    parts = [_safe_stem(title), _safe_stem(period_label), "บัญชี4ตาราง"]
    return "_".join(p for p in parts if p) + ".xlsx"


def _doc_summary(doc: SalesDoc) -> Dict[str, Any]:
    return {
        "invoice_number": doc.invoice_number,
        "date": doc.doc_date.isoformat() if doc.doc_date else "",
        "date_source": doc.date_source,
        "settlement": doc.settlement,
        "payment_method": doc.payment_method_raw,
        "customer_name": doc.customer_name,
        "line_count": len(doc.lines),
        "gross": doc.gross,
        "vat": doc.vat,
        "net": doc.net,
        "row_key": doc.row_key,
    }


def _pending_summary(doc: SalesDoc) -> Dict[str, Any]:
    out = _doc_summary(doc)
    out["reason"] = doc.pending_reason
    return out


def _protect_derived_sheets(wb) -> None:
    """派生表上锁 —— 会计要改的是明细表,在日记账里删行会断掉递推余额那条链。

    openpyxl 的 sheet protection 是弱保护(能绕),挡的是误编辑不是恶意,别当安全边界。
    """
    for name in (SHEET_JOURNAL, SHEET_LEDGER, SHEET_TRIAL_BALANCE):
        wb[name].protection.sheet = True


def _numbers(booked: Sequence[SalesDoc], legs) -> Dict[str, Any]:
    """权威数字 —— 表里那些格子是公式,openpyxl 读不到结果,回执和闸都取这里。"""
    debit, credit = entries.totals(legs)
    return {
        "doc_count": len(booked),
        "line_count": sum(len(d.lines) for d in booked),
        "gross_total": sum((d.gross for d in booked), ZERO),
        "vat_total": sum((d.vat for d in booked), ZERO),
        "net_total": sum((d.net for d in booked), ZERO),
        "debit_total": debit,
        "credit_total": credit,
    }


def build(
    docs: Sequence[SalesDoc],
    *,
    chart: Optional[ChartResolution] = None,
    title: str = "",
    period_label: str = "",
) -> RecipeResult:
    """一批销售票 → 四表账簿工作簿。chart 省略时按「未接账套」退回通用科目码并在表头标注。"""
    chart = chart or resolve_chart()
    docs = tuple(docs or ())
    booked = tuple(d for d in docs if d.bookable)
    pending = tuple(d for d in docs if not d.bookable)

    total_lines = sum(len(d.lines) for d in booked)
    if total_lines > MAX_LINES:
        raise ValueError(f"本批 {total_lines} 行超过单批上限 {MAX_LINES} 行,请按账期分批导出")

    legs = entries.build_legs(booked, chart)
    banner = chart.banner

    wb = Workbook()
    wb.active.title = SHEET_DETAIL
    detail_refs = write_detail_sheet(wb.active, booked, title=title, banner=banner)
    journal_refs = write_journal_sheet(
        wb.create_sheet(SHEET_JOURNAL), legs, detail_refs, title=title, banner=banner
    )
    write_ledger_sheet(wb.create_sheet(SHEET_LEDGER), legs, detail_refs, title=title, banner=banner)
    write_trial_balance_sheet(
        wb.create_sheet(SHEET_TRIAL_BALANCE),
        entries.ledger_accounts(legs),
        journal_refs,
        title=title,
        banner=banner,
        period_label=period_label,
    )
    write_pending_sheet(wb.create_sheet(SHEET_PENDING), pending, title=title, banner=banner)
    _protect_derived_sheets(wb)

    buf = io.BytesIO()
    wb.save(buf)
    return RecipeResult(
        code=RECIPE_CODE,
        filename=build_filename(title, period_label),
        content=buf.getvalue(),
        sheets=SHEETS,
        numbers=_numbers(booked, legs),
        documents=tuple(_doc_summary(d) for d in booked),
        pending=tuple(_pending_summary(d) for d in pending),
        chart_source=chart.source,
        chart_unresolved=chart.unresolved,
        warnings=(banner,) if chart.unresolved or not chart.erp_type else (),
    )


SPEC = register(
    RecipeSpec(
        code=RECIPE_CODE,
        title_th="สมุดบัญชีจากใบเสร็จขาย (4 ตาราง)",
        title_zh="销售票四表账簿",
        sheets=SHEETS,
        build=build,
        input_kind="sales_docs",
        notes=(
            "跨表引用一律按票号 SUMIF,会计删行不产生 #REF!",
            "试算平衡里的自检格只给眼睛看,权威判定由回导时代码重算",
        ),
    )
)
