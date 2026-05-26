# -*- coding: utf-8 -*-
"""
v118.27.8.1.2 · 导出模板「泰国销售明细」
=========================================================
按用户提供的 Excel 格式生成 · 跟泰国本地销售清单习惯一致

v27.8.1.2 修复(MR.ERP 团队 Korn 反馈 · 2026-05-10):
  - 列 7 (G) 公式从 `=F*G`(BUG · F 是单价 G 是自己 · 循环引用)
    改为 `=E*F`(数量×单价)
  - 列 8 (H) `รวมจำนวนเงิน` 改成每行都填 =E*F(MR.ERP 按行算 · 不再只第 1 行)
  - 列 9 (I) `รวมจำนวนเงินก่อนVAT` 同上 · 每行 =E*F
  - 列 10 (J) `VAT` 每行 =I*0.07 · 不再只第 1 行

特点:
  - 12 列泰文表头(วันที่ / เลขที่ / รหัสลูกค้า / รหัสสินค้า / จำนวน / ...)
  - 每张发票按 items 拆 N 行(共享单据级字段:日期/单号/客户)
  - 列 7 (G) จำนวนเงิน  = =E*F(数量×单价)
  - 列 8 (H) รวมจำนวนเงิน = =E*F
  - 列 9 (I) รวมก่อน VAT  = =E*F
  - 列 10 (J) VAT         = =I*0.07
  - 列 11/12(备注 / Status)留空给会计填

数据来源:
  records: list of {filename, engine, merged_fields}
  - merged_fields = 单张发票的合并字段(v27.5.1 已确保多发票按张拆)
  - merged_fields.items = 商品明细数组
"""

import io
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side
except ImportError:
    Workbook = None  # type: ignore

logger = logging.getLogger(__name__)


# 12 列泰文表头(完全照用户给的模板)
HEADERS_TH = [
    "วันที่",  # 1  date
    "เลขที่",  # 2  invoice_no
    "รหัสลูกค้า",  # 3  customer
    "รหัสสินค้า",  # 4  product / description
    "จำนวน",  # 5  qty
    "ราคาต่อหน่วย",  # 6  unit_price
    "จำนวนเงิน",  # 7  amount = F*G(公式)
    "รวมจำนวนเงิน",  # 8  total(发票第 1 行)
    "รวมจำนวนเงินก่อนVAT",  # 9  subtotal(发票第 1 行)
    "VAT",  # 10 = J*0.07(发票第 1 行 · 公式)
    "อ้างอิง",  # 11 ref / note(留空)
    "Status",  # 12 留空给会计填
]

COLUMN_WIDTHS = [12, 14, 30, 22, 8, 12, 14, 14, 18, 10, 12, 10]


def _norm_date(raw: Any) -> Optional[datetime]:
    """日期字符串 → datetime · 让 cell.number_format='yyyy/m/d' 自动渲染"""
    if not raw:
        return None
    if isinstance(raw, datetime):
        return raw
    s = str(raw).strip()
    if not s:
        return None
    # 容错多种格式 · 含泰国佛历(2569 → 公历 2026)
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            d = datetime.strptime(s[:10], fmt)
            # 泰国佛历转西历(2569 → 2026)· year > 2400 视为佛历
            if d.year > 2400:
                d = d.replace(year=d.year - 543)
            return d
        except Exception:
            pass
    return None


def _to_float(raw: Any) -> Optional[float]:
    if raw is None or raw == "":
        return None
    try:
        s = str(raw).replace(",", "").replace("฿", "").replace("THB", "").strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


def _str(raw: Any) -> str:
    if raw is None:
        return ""
    s = str(raw).strip()
    return "" if s.lower() in ("none", "null", "nan") else s


def build_sales_detail_xlsx(records: List[Dict[str, Any]], lang: str = "zh") -> bytes:
    """
    返回 .xlsx 二进制

    records: list of {filename, engine, merged_fields}
    每张发票按 items 拆 N 行 · 共享单据级字段
    """
    if Workbook is None:
        raise RuntimeError("openpyxl not installed")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # 表头(R1)
    for i, h in enumerate(HEADERS_TH, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = bold
        c.alignment = center
        c.border = border

    # 列宽
    for i, w in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    row = 2
    for rec in records or []:
        if not isinstance(rec, dict):
            continue
        f = rec.get("merged_fields") or {}
        if not isinstance(f, dict):
            continue

        date_v = _norm_date(f.get("date"))
        invoice_no = _str(f.get("invoice_number"))
        customer = _str(f.get("buyer_name")) or _str(f.get("customer_name"))
        total = _to_float(f.get("total_amount"))
        subtotal = _to_float(f.get("subtotal"))
        if subtotal is None and total is not None:
            # 从 total 反推 subtotal(假设 7% VAT)· 用户可在 Excel 改公式
            subtotal = round(total / 1.07, 2)

        items = f.get("items")
        if not isinstance(items, list) or len(items) == 0:
            # 没明细 · 单行兜底(空商品 + 总额)
            items = [{"description": "", "qty": None, "unit_price": None, "amount": total}]

        for it_idx, it in enumerate(items):
            if not isinstance(it, dict):
                it = {}
            qty = _to_float(it.get("qty") or it.get("quantity"))
            unit_price = _to_float(it.get("unit_price") or it.get("price"))
            description = (
                _str(it.get("description")) or _str(it.get("product_name")) or _str(it.get("name"))
            )
            line_amount = _to_float(it.get("amount") or it.get("total"))

            # 1 日期(datetime · cell 格式 yyyy/m/d)
            cell_date = ws.cell(row=row, column=1, value=date_v)
            if date_v:
                cell_date.number_format = "yyyy/m/d"
            # 2 单号(纯文本 · 防 Excel 把 IV69/00271 解释成日期)
            c2 = ws.cell(row=row, column=2, value=invoice_no)
            c2.number_format = "@"
            # 3 客户
            ws.cell(row=row, column=3, value=customer)
            # 4 商品
            ws.cell(row=row, column=4, value=description)
            # 5 数量
            ws.cell(row=row, column=5, value=qty if qty is not None else None)
            # 6 单价
            ws.cell(row=row, column=6, value=unit_price if unit_price is not None else None)

            # v27.8.1.2 修复(Korn 反馈):
            # G/H/I 三列每行都填 = E*F(数量×单价)· J 每行 = I*0.07
            # 旧版(v27.5.3)bug:G 公式写成 =F*G 是循环引用 · 而且 H/I/J 只在第 1 行写
            if qty is not None and unit_price is not None:
                ws.cell(row=row, column=7, value=f"=E{row}*F{row}")  # G  จำนวนเงิน
                ws.cell(row=row, column=8, value=f"=E{row}*F{row}")  # H  รวมจำนวนเงิน
                ws.cell(row=row, column=9, value=f"=E{row}*F{row}")  # I  รวมก่อน VAT
                ws.cell(row=row, column=10, value=f"=I{row}*0.07")  # J  VAT
            elif line_amount is not None:
                # qty/price 缺(OCR 没拆出来)· 用预算好的 line_amount 死值兜底
                ws.cell(row=row, column=7, value=line_amount)
                ws.cell(row=row, column=8, value=line_amount)
                ws.cell(row=row, column=9, value=line_amount)
                ws.cell(row=row, column=10, value=round(line_amount * 0.07, 2))
            # 11 อ้างอิง · 12 Status · 留空(用户后续填)

            row += 1

    # 输出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def sales_detail_filename(prefix: str = "Pearnly_SalesDetail") -> str:
    return f'{prefix}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
