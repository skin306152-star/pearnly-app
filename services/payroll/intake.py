# -*- coding: utf-8 -*-
"""工资 Excel 进料兼容层:读工作簿 → 单元格归一 → 按列映射装配结构化行(方案 §2)。

「任何工资表都能吃」= 读工作簿只出 (表头, 原始行),不假设列序;列语义由 guess.py 猜 +
调用方确认后的 column_map 决定,build_rows 只按 map 取值。手工键入(list[dict])走
rows_from_manual,与导入同一归一 + 同一校验(单一事实源,方案 §2.4)。

钱字段归一为 Decimal(禁 float);日期归一为公历 date(佛历自动折算)。解析不出不臆造:
金额/日期留 None + 原值进 raw,交 validate 逐行点名。
"""

from __future__ import annotations

import datetime as _dt
import io
from decimal import Decimal, InvalidOperation
from typing import Optional

from openpyxl import load_workbook

from services.payroll import model

_BE_OFFSET = 543
_BE_THRESHOLD = 2400


def read_workbook(data: bytes) -> tuple[list, list]:
    """xlsx 字节流 → (表头行, 数据行列表)。取活动表首行为表头,其后非空行为数据。

    只读值(data_only)、不解析公式;整行全空即跳过(合计/间隔行不当数据,交调用方处理)。
    """
    wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
    ws = wb.active
    matrix = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    if not matrix:
        return [], []
    header = matrix[0]
    rows = [r for r in matrix[1:] if any(c is not None and str(c).strip() != "" for c in r)]
    return list(header), rows


def coerce_id(value) -> str:
    """身份证/税号归一为数字串。openpyxl 把纯数字身份证读成 int(丢首位 0 风险)——
    数字型补足 13 位(仅当疑似丢首零),非纯数字原样去空格,交 mod-11 兜底判真伪。"""
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        text = str(value)
        return text.zfill(13) if len(text) == 12 else text
    return str(value).strip().replace("-", "").replace(" ", "")


def parse_amount(value) -> Optional[Decimal]:
    """金额 → Decimal(禁 float 中转)。空/不可解析 → None(不当 0,交 V 层点名)。"""
    if value is None or (isinstance(value, str) and value.strip() == ""):
        return None
    if isinstance(value, bool):
        return None
    try:
        return Decimal(str(value).replace(",", "").strip())
    except (InvalidOperation, ValueError):
        return None


def parse_paid_date(value) -> Optional[_dt.date]:
    """支付日 → 公历 date。吃 Excel 日期对象 / ddmmyyyy(佛历,金标 31052569)/ 带分隔符串。

    佛历自动折算(年 ≥ 2400 减 543)。认不出 → None,不臆造。
    """
    if value is None:
        return None
    if isinstance(value, _dt.datetime):
        return value.date()
    if isinstance(value, _dt.date):
        return value
    text = str(value).strip()
    if not text:
        return None
    compact = _parse_compact_be(text)
    if compact:
        return compact
    return _parse_separated(text)


def _to_ad_year(year: int) -> int:
    return year - _BE_OFFSET if year >= _BE_THRESHOLD else year


def _parse_compact_be(text: str) -> Optional[_dt.date]:
    """ddmmyyyy(8 位,年为佛历):金标 '31052569' → 2026-05-31。"""
    if not (text.isdigit() and len(text) == 8):
        return None
    day, month, year = int(text[:2]), int(text[2:4]), int(text[4:])
    return _safe_date(_to_ad_year(year), month, day)


def _parse_separated(text: str) -> Optional[_dt.date]:
    """dd/mm/yyyy · yyyy-mm-dd · dd-mm-yy 等带分隔符形态(佛历自动折算)。"""
    parts = [p for p in _split_date_parts(text) if p != ""]
    if len(parts) != 3 or not all(p.isdigit() for p in parts):
        return None
    a, b, c = (int(p) for p in parts)
    if len(parts[0]) == 4:  # yyyy-mm-dd
        year, month, day = a, b, c
    else:  # dd-mm-yyyy / dd-mm-yy
        day, month, year = a, b, c
        if len(parts[2]) == 2:  # 2 位年按泰国财税惯例视作佛历缩写 25xx
            year = 2500 + year
    return _safe_date(_to_ad_year(year), month, day)


def _split_date_parts(text: str) -> list:
    for sep in ("/", "-", "."):
        if sep in text:
            return text.split(sep)
    return [text]


def _safe_date(year: int, month: int, day: int) -> Optional[_dt.date]:
    try:
        return _dt.date(year, month, day)
    except ValueError:
        return None


def build_rows(
    header: list, raw_rows: list, column_map: dict, *, fixed_values: Optional[dict] = None
) -> list:
    """原始行 × 确认过的 column_map → 结构化 PayrollRow 列表(方案 §2)。

    column_map: {语义键: 0-based 列号}。fixed_values: 整表同值字段(收入码/支付日/条件等,
    方案 §2.2「整列=一个值」),覆盖率高于列映射。序号缺则按行序自动补。
    """
    fixed = fixed_values or {}
    out = []
    for idx, raw in enumerate(raw_rows, start=1):
        cell = {key: _cell_at(raw, col) for key, col in column_map.items()}
        out.append(_assemble_row(idx, cell, fixed))
    return out


def rows_from_manual(entries: list, *, fixed_values: Optional[dict] = None) -> list:
    """手工键入行(list[dict],键=语义字段)→ 结构化行,与导入同归一同校验(方案 §2.4)。"""
    fixed = fixed_values or {}
    out = []
    for idx, entry in enumerate(entries, start=1):
        out.append(_assemble_row(idx, dict(entry), fixed))
    return out


def is_employee_row(row) -> bool:
    """员工行判据:有身份证或姓名/称谓任一。合计/间隔行(只余金额或全空)不算员工。

    忘填身份证但有姓名的行仍算员工(交 V5 点名缺失),不静默丢;真正的合计行(无身份证
    无姓名,仅余总额)才被排除 —— 见 partition_rows。
    """
    return bool(row.employee_id or row.first_name or row.last_name or row.title)


def partition_rows(rows: list) -> tuple:
    """结构化行 → (员工行, 合计行申报总额)。合计行=非员工行里首个带金额者(方案 §3 V3 基准)。"""
    employees = [r for r in rows if is_employee_row(r)]
    declared_total = None
    for row in rows:
        if not is_employee_row(row) and row.paid_amount is not None:
            declared_total = row.paid_amount
            break
    return employees, declared_total


def _cell_at(raw: list, col: int):
    return raw[col] if 0 <= col < len(raw) else None


def _assemble_row(idx: int, cell: dict, fixed: dict) -> model.PayrollRow:
    def pick(key):
        return fixed[key] if key in fixed else cell.get(key)

    seq_raw = pick(model.F_SEQ)
    try:
        seq = int(seq_raw) if seq_raw not in (None, "") else idx
    except (TypeError, ValueError):
        seq = idx

    amount = parse_amount(pick(model.F_PAID_AMOUNT))
    wht = parse_amount(pick(model.F_WHT_AMOUNT))
    return model.PayrollRow(
        seq=seq,
        employee_id=coerce_id(pick(model.F_EMPLOYEE_ID)),
        title=_text(pick(model.F_TITLE)),
        first_name=_text(pick(model.F_FIRST_NAME)),
        last_name=_text(pick(model.F_LAST_NAME)),
        paid_amount=amount,
        wht_amount=wht if wht is not None else Decimal("0"),
        paid_date=parse_paid_date(pick(model.F_PAID_DATE)),
        income_code=_text(pick(model.F_INCOME_CODE)) or model.DEFAULT_INCOME_CODE,
        condition=_text(pick(model.F_CONDITION)) or model.DEFAULT_CONDITION,
        raw={
            "paid_amount": pick(model.F_PAID_AMOUNT),
            "paid_date": pick(model.F_PAID_DATE),
        },
    )


def _text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()
