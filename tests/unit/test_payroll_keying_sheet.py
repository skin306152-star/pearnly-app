# -*- coding: utf-8 -*-
"""键入底稿 xlsx:末行合计守恒(Decimal 同源 · 合成数据 Σ 复现金标口径)。"""

import datetime as dt
import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.payroll import keying_sheet, model


def _valid_id(prefix12: str) -> str:
    total = sum(int(prefix12[i]) * (13 - i) for i in range(12))
    return prefix12 + str((11 - total % 11) % 10)


def _row(seq, amount):
    return model.PayrollRow(
        seq=seq,
        employee_id=_valid_id("36501006974" + str(seq)),
        title="นาย",
        first_name="ก",
        last_name="ข",
        paid_amount=Decimal(str(amount)),
        wht_amount=Decimal("0"),
        paid_date=dt.date(2026, 5, 31),
    )


_ROWS = [_row(1, 13000), _row(2, 12040), _row(3, 14500)]


class KeyingSheetTests(unittest.TestCase):
    def test_totals_decimal_exact(self):
        self.assertEqual(keying_sheet.totals(_ROWS)["paid_amount"], Decimal("39540"))

    def test_workbook_total_row_matches_sum(self):
        wb = load_workbook(io.BytesIO(keying_sheet.build_workbook(_ROWS)))
        ws = wb.active
        total_row = len(_ROWS) + 2
        # 列 6 = 支付金额合计(见 _COLUMNS 序)。
        self.assertEqual(Decimal(str(ws.cell(row=total_row, column=6).value)), Decimal("39540"))

    def test_filename_marks_auxiliary_not_official(self):
        self.assertEqual(
            keying_sheet.build_filename(nid="0105548082417", tax_year_be="2569", tax_month="05"),
            "PND1_0105548082417_2569_05_keying.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
