# -*- coding: utf-8 -*-
"""K2 · services/fileconv/excel_in.py · Excel/CSV → ConvertResult。

GL 路四态全走真实 openpyxl 合成件(端到端,不桩 parse_gl_excel):①多科目 → 落 generic
(衍生单链余额 ≠ 原表每科目印刷余额,渲染即改数);②单科目带余额列自洽 → GL 路 ✓;
③篡改一格 debit(印刷 closing 不再吻合)→ 锚不匹配整件降 generic(R2:closing-only 锚
分不清「数字不自洽」和「多科目漏检」,后者更常见,假指控比漏检更伤);
④无余额列 → 落 generic(closing 非印刷,校验自己 = 自导自演)。

桩 parse_gl_excel 的链不平用例保留为辅助——衍生链按定义自洽,真实 xlsx 走不出
validate_ledger 的判不平分支,桩是覆盖该分支适配代码的唯一办法(独立锚已由 ③ 真件钉住)。
"""

from __future__ import annotations

import io
import unittest
from datetime import date
from unittest import mock

from openpyxl import Workbook

from services.fileconv import excel_in
from services.fileconv.model import (
    GENERIC_TABLE,
    GL_LEDGER,
    ISSUE_GL_BALANCE_CHAIN,
    STATUS_OK,
    STATUS_UNSUPPORTED_FORMAT,
)
from services.recon.bank_recon_types import GlRow


def _xlsx_bytes(rows) -> bytes:
    wb = Workbook()
    ws = wb.active
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gl_xlsx(*, debit1=500.00, accounts=("1140-01", "1140-01"), with_balance_col=True) -> bytes:
    """合成 GL:期初 1000 + 借 debit1 + 贷 200,余额列印刷值按 debit1=500 的正确账面写死
    (1500/1300)——篡改 debit1 即制造「印刷 closing 与借贷对不上」的真件。"""
    header = ["Date", "Doc No", "Description", "Account", "Debit", "Credit"]
    opening = ["Opening", "", "", "", "", ""]
    row1 = ["1/1/2026", "V001", "Sales", accounts[0], debit1, 0]
    row2 = ["2/1/2026", "V002", "Payment", accounts[1], 0, 200.00]
    if with_balance_col:
        header.append("Balance")
        opening.append(1000.00)
        row1.append(1500.00)
        row2.append(1300.00)
    return _xlsx_bytes([header, opening, row1, row2])


class GlConservedRoundtripTests(unittest.TestCase):
    """态②:单科目 + 余额列 + 借贷与印刷 closing 自洽 → GL 路,issues 空(真 ✓)。"""

    def setUp(self):
        self.result = excel_in.convert_excel(_gl_xlsx(), "gl.xlsx")

    def test_recognized_as_gl_and_conserved(self):
        self.assertEqual(self.result.doc_type, GL_LEDGER)
        self.assertEqual(self.result.status, STATUS_OK)
        self.assertTrue(self.result.conserved)
        self.assertEqual(self.result.issues, [])

    def test_stats_sums_are_faithful(self):
        stats = self.result.stats
        self.assertEqual(stats["engine"], "excel_gl")
        self.assertEqual(stats["row_count"], 2)
        self.assertEqual(stats["sum_debit"], "500.0")
        self.assertEqual(stats["sum_credit"], "200.0")
        self.assertEqual(stats["opening_balance"], "1000.0")
        self.assertEqual(stats["closing_balance"], "1300.0")
        self.assertEqual(stats["accounts"], ["1140-01"])

    def test_table_rows_carry_ledger_columns(self):
        table = self.result.tables[0]
        self.assertEqual(table.name, "GL Ledger")
        self.assertEqual(len(table.columns), 9)  # LEDGER_COLUMNS 定长契约
        self.assertEqual(len(table.rows), 2)
        self.assertEqual(table.rows[0][3], "V001")  # doc_no 落位不变
        self.assertEqual(table.rows[0][0], "1140-01")  # 真实科目码回填显示


class GlHonestyGateTests(unittest.TestCase):
    """R1 收口:GL 路门槛(单科目 + 印刷 closing)与真锚——全真件,不桩。"""

    def test_multi_account_falls_back_to_generic(self):
        """态①:多科目 → 落 generic。衍生单链余额会偏离原表每科目各自成链的印刷余额,
        进 GL 路渲染 = 悄悄改用户的数;generic 网格忠实转录原表(含它自己的余额列)。"""
        result = excel_in.convert_excel(
            _gl_xlsx(accounts=("1140-01", "5000-01")), "multi_acct.xlsx"
        )
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats["engine"], "excel_grid")
        self.assertEqual(result.issues, [])
        # 原表余额列印刷值原样在网格里,零改数。
        grid = result.tables[0].rows
        self.assertIn(1500, [c for row in grid for c in row])
        self.assertIn(1300, [c for row in grid for c in row])

    def test_anchor_mismatch_demotes_to_generic(self):
        """态③(R2):篡改一格 debit(500→600)→ 衍生链末值 1400 ≠ 印刷 closing 1300 →
        锚不匹配整件降 generic(带 gl_demote 留痕),网格余额列保持原表印刷值零改数。
        closing-only 锚分不清「数字不自洽」和「多科目漏检」,不指控只降级。"""
        result = excel_in.convert_excel(_gl_xlsx(debit1=600.00), "tampered.xlsx")
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats["engine"], "excel_grid")
        self.assertEqual(result.stats["gl_demote"], "closing_anchor_mismatch")
        self.assertEqual(result.issues, [])
        grid = result.tables[0].rows
        self.assertIn(1300, [c for row in grid for c in row])  # 原表印刷 closing 原样

    def test_refed_ledger_layout_with_unmapped_accounts_stays_faithful(self):
        """态⑤(R2 金标漏网复刻):多科目件但科目码 parse_gl_excel 哪都认不出(不在可映射
        列、doc_no/description 也提不出)→ accounts 空集绕过「单科目」门;每科目余额各自
        成链,衍生末值对不上印刷 closing → 锚守门降 generic,余额列印刷值逐字保留。
        修前这类件(主窗金标:735 行 MR.ERP 件)混进 GL 路被改数+假指控「不平」。"""
        header = ["วันที่", "เลขที่", "รายละเอียด", "เดบิต", "เครดิต", "คงเหลือ"]
        rows = [
            header,
            ["1/1/2026", "A1", "เงินฝากธนาคาร", 500.00, 0, 1500.00],
            ["2/1/2026", "A2", "ขายสินค้า", 0, 200.00, 800.00],  # 换科目段,余额换链
        ]
        result = excel_in.convert_excel(_xlsx_bytes(rows), "refed.xlsx")
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats.get("gl_demote"), "closing_anchor_mismatch")
        grid = result.tables[0].rows
        flat = [c for row in grid for c in row]
        self.assertIn(1500, flat)
        self.assertIn(800, flat)  # 两条链的印刷余额都原样,零改数

    def test_no_balance_column_falls_back_to_generic(self):
        """态④:无余额列 → closing 是借贷衍生值(closing_printed=False),拿它当锚校验
        自己 = 自导自演 → 拒进 GL 路,落 generic 诚实「未校验」。"""
        result = excel_in.convert_excel(_gl_xlsx(with_balance_col=False), "no_bal.xlsx")
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats["engine"], "excel_grid")
        self.assertEqual(result.issues, [])


class GlUnbalancedStubTests(unittest.TestCase):
    """辅助(桩 parse_gl_excel):覆盖 validate_ledger 链不平分支的适配代码——真实 xlsx
    的衍生链按定义自洽走不到这里;真锚路径由 GlHonestyGateTests 真件钉住。"""

    def test_broken_chain_is_named_with_line_and_expected_actual(self):
        rows = [
            GlRow(
                date=date(2026, 1, 1),
                doc_no="V1",
                account_code="1140-01",
                description="ok",
                debit=100.0,
                credit=0.0,
                balance=1100.0,
            ),
            GlRow(
                date=date(2026, 1, 2),
                doc_no="V2",
                account_code="1140-01",
                description="bad",
                debit=0.0,
                credit=50.0,
                balance=999.0,  # 应为 1050
            ),
        ]
        parsed = {
            "ok": True,
            "rows": rows,
            "accounts": ["1140-01"],
            "opening": 1000.0,
            "closing": 999.0,
            "closing_printed": True,
            "row_count": 2,
        }
        with mock.patch("services.recon.bank_gl_excel.parse_gl_excel", return_value=parsed):
            result = excel_in.convert_excel(b"stubbed-parser-ignores-bytes", "bad.xlsx")

        self.assertFalse(result.conserved)
        self.assertEqual(len(result.issues), 1)
        issue = result.issues[0]
        self.assertEqual(issue.kind, ISSUE_GL_BALANCE_CHAIN)
        self.assertEqual(issue.line_no, 2)
        self.assertEqual(issue.expected, "1050.0")
        self.assertEqual(issue.actual, "999.0")

    def test_true_account_code_still_shown_in_table_despite_single_chain_key(self):
        """链 key 统一单链(见 excel_in._gl_row 注),但显示列必须换回真实科目码。"""
        rows = [
            GlRow(
                date=date(2026, 1, 1),
                doc_no="V1",
                account_code="1140-01",
                description="a",
                debit=100.0,
                credit=0.0,
                balance=1100.0,
            ),
        ]
        parsed = {
            "ok": True,
            "rows": rows,
            "accounts": ["1140-01"],
            "opening": 1000.0,
            "closing": 1100.0,
            "closing_printed": True,
            "row_count": 1,
        }
        with mock.patch("services.recon.bank_gl_excel.parse_gl_excel", return_value=parsed):
            result = excel_in.convert_excel(b"stub", "a.xlsx")
        self.assertEqual(result.tables[0].rows[0][0], "1140-01")
        self.assertEqual(result.stats["accounts"], ["1140-01"])


class GenericFallbackTests(unittest.TestCase):
    """认不出 GL 结构(无日期/借贷表头)→ generic 网格,诚实无守恒。"""

    def test_no_gl_headers_falls_back_to_generic_grid(self):
        data = _xlsx_bytes([["Name", "Amount", "Note"], ["Alice", 100, "hi"], ["Bob", 200, "yo"]])
        result = excel_in.convert_excel(data, "sheet.xlsx")
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.status, STATUS_OK)
        self.assertEqual(result.issues, [])
        self.assertEqual(result.stats["engine"], "excel_grid")
        self.assertEqual(
            result.tables[0].rows,
            [["Name", "Amount", "Note"], ["Alice", 100, "hi"], ["Bob", 200, "yo"]],
        )

    def test_merged_cell_fills_top_left_value(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Amount"])
        ws.append(["merged-a", 5])
        ws.append(["merged-b", 6])
        ws.merge_cells("A2:A3")
        buf = io.BytesIO()
        wb.save(buf)
        # openpyxl 写合并格时只有左上角(A2)有值,A3 为空——写完再手动确认场景合理。
        result = excel_in.convert_excel(buf.getvalue(), "merged.xlsx")
        rows = result.tables[0].rows
        self.assertEqual(rows[1][0], "merged-a")
        self.assertEqual(rows[2][0], "merged-a")  # 合并格左上值回填,非空白

    def test_empty_sheet_is_skipped_not_crashed(self):
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Empty"
        ws2 = wb.create_sheet("Data")
        ws2.append(["A", "B"])
        ws2.append([1, 2])
        buf = io.BytesIO()
        wb.save(buf)
        result = excel_in.convert_excel(buf.getvalue(), "two_sheets.xlsx")
        self.assertEqual(len(result.tables), 1)
        self.assertEqual(result.tables[0].name, "Data")


class CsvFallbackTests(unittest.TestCase):
    def test_generic_csv_grid(self):
        data = "Name,Amount\r\nAlice,100\r\nBob,200\r\n".encode("utf-8")
        result = excel_in.convert_excel(data, "sheet.csv")
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats["engine"], "excel_grid_csv")
        self.assertEqual(result.tables[0].rows[0], ["Name", "Amount"])


class XlsFallbackTests(unittest.TestCase):
    """老式 .xls:openpyxl 打不开,GL 路认不出时落 generic → 单独走 xlrd。"""

    def test_xls_rejected_when_xlrd_not_installed(self):
        real_import = __import__

        def fake_import(name, *a, **k):
            if name == "xlrd":
                raise ImportError("simulated: xlrd not installed")
            return real_import(name, *a, **k)

        with mock.patch("services.recon.bank_gl_excel.parse_gl_excel", return_value={"ok": False}):
            with mock.patch("builtins.__import__", side_effect=fake_import):
                result = excel_in.convert_excel(b"any bytes, xlrd import fails first", "old.xls")

        self.assertEqual(result.status, STATUS_UNSUPPORTED_FORMAT)
        self.assertEqual(result.stats["reason"], "xls_requires_resave")

    def test_xls_grid_read_via_xlrd_when_available(self):
        """xlrd 装了但 GL 认不出结构 → generic 网格,合并格同样左上值回填。"""

        class _FakeSheet:
            name = "Sheet1"
            nrows = 3
            ncols = 2
            merged_cells = [(1, 3, 0, 1)]  # 行 1-2(0-based)、列 0 合并

            def cell_value(self, r, c):
                grid = [["Name", "Amount"], ["merged", 5], ["", 6]]
                return grid[r][c]

        class _FakeWorkbook:
            def sheets(self):
                return [_FakeSheet()]

        with mock.patch("services.recon.bank_gl_excel.parse_gl_excel", return_value={"ok": False}):
            with mock.patch("xlrd.open_workbook", return_value=_FakeWorkbook()):
                result = excel_in.convert_excel(b"fake xls bytes", "old.xls")

        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.stats["engine"], "excel_grid_xls")
        rows = result.tables[0].rows
        self.assertEqual(rows[1][0], "merged")
        self.assertEqual(rows[2][0], "merged")  # 合并格左上值回填


class RejectTests(unittest.TestCase):
    def test_empty_file_is_rejected_honestly(self):
        result = excel_in.convert_excel(b"", "empty.xlsx")
        self.assertEqual(result.status, STATUS_UNSUPPORTED_FORMAT)
        self.assertEqual(result.stats["reason"], "empty_file")

    def test_unsupported_extension_is_rejected(self):
        result = excel_in.convert_excel(b"whatever", "notes.docx")
        self.assertEqual(result.status, STATUS_UNSUPPORTED_FORMAT)
        self.assertIn("unsupported_ext", result.stats["reason"])

    def test_corrupt_xlsx_bytes_do_not_crash(self):
        result = excel_in.convert_excel(b"not a real spreadsheet at all", "corrupt.xlsx")
        self.assertEqual(result.status, STATUS_UNSUPPORTED_FORMAT)
        self.assertEqual(result.tables, [])


if __name__ == "__main__":
    unittest.main()
