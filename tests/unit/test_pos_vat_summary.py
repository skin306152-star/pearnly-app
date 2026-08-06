# -*- coding: utf-8 -*-
"""POS 销项月度汇总纯逻辑守门测试(POS 项目 · G3)。

不连库:金额求和/装配形状/月份校验/xlsx sheet 结构。真聚合 SQL 的笛卡尔积防护、跨月升级
在真库上的行为由 E2E/真库回归覆盖;此处锁的是「不重复计 VAT」的结构性前提——聚合查询不按
full_invoice_id 过滤(计一次)、全式票附录按 sold_at 而非 issue_date 归月(跨月不漏不重)。
"""

import io
import unittest
from datetime import date
from decimal import Decimal

from openpyxl import load_workbook

from services.pos import report, vat_summary, vat_summary_xlsx


class _Cur:
    """按 execute 顺序回放预置结果:每次 execute pop 一组 (one, all)(同 test_pos_report.py)。"""

    def __init__(self, scripted):
        self._scripted = list(scripted)
        self.queries = []

    def execute(self, sql, params=None):
        self.queries.append((sql, params))
        self._cur = self._scripted.pop(0)

    def fetchone(self):
        return self._cur[0]

    def fetchall(self):
        return self._cur[1]


class MonthParseTests(unittest.TestCase):
    def test_valid_month_resolves_first_and_last_day(self):
        self.assertEqual(vat_summary.parse_month("2026-06"), (date(2026, 6, 1), date(2026, 6, 30)))
        self.assertEqual(vat_summary.parse_month("2026-02"), (date(2026, 2, 1), date(2026, 2, 28)))

    def test_invalid_formats_rejected(self):
        for bad in ("", "2026-13", "2026-00", "2026/06", "26-06", "2026-6"):
            with self.assertRaises(vat_summary.MonthInvalid):
                vat_summary.parse_month(bad)


class DaysAndTotalsPrecisionTests(unittest.TestCase):
    """① 精确到分的求和:DB 侧已 SUM,本层只管字符串化,断言原样保留两位小数。"""

    def test_days_money_split_to_the_cent(self):
        cur = _Cur(
            [
                (
                    None,
                    [
                        {
                            "d": date(2026, 6, 1),
                            "sales_count": 3,
                            "subtotal": Decimal("1000.35"),
                            "discount_total": Decimal("50.15"),
                            "vat_amount": Decimal("66.51"),
                            "grand_total": Decimal("1016.71"),
                        }
                    ],
                )
            ]
        )
        out = vat_summary._days(cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(
            out[0],
            {
                "date": "2026-06-01",
                "sales_count": 3,
                "subtotal": "1000.35",
                "discount_total": "50.15",
                "vat_amount": "66.51",
                "gross": "1016.71",
            },
        )

    def test_totals_filter_shape_matches_report_kpi(self):
        """④ 退款口径与 report._kpi 同一 FILTER 形状(sale 拆列营收 · refund 单独净额)。"""
        cur = _Cur(
            [
                (
                    {
                        "subtotal": Decimal("10000"),
                        "discount_total": Decimal("200"),
                        "vat_amount": Decimal("700"),
                        "gross": Decimal("10500"),
                        "sales_count": 40,
                        "refund": Decimal("300"),
                    },
                    None,
                )
            ]
        )
        out = vat_summary._totals(cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        sql = cur.queries[0][0]
        self.assertIn("FILTER (WHERE sale_type='sale')", sql)
        self.assertIn("COALESCE(-SUM(grand_total) FILTER (WHERE sale_type='refund'),0)", sql)
        self.assertEqual(out["refund"], "300.00")
        self.assertEqual(out["gross"], "10500.00")

        # 同一份数字喂给 report._kpi:两处对同一个月不会报出两套「退款」。
        kpi_cur = _Cur(
            [
                ({"gross": Decimal("10500"), "sales_count": 40, "refund": Decimal("300")}, None),
                ({"cost": Decimal("0"), "complete": True}, None),
            ]
        )
        kpi = report._kpi(kpi_cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(kpi["refund"], out["refund"])
        self.assertEqual(kpi["gross"], out["gross"])

    def test_totals_and_days_do_not_filter_by_full_invoice_id(self):
        """② 已升级小票只计一次的结构性前提:聚合 SQL 不按 full_invoice_id 过滤——
        升级只是回填标记,金额留在 pos_sales 原行,不会被这两句悄悄排掉或算两次。"""
        totals_cur = _Cur(
            [
                (
                    {
                        "subtotal": 0,
                        "discount_total": 0,
                        "vat_amount": 0,
                        "gross": 0,
                        "sales_count": 0,
                        "refund": 0,
                    },
                    None,
                )
            ]
        )
        vat_summary._totals(totals_cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        days_cur = _Cur([(None, [])])
        vat_summary._days(days_cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        for cur in (totals_cur, days_cur):
            self.assertNotIn("full_invoice_id", cur.queries[0][0])


class ByMethodWithCountsTests(unittest.TestCase):
    def test_merges_report_amounts_with_own_counts(self):
        cur = _Cur(
            [
                (
                    None,
                    [{"method": "cash", "amount": Decimal("200")}],
                ),  # report._by_method tendered
                ({"chg": Decimal("0")}, None),  # report._by_method change
                (None, [{"method": "cash", "n": 5}, {"method": "promptpay", "n": 2}]),  # 笔数
            ]
        )
        out = vat_summary._by_method_with_counts(cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        self.assertEqual(out["cash"], {"amount": "200.00", "count": 5})
        # promptpay 没在 report._by_method 输出里(该桶净额为 0 会被 SQL 端过滤掉)→ 不出现在
        # 结果里(count 无所依附的方式本身没意义)。
        self.assertNotIn("promptpay", out)


class UpgradedSaleAndCrossMonthTests(unittest.TestCase):
    """② 已升级小票只计一次 + ③ 跨月升级:附录按原单 sold_at 归月,不按 issue_date。"""

    def test_full_invoice_follows_original_sold_month_not_issue_date(self):
        cur = _Cur(
            [
                (
                    None,
                    [
                        {
                            "doc_number": "INV2026-00042",
                            "issue_date": date(2026, 7, 3),  # 升级发生在下月
                            "source_receipt_no": "RCP-T1-2026-00010",
                            "buyer_name": "บริษัท เอบีซี จำกัด",
                            "buyer_tax_id": "0105500000001",
                            "subtotal": Decimal("500"),
                            "discount_total": Decimal("0"),
                            "vat_amount": Decimal("35"),
                            "grand_total": Decimal("535"),
                        }
                    ],
                )
            ]
        )
        out = vat_summary._full_invoices(cur, ("t", 9), date(2026, 6, 1), date(2026, 6, 30))
        sql, params = cur.queries[0]
        # 归属判据是 s.sold_at(原单),不是 d.issue_date——即便票开在 7 月,原单在 6 月的
        # 窗口参数(半开 [6/1, 7/1))仍把它筛进来。
        self.assertIn("s.sold_at", sql)
        self.assertNotIn("d.issue_date >=", sql)
        self.assertIn(date(2026, 6, 1), params)
        self.assertIn(date(2026, 7, 1), params)
        self.assertEqual(out[0]["doc_number"], "INV2026-00042")
        self.assertEqual(out[0]["issued_date"], "2026-07-03")

    def test_july_window_excludes_june_sold_invoice(self):
        """真库里同一张票不会同时满足 6 月与 7 月两个半开窗口(s.sold_at 只有一个值),
        此处用空回放模拟「7 月窗口查不到这张 6 月售出的票」,与上一条互为对照。"""
        cur = _Cur([(None, [])])
        out = vat_summary._full_invoices(cur, ("t", 9), date(2026, 7, 1), date(2026, 7, 31))
        params = cur.queries[0][1]
        self.assertIn(date(2026, 7, 1), params)
        self.assertIn(date(2026, 8, 1), params)
        self.assertEqual(out, [])


class MonthSummaryAssemblyTests(unittest.TestCase):
    """整装:一次 month_summary 调用要按序打出 7 句 SQL(days 1 + by_method 3 + totals 1 +
    abb 1 + full_invoices 1),形状与各分区单测一致。"""

    def test_shape(self):
        scripted = [
            (
                None,
                [
                    {
                        "d": date(2026, 6, 1),
                        "sales_count": 1,
                        "subtotal": Decimal("100.00"),
                        "discount_total": Decimal("0.00"),
                        "vat_amount": Decimal("7.00"),
                        "grand_total": Decimal("107.00"),
                    }
                ],
            ),  # days
            (None, [{"method": "cash", "amount": Decimal("107.00")}]),  # by_method tendered
            ({"chg": Decimal("0")}, None),  # by_method change
            (None, [{"method": "cash", "n": 1}]),  # by_method counts
            (
                {
                    "subtotal": Decimal("100.00"),
                    "discount_total": Decimal("0.00"),
                    "vat_amount": Decimal("7.00"),
                    "gross": Decimal("107.00"),
                    "sales_count": 1,
                    "refund": Decimal("0.00"),
                },
                None,
            ),  # totals
            (
                None,
                [
                    {
                        "d": date(2026, 6, 1),
                        "receipt_min": "RCP-T1-2026-00001",
                        "receipt_max": "RCP-T1-2026-00001",
                        "n": 1,
                    }
                ],
            ),  # abb_ranges
            (None, []),  # full_invoices
        ]
        cur = _Cur(scripted)
        out = vat_summary.month_summary(cur, tenant_id="t", workspace_client_id=9, month="2026-06")
        self.assertEqual(len(cur.queries), 7)
        self.assertEqual(out["month"], "2026-06")
        self.assertEqual(out["date_from"], "2026-06-01")
        self.assertEqual(out["date_to"], "2026-06-30")
        self.assertEqual(out["totals"]["gross"], "107.00")
        self.assertEqual(out["by_method"]["cash"], {"amount": "107.00", "count": 1})
        self.assertEqual(out["abb_ranges"][0]["count"], 1)
        self.assertEqual(out["full_invoices"], [])

    def test_invalid_month_raises_before_touching_cursor(self):
        cur = _Cur([])
        with self.assertRaises(vat_summary.MonthInvalid):
            vat_summary.month_summary(cur, tenant_id="t", workspace_client_id=9, month="bad")
        self.assertEqual(cur.queries, [])


class XlsxGoldenTests(unittest.TestCase):
    """⑥ xlsx sheet 结构 golden:sheet 名与表头列锁死,散布式改动会在这里第一时间炸。"""

    def _sample_data(self) -> dict:
        return {
            "month": "2026-06",
            "date_from": "2026-06-01",
            "date_to": "2026-06-30",
            "days": [
                {
                    "date": "2026-06-01",
                    "sales_count": 2,
                    "subtotal": "100.00",
                    "discount_total": "0.00",
                    "vat_amount": "7.00",
                    "gross": "107.00",
                }
            ],
            "by_method": {"cash": {"amount": "107.00", "count": 2}},
            "totals": {
                "subtotal": "100.00",
                "discount_total": "0.00",
                "vat_amount": "7.00",
                "gross": "107.00",
                "sales_count": 2,
                "refund": "0.00",
            },
            "abb_ranges": [
                {
                    "date": "2026-06-01",
                    "receipt_min": "RCP-T1-2026-00001",
                    "receipt_max": "RCP-T1-2026-00002",
                    "count": 2,
                }
            ],
            "full_invoices": [
                {
                    "doc_number": "INV2026-00042",
                    "issued_date": "2026-06-03",
                    "source_receipt_no": "RCP-T1-2026-00001",
                    "buyer_name": "บริษัท เอบีซี จำกัด",
                    "buyer_tax_id": "0105500000001",
                    "subtotal": "100.00",
                    "discount_total": "0.00",
                    "vat_amount": "7.00",
                    "gross": "107.00",
                }
            ],
        }

    def test_sheet_names_and_headers(self):
        wb = load_workbook(io.BytesIO(vat_summary_xlsx.build_xlsx(self._sample_data())))
        labels = vat_summary_xlsx._LABELS
        self.assertEqual(
            wb.sheetnames,
            [
                labels["sheet_days"],
                labels["sheet_method"],
                labels["sheet_invoices"],
                labels["sheet_abb"],
            ],
        )
        days_ws = wb[labels["sheet_days"]]
        self.assertEqual(
            [c.value for c in next(days_ws.iter_rows(min_row=1, max_row=1))],
            [
                labels[k]
                for k in (
                    "col_date",
                    "col_count",
                    "col_subtotal",
                    "col_discount",
                    "col_vat",
                    "col_gross",
                )
            ],
        )
        method_ws = wb[labels["sheet_method"]]
        self.assertEqual(
            [c.value for c in next(method_ws.iter_rows(min_row=1, max_row=1))],
            [labels[k] for k in ("col_method", "col_count", "col_amount")],
        )
        inv_ws = wb[labels["sheet_invoices"]]
        self.assertEqual(inv_ws.cell(row=2, column=1).value, "INV2026-00042")
        abb_ws = wb[labels["sheet_abb"]]
        self.assertEqual(
            [c.value for c in next(abb_ws.iter_rows(min_row=1, max_row=1))],
            [labels[k] for k in ("col_date", "col_receipt_min", "col_receipt_max", "col_count")],
        )

    def test_sheet_names_within_excel_limit(self):
        # Excel 硬限 31 字符,超限 openpyxl 直接抛异常——这是防回归闸,不是抽样检查。
        for key in ("sheet_days", "sheet_method", "sheet_invoices", "sheet_abb"):
            self.assertLessEqual(len(vat_summary_xlsx._LABELS[key]), 31)


if __name__ == "__main__":
    unittest.main()
