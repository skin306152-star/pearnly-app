# -*- coding: utf-8 -*-
"""fileconv 编排 + no_text_layer 拒绝 + xlsx 输出闭环。"""

import io
import unittest

import openpyxl

from services.fileconv.convert import convert_pages, convert_pdf
from services.fileconv.text_layer import has_text_layer
from services.fileconv.xlsx_out import build_xlsx
from services.fileconv.model import GL_LEDGER, STATUS_OK, STATUS_NO_TEXT_LAYER

_GL = """รายงานสมุดแยกประเภท
วันที่ เดบิต เครดิต ยอดคงเหลือ
1113-01 เงินฝาก -100.00
01/01/2569 JV JV001 pay 50.00 0.00 -50.00
02/01/2569 JV JV002 fee 0.00 30.00 -80.00"""


class NoTextLayerTests(unittest.TestCase):
    def test_has_text_layer_decision(self):
        self.assertFalse(has_text_layer([""]))
        self.assertFalse(has_text_layer(None))
        self.assertTrue(has_text_layer(["x" * 500]))

    def test_blank_pdf_rejected(self):
        from pypdf import PdfWriter

        writer = PdfWriter()
        writer.add_blank_page(width=200, height=200)
        buf = io.BytesIO()
        writer.write(buf)

        result = convert_pdf(buf.getvalue(), "blank.pdf")
        self.assertEqual(result.status, STATUS_NO_TEXT_LAYER)
        self.assertEqual(result.tables, [])


class ConvertPagesTests(unittest.TestCase):
    def test_gl_dispatch_and_conserved(self):
        result = convert_pages([_GL], "gl.pdf")
        self.assertEqual(result.doc_type, GL_LEDGER)
        self.assertEqual(result.status, STATUS_OK)
        self.assertTrue(result.conserved)
        self.assertEqual(result.stats["row_count"], 2)

    def test_broken_chain_not_conserved(self):
        result = convert_pages([_GL.replace("-80.00", "-79.00")], "gl.pdf")
        self.assertFalse(result.conserved)
        self.assertEqual(len(result.issues), 1)


class XlsxOutTests(unittest.TestCase):
    def test_xlsx_roundtrip_has_sheets(self):
        result = convert_pages([_GL], "gl.pdf")
        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(result)))
        self.assertIn("Issues", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)

    def test_no_text_layer_xlsx_marks_rejection(self):
        from pypdf import PdfWriter

        writer = PdfWriter()
        writer.add_blank_page(width=200, height=200)
        buf = io.BytesIO()
        writer.write(buf)
        result = convert_pdf(buf.getvalue(), "blank.pdf")

        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(result)))
        self.assertIn("Rejected", wb.sheetnames)
        self.assertEqual(wb["Rejected"]["A1"].value, "no_text_layer")


if __name__ == "__main__":
    unittest.main()
