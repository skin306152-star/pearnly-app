# -*- coding: utf-8 -*-
"""固定对账模板(新版)· 识别 + 多场景对账守门。

锁住:用户照『我们的固定模板』(期初 ยอดยกมา 行 + 余额列 + GL 借贷标注 + GL 去科目码)
填真实数据后——① 四语模板自产能被识别(期初读对)② 对账对的判对、错的判错(不漏配/不误配)。
不验对账匹配引擎本身(那套已调好 · 见 bank_recon_reconcile),只验固定模板这条链。
"""

import io
import unittest

import openpyxl

from services.recon import recon_templates as rt
from services.recon.bank_stmt_xlsx import parse_bank_stmt_xlsx_direct
from services.recon.bank_gl_excel import parse_gl_excel
from services.recon.bank_recon_reconcile import reconcile

STMT_OPEN = 800000.00
GL_OPEN = 0.0


def _tx(date, voucher, desc, amount, kind):
    return {"date": date, "voucher": voucher, "desc": desc, "amount": amount, "kind": kind}


# 基准:02/05/2026 · 泰语摘要 · 8 笔入账 + 1 笔 SIM 费出账(取自 TTB 原型口径)
BASE = [
    _tx("02/05/2026", "SVTAX26004029", "QR / SCB X4184 มัญชริณ", 1727.46, "in"),
    _tx("02/05/2026", "SVTAX26004030", "QR / GHB X4065 ไกล้รุ่ง", 3576.18, "in"),
    _tx("02/05/2026", "SVTAX26004031", "QR / SCB X6253 ดรุณี", 1361.04, "in"),
    _tx("02/05/2026", "SVTAX26004032", "QR / SCB X7027 ภาณุวัฒน์", 2439.17, "in"),
    _tx("02/05/2026", "SVTAX26004033", "QR / KBANK X8949 จาตุรนต์", 4373.40, "in"),
    _tx("02/05/2026", "SVTAX26004039", "QR / SCB X3307 ประภาศิริ", 2975.19, "in"),
    _tx("02/05/2026", "SVTAX26004046", "QR / SCB X8706 วราห์", 374.50, "in"),
    _tx("02/05/2026", "SVTAX26004048", "QR / SCB X2423 เสกสรร", 5656.45, "in"),
    _tx("02/05/2026", "FEE-SIM", "ค่า SIM เครื่องรูดการ์ด", 107.00, "out"),
]


def _clone(txns):
    return [dict(t) for t in txns]


def _stmt_xlsx(txns, opening=STMT_OPEN):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pearnly-STATEMENT"
    ws.append(["วันที่", "รายละเอียด", "ถอน", "ฝาก", "ยอดคงเหลือ"])
    ws.append(["ยอดยกมา", None, None, None, round(opening, 2)])
    bal = opening
    for t in txns:
        amt = t["amount"]
        if t["kind"] == "in":
            bal = round(bal + amt, 2)
            ws.append([t["date"], t["desc"], None, amt, bal])
        else:
            bal = round(bal - amt, 2)
            ws.append([t["date"], t["desc"], amt, None, bal])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _gl_xlsx(txns, opening=GL_OPEN):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Pearnly-GL"
    ws.append(
        ["วันที่", "เลขที่ใบสำคัญ", "รายละเอียด", "เดบิต=ฝากเงิน", "เครดิต=ถอนเงิน", "คงเหลือ"]
    )
    ws.append(["ยอดยกมา", None, None, None, None, round(opening, 2)])
    bal = opening
    for t in txns:
        amt = t["amount"]
        kind = t.get("gl_kind", t["kind"])  # 故意写反方向的场景用 gl_kind 覆盖
        if kind == "in":
            bal = round(bal + amt, 2)
            ws.append([t["date"], t["voucher"], t["desc"], amt, None, bal])
        else:
            bal = round(bal - amt, 2)
            ws.append([t["date"], t["voucher"], t["desc"], None, amt, bal])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _reconcile(stmt_txns, gl_txns):
    sres = parse_bank_stmt_xlsx_direct(_stmt_xlsx(stmt_txns), "stm.xlsx")
    gres = parse_gl_excel(_gl_xlsx(gl_txns), "gl.xlsx")
    assert sres.get("ok"), sres.get("error_code")
    assert gres.get("ok"), gres.get("error_code")
    _rows, summary = reconcile(
        sres["rows"],
        gres["rows"],
        stmt_opening=sres.get("opening", 0.0),
        gl_opening=gres.get("opening", 0.0),
    )
    return summary


class FixedTemplateRecognitionTests(unittest.TestCase):
    def test_generated_statement_has_opening_and_balance(self):
        wb = openpyxl.load_workbook(io.BytesIO(rt.generate_template("statement", "th")))
        ws = wb["Pearnly-STATEMENT"]
        self.assertEqual(ws.cell(1, 5).value, "ยอดคงเหลือ")  # 余额列在
        self.assertEqual(ws.cell(2, 1).value, "ยอดยกมา")  # 期初行在日期列

    def test_generated_gl_is_six_cols_no_account_with_annotation(self):
        wb = openpyxl.load_workbook(io.BytesIO(rt.generate_template("gl", "th")))
        ws = wb["Pearnly-GL"]
        headers = [ws.cell(1, c).value for c in range(1, 8)]
        self.assertEqual(headers[3], "เดบิต=ฝากเงิน")  # 借方带标注
        self.assertEqual(headers[4], "เครดิต=ถอนเงิน")  # 贷方带标注
        self.assertEqual(headers[5], "คงเหลือ")  # 第 6 列 = 余额
        self.assertIsNone(headers[6])  # 没有第 7 列(科目码已去除)
        self.assertNotIn("รหัสบัญชี", [h for h in headers if h])

    def test_four_lang_opening_roundtrip(self):
        # 四语模板自产 → 期初(放日期列)都能被识别读出 100(zh/ja 此前读 0 的回归)
        for lang in ("zh", "en", "th", "ja"):
            g = parse_gl_excel(rt.generate_template("gl", lang), "t.xlsx")
            s = parse_bank_stmt_xlsx_direct(rt.generate_template("statement", lang), "t.xlsx")
            self.assertEqual(g.get("opening"), 100.0, f"GL opening {lang}")
            self.assertEqual(s.get("opening"), 100.0, f"STMT opening {lang}")

    def test_four_lang_amount_columns_parsed(self):
        # 四语列头的存/取(statement)、借/贷(GL)金额列都要真解析出来,
        # 不能像 zh『存入』那样静默丢一列(对的被说错)· 锁死该回归。
        def _hdr(doc, lang):
            return [c[1].get(lang) or c[1]["en"] for c in rt._COLS[doc]]

        for lang in ("zh", "en", "th", "ja"):
            olabel = rt._OPENING_LABEL[lang]
            swb = openpyxl.Workbook()
            sws = swb.active
            sws.title = "Pearnly-STATEMENT"
            sws.append(_hdr("statement", lang))
            sws.append([olabel, None, None, None, 100])
            sws.append(["02/05/2026", "in", None, 1727.46, 1827.46])
            sws.append(["02/05/2026", "fee", 107.0, None, 1720.46])
            sbuf = io.BytesIO()
            swb.save(sbuf)
            sr = parse_bank_stmt_xlsx_direct(sbuf.getvalue(), "t.xlsx")
            self.assertAlmostEqual(sum(r.deposit for r in sr["rows"]), 1727.46, msg=f"dep {lang}")
            self.assertAlmostEqual(sum(r.withdrawal for r in sr["rows"]), 107.0, msg=f"wd {lang}")

            gwb = openpyxl.Workbook()
            gws = gwb.active
            gws.title = "Pearnly-GL"
            gws.append(_hdr("gl", lang))
            gws.append([olabel, None, None, None, None, 0])
            gws.append(["02/05/2026", "V1", "in", 1727.46, None, 1727.46])
            gws.append(["02/05/2026", "V2", "fee", None, 107.0, 1620.46])
            gbuf = io.BytesIO()
            gwb.save(gbuf)
            gr = parse_gl_excel(gbuf.getvalue(), "t.xlsx")
            self.assertAlmostEqual(sum(r.debit for r in gr["rows"]), 1727.46, msg=f"debit {lang}")
            self.assertAlmostEqual(sum(r.credit for r in gr["rows"]), 107.0, msg=f"credit {lang}")


class ReconcileScenarioTests(unittest.TestCase):
    """对的判对、错的判错(不漏配/不误配)。"""

    def test_s1_all_match(self):
        s = _reconcile(BASE, BASE)
        self.assertEqual(s.matched_count, 9)
        self.assertEqual(s.stmt_deposit_only_count, 0)
        self.assertEqual(s.gl_debit_only_count, 0)
        self.assertEqual(s.gl_credit_only_count, 0)

    def test_s2_gl_missing_one_flags_stmt_orphan(self):
        s = _reconcile(BASE, _clone(BASE[:-2] + BASE[-1:]))  # 去掉第 8 笔入账
        self.assertEqual(s.matched_count, 8)
        self.assertEqual(s.stmt_deposit_only_count, 1)
        self.assertEqual(s.gl_debit_only_count, 0)

    def test_s3_gl_extra_one_flags_gl_orphan(self):
        extra = _tx("02/05/2026", "SVTAX26004099", "QR / SCB X9999 ผี", 9999.99, "in")
        s = _reconcile(BASE, _clone(BASE) + [extra])
        self.assertEqual(s.matched_count, 9)
        self.assertEqual(s.gl_debit_only_count, 1)
        self.assertEqual(s.stmt_deposit_only_count, 0)

    def test_s4_amount_typo_not_falsely_matched(self):
        gl = _clone(BASE)
        gl[0]["amount"] = 1728.46  # 1727.46 → 1728.46(差 1 元)
        s = _reconcile(BASE, gl)
        self.assertEqual(s.matched_count, 8)
        self.assertEqual(s.stmt_deposit_only_count, 1)
        self.assertEqual(s.gl_debit_only_count, 1)

    def test_s5_wrong_direction_not_falsely_matched(self):
        gl = _clone(BASE)
        gl[0]["gl_kind"] = "out"  # 入账被 GL 记成贷方(取出)
        s = _reconcile(BASE, gl)
        self.assertEqual(s.matched_count, 8)
        self.assertEqual(s.stmt_deposit_only_count, 1)
        self.assertEqual(s.gl_credit_only_count, 1)

    def test_s6_date_within_tolerance_still_matched(self):
        gl = _clone(BASE)
        gl[0]["date"] = "04/05/2026"  # +2 天 · 容差内
        s = _reconcile(BASE, gl)
        self.assertEqual(s.matched_count, 9)
        self.assertEqual(s.stmt_deposit_only_count, 0)
        self.assertEqual(s.gl_debit_only_count, 0)

    def test_s7_duplicate_amount_paired_no_double_count(self):
        dup = _tx("02/05/2026", "SVTAX26004060", "QR / SCB X1111 ซ้ำ", 1727.46, "in")
        s = _reconcile(_clone(BASE) + [dup], _clone(BASE) + [dup])
        self.assertEqual(s.matched_count, 10)
        self.assertEqual(s.stmt_deposit_only_count, 0)
        self.assertEqual(s.gl_debit_only_count, 0)


if __name__ == "__main__":
    unittest.main()
