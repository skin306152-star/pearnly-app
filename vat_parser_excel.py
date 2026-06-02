# -*- coding: utf-8 -*-
"""VAT 报告解析 · Excel(.xlsx/.xls)路径(零成本 openpyxl 表头识别)。"""

import io
import logging
from typing import List, Dict, Any

from field_comparator import normalize_tax_id, normalize_branch, parse_date

from vat_parser_common import (
    _hit,
    _to_float,
    PARSER_VERSION,
    _DATE_H,
    _NO_H,
    _REF_H,
    _NAME_H,
    _TAX_H,
    _BRANCH_H,
    _NET_H,
    _VAT_H,
    _TOTAL_H,
    _SKIP_H,
)

logger = logging.getLogger(__name__)


# ======================================================================
# 1. Excel 解析(零成本)
# ======================================================================


def parse_excel(file_bytes: bytes) -> Dict[str, Any]:
    try:
        import openpyxl
    except ImportError:
        return {"ok": False, "error": "openpyxl 未安装", "rows": []}

    rows: List[Dict] = []
    meta: Dict[str, Any] = {}
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        ws = wb.active

        header_idx = None
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), 1):
            text = " ".join(str(c or "").lower() for c in row)
            if "เลขที่" in text or "invoice no" in text or "วันที่" in text:
                header_idx = i
                break
        if header_idx is None:
            return {"ok": False, "error": "找不到表头行", "rows": []}

        headers = tuple(str(c.value or "") for c in ws[header_idx])
        col_map = _map_columns(headers)

        row_no = 0
        for raw in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            cells = list(raw)
            if not any(c for c in cells if c not in (None, "")):
                continue
            first = str(cells[0] or "").strip().lower()
            if any(k in first for k in _SKIP_H):
                if "amount_pre_vat" in col_map:
                    meta["total_amount_pre_vat"] = _to_float(cells[col_map["amount_pre_vat"]])
                if "vat_amount" in col_map:
                    meta["total_vat"] = _to_float(cells[col_map["vat_amount"]])
                continue
            row_no += 1
            rows.append(_build_row(row_no, cells, col_map))

        return {
            "ok": True,
            "rows": rows,
            "meta": meta,
            "warnings": [],
            "parser_version": PARSER_VERSION,
            "row_count": len(rows),
            "method": "excel",
        }
    except Exception as e:
        logger.error(f"parse_excel failed: {e}")
        return {"ok": False, "error": str(e), "rows": []}


def _map_columns(headers) -> Dict[str, int]:
    col_map: Dict[str, int] = {}
    for i, h in enumerate(headers):
        h = (str(h) or "").strip()
        if not h:
            continue
        if _hit(h, _DATE_H) and "date" not in col_map:
            col_map["date"] = i
        # v118.32.5 · 参考单据号优先检测（关键词更specific，避免被 _NO_H 抢走）
        elif _hit(h, _REF_H) and "ref_no" not in col_map:
            col_map["ref_no"] = i
        elif _hit(h, _NO_H) and "invoice_no" not in col_map:
            col_map["invoice_no"] = i
        elif _hit(h, _NAME_H) and "buyer_name" not in col_map:
            col_map["buyer_name"] = i
        elif _hit(h, _TAX_H) and "buyer_tax_id" not in col_map:
            col_map["buyer_tax_id"] = i
        elif _hit(h, _BRANCH_H) and "buyer_branch" not in col_map:
            col_map["buyer_branch"] = i
        elif _hit(h, _NET_H) and "amount_pre_vat" not in col_map:
            col_map["amount_pre_vat"] = i
        elif _hit(h, _VAT_H) and "vat_amount" not in col_map:
            col_map["vat_amount"] = i
        elif _hit(h, _TOTAL_H) and "total_amount" not in col_map:
            col_map["total_amount"] = i
    return col_map


def _build_row(row_no: int, cells: list, col_map: Dict[str, int]) -> Dict[str, Any]:
    parsed: Dict[str, Any] = {"row_no": row_no}
    for field, ci in col_map.items():
        raw = cells[ci] if ci < len(cells) else None
        val = str(raw).strip() if raw is not None else ""
        if field == "date":
            d = parse_date(val)
            parsed["report_date"] = d.isoformat() if d else val
        elif field == "invoice_no":
            parsed["report_invoice_no"] = val
        elif field == "ref_no":
            parsed["report_ref_no"] = val  # v118.32.5 · GL对账匹配键
        elif field == "buyer_name":
            parsed["report_buyer_name"] = val
        elif field == "buyer_tax_id":
            parsed["report_buyer_tax_id"] = normalize_tax_id(val)
        elif field == "buyer_branch":
            parsed["report_buyer_branch"] = normalize_branch(val)
        elif field == "amount_pre_vat":
            parsed["report_amount_pre_vat"] = _to_float(val)
        elif field == "vat_amount":
            parsed["report_vat_amount"] = _to_float(val)
        elif field == "total_amount":
            parsed["report_amount"] = _to_float(val)
    parsed["is_individual"] = not bool(parsed.get("report_buyer_tax_id"))
    return parsed
