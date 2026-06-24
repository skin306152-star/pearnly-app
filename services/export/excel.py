# -*- coding: utf-8 -*-
"""导出行 → xlsx 内存流(openpyxl · 零授权兜底 · 契约 04 §七 / 05 §2.2)。

Excel 是外流最稳的零依赖兜底:不需 Google 授权、本地直接下载。列序复用 rows.COLUMNS
(与 Sheet 同一真源)。证据列若是 URL 渲成超链。Decimal → float 写数值格(可在 Excel 求和)。
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font

from services.export import rows as rows_mod
from services.export.rows import COLUMN_KEYS

_HEADER_FONT = Font(bold=True)
_LINK_FONT = Font(color="0563C1", underline="single")


def _cell_value(v):
    """Decimal → float(Excel 可求和);None → "";其余原样。"""
    if isinstance(v, Decimal):
        return float(v)
    if v is None:
        return ""
    return v


def build_workbook(rows: list, *, sheet_title: Optional[str] = None, lang: str = "th") -> bytes:
    """一行一明细导出行 → xlsx 字节流。表头/证据文案随 lang;evidence 是 URL 渲超链。"""
    wb = Workbook()
    ws = wb.active
    # openpyxl sheet 标题上限 31 字符且禁 []:*?/\
    ws.title = (sheet_title or rows_mod.sheet_title(lang) or "Sheet1")[:31]

    keys = COLUMN_KEYS
    ws.append(rows_mod.headers(lang))
    for cell in ws[1]:
        cell.font = _HEADER_FONT

    evidence_col = keys.index("evidence") + 1  # 1-based
    for row in rows or []:
        ws.append([_cell_value(row.get(k)) for k in keys])
        r = ws.max_row
        ev = row.get("evidence") or ""
        if isinstance(ev, str) and ev.startswith(("http://", "https://", "/api/")):
            cell = ws.cell(row=r, column=evidence_col)
            cell.hyperlink = ev
            cell.value = rows_mod.view_label(lang)
            cell.font = _LINK_FONT

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
