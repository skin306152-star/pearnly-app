# -*- coding: utf-8 -*-
"""
BUG-FIX-T1 v118.35.0.42 · 守门测试 · _parse_date 佛历 BE → 公历 CE 转换 + datetime 直通

紧急 BUG hotfix(铁律 #18 整顿期破例)· 触发:付费用户上传 GL Excel 报『0 แถว』
docs/audits/2026-05-23-bug-fix-t1-gl-excel-be-year.md(根因分析)

锁定 8 个契约 + 真用户文件场景复现:
  1-4. _parse_date 4 种输入类型都正确转 CE(datetime BE / date BE / ISO str BE / yyyy-mm-dd BE)
  5. ISO 字符串带时分秒(`2568-12-31 00:00:00`)正确去掉时分秒
  6. CE 字符串原样通过(不误改非佛历日期)
  7. dd/mm/yy 短年仍走老路径(BE 67/68/69 → CE 2024/2025/2026)
  8. parse_gl_excel 跑 in-memory 构造的客户同款 xlsx · 至少 1 row · date 是公历
"""

import io
import unittest
from datetime import date, datetime

import openpyxl

from services.recon import bank_recon_v2 as brv2


class ParseDateBuddhistEraTests(unittest.TestCase):

    def test_datetime_be_year_converts_to_ce(self):
        """BE 2568 datetime → CE 2025 date"""
        self.assertEqual(brv2._parse_date(datetime(2568, 12, 31)), date(2025, 12, 31))
        self.assertEqual(brv2._parse_date(datetime(2569, 1, 1)), date(2026, 1, 1))

    def test_date_be_year_converts_to_ce(self):
        """BE date → CE date"""
        self.assertEqual(brv2._parse_date(date(2568, 1, 15)), date(2025, 1, 15))

    def test_iso_string_be_with_time_strips_time_and_converts(self):
        """'2568-12-31 00:00:00' → CE 2025-12-31 · ISO T 同款"""
        self.assertEqual(brv2._parse_date("2568-12-31 00:00:00"), date(2025, 12, 31))
        self.assertEqual(brv2._parse_date("2568-12-31T00:00:00"), date(2025, 12, 31))
        self.assertEqual(brv2._parse_date("2568-12-31T00:00:00.123"), date(2025, 12, 31))

    def test_yyyy_mm_dd_be_string_converts(self):
        """'2568-12-31' yyyy-mm-dd 路径 BE → CE"""
        self.assertEqual(brv2._parse_date("2568-12-31"), date(2025, 12, 31))

    def test_ce_date_passes_through(self):
        """公历 2025-12-31 原样通过 · 不被误改"""
        self.assertEqual(brv2._parse_date("2025-12-31"), date(2025, 12, 31))
        self.assertEqual(brv2._parse_date(date(2025, 12, 31)), date(2025, 12, 31))

    def test_dd_mm_yy_short_thai_year_still_works(self):
        """dd/mm/yy 短佛历年 67/68/69 → CE 2024/2025/2026(老逻辑保留)"""
        self.assertEqual(brv2._parse_date("31/12/68"), date(2025, 12, 31))
        self.assertEqual(brv2._parse_date("01/01/69"), date(2026, 1, 1))

    def test_garbage_and_none_return_none(self):
        """None / 垃圾字符串 → None(不爆 · 不返默认值)"""
        self.assertIsNone(brv2._parse_date(None))
        self.assertIsNone(brv2._parse_date(""))
        self.assertIsNone(brv2._parse_date("244319"))
        self.assertIsNone(brv2._parse_date("not a date"))

    def test_parse_gl_excel_handles_buddhist_datetime_cells(self):
        """BUG-FIX-T1 复现 · in-memory 构造客户同款 xlsx(BE datetime cells)· 验证至少 1 row + CE date

        客户文件结构(已隐私化):
        - Row 1 header (date/doc_no/account/desc/debit/credit/balance/source 8 列泰文)
        - Row 2 期初 ยอดยกมา (BE datetime · 余额列有值)
        - Row 3 1 笔交易 (BE datetime · debit 列有值)

        修法前:datetime cell str() 化 '2568-12-31 00:00:00' · _parse_date split 4 parts 拒绝 · 0 rows
        修法后:datetime 直通 + ISO 字符串 fallback · 至少 1 row + date 是 CE 公历
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header (Thai)
        ws.append(
            [
                "วันที่",
                "เลขที่เอกสาร",
                "รหัสบัญชี",
                "รายการ",
                "เดบิต",
                "เครดิต",
                "คงเหลือ",
                "ไฟล์ต้นทาง",
            ]
        )
        # Row 2 · 期初(BE datetime + 余额列填)
        ws.append([datetime(2568, 12, 1), None, None, "ยอดยกมา", None, None, 1000.00, None])
        # Row 3 · 1 笔(BE datetime + 借方填 · 余额公式)
        ws.append(
            [datetime(2568, 12, 31), "TST001", "1112-10", "anon test", 500.00, 0, 1500.00, None]
        )

        buf = io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()

        result = brv2.parse_gl_excel(data, "be_test.xlsx", "")
        self.assertTrue(result.get("ok"), f"parse failed: {result.get('error')}")
        self.assertGreaterEqual(
            result.get("row_count", 0), 1, "至少 1 row 应被识别(BUG-FIX-T1 主修)"
        )
        rows = result.get("rows") or []
        self.assertTrue(rows)
        # date 必须是 CE 公历(2025-12-31)· 不是 BE (2568-12-31)
        self.assertEqual(
            rows[0].date,
            date(2025, 12, 31),
            f"date 应转 CE 公历 · 实际 {rows[0].date}(若为 2568-* 则 BE → CE 转换没起效)",
        )
        self.assertEqual(rows[0].account_code, "1112-10")
        self.assertEqual(rows[0].debit, 500.00)

    def test_parse_gl_excel_reads_opening_and_closing_from_balance_column(self):
        """BUG-FIX-T2 v118.35.0.43 · 期初余额读 balance 列 + 期末读最后一笔 balance 列

        老逻辑:opening 检测只读 debit/credit · Row 2 期初 ยอดยกมา 借贷列空 → opening=0
        老逻辑:closing = opening + credit - debit · 对资产科目方向反 → closing 错
        修法:_map_gl_cols 识别 คงเหลือ/balance · opening 优先读 balance · closing 优先读最后一笔 balance
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(
            [
                "วันที่",
                "เลขที่เอกสาร",
                "รหัสบัญชี",
                "รายการ",
                "เดบิต",
                "เครดิต",
                "คงเหลือ",
                "ไฟล์ต้นทาง",
            ]
        )
        ws.append([datetime(2568, 12, 1), None, None, "ยอดยกมา", None, None, 39749.85, None])
        ws.append([datetime(2568, 12, 31), "JV001", "1112-10", "anon", 24472.92, 0, 64222.77, None])
        buf = io.BytesIO()
        wb.save(buf)

        result = brv2.parse_gl_excel(buf.getvalue(), "be_test_with_bal.xlsx", "")
        self.assertTrue(result.get("ok"))
        # opening · 从 balance 列 Row 2 读到(不是 0)
        self.assertEqual(
            result.get("opening"),
            39749.85,
            f"opening 应从 balance 列读 39749.85 · 实际 {result.get('opening')}",
        )
        # closing · 从最后一笔 balance 读到(不走公式)
        self.assertEqual(
            result.get("closing"),
            64222.77,
            f"closing 应从最后一笔 balance 读 64222.77 · 实际 {result.get('closing')}",
        )

    def test_parse_gl_excel_no_balance_column_falls_back_to_formula(self):
        """BUG-FIX-T2 regression · 老格式 GL(没 balance 列)走老公式不变"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Date", "Doc", "Acct", "Desc", "Debit", "Credit"])
        ws.append(["2025-01-15", "V001", "4001", "sale1", 0, 1000])
        ws.append(["2025-01-20", "V002", "4001", "sale2", 0, 2000])
        buf = io.BytesIO()
        wb.save(buf)

        result = brv2.parse_gl_excel(buf.getvalue(), "old_no_bal.xlsx", "")
        self.assertTrue(result.get("ok"))
        self.assertEqual(result.get("row_count"), 2)
        # 老公式: 0 + 3000 - 0 = 3000
        self.assertEqual(result.get("closing"), 3000.0)


class M3GlVatBuddhistEraTests(unittest.TestCase):
    """BUG-FIX-T5 v118.35.0.46 · M3 (收入对账) parse_gl_excel 也加 datetime + BE→CE 处理
    根因:M3 用独立的 parse_gl_excel(gl_vat_reconciler.py)· _get('date') 直接 str(datetime)
          → "2568-12-31 00:00:00" garbage 字符串显示 + Excel/历史详情看着别扭
    修法:_get 加 isinstance(v, datetime) 处理 + BE→CE(同 M4 T1 修法)
    影响:M3 GlRow.date 字段从 garbage 变干净 ISO date 字符串
    业务影响:M3 用 doc_no 匹配 · date 仅 sort key · 修前不影响匹配 · 修后显示干净
    """

    def test_m3_parse_gl_excel_datetime_be_year_converts(self):
        """M3 parse_gl_excel 接收 BE datetime cell → 输出 CE ISO date 字符串"""
        from services.recon import gl_vat_reconciler as glvr

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["วันที่", "ใบสำคัญ", "รหัสบัญชี", "รายการ", "เดบิต", "เครดิต"])
        ws.append([datetime(2568, 12, 31), "JV001", "4110-01", "sale1", 0, 1000])
        ws.append([datetime(2569, 1, 15), "JV002", "4110-01", "sale2", 0, 2000])
        buf = io.BytesIO()
        wb.save(buf)

        result = glvr.parse_gl_excel(buf.getvalue())
        self.assertTrue(result.get("ok"))
        self.assertEqual(result.get("row_count"), 2)
        rows = result.get("rows") or []
        # date 必须是 CE ISO 字符串 · 不是 "2568-12-31 00:00:00" garbage
        self.assertEqual(
            rows[0].date, "2025-12-31", f"M3 BE 2568 应转 CE 2025 · 实际 {rows[0].date!r}"
        )
        self.assertEqual(
            rows[1].date, "2026-01-15", f"M3 BE 2569 应转 CE 2026 · 实际 {rows[1].date!r}"
        )

    def test_m3_parse_gl_excel_string_date_passes_through(self):
        """M3 字符串 date 原样通过(不破坏老格式 GL · 用 M3 真实 header 关键词)"""
        from services.recon import gl_vat_reconciler as glvr

        wb = openpyxl.Workbook()
        ws = wb.active
        # 用 M3 _GL_*_H 字典里真实关键词: voucher / gl account / debit / credit
        ws.append(["Date", "Voucher", "GL Account", "Description", "Debit", "Credit"])
        ws.append(["2025-01-15", "V001", "4110-01", "sale1", 0, 1000])
        buf = io.BytesIO()
        wb.save(buf)

        result = glvr.parse_gl_excel(buf.getvalue())
        self.assertTrue(result.get("ok"))
        rows = result.get("rows") or []
        self.assertGreaterEqual(len(rows), 1, f"应识别 ≥1 row · 实际 {len(rows)}")
        self.assertEqual(rows[0].date, "2025-01-15", "M3 string date 应原样通过 · 不被改")


if __name__ == "__main__":
    unittest.main()
