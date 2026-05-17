# -*- coding: utf-8 -*-
"""
Mr.Pilot · 统一报表引擎 v109.0
==============================
职责:
  - 4 个内置模板(input_vat / standard / erp / print)
  - 统一 build_report(template_code, rows, meta, lang) → bytes
  - 统一 list_templates(lang) → 给前端弹窗
  - 多 Sheet 输出(主明细 + 汇总分析)
  - 4 语言跟随 UI
  - 专业样式(深蓝表头 / 斑马行 / 千位分隔 / SUM 公式 / 冻结首行)

老 excel_export.build_xlsx 保留兼容 · 不删。
"""

import io
import re
from datetime import datetime
from typing import List, Dict, Any, Optional
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from i18n_reports import i18n_get, i18n_format


# ============================================================
# 常量样式
# ============================================================
COLOR_HEADER_BG = "1E3A8A"      # 深蓝表头(更专业 · 比老 #2C5282 深一档)
COLOR_HEADER_FG = "FFFFFF"
COLOR_ZEBRA = "F9FAFB"          # 斑马行灰
COLOR_BORDER = "D1D5DB"
COLOR_TOTAL_BG = "DBEAFE"       # 合计行浅蓝
COLOR_TITLE_FG = "111827"       # 顶部大标题深灰
COLOR_INFO_FG = "374151"        # 信息块灰
COLOR_WARN_AMOUNT = "DC2626"    # 大额警告红(>50 万)

# 字体优先级:Tahoma 泰文支持好;退回 Calibri / Arial
FONT_NAME = "Tahoma"
FONT_NAME_NUM = "Calibri"

THIN = Side(style="thin", color=COLOR_BORDER)
THICK = Side(style="medium", color="6B7280")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOP_THICK = Border(left=THIN, right=THIN, top=THICK, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor=COLOR_HEADER_BG)
ZEBRA_FILL = PatternFill("solid", fgColor=COLOR_ZEBRA)
TOTAL_FILL = PatternFill("solid", fgColor=COLOR_TOTAL_BG)

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_HEADER_FG)
CELL_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT = Font(name=FONT_NAME, size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_TITLE_FG)
INFO_FONT = Font(name=FONT_NAME, size=10, color=COLOR_INFO_FG)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)


# ============================================================
# 模板定义
# ============================================================
REPORT_TEMPLATES = {
    # ----- 模板 1 · ภ.พ.30 进项税明细(法定 · 默认) -----
    "input_vat": {
        "code": "input_vat",
        "name_key": "tpl-input-vat",
        "desc_key": "tpl-input-vat-desc",
        "category_key": "tpl-cat-tax",
        "title_key": "tpl-input-vat-title",
        "recommended": True,
        "show_info_block": True,        # 顶部信息块(公司名 / 税号 / 月份 / 共 X 张)
        "show_summary_sheet": True,     # 第二 Sheet 汇总
        "show_signature_block": True,   # 底部签名栏
        "freeze_panes": "A4",           # 冻结到信息块下面
        "columns": [
            {"key": "no",                "header_key": "col-no",                "width":  6, "align": "center", "type": "int"},
            {"key": "invoice_date",      "header_key": "col-invoice-date",      "width": 12, "align": "center", "type": "date"},
            {"key": "invoice_no",        "header_key": "col-invoice-no",        "width": 18, "align": "left",   "type": "text"},
            {"key": "seller_name",       "header_key": "col-seller-name",       "width": 32, "align": "left",   "type": "text"},
            {"key": "seller_tax_id",     "header_key": "col-seller-tax-id",     "width": 18, "align": "center", "type": "tax_id"},
            {"key": "seller_branch",     "header_key": "col-seller-branch",     "width": 14, "align": "center", "type": "branch"},
            {"key": "amount_before_vat", "header_key": "col-amount-before-vat", "width": 14, "align": "right",  "type": "money"},
            {"key": "vat_amount",        "header_key": "col-vat-amount",        "width": 12, "align": "right",  "type": "money"},
            {"key": "total_amount",      "header_key": "col-total-amount",      "width": 14, "align": "right",  "type": "money"},
            {"key": "category",          "header_key": "col-category",          "width": 12, "align": "center", "type": "text"},
            {"key": "source",            "header_key": "col-source",            "width": 10, "align": "center", "type": "source"},
        ],
        "total_columns": ["amount_before_vat", "vat_amount", "total_amount"],
    },

    # ----- 模板 2 · 标准明细(老板 / 内部审核) -----
    "standard": {
        "code": "standard",
        "name_key": "tpl-standard",
        "desc_key": "tpl-standard-desc",
        "category_key": "tpl-cat-internal",
        "title_key": "tpl-standard-title",
        "recommended": False,
        "show_info_block": True,
        "show_summary_sheet": True,
        "show_signature_block": False,
        "freeze_panes": "A4",
        "auto_filter": True,            # 列首加 Excel 自动筛选
        "warn_amount_threshold": 500000, # 大额警告阈值
        "columns": [
            {"key": "no",                "header_key": "col-no",                "width":  6, "align": "center", "type": "int"},
            {"key": "invoice_date",      "header_key": "col-invoice-date",      "width": 12, "align": "center", "type": "date"},
            {"key": "invoice_no",        "header_key": "col-invoice-no",        "width": 18, "align": "left",   "type": "text"},
            {"key": "filename",          "header_key": "col-filename",          "width": 22, "align": "left",   "type": "text"},
            {"key": "seller_name",       "header_key": "col-seller-name",       "width": 30, "align": "left",   "type": "text"},
            {"key": "seller_tax_id",     "header_key": "col-seller-tax-id",     "width": 18, "align": "center", "type": "tax_id"},
            {"key": "amount_before_vat", "header_key": "col-amount-before-vat", "width": 14, "align": "right",  "type": "money"},
            {"key": "vat_amount",        "header_key": "col-vat-amount",        "width": 12, "align": "right",  "type": "money"},
            {"key": "total_amount",      "header_key": "col-total-amount",      "width": 14, "align": "right",  "type": "money"},
            {"key": "category",          "header_key": "col-category",          "width": 12, "align": "center", "type": "text"},
            {"key": "source",            "header_key": "col-source",            "width": 10, "align": "center", "type": "source"},
            {"key": "notes",             "header_key": "col-notes",             "width": 20, "align": "left",   "type": "text"},
        ],
        "total_columns": ["amount_before_vat", "vat_amount", "total_amount"],
    },

    # ----- 模板 3 · ERP 录入格式(数据导向) -----
    "erp": {
        "code": "erp",
        "name_key": "tpl-erp",
        "desc_key": "tpl-erp-desc",
        "category_key": "tpl-cat-erp",
        "title_key": "tpl-erp-title",
        "recommended": False,
        "show_info_block": False,       # ERP 不要装饰
        "show_summary_sheet": False,
        "show_signature_block": False,
        "freeze_panes": "A2",
        "ascii_only": True,             # 列名严格 ASCII
        "iso_date": True,               # 日期 ISO 格式
        "no_thousand_sep": True,        # 金额不带千位分隔
        "first_row_marker": "<!-- ERP_IMPORT_FORMAT_v1 · Mr.Pilot -->",
        "columns": [
            {"key": "invoice_date",      "header_key": "col-invoice-date",      "header_ascii": "Date",        "width": 12, "align": "center", "type": "date_iso"},
            {"key": "invoice_no",        "header_key": "col-doc-no",            "header_ascii": "DocNo",       "width": 18, "align": "left",   "type": "text"},
            {"key": "seller_name",       "header_key": "col-vendor-name",       "header_ascii": "VendorName",  "width": 30, "align": "left",   "type": "text"},
            {"key": "seller_tax_id",     "header_key": "col-vendor-tax",        "header_ascii": "VendorTaxID", "width": 16, "align": "left",   "type": "text"},
            {"key": "seller_branch",     "header_key": "col-seller-branch",     "header_ascii": "Branch",      "width": 12, "align": "center", "type": "branch_code"},
            {"key": "amount_before_vat", "header_key": "col-net-amount",        "header_ascii": "NetAmount",   "width": 12, "align": "right",  "type": "money_raw"},
            {"key": "vat_amount",        "header_key": "col-vat-amount",        "header_ascii": "VAT",         "width": 12, "align": "right",  "type": "money_raw"},
            {"key": "total_amount",      "header_key": "col-total-amount",      "header_ascii": "TotalAmount", "width": 12, "align": "right",  "type": "money_raw"},
            {"key": "category",          "header_key": "col-category",          "header_ascii": "Category",    "width": 12, "align": "left",   "type": "text"},
        ],
        "total_columns": [],            # ERP 不要合计行
    },

    # ----- 模板 4 · 凭证装订清单(打印) -----
    "print": {
        "code": "print",
        "name_key": "tpl-print",
        "desc_key": "tpl-print-desc",
        "category_key": "tpl-cat-print",
        "title_key": "tpl-print-title",
        "recommended": False,
        "show_info_block": True,
        "show_summary_sheet": False,
        "show_signature_block": True,
        "freeze_panes": "A4",
        "row_height": 32,               # 行高大 · 签名空间
        "page_setup": "A4_landscape",
        "columns": [
            {"key": "no",                "header_key": "col-no",                "width":  6, "align": "center", "type": "int"},
            {"key": "invoice_date",      "header_key": "col-invoice-date",      "width": 14, "align": "center", "type": "date"},
            {"key": "invoice_no",        "header_key": "col-invoice-no",        "width": 22, "align": "left",   "type": "text"},
            {"key": "seller_name",       "header_key": "col-seller-name",       "width": 36, "align": "left",   "type": "text"},
            {"key": "total_amount",      "header_key": "col-total-amount",      "width": 16, "align": "right",  "type": "money"},
            {"key": "signature",         "header_key": "col-signature",         "width": 20, "align": "center", "type": "blank"},
        ],
        "total_columns": ["total_amount"],
    },
}


# ============================================================
# 工具函数
# ============================================================
def _to_float(v) -> float:
    """安全转 float"""
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(",", "").replace("฿", "").replace(" ", "").strip())
    except (ValueError, TypeError):
        return 0.0


def _format_tax_id(tax_id: str) -> str:
    """泰国税号 13 位 · 自动按 1-4-5-2-1 分组"""
    if not tax_id:
        return ""
    digits = re.sub(r"\D", "", str(tax_id))
    if len(digits) == 13:
        return f"{digits[0]}-{digits[1:5]}-{digits[5:10]}-{digits[10:12]}-{digits[12]}"
    return str(tax_id)


def _format_branch(branch: str, lang: str) -> str:
    """分公司:空 → 总公司;'00000' → 总公司"""
    if not branch or str(branch).strip() in ("", "0", "00000", "head", "head_office"):
        return i18n_get(lang, "info-head-office")
    s = str(branch).strip()
    # 已经是泰文 'สำนักงานใหญ่' / 'สาขา' 直接返回
    if "สำนักงาน" in s or "สาขา" in s:
        return s
    # 数字分公司编号
    if s.isdigit():
        return f"{i18n_get(lang, 'info-branch-office')} {int(s):05d}"
    return s


def _format_branch_code(branch: str) -> str:
    """ERP 用:输出 5 位代码"""
    if not branch or str(branch).strip() in ("", "head", "head_office"):
        return "00000"
    s = str(branch).strip()
    if "สำนักงานใหญ่" in s:
        return "00000"
    digits = re.sub(r"\D", "", s)
    if digits:
        return f"{int(digits):05d}"
    return "00000"


def _format_date(v, iso: bool = False) -> str:
    """日期归一化"""
    if not v:
        return ""
    s = str(v).strip()
    # 已 ISO
    m = re.match(r"^(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = m.groups()
        if iso:
            return f"{y}-{int(mo):02d}-{int(d):02d}"
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    # 兜底原样返回
    return s


def _source_label(src: str, lang: str) -> str:
    """来源标记 i18n"""
    if not src:
        return ""
    src_l = str(src).lower()
    mapping = {
        "email": "source-email", "imap": "source-email",
        "folder": "source-folder", "watch": "source-folder",
        "scan": "source-scan", "scanner": "source-scan",
        "upload": "source-upload",
        "manual": "source-manual",
        "api": "source-api",
        "line": "source-line",
    }
    key = mapping.get(src_l, None)
    return i18n_get(lang, key) if key else str(src)


# ============================================================
# 行数据归一化(从 ocr_history 或前端 records 都能用)
# ============================================================
def _normalize_row(rec: Dict[str, Any], idx: int) -> Dict[str, Any]:
    """
    把任意来源的发票数据 → 统一字段
    支持来源:
      - ocr_history 表行(seller_name / invoice_no / total_amount 等)
      - 前端 records(merged_fields 嵌套)
      - clients/{id}/export 后端拼装
    """
    # 双层尝试:rec 直接拿 + merged_fields / pages 兜底
    mf = rec.get("merged_fields") or {}
    # pages 第一页
    pages = rec.get("pages") or []
    p0 = pages[0] if isinstance(pages, list) and pages else {}
    if isinstance(p0, dict):
        pf = p0.get("fields") or p0
    else:
        pf = {}

    def pick(*keys):
        for k in keys:
            for source in (rec, mf, pf):
                if not isinstance(source, dict):
                    continue
                v = source.get(k)
                if v not in (None, "", []):
                    return v
        return ""

    subtotal = _to_float(pick("amount_before_vat", "subtotal", "amount_before_tax"))
    vat = _to_float(pick("vat_amount", "vat", "tax_amount"))
    total = _to_float(pick("total_amount", "total", "grand_total"))

    # VAT 逆算兜底:有 total 没 subtotal/vat → 按 7% 还原
    if total > 0 and subtotal == 0 and vat == 0:
        subtotal = round(total * 100 / 107, 2)
        vat = round(total - subtotal, 2)

    return {
        "no": idx,
        "filename":          pick("filename", "archive_name", "file_name") or "",
        "invoice_no":        pick("invoice_no", "invoice_number", "doc_no") or "",
        "invoice_date":      pick("invoice_date", "date") or "",
        "seller_name":       pick("seller_name", "vendor_name") or "",
        "seller_tax_id":     pick("seller_tax_id", "seller_tax", "vendor_tax_id") or "",
        "seller_branch":     pick("seller_branch", "branch") or "",
        "seller_addr":       pick("seller_addr", "seller_address") or "",
        "buyer_name":        pick("buyer_name") or "",
        "buyer_tax_id":      pick("buyer_tax_id", "buyer_tax") or "",
        "buyer_addr":        pick("buyer_addr", "buyer_address") or "",
        "amount_before_vat": subtotal,
        "vat_amount":        vat,
        "total_amount":      total,
        "wht_rate":          pick("wht_rate") or "",
        "wht_amount":        _to_float(pick("wht_amount")),
        "category":          pick("category", "category_tag") or "",
        "source":            pick("source") or "",
        "notes":             pick("notes", "remark") or "",
        "items":             pick("items") or [],
    }


# ============================================================
# 主渲染函数
# ============================================================
def _render_cell_value(value: Any, col_type: str, lang: str, no_thousand: bool = False) -> Any:
    """根据列类型加工单元格值"""
    if col_type == "int":
        try:
            return int(value)
        except (ValueError, TypeError):
            return value
    if col_type in ("money", "money_raw"):
        return _to_float(value)
    if col_type == "tax_id":
        return _format_tax_id(value) if value else ""
    if col_type == "branch":
        return _format_branch(value, lang)
    if col_type == "branch_code":
        return _format_branch_code(value)
    if col_type == "date":
        return _format_date(value, iso=False)
    if col_type == "date_iso":
        return _format_date(value, iso=True)
    if col_type == "source":
        return _source_label(value, lang)
    if col_type == "blank":
        return ""
    return value if value not in (None,) else ""


def _apply_cell_style(cell, col_def: dict, is_zebra: bool = False, warn_amount: bool = False, no_thousand: bool = False):
    """单元格样式"""
    cell.font = CELL_FONT
    cell.border = BORDER_ALL
    align = col_def.get("align", "left")
    if align == "center":
        cell.alignment = ALIGN_CENTER
    elif align == "right":
        cell.alignment = ALIGN_RIGHT
    else:
        cell.alignment = ALIGN_LEFT
    if is_zebra:
        cell.fill = ZEBRA_FILL
    col_type = col_def.get("type", "")
    if col_type == "money":
        cell.number_format = "#,##0.00" if not no_thousand else "0.00"
    elif col_type == "money_raw":
        cell.number_format = "0.00"
    if warn_amount:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WARN_AMOUNT)


def _write_info_block(ws: Worksheet, template: dict, meta: dict, lang: str, n_cols: int) -> int:
    """
    顶部信息块(标题 + 元信息)
    返回:数据起始行号(1-indexed)
    """
    if not template.get("show_info_block"):
        return 1

    # 第 1 行:大标题(合并)
    title = i18n_format(
        lang, template["title_key"],
        client=meta.get("client_name") or i18n_get(lang, "client-default"),
        month=meta.get("period_label") or i18n_get(lang, "month-all"),
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 第 2 行:元信息(税号 + 分公司 + 张数 + 生成时间)
    info_parts = []
    if meta.get("client_tax_id"):
        info_parts.append(f"{i18n_get(lang, 'info-tax-id')}: {_format_tax_id(meta['client_tax_id'])}")
    if meta.get("client_branch"):
        info_parts.append(f"{i18n_get(lang, 'info-branch')}: {_format_branch(meta['client_branch'], lang)}")
    if meta.get("doc_count"):
        info_parts.append(i18n_format(lang, "info-doc-count", n=meta["doc_count"]))
    info_parts.append(f"{i18n_get(lang, 'info-generated-at')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}")

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    cell = ws.cell(row=2, column=1, value="    ".join(info_parts))
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 第 3 行留空(分隔)
    ws.row_dimensions[3].height = 8

    return 4  # 表头从第 4 行开始


def _write_main_sheet(wb: Workbook, template: dict, rows: List[Dict], meta: dict, lang: str):
    """主明细 Sheet"""
    sheet_name = i18n_get(lang, template["name_key"])[:31]  # Excel sheet name max 31
    # 替换非法字符
    sheet_name = re.sub(r"[\\/?*\[\]:]", "·", sheet_name)
    ws = wb.active
    ws.title = sheet_name

    columns = template["columns"]
    n_cols = len(columns)
    no_thousand = template.get("no_thousand_sep", False)
    ascii_only = template.get("ascii_only", False)
    warn_threshold = template.get("warn_amount_threshold", 0)

    # ERP 模板:第一行写隐藏标记
    start_row = 1
    if template.get("first_row_marker"):
        ws.cell(row=1, column=1, value=template["first_row_marker"])
        ws.row_dimensions[1].hidden = True
        start_row = 2

    # 信息块(input_vat / standard / print 才有)
    if not ascii_only and template.get("show_info_block"):
        start_row = _write_info_block(ws, template, meta, lang, n_cols)

    header_row = start_row

    # 表头
    for col_idx, col in enumerate(columns, start=1):
        if ascii_only and col.get("header_ascii"):
            label = col["header_ascii"]
        else:
            label = i18n_get(lang, col["header_key"])
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]
    ws.row_dimensions[header_row].height = 28

    # 数据行
    data_start = header_row + 1
    for i, row_data in enumerate(rows):
        excel_row = data_start + i
        is_zebra = (i % 2 == 1)
        for col_idx, col in enumerate(columns, start=1):
            raw_value = row_data.get(col["key"], "")
            value = _render_cell_value(raw_value, col["type"], lang, no_thousand)
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            warn = (
                warn_threshold > 0
                and col["type"] == "money"
                and isinstance(value, (int, float))
                and value > warn_threshold
            )
            _apply_cell_style(cell, col, is_zebra=is_zebra, warn_amount=warn, no_thousand=no_thousand)
        if template.get("row_height"):
            ws.row_dimensions[excel_row].height = template["row_height"]

    # 合计行(SUM 公式)
    total_cols = template.get("total_columns", [])
    if total_cols and rows:
        total_row = data_start + len(rows)
        # 合计标签放在第 1 列
        total_label_col = 1
        # 找第一个数字列前面所有非数字列合并起来放标签
        first_money_idx = None
        for col_idx, col in enumerate(columns, start=1):
            if col["key"] in total_cols:
                first_money_idx = col_idx
                break

        if first_money_idx and first_money_idx > 1:
            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=first_money_idx - 1)
            cell = ws.cell(row=total_row, column=1, value=i18n_get(lang, "report-grand-total"))
            cell.font = TOTAL_FONT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = TOTAL_FILL
            cell.border = BORDER_TOP_THICK

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=total_row, column=col_idx)
            if col["key"] in total_cols:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})"
                cell.value = formula
                cell.number_format = "#,##0.00" if not no_thousand else "0.00"
                cell.alignment = ALIGN_RIGHT
            elif col_idx >= (first_money_idx or 0):
                # 数字列右侧空合计列
                pass
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER_TOP_THICK
        ws.row_dimensions[total_row].height = 26

    # 冻结
    if template.get("freeze_panes"):
        # freeze_panes 配置是相对值;实际取决于 header_row + 1
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # 自动筛选(标准模板)
    if template.get("auto_filter") and rows:
        last_col = get_column_letter(n_cols)
        last_row = data_start + len(rows) - 1
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    # 签名栏
    if template.get("show_signature_block"):
        sig_row = data_start + len(rows) + (2 if total_cols else 1) + 2
        labels = [
            i18n_get(lang, "sig-prepared-by"),
            i18n_get(lang, "sig-reviewed-by"),
            i18n_get(lang, "sig-approved-by"),
        ]
        col_step = max(1, n_cols // 3)
        for i, label in enumerate(labels):
            target_col = min(1 + i * col_step, n_cols - 1)
            cell = ws.cell(row=sig_row, column=target_col, value=f"{label}: ____________")
            cell.font = INFO_FONT
            cell.alignment = ALIGN_LEFT
        ws.row_dimensions[sig_row].height = 28

    # 打印设置(print 模板)
    if template.get("page_setup") == "A4_landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = f"{header_row}:{header_row}"  # 每页打印重复表头


def _write_summary_sheet(wb: Workbook, rows: List[Dict], lang: str):
    """第二 Sheet · 汇总分析(老板视角)"""
    ws = wb.create_sheet(title=i18n_get(lang, "summary-sheet-name")[:31])

    # ===== 区域 1 · 按卖方 TOP 10 =====
    by_seller = defaultdict(lambda: {"count": 0, "amount": 0.0})
    total_amount = 0.0
    for r in rows:
        name = (r.get("seller_name") or "—").strip() or "—"
        amt = _to_float(r.get("total_amount"))
        by_seller[name]["count"] += 1
        by_seller[name]["amount"] += amt
        total_amount += amt
    top_sellers = sorted(by_seller.items(), key=lambda x: x[1]["amount"], reverse=True)[:10]

    cur_row = 1
    # 标题
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-seller"))
    c.font = TITLE_FONT
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    headers = [
        i18n_get(lang, "summary-rank"),
        i18n_get(lang, "col-seller-name"),
        i18n_get(lang, "summary-count"),
        i18n_get(lang, "summary-amount"),
        i18n_get(lang, "summary-percent"),
    ]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=cur_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[cur_row].height = 24
    cur_row += 1

    for rank, (name, agg) in enumerate(top_sellers, start=1):
        pct = (agg["amount"] / total_amount * 100) if total_amount > 0 else 0
        cells = [
            (rank, "center", "int"),
            (name, "left", "text"),
            (agg["count"], "center", "int"),
            (round(agg["amount"], 2), "right", "money"),
            (f"{pct:.1f}%", "right", "text"),
        ]
        for ci, (v, align, t) in enumerate(cells, start=1):
            c = ws.cell(row=cur_row, column=ci, value=v)
            c.font = CELL_FONT
            c.border = BORDER_ALL
            c.alignment = ALIGN_CENTER if align == "center" else (ALIGN_RIGHT if align == "right" else ALIGN_LEFT)
            if t == "money":
                c.number_format = "#,##0.00"
        cur_row += 1

    cur_row += 2  # 空 2 行

    # ===== 区域 2 · 按类目 =====
    by_cat = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for r in rows:
        cat = (r.get("category") or "—").strip() or "—"
        by_cat[cat]["count"] += 1
        by_cat[cat]["amount"] += _to_float(r.get("total_amount"))

    if len(by_cat) > 1 or (by_cat and "—" not in by_cat):
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-category"))
        c.font = TITLE_FONT
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=cur_row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        for rank, (cat, agg) in enumerate(sorted(by_cat.items(), key=lambda x: x[1]["amount"], reverse=True), start=1):
            pct = (agg["amount"] / total_amount * 100) if total_amount > 0 else 0
            cells = [rank, cat, agg["count"], round(agg["amount"], 2), f"{pct:.1f}%"]
            aligns = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_RIGHT]
            for ci, (v, a) in enumerate(zip(cells, aligns), start=1):
                c = ws.cell(row=cur_row, column=ci, value=v)
                c.font = CELL_FONT
                c.border = BORDER_ALL
                c.alignment = a
                if ci == 4:
                    c.number_format = "#,##0.00"
            cur_row += 1

    cur_row += 2

    # ===== 区域 3 · 按来源(自动化覆盖率)=====
    by_src = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for r in rows:
        src = _source_label(r.get("source") or "—", lang) or "—"
        by_src[src]["count"] += 1
        by_src[src]["amount"] += _to_float(r.get("total_amount"))

    if by_src:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-source"))
        c.font = TITLE_FONT
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=cur_row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        total_count = sum(v["count"] for v in by_src.values())
        for rank, (src, agg) in enumerate(sorted(by_src.items(), key=lambda x: x[1]["count"], reverse=True), start=1):
            pct = (agg["count"] / total_count * 100) if total_count > 0 else 0
            cells = [rank, src, agg["count"], round(agg["amount"], 2), f"{pct:.1f}%"]
            aligns = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_RIGHT]
            for ci, (v, a) in enumerate(zip(cells, aligns), start=1):
                c = ws.cell(row=cur_row, column=ci, value=v)
                c.font = CELL_FONT
                c.border = BORDER_ALL
                c.alignment = a
                if ci == 4:
                    c.number_format = "#,##0.00"
            cur_row += 1

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12


# ============================================================
# 公开 API
# ============================================================
def list_templates(lang: str = "zh") -> List[Dict[str, Any]]:
    """前端模板选择弹窗用"""
    out = []
    for code, t in REPORT_TEMPLATES.items():
        out.append({
            "code": code,
            "name": i18n_get(lang, t["name_key"]),
            "desc": i18n_get(lang, t["desc_key"]),
            "category": i18n_get(lang, t["category_key"]),
            "category_code": t["category_key"].replace("tpl-cat-", ""),
            "recommended": bool(t.get("recommended")),
        })
    return out


def build_report(
    template_code: str,
    rows: List[Dict[str, Any]],
    meta: Optional[Dict[str, Any]] = None,
    lang: str = "zh",
) -> bytes:
    """
    统一报表生成器
    template_code: input_vat / standard / erp / print
    rows: 任意来源的发票数据(自动归一化)
    meta: {client_name, client_tax_id, client_branch, period_label, doc_count}
    lang: zh / th / en / ja
    返回:xlsx bytes
    """
    if template_code not in REPORT_TEMPLATES:
        raise ValueError(f"Unknown template: {template_code}")
    if lang not in ("zh", "th", "en", "ja"):
        lang = "zh"
    template = REPORT_TEMPLATES[template_code]
    meta = dict(meta or {})

    # 归一化行
    norm_rows = [_normalize_row(r, idx + 1) for idx, r in enumerate(rows or [])]
    if "doc_count" not in meta:
        meta["doc_count"] = len(norm_rows)

    wb = Workbook()
    _write_main_sheet(wb, template, norm_rows, meta, lang)

    if template.get("show_summary_sheet") and norm_rows:
        _write_summary_sheet(wb, norm_rows, lang)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def default_filename(template_code: str = "standard", client: str = "", period: str = "") -> str:
    """生成默认文件名"""
    parts = ["mrpilot"]
    if client:
        # 清理文件名非法字符
        safe_client = re.sub(r"[\\/:*?\"<>|]", "_", client.strip())[:40]
        if safe_client:
            parts.append(safe_client)
    if period:
        parts.append(period.replace("/", "-"))
    parts.append(template_code)
    parts.append(datetime.now().strftime("%Y%m%d-%H%M%S"))
    return "-".join(parts) + ".xlsx"
