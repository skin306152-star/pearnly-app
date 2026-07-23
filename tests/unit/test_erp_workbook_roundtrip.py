# -*- coding: utf-8 -*-
"""复核工作簿的整环守门:写出去 → 模拟会计在 Excel 里改 → 读回来。

这个环是会计真实工作流的骨架(导出 → 挑错 → 删对的留错的 → 去 ERP 删错单 →
改后回导重推),所以守的都是「改了会出人命」的行为,不是字段拼写。
"""

import io
import unittest

from openpyxl import load_workbook

from services.excel import erp_roundtrip as rt
from services.excel.erp_workbook import build_review_workbook
from services.excel.erp_roundtrip_reader import (
    RoundtripParseError,
    parse_roundtrip_workbook,
)


def _sales_rec(hid, inv, buyer, items, **extra):
    f = {
        "history_id": hid,
        "invoice_number": inv,
        "date": "2026-07-22",
        "buyer_name": buyer,
        "items": items,
    }
    f.update(extra)
    return {"filename": f"{inv}.png", "merged_fields": f}


def _purchase_rec(hid, inv, seller, base, vat, **extra):
    f = {
        "history_id": hid,
        "invoice_number": inv,
        "date": "2026-07-20",
        "seller_name": seller,
        "seller_tax": "0105533000202",
        "amount_before_vat": base,
        "vat_amount": vat,
        "total_amount": base + vat,
    }
    f.update(extra)
    return {"filename": f"{inv}.png", "merged_fields": f}


SALES = [
    _sales_rec(
        "hid-s1",
        "SA1-0722",
        "หจก.กิจสมบูรณ์ออยล์(สาขา1)",
        [{"description": "BRAKE (0.5L)", "qty": 3, "unit_price": 250, "erp_item_code": "500603"}],
        erp_docnum="SA1-0722",
        erp_party_code="ก001",
        push_status="success",
    )
]
PURCHASE = [
    _purchase_rec(
        "hid-p1",
        "PV-001",
        "บริษัท ซัพพลาย จำกัด",
        1000.0,
        70.0,
        erp_docnum="PV-001",
        erp_party_code="ซ012",
        push_status="success",
    )
]
PENDING = [
    {
        "merged_fields": {"history_id": "hid-x1", "invoice_number": "UNK-9", "total_amount": 500},
        "reason": "ตรวจไม่พบเลขภาษีของกิจการ",
    }
]


class SheetLayoutTests(unittest.TestCase):
    def test_all_four_sheets_exist_even_when_empty(self):
        """空表也要在。少一张表,会计分不清"没有待判的"和"这版没这功能"。"""
        wb = load_workbook(io.BytesIO(build_review_workbook()))
        self.assertEqual(
            wb.sheetnames,
            [rt.SHEET_SALES, rt.SHEET_PURCHASE, rt.SHEET_PENDING, rt.SHEET_SUMMARY],
        )

    def test_summary_sheet_is_not_parsed_back(self):
        """汇总是给人看的清理清单 · 绝不能被当数据回导重推。"""
        out = parse_roundtrip_workbook(
            build_review_workbook(sales=SALES, created_masters=[{"kind": "item", "code": "PN1"}])
        )
        self.assertNotIn(rt.SHEET_SUMMARY, out["sheets"])


class DirectionFromSheetTests(unittest.TestCase):
    def setUp(self):
        self.raw = build_review_workbook(sales=SALES, purchase=PURCHASE, pending=PENDING)

    def test_direction_comes_from_sheet(self):
        out = parse_roundtrip_workbook(self.raw)
        by_inv = {d["fields"]["invoice_number"]: d for d in out["documents"]}
        self.assertEqual(by_inv["SA1-0722"]["direction"], "sales")
        self.assertEqual(by_inv["PV-001"]["direction"], "purchase")

    def test_direction_also_lands_in_fields_for_explicit_override(self):
        """必须写进 fields.direction —— express_push.direction.explicit_direction 认这个键,
        它优先于税号自动判定,会计挪行改分类靠的就是这条。"""
        out = parse_roundtrip_workbook(self.raw)
        for d in out["documents"]:
            self.assertEqual(d["fields"]["direction"], d["direction"])

    def test_pending_rows_are_never_auto_pushed(self):
        """留在待判表 = 会计还没裁决。绝不能被当成任一方向推出去。"""
        out = parse_roundtrip_workbook(self.raw)
        self.assertEqual(len(out["pending"]), 1)
        self.assertEqual(out["pending"][0]["fields"]["invoice_number"], "UNK-9")
        self.assertNotIn("UNK-9", [d["fields"]["invoice_number"] for d in out["documents"]])

    def test_accountant_moving_a_row_changes_classification(self):
        """把销项那行剪到进项表 = 会计判定我们分错了 → 回导必须按进项走。"""
        wb = load_workbook(io.BytesIO(self.raw))
        src, dst = wb[rt.SHEET_SALES], wb[rt.SHEET_PURCHASE]
        hmap_s = {src.cell(row=1, column=c).value: c for c in range(1, src.max_column + 1)}
        hmap_d = {dst.cell(row=1, column=c).value: c for c in range(1, dst.max_column + 1)}
        moved_row = dst.max_row + 1
        for header, col_d in hmap_d.items():
            col_s = hmap_s.get(header)
            dst.cell(
                row=moved_row,
                column=col_d,
                value=src.cell(row=2, column=col_s).value if col_s else None,
            )
        # 进项表用「ชื่อผู้ขาย」列装对手方 · 从销项的客户列搬过来
        dst.cell(
            row=moved_row,
            column=hmap_d[rt.PURCHASE_COL_PARTY],
            value=src.cell(row=2, column=hmap_s[rt.SALES_COL_PARTY]).value,
        )
        dst.cell(
            row=moved_row,
            column=hmap_d[rt.PURCHASE_COL_INVOICE],
            value=src.cell(row=2, column=hmap_s[rt.SALES_COL_INVOICE]).value,
        )
        src.delete_rows(2)
        buf = io.BytesIO()
        wb.save(buf)

        out = parse_roundtrip_workbook(buf.getvalue())
        by_inv = {d["fields"]["invoice_number"]: d for d in out["documents"]}
        self.assertEqual(by_inv["SA1-0722"]["direction"], "purchase")
        self.assertEqual(by_inv["SA1-0722"]["fields"]["direction"], "purchase")


class AccountantEditsTests(unittest.TestCase):
    def _edit(self, mutate):
        wb = load_workbook(io.BytesIO(build_review_workbook(sales=SALES)))
        ws = wb[rt.SHEET_SALES]
        hmap = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        mutate(ws, hmap)
        buf = io.BytesIO()
        wb.save(buf)
        return parse_roundtrip_workbook(buf.getvalue())

    def test_changed_invoice_no_still_links_back(self):
        """改票号是这个流程的主要目的之一 —— 行键不含票号,改完仍认得回原记录。"""
        out = self._edit(
            lambda ws, h: ws.cell(row=2, column=h[rt.SALES_COL_INVOICE], value="SA1-CORRECTED")
        )
        f = out["documents"][0]["fields"]
        self.assertEqual(f["invoice_number"], "SA1-CORRECTED")
        self.assertEqual(f["history_id"], "hid-s1")

    def test_changed_qty_recomputes_money(self):
        """会计把数量 3 改成 5 → 金额必须跟着变。公式无缓存值,读侧按合同重算。"""
        out = self._edit(lambda ws, h: ws.cell(row=2, column=h[rt.SALES_COL_QTY], value=5))
        f = out["documents"][0]["fields"]
        self.assertEqual(f["items"][0]["amount"], 1250.0)
        self.assertEqual(f["amount_before_vat"], 1250.0)
        self.assertEqual(f["vat_amount"], 87.5)

    def test_literal_override_beats_contract_formula(self):
        """会计把金额直接改成死值 → 那是他的裁决,压过我们的公式。"""
        out = self._edit(lambda ws, h: ws.cell(row=2, column=h[rt.SALES_COL_AMOUNT], value=999.0))
        self.assertEqual(out["documents"][0]["fields"]["items"][0]["amount"], 999.0)

    def test_deleting_correct_rows_leaves_only_the_rest(self):
        """会计把对的行删掉、只留要重做的 —— 回导就只该看到剩下那些。"""
        two = [
            *SALES,
            _sales_rec(
                "hid-s2", "SA2-0722", "X", [{"description": "Y", "qty": 1, "unit_price": 10}]
            ),
        ]
        wb = load_workbook(io.BytesIO(build_review_workbook(sales=two)))
        wb[rt.SHEET_SALES].delete_rows(2)  # 删掉第一张(已核对无误)
        buf = io.BytesIO()
        wb.save(buf)
        out = parse_roundtrip_workbook(buf.getvalue())
        self.assertEqual([d["fields"]["invoice_number"] for d in out["documents"]], ["SA2-0722"])

    def test_inserted_column_does_not_break_parsing(self):
        """会计自己插一列做笔记 —— 解析按列名取值,不该因此错位。"""
        wb = load_workbook(io.BytesIO(build_review_workbook(sales=SALES)))
        ws = wb[rt.SHEET_SALES]
        ws.insert_cols(3)
        ws.cell(row=1, column=3, value="โน้ตของฉัน")
        ws.cell(row=2, column=3, value="ตรวจแล้ว")
        buf = io.BytesIO()
        wb.save(buf)
        out = parse_roundtrip_workbook(buf.getvalue())
        f = out["documents"][0]["fields"]
        self.assertEqual(f["invoice_number"], "SA1-0722")
        self.assertEqual(f["items"][0]["qty"], 3.0)


class ForeignWorkbookTests(unittest.TestCase):
    def test_unrelated_workbook_is_rejected_not_half_parsed(self):
        """别人的表格必须整体拒绝,回落通用路 —— 半解析出来的假数据比解析不了更危险。"""
        from openpyxl import Workbook

        wb = Workbook()
        wb.active.append(["日期", "金额"])
        wb.active.append(["2026-07-01", 100])
        buf = io.BytesIO()
        wb.save(buf)
        with self.assertRaises(RoundtripParseError):
            parse_roundtrip_workbook(buf.getvalue())


if __name__ == "__main__":
    unittest.main()
