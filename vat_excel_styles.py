# -*- coding: utf-8 -*-
"""VAT 对账 Excel 共享样式(字体/填充/边框/对齐)+ 表头/斑马 helper · vat_excel_build 拆分 leaf。"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 通用样式(Sarabun 优先 · 泰文友好) ──
FONT_NAME = "Sarabun"
F_HEAD = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
F_NORM = Font(name=FONT_NAME, size=10)
F_BOLD = Font(name=FONT_NAME, size=10, bold=True)
F_TITLE = Font(name=FONT_NAME, size=18, bold=True, color="111827")
# F_KPI_LBL / F_KPI_VAL removed in v4.10.23 (replaced by Korn-style inline KPI)
F_DIFF_RED = Font(name=FONT_NAME, size=10, color="DC2626")

FILL_HEAD = PatternFill("solid", fgColor="2563EB")  # 蓝表头
FILL_SUM = PatternFill("solid", fgColor="EFF6FF")
FILL_ZEBRA = PatternFill("solid", fgColor="F9FAFB")  # 斑马偶数行
FILL_OK = PatternFill("solid", fgColor="DCFCE7")  # 绿底匹配
FILL_DIFF = PatternFill("solid", fgColor="FEF3C7")  # 黄底差异
FILL_MISS = PatternFill("solid", fgColor="FEE2E2")  # 红底缺一边
FILL_FUZZY = PatternFill("solid", fgColor="DBEAFE")  # 蓝底疑似
FILL_OCRMSG = PatternFill("solid", fgColor="FED7AA")  # 橙底 OCR 漏抽
# FILL_KPI_B/G/R/O removed in v4.10.23 (KPI row now uses EFF6FF bg + colored text)

BORDER_TH = Border(
    left=Side(style="thin", color="E5E7EB"),
    right=Side(style="thin", color="E5E7EB"),
    top=Side(style="thin", color="E5E7EB"),
    bottom=Side(style="thin", color="E5E7EB"),
)
AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
AL_R = Alignment(horizontal="right", vertical="center")
AL_L = Alignment(horizontal="left", vertical="center", wrap_text=True)

HEAD_HEIGHT = 32  # 表头行高 32
ROW_HEIGHT = 22


def _style_header(ws, row, num_cols):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = AL_C
        cell.border = BORDER_TH
    ws.row_dimensions[row].height = HEAD_HEIGHT


def _zebra(ws, start_row, end_row, num_cols):
    """偶数行(相对 start_row)斑马条纹"""
    for r in range(start_row, end_row + 1):
        if (r - start_row) % 2 == 1:
            for c in range(1, num_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.fill.fgColor and cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = FILL_ZEBRA
