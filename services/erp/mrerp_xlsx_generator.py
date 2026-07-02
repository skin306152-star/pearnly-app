# -*- coding: utf-8 -*-
"""
v118.27.4.2 · MR.ERP 适配器(方案 C · xlsx 一键导出)
=========================================================
铁律 28-31(2026-05-09 锁定 · 28 于 v118.27.4.2 修订)

格式:
  - .xlsx 不接受 .csv
  - 🔴 铁律 28 修订(v118.27.4.2):**sheet 数随模板动态 1-4**
      · 1 sheet → 名 `Worksheet`(单独 sheet 不带数字 · 客户/商品档案)
      · 2 sheet → `Worksheet` / `Worksheet 1`(会计凭证)
      · 3 sheet → `Worksheet` / `Worksheet 1` / `Worksheet 2`(销售/采购/收款)
      · 4 sheet → `Worksheet` / `Worksheet 1` / `Worksheet 2` / `Worksheet 3`(付款带代扣税)
      · 命名严格 = 空格分隔(不是 Sheet1/Sheet2)
  - 9 种文件类型(销售-赊销/销售-现金/采购-赊购/采购-现金/客户档案/商品档案/财务收款/财务付款/会计凭证)

字段:
  - 日期 = 字符串 YYYY-MM-DD(cell `@` 文本)
  - 单号 = 字符串(cell `@` 文本)
  - 金额 = 数字 cell · 上限 999,999,999.99
  - 客户代码 = 任意字符串(三段式 + 含泰文)· 例 `01-อนุรักษ์-001`
  - 税率 = 字符串枚举 `"7%"` / `"0%"` / `"นอกระบบ"` / `"ยกเว้น"` 等
  - 字段长度 trim 14-50 字符

模块拆分(2026-06-02 · REFACTOR-WB-modularize · 本文件 = facade):
  - mrerp_xlsx_fmt          ← 字段格式工具 + 校验上限常量
  - mrerp_xlsx_lookups      ← 商品归一化/映射 + 客户/科目/税种查表 + tax_kind 推断
  - mrerp_xlsx_schemas      ← 9 种 sheet schema + 错误码友好提示 + sheet 收集/命名
  - mrerp_xlsx_sharedstrings← openpyxl 输出转 PhpSpreadsheet 兼容(非 sales_credit 路径)
  - mrerp_xlsx_sales_credit ← sales_credit row/detail/tail 装配 + preflight 校验 + Korn 克隆
  本文件保留:derive_mrerp_invoice_no(monkeypatch 目标)+ generate_xlsx 编排器 + make_filename。

stub-first 口子(物料到了改这几处 · 不动其他代码):
  1. MRERP_SHEET_SCHEMAS['sales_cash']        ← 物料 1(7 列收款方式)
  2. MRERP_SHEET_SCHEMAS['purchase_credit']
     MRERP_SHEET_SCHEMAS['purchase_cash']     ← 物料 2
  3. MRERP_SHEET_SCHEMAS['journal']           ← 物料 3
  4. MRERP_ERROR_FRIENDLY                     ← 物料 4
"""

import io
import logging
from datetime import date, datetime
from typing import Dict, Any, List, Optional, Tuple  # noqa: F401  facade public re-export

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font
except ImportError:
    Workbook = None  # type: ignore

# 字段格式工具 + 校验上限常量 · derive_mrerp_invoice_no 用 fmt_date · 余为 re-export 契约
from services.erp.mrerp_xlsx_fmt import (  # noqa: F401  facade public re-export
    MAX_AMOUNT,
    MRERP_INVOICE_NO_MAX,
    MRERP_BILL_NO_MAX,
    MRERP_CUSTOMER_CODE_MAX,
    MRERP_CUSTOMER_BILL_MAX,
    MRERP_VALID_TAX_KINDS_SC,
    MRERP_DATE_FUTURE_HARD_REJECT_DAYS,
    MRERP_DATE_FUTURE_WARN_DAYS,
    MRERP_DATE_PAST_WARN_DAYS,
    fmt_date,
    fmt_str,
    fmt_number,
    fmt_number_strict,
    history_number,
)

# 商品/客户/科目/税种查表 + tax_kind 推断 · masterdata adapter 经 _gen.X 调 · re-export 契约
from services.erp.mrerp_xlsx_lookups import (  # noqa: F401  facade public re-export
    _PRODUCT_NAME_NORM_RE,
    _norm_product_name,
    _build_product_lookup,
    _resolve_product_code,
    lookup_customer_code,
    lookup_account_code,
    lookup_tax_code,
    derive_tax_kind,
)

# sheet schema 数据 + sheet 收集/命名 · generate_xlsx/make_filename 用 · 余为 re-export 契约
from services.erp.mrerp_xlsx_schemas import (  # noqa: F401  facade public re-export
    MRERP_SHEET_SCHEMAS,
    MRERP_ERROR_FRIENDLY,
    get_error_friendly,
    _collect_sheet_groups,
    _sheet_title,
)

# PhpSpreadsheet 兼容后处理(非 sales_credit 路径)· generate_xlsx 用
from services.erp.mrerp_xlsx_sharedstrings import _convert_inline_to_shared_strings

# sales_credit 装配/Korn 克隆 · 必在上方 leaf import 之后(sales_credit 反向 import 本模块)
from services.erp.mrerp_xlsx_sales_credit import (  # noqa: F401  facade public re-export
    build_sales_credit_row,
    build_sales_credit_detail_rows,
    build_sales_credit_tail_row,
    _generate_xlsx_sales_credit_korn_clone,
    _format_num,
)

logger = logging.getLogger(__name__)


def derive_mrerp_invoice_no(history: Dict[str, Any]) -> str:
    """v27.8.1.5 · 根据 OCR invoice_date 生成 MR.ERP 期望格式 YYMMDD-NNN
    Korn 真样本:'690507-001' = BE 年(西历+543 末 2 位)+ 月 + 日 + 序号

    v118.34.28 (Zihao 2026-05-19 拍板) · 修死链:之前 seq 写死 "001" ·
    所有 push 都 derive 出 YYMMDD-001 · MR.ERP 服务端不会自动 increment ·
    第一笔成功后第二笔重复就 "เลขที่ดังกล่าวมีอยู่ในระบบแล้ว".
    序号现在按 docstring 设计取 history.id 末 3 位 hex 转 dec mod 1000 ·
    同一 history 重传幂等 · 不同 history 序号不同."""
    inv_date_str = fmt_date(history.get("invoice_date"))
    if inv_date_str and len(inv_date_str) >= 10:
        try:
            ce_year = int(inv_date_str[:4])
            be_year = (ce_year + 543) % 100
            mm = inv_date_str[5:7]
            dd = inv_date_str[8:10]
            date_part = f"{be_year:02d}{mm}{dd}"
        except Exception:
            date_part = "690101"
    else:
        from datetime import datetime as _dt

        d = _dt.utcnow()
        date_part = f"{(d.year + 543) % 100:02d}{d.month:02d}{d.day:02d}"

    # v118.34.31 (Zihao 2026-05-19 拍板) · seq 升级到 history.id 末 8 位 hex
    # mod 999999 + 1 (1-999999 · 6 位 seq) · MR.ERP test 数据库累计 push 多 ·
    # 5 位也撞 · 6 位几乎不撞 (撞率 ~ 1/百万). 同 history 仍幂等 (确定算法).
    # invoice_no 总长 6+1+6 = 13 char · 还在 18 字符上限内.
    hist_id = str(history.get("id") or "").replace("-", "")
    if hist_id:
        try:
            seq_int = int(hist_id[-8:], 16) % 999999 + 1  # 1-999999
            seq = f"{seq_int:06d}"
        except ValueError:
            seq = "000001"
    else:
        seq = "000001"
    return f"{date_part}-{seq}"


# 泰国标准 VAT=7%。税基取 total−vat(对折扣稳健:折扣已含在 total,不误杀 7-11 折扣票)。
# 票面 VAT 隐含税率落在 [MIN, MAX] 外(如 10%)= 异常 → 转人工,不自动过账错税。
MRERP_VAT_RATE_MIN = 0.05
MRERP_VAT_RATE_MAX = 0.09


def vat_rate_anomaly(history: Dict[str, Any]) -> bool:
    """票面 VAT 与应税基的隐含税率是否异常(非 ≈7%)· 对抗票 06(10% 内部自洽仍应转人工)。

    仅在能算(total>0 且 vat>0)时判;vat≈0(免税/零税率)不判。税基=total−vat
    → 折扣已含在 total,不误杀折扣票;偏离 7% 太多(默认 [5%,9%] 外)→ True(异常)。

    金额走 history_number(顶层→fields 兜底):真实推送流 history 经 flatten,
    vat 只在 fields 里——只读顶层曾让本闸对所有真实单据空转(2026-07-02 复测)。
    """
    total = history_number(history, "total_amount")
    vat = history_number(history, "vat")
    if not total or total <= 0 or not vat or vat <= 0:
        return False
    base = total - vat
    if base <= 0:
        return False
    rate = vat / base
    return rate < MRERP_VAT_RATE_MIN or rate > MRERP_VAT_RATE_MAX


# ============================================================
# 验证:发票装得出有效 sales_credit 行吗?
# 留 facade(读 facade 级 derive_mrerp_invoice_no + MRERP_*_MAX · 二者皆 monkeypatch 目标)
# ============================================================
def validate_history_for_sales_credit(
    history: Dict[str, Any],
    mappings: Dict[str, Any],
) -> Tuple[bool, Optional[str], List[str]]:
    """Preflight check before sending a history through the MR.ERP sales_credit
    upload. Returns (ok, error_code, warnings).

    error_code (when ok=False) — one of:
        ERR_NO_HISTORY            history dict empty
        ERR_NO_CLIENT             client_id missing
        ERR_NO_CUSTOMER_MAPPING   erp_client_mappings has no mrerp row
        ERR_NO_INVOICE_NO         neither invoice_no nor invoice_number set
        ERR_NO_INVOICE_DATE       invoice_date missing
        ERR_NO_TOTAL_AMOUNT       total_amount missing or 0
        ERR_NEGATIVE_AMOUNT       total_amount < 0  (P1-A §3.3)
        ERR_INVOICE_NO_TOO_LONG   derived invoice_no > 18 chars  (P1-A §3.1)
        ERR_BILL_NO_TOO_LONG      SI+invoice_no > 20 chars       (P1-A §3.1)
        ERR_CUSTOMER_CODE_TOO_LONG       customer_code > 20 chars  (P1-A §3.1)
        ERR_CUSTOMER_BILL_TOO_LONG       customer_bill > 20 chars  (P1-A §3.1)
        ERR_TAX_RATE_INVALID      derive_tax_kind ∉ MRERP_VALID_TAX_KINDS_SC  (P1-A §3.2)
        ERR_DATE_FUTURE           invoice_date > today + 30 days  (P1-A §3.5)

    warnings (non-blocking; caller can push these into the exceptions table):
        WARN_DATE_NEAR_FUTURE     invoice_date in (today + 7d, today + 30d]
        WARN_DATE_TOO_OLD         invoice_date < today - 730 days

    The check order intentionally matches MR.ERP's own server-side ordering
    (length → existence → enum → date) so error messages line up.
    """
    if not history:
        return False, "ERR_NO_HISTORY", []
    cid = history.get("client_id") or 0
    if not cid:
        return False, "ERR_NO_CLIENT", []
    customer_code = lookup_customer_code(cid, mappings)
    if not customer_code:
        return False, "ERR_NO_CUSTOMER_MAPPING", []
    if not (history.get("invoice_no") or history.get("invoice_number")):
        return False, "ERR_NO_INVOICE_NO", []
    if not history.get("invoice_date"):
        return False, "ERR_NO_INVOICE_DATE", []

    # Amount: strict mode rejects negative + overflow.
    try:
        total = fmt_number_strict(history.get("total_amount"))
    except ValueError as e:
        msg = str(e).lower()
        if "negative" in msg:
            return False, "ERR_NEGATIVE_AMOUNT", []
        if "exceeds" in msg or "missing" in msg:
            return False, "ERR_NO_TOTAL_AMOUNT", []
        return False, "ERR_NO_TOTAL_AMOUNT", []
    if total <= 0:
        return False, "ERR_NO_TOTAL_AMOUNT", []

    # Length pre-flight (P1-A §3.1).
    invoice_no = derive_mrerp_invoice_no(history)
    if len(invoice_no) > MRERP_INVOICE_NO_MAX:
        return False, "ERR_INVOICE_NO_TOO_LONG", []
    bill_no = "SI" + invoice_no
    if len(bill_no) > MRERP_BILL_NO_MAX:
        return False, "ERR_BILL_NO_TOO_LONG", []
    if len(customer_code) > MRERP_CUSTOMER_CODE_MAX:
        return False, "ERR_CUSTOMER_CODE_TOO_LONG", []
    # customer_bill defaults to customer_code; allow override via mapping
    # if a future schema introduces it.
    customer_bill = customer_code
    if len(customer_bill) > MRERP_CUSTOMER_BILL_MAX:
        return False, "ERR_CUSTOMER_BILL_TOO_LONG", []

    # Tax rate enum gate (P1-A §3.2).
    tax_kind = derive_tax_kind(history)
    if tax_kind not in MRERP_VALID_TAX_KINDS_SC:
        return False, "ERR_TAX_RATE_INVALID", []

    # VAT 税率合理性(对抗票 06):票面 VAT 隐含税率非 ≈7% → 转人工,不自动过账错税。
    if vat_rate_anomaly(history):
        return False, "ERR_VAT_RATE_ANOMALY", []

    # Date sanity (P1-A §3.5).
    warnings: List[str] = []
    inv_date_str = fmt_date(history.get("invoice_date"))
    parsed_inv: Optional[date] = None
    if inv_date_str and len(inv_date_str) >= 10:
        try:
            parsed_inv = datetime.strptime(inv_date_str[:10], "%Y-%m-%d").date()
        except Exception:
            parsed_inv = None
    if parsed_inv is not None:
        today = date.today()
        delta = (parsed_inv - today).days
        if delta > MRERP_DATE_FUTURE_HARD_REJECT_DAYS:
            return False, "ERR_DATE_FUTURE", []
        if delta > MRERP_DATE_FUTURE_WARN_DAYS:
            warnings.append("WARN_DATE_NEAR_FUTURE")
        if delta < -MRERP_DATE_PAST_WARN_DAYS:
            warnings.append("WARN_DATE_TOO_OLD")

    return True, None, warnings


# ============================================================
# 主入口:生成 .xlsx bytes(sheet 数动态 · 铁律 28 修订)
# ============================================================
def generate_xlsx(
    histories: List[Dict[str, Any]],
    mappings: Dict[str, Any],
    sheet_kind: str = "sales_credit",
) -> bytes:
    """
    返回 .xlsx 二进制字节 · 调用方负责发给前端

    histories: List of OCR history dicts(1 个文件可多行 · 同 sheet 类型)
    mappings:  {'clients': [...], 'accounts': [...], 'taxes': [...]}(从 db 拿)
    sheet_kind: 9 选 1 · 默认销售-赊销

    🔴 铁律 28 修订(v118.27.4.2):
      根据 schema 实际有的 columns 段动态创建 1-4 sheet · 不再硬编码 3 sheet
      命名严格 = `Worksheet` / `Worksheet 1` / `Worksheet 2` / `Worksheet 3`(空格分隔)
    """
    if Workbook is None:
        raise RuntimeError("openpyxl not installed · pip install openpyxl")

    # sales_cash 走官方模板克隆(同 sales_credit 思路 · 绕过 stub schema)· 见 mrerp_xlsx_sales_cash
    if sheet_kind == "sales_cash":
        from services.erp.mrerp_xlsx_sales_cash import generate_xlsx_sales_cash

        return generate_xlsx_sales_cash(histories, mappings)

    # master_customer 自建客户(imparmas)· 官方模板克隆 · 见 mrerp_xlsx_customer
    if sheet_kind == "customer_master":
        from services.erp.mrerp_xlsx_customer import generate_xlsx_customer

        return generate_xlsx_customer(histories, mappings)

    # master_product 自建商品(impstkmas)· 官方模板克隆 · 见 mrerp_xlsx_product
    if sheet_kind == "product_master":
        from services.erp.mrerp_xlsx_product import generate_xlsx_product

        return generate_xlsx_product(histories, mappings)

    # purchase 采购(impaptran)· 3-sheet 官方模板克隆 · 见 mrerp_xlsx_purchase
    if sheet_kind == "purchase":
        from services.erp.mrerp_xlsx_purchase import generate_xlsx_purchase

        return generate_xlsx_purchase(histories, mappings)

    # master_supplier 自建供应商(impapmas)· 官方模板克隆 · 见 mrerp_xlsx_supplier
    if sheet_kind == "supplier_master":
        from services.erp.mrerp_xlsx_supplier import generate_xlsx_supplier

        return generate_xlsx_supplier(histories, mappings)

    # 库存进出(impstktranrec/impstktraniss · 仅数量)· 官方模板克隆 · 见 mrerp_xlsx_stock
    if sheet_kind in ("stock_receive", "stock_issue"):
        from services.erp.mrerp_xlsx_stock import generate_xlsx_stock

        return generate_xlsx_stock(
            histories, mappings, kind="receive" if sheet_kind == "stock_receive" else "issue"
        )

    schema = MRERP_SHEET_SCHEMAS.get(sheet_kind)
    if not schema:
        raise ValueError(f"unknown sheet_kind: {sheet_kind}")
    if schema.get("stub") and not schema.get("header_columns"):
        raise RuntimeError(f"sheet_kind={sheet_kind} stub schema has no header_columns yet")

    # sales_credit 走 Korn 真样本克隆路径(100% PhpSpreadsheet 兼容)
    # 详见 docs/integrations/mrerp-known-facts.md §6 xlsx 字节级冷知识
    if sheet_kind == "sales_credit":
        try:
            return _generate_xlsx_sales_credit_korn_clone(histories, mappings)
        except FileNotFoundError as e:
            logger.warning(f"Korn template missing (fallback to openpyxl): {e}")
        except Exception as e:
            logger.warning(f"Korn clone failed (fallback to openpyxl): {e}")

    groups = _collect_sheet_groups(schema)
    if not groups:
        raise RuntimeError(f"sheet_kind={sheet_kind} has no sheet groups configured")

    wb = Workbook()
    bold = Font(bold=True)
    sheets: List[Tuple[str, Any, List[Dict[str, Any]]]] = []

    # 动态建 sheet · 第 1 个用 active 改名 · 后续用 create_sheet
    for idx, (group_name, cols) in enumerate(groups):
        title = _sheet_title(idx)
        if idx == 0:
            ws = wb.active
            ws.title = title
        else:
            ws = wb.create_sheet(title)
        # 标题行
        for i, col in enumerate(cols, start=1):
            c = ws.cell(row=1, column=i, value=col["label"])
            c.font = bold
        sheets.append((group_name, ws, cols))

    # v27.8.1.11 · 空 cell 直接返 None · 让 openpyxl skip cell
    # 后处理时再把缺的 cell 补成完全空 cell `<c r="X#"/>`(对齐 Korn 风格)
    def _safe_val(val, col):
        if val is None or val == "":
            val = col.get("default", None)
        if val is None or val == "":
            return None  # ← skip cell · 后处理插完全空 cell
        return val

    for group_name, ws, cols in sheets:
        if sheet_kind == "sales_credit" and group_name == "header":
            for row_idx, history in enumerate(histories, start=2):
                row_data = build_sales_credit_row(history, mappings)
                for col_idx, col in enumerate(cols, start=1):
                    val = _safe_val(row_data.get(col["key"]), col)
                    if val is None:
                        continue  # v27.8.1.11 · 完全 skip · 后处理补完全空 cell
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    if col["type"] in ("str", "date"):
                        cell.number_format = "@"
        elif sheet_kind == "sales_credit" and group_name == "detail":
            cur_row = 2
            for history in histories:
                detail_rows = build_sales_credit_detail_rows(history, mappings)
                for row_data in detail_rows:
                    for col_idx, col in enumerate(cols, start=1):
                        val = _safe_val(row_data.get(col["key"]), col)
                        if val is None:
                            continue
                        cell = ws.cell(row=cur_row, column=col_idx, value=val)
                        if col["type"] in ("str", "date"):
                            cell.number_format = "@"
                    cur_row += 1
        elif sheet_kind == "sales_credit" and group_name == "tail":
            # v27.8.1.11 · Korn 真样本 tail sheet 只 header · 没 data row
            # MR.ERP 期望 tail 是「条件可选」(押金/已开销售单)· 没数据就别写
            pass
        # 其他 sheet_kind:首版只写表头(等 v27.8.2)

    # v27.8.1.11 · spacer 占位改在后处理阶段插「完全空 cell」<c r="S2"/>
    # 这里只生成 18 列 · 后处理时 expand 到 19 列(跟 Korn 真样本完全一致风格)

    buf = io.BytesIO()
    wb.save(buf)
    out = buf.getvalue()

    # v27.8.1.9 + v27.8.1.11 · 关键!后处理:
    # 1. inlineStr → sharedStrings(PhpSpreadsheet 不识别 inline)
    # 2. 缺失 cell 补「完全空 cell <c r='X#'/>」(让 row 显式声明每列存在)
    # 3. row 加 spans 属性(跟 Korn 真样本对齐)
    # 4. sheet1 末尾 col 19 加完全空 cell(让 dim=A1:S2)
    if sheet_kind == "sales_credit":
        try:
            sheet_col_map = {}
            # idx 对应 sheets 列表 · sheet1=header sheet2=detail sheet3=tail
            for idx, (group_name, _ws, cols) in enumerate(sheets):
                sheet_col_map[f"sheet{idx+1}"] = len(cols)
            out = _convert_inline_to_shared_strings(out, sheet_col_map=sheet_col_map)
        except Exception as e:
            logger.warning(f"_convert_inline_to_shared_strings failed (使用原 xlsx): {e}")

    return out


def make_filename(sheet_kind: str, history_id: str) -> str:
    """
    .xlsx 文件名 · stub sheet 加 -DRAFT 提醒用户字段是猜测
    例:Pearnly_MRERP_sales_credit_20260510_a1b2c3.xlsx
    """
    schema = MRERP_SHEET_SCHEMAS.get(sheet_kind, {})
    is_stub = bool(schema.get("stub"))
    today = datetime.utcnow().strftime("%Y%m%d")
    sid = (history_id or "")[:8]
    suffix = "-DRAFT" if is_stub else ""
    return f"Pearnly_MRERP_{sheet_kind}_{today}_{sid}{suffix}.xlsx"
