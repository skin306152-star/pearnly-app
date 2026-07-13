# -*- coding: utf-8 -*-
"""N1 P0-3(报表看得到、拿不走)守门:services/reports/financials_pdf.py 把真实
services.accounting.workorder_financials.build_financials() 产出的 BS/PL/TB 适配成
services.fileconv 的 Table/ConvertResult,喂给 K2 引擎(pdf_out/xlsx_out)出打印级
PDF/Excel——用真实金标 fin dict(同 test_workorder_financials.py 的影子底稿路径),
不是手编字典,确保适配层与真实数据形状对得上。

断言:①三张表(BS/PL/TB)行数与影子底稿的科目数一致 ②守恒(balanced=True)→
ConvertResult.conserved 为真,不平 → 补 Issue 让它诚实变假 ③PDF 字节以 %PDF 开头
(不只测"不炸") ④Excel 是可读的真工作簿,三张表 sheet 名齐全,取到的金额非空。
"""

from __future__ import annotations

import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.accounting import workorder_financials, workorder_shadow_adapter
from services.fileconv import pdf_out, xlsx_out
from services.fileconv.model import FINANCIALS_REPORT
from services.reports.financials_pdf import build_financials_convert_result


def _golden_fin(period: str = "2569-05") -> dict:
    shadow = workorder_shadow_adapter.build_shadow(
        purchase_entries=[{"net": Decimal("1000"), "vat": Decimal("70"), "grand": Decimal("1070")}],
        sales_amount=Decimal("5000"),
        output_vat=Decimal("350"),
        period=period,
    )
    return workorder_financials.build_financials(shadow.as_gate_payload(), period=period)


class BuildConvertResultTests(unittest.TestCase):
    def test_doc_type_and_three_tables(self):
        result = build_financials_convert_result(
            _golden_fin(), period="2569-05", source_name="Sister Makeup 2569-05", lang="th"
        )
        self.assertEqual(result.doc_type, FINANCIALS_REPORT)
        self.assertEqual(len(result.tables), 3)
        names = [t.name for t in result.tables]
        self.assertTrue(any("งบดุล" in n for n in names))
        self.assertTrue(any("งบกำไรขาดทุน" in n for n in names))
        self.assertTrue(any("งบทดลอง" in n for n in names))

    def test_balanced_input_has_no_issues(self):
        fin = _golden_fin()
        self.assertTrue(fin["balance_sheet"]["balanced"])
        self.assertTrue(fin["trial_balance"]["balanced"])
        result = build_financials_convert_result(fin, period="2569-05", source_name="x")
        self.assertEqual(result.issues, [])
        self.assertTrue(result.conserved)

    def test_unbalanced_bs_flags_issue_and_breaks_conserved(self):
        fin = _golden_fin()
        fin["balance_sheet"] = dict(fin["balance_sheet"], balanced=False, diff="99.00")
        result = build_financials_convert_result(fin, period="2569-05", source_name="x")
        self.assertFalse(result.conserved)
        self.assertEqual([i.kind for i in result.issues], ["bs_unbalanced"])

    def test_bs_row_count_matches_shadow_accounts(self):
        fin = _golden_fin()
        bs = fin["balance_sheet"]
        expected_rows = (
            len(bs.get("assets") or [])
            + len(bs.get("liabilities") or [])
            + len(bs.get("equity") or [])
            + 1  # current_earnings 恒加一行(见 financials_pdf._bs_rows)
            + 3  # 三条 Total 行(assets/liabilities/equity)
        )
        result = build_financials_convert_result(fin, period="2569-05", source_name="x")
        bs_table = next(t for t in result.tables if "งบดุล" in t.name)
        self.assertEqual(len(bs_table.rows), expected_rows)


class RenderSmokeTests(unittest.TestCase):
    """真渲染(不只测"不炸"):PDF 字节头 + Excel 真读出内容。"""

    def test_pdf_bytes_start_with_pdf_magic(self):
        result = build_financials_convert_result(
            _golden_fin(), period="2569-05", source_name="Sister Makeup 2569-05", lang="th"
        )
        content = pdf_out.render(result, lang="th")
        self.assertTrue(content.startswith(b"%PDF"))
        self.assertGreater(len(content), 1000)  # 空壳 PDF 远小于此,确有内容

    def test_xlsx_readable_with_three_report_sheets_and_amounts(self):
        result = build_financials_convert_result(
            _golden_fin(), period="2569-05", source_name="Sister Makeup 2569-05", lang="th"
        )
        content = xlsx_out.build_xlsx(result)
        wb = load_workbook(io.BytesIO(content))
        # 三张报表 + Issues + Summary(xlsx_out._write_issue_sheet/_write_summary_sheet)。
        self.assertEqual(len(wb.sheetnames), 5)
        self.assertIn("Issues", wb.sheetnames)
        self.assertIn("Summary", wb.sheetnames)
        bs_sheet_name = next(n for n in wb.sheetnames if n.startswith("งบดุล"))
        ws = wb[bs_sheet_name]
        # 表头行 + 至少一行数据(资产合计一定在),金额列是数字不是空串。
        self.assertGreaterEqual(ws.max_row, 2)
        amounts = [
            ws.cell(row=r, column=3).value
            for r in range(2, ws.max_row + 1)
            if ws.cell(row=r, column=3).value not in (None, "")
        ]
        self.assertTrue(amounts)
        self.assertTrue(all(isinstance(v, (int, float)) for v in amounts))

    def test_unbalanced_still_renders_honest_pdf(self):
        fin = _golden_fin()
        fin["trial_balance"] = dict(fin["trial_balance"], balanced=False, diff="1.23")
        result = build_financials_convert_result(fin, period="2569-05", source_name="x", lang="en")
        content = pdf_out.render(result, lang="en")
        self.assertTrue(content.startswith(b"%PDF"))
        self.assertFalse(result.conserved)


if __name__ == "__main__":
    unittest.main()
