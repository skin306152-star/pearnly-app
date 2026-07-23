# -*- coding: utf-8 -*-
"""收料口回导短路守门 · 会计把改过的表格丢回来重录这条路。

守的是:自家表走确定性解析(零成本零误读)、别人的表不抢、待判票不静默丢。
"""

import io
import unittest

from openpyxl import Workbook, load_workbook

from services.excel import erp_roundtrip as rt
from services.excel.erp_workbook import build_review_workbook
from services.ocr.roundtrip_intake import ENGINE, try_parse_roundtrip


def _sales(hid, inv, items, **extra):
    f = {
        "history_id": hid,
        "invoice_number": inv,
        "date": "2026-07-22",
        "buyer_name": "หจก.กิจสมบูรณ์ออยล์(สาขา1)",
        "items": items,
    }
    f.update(extra)
    return {"filename": f"{inv}.png", "merged_fields": f}


SALES = [
    _sales(
        "h1",
        "SA1-0722",
        [{"description": "BRAKE (0.5L)", "qty": 3, "unit_price": 250}],
        erp_docnum="SA1-0722",
        push_status="success",
    )
]
PURCHASE = [
    {
        "merged_fields": {
            "history_id": "h2",
            "invoice_number": "PV-001",
            "date": "2026-07-20",
            "seller_name": "บริษัท ซัพพลาย จำกัด",
            "seller_tax": "0105533000202",
            "amount_before_vat": 1000.0,
            "vat_amount": 70.0,
            "total_amount": 1070.0,
        }
    }
]
PENDING = [
    {
        "merged_fields": {"history_id": "h3", "invoice_number": "UNK-9", "total_amount": 500},
        "reason": "ตรวจไม่พบเลขภาษีของกิจการ",
    }
]


class ShortCircuitTests(unittest.TestCase):
    def test_own_workbook_parsed_without_model(self):
        """自家表逐格读:引擎标识如实、成本 0、置信度 1.0 —— 不冒充跑过 OCR。"""
        res = try_parse_roundtrip(build_review_workbook(sales=SALES), "review.xlsx")
        self.assertIsNotNone(res)
        self.assertEqual(res.engine, ENGINE)
        self.assertEqual(res.estimated_cost_thb, 0.0)
        self.assertEqual(res.pages[0].layer_chain, [ENGINE])
        self.assertEqual(res.pages[0].layer1_avg_confidence, 1.0)

    def test_fields_and_lines_land(self):
        res = try_parse_roundtrip(build_review_workbook(sales=SALES), "review.xlsx")
        inv = res.pages[0].invoice
        self.assertEqual(inv.invoice_number, "SA1-0722")
        self.assertEqual(inv.buyer_name, "หจก.กิจสมบูรณ์ออยล์(สาขา1)")
        self.assertEqual(len(inv.items), 1)
        self.assertEqual(inv.items[0].name, "BRAKE (0.5L)")
        # 数字存字符串是 ThaiInvoice 的既有约定 · 且不能带 "3.0" 这种尾巴
        self.assertEqual(inv.items[0].qty, "3")
        self.assertEqual(inv.items[0].subtotal, "750")

    def test_both_directions_come_through(self):
        res = try_parse_roundtrip(
            build_review_workbook(sales=SALES, purchase=PURCHASE), "review.xlsx"
        )
        nos = {p.invoice.invoice_number for p in res.pages}
        self.assertEqual(nos, {"SA1-0722", "PV-001"})

    def test_pending_rows_flagged_not_dropped(self):
        """会计没裁决方向的票必须带出来并标人工 —— 静默丢掉 = 漏票。"""
        res = try_parse_roundtrip(
            build_review_workbook(sales=SALES, pending=PENDING), "review.xlsx"
        )
        by_no = {p.invoice.invoice_number: p for p in res.pages}
        self.assertIn("UNK-9", by_no)
        self.assertTrue(by_no["UNK-9"].needs_manual_review)
        self.assertFalse(by_no["SA1-0722"].needs_manual_review)


class FallbackTests(unittest.TestCase):
    def _foreign_xlsx(self):
        wb = Workbook()
        wb.active.append(["วันที่", "จำนวนเงิน"])
        wb.active.append(["2026-07-01", 100])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_foreign_workbook_falls_back(self):
        """别人的表格返回 None → 调用方回落通用表格路,不抢活也不半解析。"""
        self.assertIsNone(try_parse_roundtrip(self._foreign_xlsx(), "supplier.xlsx"))

    def test_non_table_extension_ignored(self):
        self.assertIsNone(try_parse_roundtrip(b"\x89PNG", "scan.png"))
        self.assertIsNone(try_parse_roundtrip(b"", "empty.xlsx"))

    def test_corrupt_bytes_fall_back_not_crash(self):
        """解析崩了也只该降级 —— 不能让整次上传失败。"""
        self.assertIsNone(try_parse_roundtrip(b"not really a workbook", "x.xlsx"))

    def test_workbook_with_only_headers_falls_back(self):
        """空工作簿不返回 0 页的假成功。"""
        self.assertIsNone(try_parse_roundtrip(build_review_workbook(), "empty_review.xlsx"))


class AccountantEditThenReimportTests(unittest.TestCase):
    def test_edited_workbook_reflects_corrections(self):
        """整条真实动作:改票号 + 改数量 → 回导拿到的是改后的值。"""
        raw = build_review_workbook(sales=SALES)
        wb = load_workbook(io.BytesIO(raw))
        ws = wb[rt.SHEET_SALES]
        hmap = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}
        ws.cell(row=2, column=hmap[rt.SALES_COL_INVOICE], value="SA1-FIXED")
        ws.cell(row=2, column=hmap[rt.SALES_COL_QTY], value=5)
        buf = io.BytesIO()
        wb.save(buf)

        res = try_parse_roundtrip(buf.getvalue(), "review.xlsx")
        inv = res.pages[0].invoice
        self.assertEqual(inv.invoice_number, "SA1-FIXED")
        self.assertEqual(inv.items[0].qty, "5")
        self.assertEqual(inv.items[0].subtotal, "1250")
        self.assertEqual(inv.subtotal, "1250")
        self.assertEqual(inv.vat, "87.5")


if __name__ == "__main__":
    unittest.main()
