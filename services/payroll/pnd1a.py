# -*- coding: utf-8 -*-
"""ภ.ง.ด.1ก 年度聚合(批次 H 收尾件 · 只聚合不重算)。

事务所 2 月底前须交 1ก(全年工资预扣年报:逐员工全年支付总额/预扣总额)。数据源是
services/payroll/store.py 里 H1b 工具卡逐月已提交的 client_payroll_rows(整体替换 ·
单一事实源)——本模块按客户 + 佛历年度读出该年度所有月度行,按员工身份证聚合出年度
Σ支付/Σ预扣,供会计对着 e-Filing 网页键入年报。与 services/payroll/validate.py 同边界:
只验不算,不重算个税(累进/减免/全年平均法是 HR 范畴)。

聚合前三项复验(诚实点名,不吞):
  ① 逐行 mod-11 身份证复验(月度提交时校验过,但落库不阻断——commit_endpoint 的
     issues 不阻断 save_period_rows,坏行可能已进库,年度聚合前再查一遍)
  ② 跨月同身份证不同姓名 → 点名,出件按最新月姓名(不静默挑一个)
  ③ Σ员工年度额 = Σ月度已提交额守恒断言(防聚合实现有 bug 悄悄丢钱,不平点名月份)

上传件降级(**格式禁臆造**):ใบแนบ/FORMAT กลาง 的 1ก 年报变体字段与月报不同(官方
DETAIL 逐笔支付日结构如何映射到「年度总额、无单笔支付日」的行,services/tax/
rdprep_pnd1.py 现有 DETAIL_FIELDS 是照月报官方 PDF 逐字段核实的,1ก 版本本仓库既无
官方 PDF 也无金标样本佐证)——本模块只产出键入底稿 xlsx,不装配上传件;
routes/payroll_routes.py 的年报端点同步只开放 kind=keying,attach/central 诚实拒绝。
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal
from typing import Iterable, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from services.payroll import model, validate

_EPS = Decimal("0.01")

_TWO_DP = Decimal("0.01")
_HEADER_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")

# (AnnualRow 属性名, 表头文案, 列宽)——年度口径:无支付日/收入码列(整年多笔已聚合,
# 单笔层面的这些字段年度视图无意义),换成涉及月数供会计核对完整性。
_COLUMNS: tuple[tuple[str, str, int], ...] = (
    ("seq", "ลำดับ (序号)", 6),
    ("employee_id", "เลข 13 หลัก (身份证)", 18),
    ("_name", "ชื่อ-สกุล (姓名)", 30),
    ("months_count", "จำนวนเดือนที่จ่าย (涉及月数)", 16),
    ("paid_amount", "จำนวนเงินที่จ่ายทั้งปี (年度Σ支付)", 20),
    ("wht_amount", "ภาษีที่หักทั้งปี (年度Σ预扣)", 18),
)

_AMOUNT_COLS = {"paid_amount", "wht_amount"}


@dataclass
class AnnualRow:
    """单员工年度聚合行。paid_by_period 只供 _check_conservation 诊断用,不进键入底稿。"""

    seq: int
    employee_id: str
    title: str
    first_name: str
    last_name: str
    paid_amount: Decimal
    wht_amount: Decimal
    months_count: int
    periods: tuple
    paid_by_period: dict = field(default_factory=dict)


def aggregate_year(rows: list) -> tuple:
    """月度进料行(store.load_year_rows 的 dict 列表)→ (年度聚合行, 校验 Issue 列表)。

    按员工身份证分组、按时间序(period 字符串 "YYYY-MM" 佛历,零填充可直接字典序排)
    取最新月姓名。不丢任何行——坏身份证/跨月改名行仍计入 Σ,只点名不排除(方案 §校验:
    只验不算)。
    """
    issues: list = []
    groups: dict = {}
    order: list = []

    for row in rows:
        issues.extend(_check_id(row))
        emp_id = row.get("employee_id") or ""
        if emp_id not in groups:
            groups[emp_id] = {
                "paid_amount": Decimal("0"),
                "wht_amount": Decimal("0"),
                "paid_by_period": {},
                "name_by_period": {},
            }
            order.append(emp_id)
        acc = groups[emp_id]
        period = row["period"]
        amount = row["paid_amount"] or Decimal("0")
        acc["paid_amount"] += amount
        acc["wht_amount"] += row["wht_amount"] or Decimal("0")
        acc["paid_by_period"][period] = acc["paid_by_period"].get(period, Decimal("0")) + amount
        acc["name_by_period"][period] = (row["title"], row["first_name"], row["last_name"])

    annual_rows: list = []
    for seq, emp_id in enumerate(order, start=1):
        acc = groups[emp_id]
        periods_sorted = tuple(sorted(acc["paid_by_period"]))
        latest = periods_sorted[-1]
        title, first, last = acc["name_by_period"][latest]
        if len(set(acc["name_by_period"].values())) > 1:
            issues.append(
                model.Issue(
                    kind=model.ISSUE_NAME_MISMATCH,
                    field=model.F_EMPLOYEE_ID,
                    message="同身份证跨月姓名不一致,已按最新月姓名出件",
                    row_no=None,
                    value=f"id={emp_id} months={','.join(periods_sorted)}",
                )
            )
        annual_rows.append(
            AnnualRow(
                seq=seq,
                employee_id=emp_id,
                title=title,
                first_name=first,
                last_name=last,
                paid_amount=acc["paid_amount"],
                wht_amount=acc["wht_amount"],
                months_count=len(periods_sorted),
                periods=periods_sorted,
                paid_by_period=acc["paid_by_period"],
            )
        )

    issues.extend(_check_conservation(rows, annual_rows))
    return annual_rows, issues


def annual_totals(rows: Iterable) -> dict:
    """Σ年度支付 / Σ年度扣税(Decimal 精确值,供响应 JSON 与守恒断言同源)。"""
    paid = Decimal("0")
    wht = Decimal("0")
    for row in rows:
        paid += row.paid_amount or Decimal("0")
        wht += row.wht_amount or Decimal("0")
    return {"paid_amount": paid, "wht_amount": wht}


def build_annual_keying_workbook(rows: Iterable) -> bytes:
    """年度聚合行 → 键入底稿 xlsx(照 keying_sheet.build_workbook 范式 · 末行 Σ合计)。"""
    rows = list(rows)
    wb = Workbook()
    ws = wb.active
    ws.title = "PND1A keying"

    for col, (_, label, width) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=label)
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"

    for idx, row in enumerate(rows, start=2):
        _write_row(ws, idx, row)

    row_totals = annual_totals(rows)
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="รวม (合计)").font = _HEADER_FONT
    _write_amount(ws, total_row, 5, row_totals["paid_amount"], bold=True)
    _write_amount(ws, total_row, 6, row_totals["wht_amount"], bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_annual_filename(*, nid: str, tax_year_be: str) -> str:
    """年度键入底稿文件名(辅助件,不套官方 §3 规约——同 keying_sheet.build_filename 先例)。"""
    return f"PND1A_{nid}_{tax_year_be}_keying.xlsx"


def _write_row(ws, row_no: int, row: AnnualRow) -> None:
    for col, (attr, _, _) in enumerate(_COLUMNS, start=1):
        if attr in _AMOUNT_COLS:
            _write_amount(ws, row_no, col, getattr(row, attr))
        elif attr == "_name":
            ws.cell(
                row=row_no,
                column=col,
                value=f"{row.title} {row.first_name} {row.last_name}".strip(),
            )
        else:
            ws.cell(row=row_no, column=col, value=getattr(row, attr))


def _write_amount(
    ws, row_no: int, col: int, value: Optional[Decimal], *, bold: bool = False
) -> None:
    cell = ws.cell(row=row_no, column=col, value=_money(value))
    cell.alignment = _RIGHT
    cell.number_format = "#,##0.00"
    if bold:
        cell.font = _HEADER_FONT


def _money(value: Optional[Decimal]) -> float:
    """Decimal → 两位小数 float,仅供 xlsx 显示;真值取 annual_totals()/上游 Decimal。"""
    if value is None:
        return 0.0
    return float(Decimal(str(value)).quantize(_TWO_DP, rounding=ROUND_HALF_UP))


def _check_id(row: dict) -> list:
    """mod-11 复验(月度提交时验过,但 commit_endpoint 的 issues 不阻断落库——坏行可能
    已进库,年度聚合前再查一遍,不能假设库里的都是干净数据)。"""
    tid = row.get("employee_id") or ""
    if validate.is_valid_employee_id(tid):
        return []
    return [
        model.Issue(
            kind=model.ISSUE_INVALID_ID,
            field=model.F_EMPLOYEE_ID,
            message="身份证非 13 位数字或 mod-11 校验位不符(年度聚合复验)",
            row_no=row.get("seq"),
            value=f"period={row.get('period')} id={tid}",
        )
    ]


def _check_conservation(rows: list, annual_rows: list) -> list:
    """Σ员工年度额 = Σ月度已提交额(两条独立路径同源比对,不平点名具体月份)。

    正常路径下必然守恒(聚合不丢行、不筛数据),此断言防的是未来实现改动悄悄漏计——
    同 services/payroll/validate.py 的 V3 sum_mismatch 一样,只验不算,不做静默容错。
    """
    raw_by_period: dict = {}
    for row in rows:
        period = row["period"]
        raw_by_period[period] = raw_by_period.get(period, Decimal("0")) + (
            row["paid_amount"] or Decimal("0")
        )

    agg_by_period: dict = {}
    for annual_row in annual_rows:
        for period, amount in annual_row.paid_by_period.items():
            agg_by_period[period] = agg_by_period.get(period, Decimal("0")) + amount

    issues = []
    for period, raw_total in raw_by_period.items():
        agg_total = agg_by_period.get(period, Decimal("0"))
        if abs(raw_total - agg_total) > _EPS:
            issues.append(
                model.Issue(
                    kind=model.ISSUE_YEAR_SUM_MISMATCH,
                    field=model.F_PAID_AMOUNT,
                    message=f"年度聚合与月度进料不守恒,点名月份 {period}",
                    row_no=None,
                    value=f"period={period} 月度Σ={raw_total} 聚合Σ={agg_total} 差={raw_total - agg_total}",
                )
            )
    return issues
