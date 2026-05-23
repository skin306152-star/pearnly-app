# -*- coding: utf-8 -*-
"""
P1.2-M2 v118.35.0.47 · 守门测试 · vat_excel_exporter 发票侧字段校正标黄 + 痕迹 section

锁定 4 个契约(同 M4 P0.2 套路):
  1. 无 field_overrides → Excel 不含『手动修改痕迹』section · 老 layout 不变
  2. 有 field_overrides → 末尾含『手动修改痕迹』section(en: Manual Edit Trail)
  3. 被改的发票侧 cell 标黄(FFE082)+ comment 含 OCR / User
  4. 发票侧列显示用户校正值(覆盖 OCR 原值)
"""
import io
import unittest

import openpyxl

from vat_excel_exporter import export_recon_task


def _task():
    return {"period_month": 3, "period_year": 2025, "client_id": 1}


def _client():
    return {"name": "Test Co", "tax_id": "0105500000000", "address": "Bangkok"}


def _row(field_overrides=None):
    return {
        "id": 1, "invoice_no": "INV-001", "invoice_date": "2025-03-15",
        "buyer_name": "ABC Co", "buyer_tax_id": "0100000000001",
        "buyer_branch": "00000", "amount_pre_vat": 1000.0, "vat_amount": 70.0,
        "total_amount": 1070.0, "invoice_filename": "a.pdf",
        "diff_fields": {}, "diff_categories": "", "status": "matched",
        "field_overrides": field_overrides or {},
    }


def _all_text(ws):
    out = []
    for row in ws.iter_rows(values_only=True):
        for v in row:
            if isinstance(v, str):
                out.append(v)
    return " ".join(out)


class VatExportFieldOverrideTests(unittest.TestCase):

    def test_no_overrides_has_no_manual_section(self):
        b = export_recon_task(_task(), [_row()], _client(), lang="en")
        wb = openpyxl.load_workbook(io.BytesIO(b))
        self.assertNotIn("Manual Edit Trail", _all_text(wb.worksheets[0]))

    def test_with_overrides_has_manual_section(self):
        ov = {"buyer_name": {"ocr": "ABC Co", "user": "ABC Company Ltd", "ts": "x"}}
        b = export_recon_task(_task(), [_row(ov)], _client(), lang="en")
        wb = openpyxl.load_workbook(io.BytesIO(b))
        self.assertIn("Manual Edit Trail", _all_text(wb.worksheets[0]))

    def test_overridden_cell_yellow_with_comment(self):
        ov = {"buyer_name": {"ocr": "ABC Co", "user": "ABC Company Ltd", "ts": "x"}}
        b = export_recon_task(_task(), [_row(ov)], _client(), lang="en")
        wb = openpyxl.load_workbook(io.BytesIO(b))
        ws = wb.worksheets[0]
        found = False
        for row in ws.iter_rows():
            for cell in row:
                if cell.value == "ABC Company Ltd" and cell.fill and cell.fill.fgColor:
                    fg = cell.fill.fgColor.rgb or ""
                    if fg.endswith("FFE082") and cell.comment:
                        txt = cell.comment.text or ""
                        self.assertIn("OCR:", txt)
                        self.assertIn("User:", txt)
                        found = True
        self.assertTrue(found, "被改的发票侧 cell 未标 FFE082 黄底 + comment")

    def test_invoice_column_shows_user_value(self):
        """发票侧列显示用户校正值(而非 OCR 原值 ABC Co)"""
        ov = {"buyer_name": {"ocr": "ABC Co", "user": "ABC Company Ltd", "ts": "x"}}
        b = export_recon_task(_task(), [_row(ov)], _client(), lang="en")
        wb = openpyxl.load_workbook(io.BytesIO(b))
        joined = _all_text(wb.worksheets[0])
        self.assertIn("ABC Company Ltd", joined)


if __name__ == "__main__":
    unittest.main()
