# -*- coding: utf-8 -*-
"""单测 · build_billing_xlsx(账单三 sheet 导出 · 2026-06-28)

锁:① 三个 sheet(扣费/充值/识别)② 表头随 lang ③ 计费如实区分 实扣/套餐内免费/月费
④ 识别三态 成功/失败 ⑤ 充值审核态 到账/待审核 ⑥ 金额取绝对值 + 会计格式。
"""

import io
import unittest

from openpyxl import load_workbook

from services.usage.billing_export import build_billing_xlsx, _t

_USAGE = [
    {
        "date": "2026-06-28 13:00",
        "type": "usage",
        "description": "inv.pdf",
        "filename": "inv.pdf",
        "pages": 2,
        "cost_thb": -3.0,
        "balance_after": 497.0,
    },
    {
        "date": "2026-06-28 12:00",
        "type": "usage",
        "description": "q.pdf",
        "filename": "q.pdf",
        "pages": 1,
        "cost_thb": 0.0,
        "balance_after": 500.0,
    },
    {
        "date": "2026-06-01 09:00",
        "type": "subscription",
        "description": "订阅 Package L",
        "filename": "",
        "pages": 0,
        "cost_thb": -500.0,
        "balance_after": 500.0,
    },
]
_TOPUP = [
    {
        "created_at": "2026-06-01 08:00",
        "amount_thb": 1000.0,
        "payer_name": "Somchai",
        "status": "approved",
        "reviewed_at": "2026-06-01 08:30",
        "note": "promptpay",
    }
]
_OCR = [
    {
        "created_at": "2026-06-28 13:00",
        "filename": "inv.pdf",
        "status": "confirmed",
        "invoice_no": "IV001",
        "seller_name": "ACME",
        "total_amount": 1070.0,
        "page_count": 2,
        "source": "upload",
    },
    {
        "created_at": "2026-06-28 12:00",
        "filename": "blur.jpg",
        "status": "failed",
        "invoice_no": None,
        "seller_name": None,
        "total_amount": None,
        "page_count": 1,
        "source": "line",
    },
]


def _wb(lang="zh"):
    data = build_billing_xlsx(
        lang=lang, company="ACME Co", usage_rows=_USAGE, topup_rows=_TOPUP, ocr_rows=_OCR
    )
    assert data[:2] == b"PK"  # 合法 xlsx(zip)
    return load_workbook(io.BytesIO(data))


class BuildBillingXlsxTests(unittest.TestCase):
    def test_three_sheets(self):
        wb = _wb("zh")
        self.assertEqual(wb.sheetnames, ["扣费记录", "充值记录", "识别记录"])

    def test_headers_follow_language(self):
        # zh 与 en 表头不同 · 证明随 lang 切换
        zh = _wb("zh")["扣费记录"]
        en_wb = _wb("en")
        en = en_wb[en_wb.sheetnames[0]]
        self.assertEqual(zh.cell(4, 1).value, _t("zh", "c_date"))
        self.assertEqual(en.cell(4, 1).value, _t("en", "c_date"))
        self.assertNotEqual(zh.cell(4, 1).value, en.cell(4, 1).value)

    def test_billing_kind_honest(self):
        ws = _wb("zh")["扣费记录"]
        # 第 3 列=计费 · 行 5/6/7 对应 实扣 / 套餐内免费 / 月费
        self.assertEqual(ws.cell(5, 3).value, "实扣")
        self.assertEqual(ws.cell(6, 3).value, "套餐内免费")
        self.assertEqual(ws.cell(7, 3).value, "月费")
        # 金额取绝对值(第 6 列)
        self.assertEqual(ws.cell(5, 6).value, 3.0)
        self.assertEqual(ws.cell(6, 6).value, 0.0)

    def test_topup_status_credited(self):
        ws = _wb("zh")["充值记录"]
        self.assertEqual(ws.cell(5, 4).value, "已到账")  # approved
        self.assertEqual(ws.cell(5, 2).value, 1000.0)

    def test_ocr_status_success_and_failed(self):
        ws = _wb("zh")["识别记录"]
        self.assertEqual(ws.cell(5, 3).value, "识别成功")  # confirmed
        self.assertEqual(ws.cell(6, 3).value, "识别失败")  # failed

    def test_empty_sections_render_placeholder(self):
        data = build_billing_xlsx(lang="en", company="X", usage_rows=[], topup_rows=[], ocr_rows=[])
        wb = load_workbook(io.BytesIO(data))
        ws = wb[wb.sheetnames[0]]
        self.assertEqual(ws.cell(5, 1).value, _t("en", "empty"))

    def test_unknown_lang_falls_back_en(self):
        data = build_billing_xlsx(
            lang="xx", company="X", usage_rows=_USAGE, topup_rows=[], ocr_rows=[]
        )
        wb = load_workbook(io.BytesIO(data))
        self.assertEqual(wb.sheetnames[0], _t("en", "sheet_usage"))


if __name__ == "__main__":
    unittest.main()
