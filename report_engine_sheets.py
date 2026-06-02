# -*- coding: utf-8 -*-
"""report_engine · sheet 写入(info block / 主表 / 汇总表)+ cell 样式应用 leaf。"""

import re
from collections import defaultdict
from datetime import datetime
from typing import List, Dict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from i18n_reports import i18n_get, i18n_format
from report_engine_styles import (
    COLOR_WARN_AMOUNT,
    FONT_NAME,
    BORDER_ALL,
    BORDER_TOP_THICK,
    HEADER_FILL,
    ZEBRA_FILL,
    TOTAL_FILL,
    HEADER_FONT,
    CELL_FONT,
    TOTAL_FONT,
    TITLE_FONT,
    INFO_FONT,
    ALIGN_CENTER,
    ALIGN_LEFT,
    ALIGN_RIGHT,
)
from report_engine_format import (
    _to_float,
    _format_tax_id,
    _format_branch,
    _render_cell_value,
    _source_label,
)


def _apply_cell_style(
    cell,
    col_def: dict,
    is_zebra: bool = False,
    warn_amount: bool = False,
    no_thousand: bool = False,
):
    """单元格样式"""
    cell.font = CELL_FONT
    cell.border = BORDER_ALL
    align = col_def.get("align", "left")
    if align == "center":
        cell.alignment = ALIGN_CENTER
    elif align == "right":
        cell.alignment = ALIGN_RIGHT
    else:
        cell.alignment = ALIGN_LEFT
    if is_zebra:
        cell.fill = ZEBRA_FILL
    col_type = col_def.get("type", "")
    if col_type == "money":
        cell.number_format = "#,##0.00" if not no_thousand else "0.00"
    elif col_type == "money_raw":
        cell.number_format = "0.00"
    if warn_amount:
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color=COLOR_WARN_AMOUNT)


def _write_info_block(ws: Worksheet, template: dict, meta: dict, lang: str, n_cols: int) -> int:
    """
    顶部信息块(标题 + 元信息)
    返回:数据起始行号(1-indexed)
    """
    if not template.get("show_info_block"):
        return 1

    # 第 1 行:大标题(合并)
    title = i18n_format(
        lang,
        template["title_key"],
        client=meta.get("client_name") or i18n_get(lang, "client-default"),
        month=meta.get("period_label") or i18n_get(lang, "month-all"),
    )
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # 第 2 行:元信息(税号 + 分公司 + 张数 + 生成时间)
    info_parts = []
    if meta.get("client_tax_id"):
        info_parts.append(
            f"{i18n_get(lang, 'info-tax-id')}: {_format_tax_id(meta['client_tax_id'])}"
        )
    if meta.get("client_branch"):
        info_parts.append(
            f"{i18n_get(lang, 'info-branch')}: {_format_branch(meta['client_branch'], lang)}"
        )
    if meta.get("doc_count"):
        info_parts.append(i18n_format(lang, "info-doc-count", n=meta["doc_count"]))
    info_parts.append(
        f"{i18n_get(lang, 'info-generated-at')}: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    )

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    cell = ws.cell(row=2, column=1, value="    ".join(info_parts))
    cell.font = INFO_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 20

    # 第 3 行留空(分隔)
    ws.row_dimensions[3].height = 8

    return 4  # 表头从第 4 行开始


def _write_main_sheet(wb: Workbook, template: dict, rows: List[Dict], meta: dict, lang: str):
    """主明细 Sheet"""
    sheet_name = i18n_get(lang, template["name_key"])[:31]  # Excel sheet name max 31
    # 替换非法字符
    sheet_name = re.sub(r"[\\/?*\[\]:]", "·", sheet_name)
    ws = wb.active
    ws.title = sheet_name

    columns = template["columns"]
    n_cols = len(columns)
    no_thousand = template.get("no_thousand_sep", False)
    ascii_only = template.get("ascii_only", False)
    warn_threshold = template.get("warn_amount_threshold", 0)

    # ERP 模板:第一行写隐藏标记
    start_row = 1
    if template.get("first_row_marker"):
        ws.cell(row=1, column=1, value=template["first_row_marker"])
        ws.row_dimensions[1].hidden = True
        start_row = 2

    # 信息块(input_vat / standard / print 才有)
    if not ascii_only and template.get("show_info_block"):
        start_row = _write_info_block(ws, template, meta, lang, n_cols)

    header_row = start_row

    # 表头
    for col_idx, col in enumerate(columns, start=1):
        if ascii_only and col.get("header_ascii"):
            label = col["header_ascii"]
        else:
            label = i18n_get(lang, col["header_key"])
        cell = ws.cell(row=header_row, column=col_idx, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_ALL
        ws.column_dimensions[get_column_letter(col_idx)].width = col["width"]
    ws.row_dimensions[header_row].height = 28

    # 数据行
    data_start = header_row + 1
    for i, row_data in enumerate(rows):
        excel_row = data_start + i
        is_zebra = i % 2 == 1
        for col_idx, col in enumerate(columns, start=1):
            raw_value = row_data.get(col["key"], "")
            value = _render_cell_value(raw_value, col["type"], lang, no_thousand)
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            warn = (
                warn_threshold > 0
                and col["type"] == "money"
                and isinstance(value, (int, float))
                and value > warn_threshold
            )
            _apply_cell_style(
                cell, col, is_zebra=is_zebra, warn_amount=warn, no_thousand=no_thousand
            )
        if template.get("row_height"):
            ws.row_dimensions[excel_row].height = template["row_height"]

    # 合计行(SUM 公式)
    total_cols = template.get("total_columns", [])
    if total_cols and rows:
        total_row = data_start + len(rows)
        # 合计标签放在第 1 列
        total_label_col = 1
        # 找第一个数字列前面所有非数字列合并起来放标签
        first_money_idx = None
        for col_idx, col in enumerate(columns, start=1):
            if col["key"] in total_cols:
                first_money_idx = col_idx
                break

        if first_money_idx and first_money_idx > 1:
            ws.merge_cells(
                start_row=total_row,
                start_column=1,
                end_row=total_row,
                end_column=first_money_idx - 1,
            )
            cell = ws.cell(row=total_row, column=1, value=i18n_get(lang, "report-grand-total"))
            cell.font = TOTAL_FONT
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.fill = TOTAL_FILL
            cell.border = BORDER_TOP_THICK

        for col_idx, col in enumerate(columns, start=1):
            cell = ws.cell(row=total_row, column=col_idx)
            if col["key"] in total_cols:
                col_letter = get_column_letter(col_idx)
                formula = f"=SUM({col_letter}{data_start}:{col_letter}{total_row - 1})"
                cell.value = formula
                cell.number_format = "#,##0.00" if not no_thousand else "0.00"
                cell.alignment = ALIGN_RIGHT
            elif col_idx >= (first_money_idx or 0):
                # 数字列右侧空合计列
                pass
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER_TOP_THICK
        ws.row_dimensions[total_row].height = 26

    # 冻结
    if template.get("freeze_panes"):
        # freeze_panes 配置是相对值;实际取决于 header_row + 1
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    # 自动筛选(标准模板)
    if template.get("auto_filter") and rows:
        last_col = get_column_letter(n_cols)
        last_row = data_start + len(rows) - 1
        ws.auto_filter.ref = f"A{header_row}:{last_col}{last_row}"

    # 签名栏
    if template.get("show_signature_block"):
        sig_row = data_start + len(rows) + (2 if total_cols else 1) + 2
        labels = [
            i18n_get(lang, "sig-prepared-by"),
            i18n_get(lang, "sig-reviewed-by"),
            i18n_get(lang, "sig-approved-by"),
        ]
        col_step = max(1, n_cols // 3)
        for i, label in enumerate(labels):
            target_col = min(1 + i * col_step, n_cols - 1)
            cell = ws.cell(row=sig_row, column=target_col, value=f"{label}: ____________")
            cell.font = INFO_FONT
            cell.alignment = ALIGN_LEFT
        ws.row_dimensions[sig_row].height = 28

    # 打印设置(print 模板)
    if template.get("page_setup") == "A4_landscape":
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins.left = 0.4
        ws.page_margins.right = 0.4
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5
        ws.print_options.horizontalCentered = True
        ws.print_title_rows = f"{header_row}:{header_row}"  # 每页打印重复表头


def _write_summary_sheet(wb: Workbook, rows: List[Dict], lang: str):
    """第二 Sheet · 汇总分析(老板视角)"""
    ws = wb.create_sheet(title=i18n_get(lang, "summary-sheet-name")[:31])

    # ===== 区域 1 · 按卖方 TOP 10 =====
    by_seller = defaultdict(lambda: {"count": 0, "amount": 0.0})
    total_amount = 0.0
    for r in rows:
        name = (r.get("seller_name") or "—").strip() or "—"
        amt = _to_float(r.get("total_amount"))
        by_seller[name]["count"] += 1
        by_seller[name]["amount"] += amt
        total_amount += amt
    top_sellers = sorted(by_seller.items(), key=lambda x: x[1]["amount"], reverse=True)[:10]

    cur_row = 1
    # 标题
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
    c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-seller"))
    c.font = TITLE_FONT
    c.alignment = ALIGN_LEFT
    ws.row_dimensions[cur_row].height = 28
    cur_row += 1

    headers = [
        i18n_get(lang, "summary-rank"),
        i18n_get(lang, "col-seller-name"),
        i18n_get(lang, "summary-count"),
        i18n_get(lang, "summary-amount"),
        i18n_get(lang, "summary-percent"),
    ]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=cur_row, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = ALIGN_CENTER
        c.border = BORDER_ALL
    ws.row_dimensions[cur_row].height = 24
    cur_row += 1

    for rank, (name, agg) in enumerate(top_sellers, start=1):
        pct = (agg["amount"] / total_amount * 100) if total_amount > 0 else 0
        cells = [
            (rank, "center", "int"),
            (name, "left", "text"),
            (agg["count"], "center", "int"),
            (round(agg["amount"], 2), "right", "money"),
            (f"{pct:.1f}%", "right", "text"),
        ]
        for ci, (v, align, t) in enumerate(cells, start=1):
            c = ws.cell(row=cur_row, column=ci, value=v)
            c.font = CELL_FONT
            c.border = BORDER_ALL
            c.alignment = (
                ALIGN_CENTER
                if align == "center"
                else (ALIGN_RIGHT if align == "right" else ALIGN_LEFT)
            )
            if t == "money":
                c.number_format = "#,##0.00"
        cur_row += 1

    cur_row += 2  # 空 2 行

    # ===== 区域 2 · 按类目 =====
    by_cat = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for r in rows:
        cat = (r.get("category") or "—").strip() or "—"
        by_cat[cat]["count"] += 1
        by_cat[cat]["amount"] += _to_float(r.get("total_amount"))

    if len(by_cat) > 1 or (by_cat and "—" not in by_cat):
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-category"))
        c.font = TITLE_FONT
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=cur_row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        for rank, (cat, agg) in enumerate(
            sorted(by_cat.items(), key=lambda x: x[1]["amount"], reverse=True), start=1
        ):
            pct = (agg["amount"] / total_amount * 100) if total_amount > 0 else 0
            cells = [rank, cat, agg["count"], round(agg["amount"], 2), f"{pct:.1f}%"]
            aligns = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_RIGHT]
            for ci, (v, a) in enumerate(zip(cells, aligns), start=1):
                c = ws.cell(row=cur_row, column=ci, value=v)
                c.font = CELL_FONT
                c.border = BORDER_ALL
                c.alignment = a
                if ci == 4:
                    c.number_format = "#,##0.00"
            cur_row += 1

    cur_row += 2

    # ===== 区域 3 · 按来源(自动化覆盖率)=====
    by_src = defaultdict(lambda: {"count": 0, "amount": 0.0})
    for r in rows:
        src = _source_label(r.get("source") or "—", lang) or "—"
        by_src[src]["count"] += 1
        by_src[src]["amount"] += _to_float(r.get("total_amount"))

    if by_src:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=5)
        c = ws.cell(row=cur_row, column=1, value=i18n_get(lang, "summary-by-source"))
        c.font = TITLE_FONT
        c.alignment = ALIGN_LEFT
        ws.row_dimensions[cur_row].height = 28
        cur_row += 1

        for ci, h in enumerate(headers, start=1):
            c = ws.cell(row=cur_row, column=ci, value=h)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = ALIGN_CENTER
            c.border = BORDER_ALL
        ws.row_dimensions[cur_row].height = 24
        cur_row += 1

        total_count = sum(v["count"] for v in by_src.values())
        for rank, (src, agg) in enumerate(
            sorted(by_src.items(), key=lambda x: x[1]["count"], reverse=True), start=1
        ):
            pct = (agg["count"] / total_count * 100) if total_count > 0 else 0
            cells = [rank, src, agg["count"], round(agg["amount"], 2), f"{pct:.1f}%"]
            aligns = [ALIGN_CENTER, ALIGN_LEFT, ALIGN_CENTER, ALIGN_RIGHT, ALIGN_RIGHT]
            for ci, (v, a) in enumerate(zip(cells, aligns), start=1):
                c = ws.cell(row=cur_row, column=ci, value=v)
                c.font = CELL_FONT
                c.border = BORDER_ALL
                c.alignment = a
                if ci == 4:
                    c.number_format = "#,##0.00"
            cur_row += 1

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 12
