# -*- coding: utf-8 -*-
"""K1c · OCR 桥(services/fileconv/ocr_bridge.py)。

注入 fake provider_call 全覆盖,不触网络。锁定:
①截断硬闸(命门):error_kind='parse' / data 非 dict / schema 残缺 → STATUS_OCR_INCOMPLETE,
  零表格输出,xlsx 只出 Rejected 页 —— 绝不把半截行集当成功;
②多页逐页调用 + 跨页余额链衔接:页间断链如实 issue;
③适配器 str→Decimal(禁 float)/deposit→debit 语义/line_no 跨页连续;
④独立锚:印刷期末余额对不上 → closing_anchor issue(治「截断+假自洽」第二道网);
⑤归因:_default_provider_call 在调用期间挂 fileconv_ocr attribution,收尾必 reset;
⑥convert_pdf 无文字层分支转 OCR 桥(K1a status 语义不再直接回 no_text_layer)。
"""

import io
import unittest
from decimal import Decimal
from unittest import mock

import openpyxl

from services.fileconv import ocr_bridge
from services.fileconv.model import (
    BANK_STATEMENT,
    GENERIC_TABLE,
    GL_LEDGER,
    ISSUE_CLOSING_ANCHOR,
    ISSUE_GL_BALANCE_CHAIN,
    STATUS_OCR_INCOMPLETE,
    STATUS_OCR_UNAVAILABLE,
    STATUS_OK,
)
from services.fileconv.xlsx_out import build_xlsx


class _Outcome:
    """ProviderOutcome 最小同形桩。"""

    def __init__(self, ok, data=None, error_kind=None):
        self.ok = ok
        self.data = data
        self.error_kind = error_kind


def _classify_data(doc_type):
    return _Outcome(True, {"document_type": doc_type})


def _bank_entry(date, desc, deposit, withdrawal, balance):
    return {
        "transaction_date": date,
        "description": desc,
        "deposit": deposit,
        "withdrawal": withdrawal,
        "balance": balance,
    }


def _bank_page(entries, opening="", closing=""):
    return _Outcome(
        True,
        {
            "document_type": "bank_statement",
            "opening_balance": opening,
            "closing_balance": closing,
            "entries": entries,
        },
    )


def _fake_call(outcomes):
    """按调用序返回预排队列(首个 = 分类,其后 = 逐页提取)。"""
    queue = list(outcomes)

    def call(prompt, image_bytes, *, tenant_id=None, api_key=None, max_tokens=0):
        return queue.pop(0)

    return call


class TruncationHardGateTests(unittest.TestCase):
    """命门:截断 = 诚实拒绝,零产出。"""

    def test_parse_error_rejects_whole_file(self):
        call = _fake_call([_classify_data("bank_statement"), _Outcome(False, error_kind="parse")])
        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=call)
        self.assertEqual(result.status, STATUS_OCR_INCOMPLETE)
        self.assertEqual(result.tables, [])

    def test_second_page_truncation_rejects_even_with_good_first_page(self):
        """首页读满、次页截断 → 整件拒绝,首页的行也绝不出件(半件比拒绝更危险)。"""
        page1 = _bank_page([_bank_entry("2026-01-01", "in", "100.00", "", "100.00")])
        call = _fake_call(
            [_classify_data("bank_statement"), page1, _Outcome(False, error_kind="parse")]
        )
        result = ocr_bridge.convert_images([b"p1", b"p2"], "scan.pdf", provider_call=call)
        self.assertEqual(result.status, STATUS_OCR_INCOMPLETE)
        self.assertEqual(result.tables, [])

    def test_non_dict_data_is_incomplete(self):
        call = _fake_call([_classify_data("bank_statement"), _Outcome(True, data="[truncated")])
        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=call)
        self.assertEqual(result.status, STATUS_OCR_INCOMPLETE)

    def test_max_tokens_truncation_is_incomplete(self):
        """provider 层(aistudio._safe_raw)把 MAX_TOKENS 截断收敛为 error_kind='parse'
        (真调实锤:resp.text 快取器在 candidates 为空时抛 ValueError,provider 层已兜住,
        桥只消费结构化结果)——桥据此拒绝整件,绝不出半件。"""
        call = _fake_call([_classify_data("bank_statement"), _Outcome(False, error_kind="parse")])
        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=call)
        self.assertEqual(result.status, STATUS_OCR_INCOMPLETE)
        self.assertEqual(result.tables, [])

    def test_other_provider_raise_is_unavailable(self):
        def raising_call(prompt, image_bytes, **kwargs):
            raise RuntimeError("connection reset")

        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=raising_call)
        self.assertEqual(result.status, STATUS_OCR_UNAVAILABLE)

    def test_provider_unreachable_is_unavailable_not_incomplete(self):
        call = _fake_call([_classify_data("bank_statement"), _Outcome(False, error_kind="auth")])
        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=call)
        self.assertEqual(result.status, STATUS_OCR_UNAVAILABLE)
        self.assertEqual(result.tables, [])

    def test_rejected_result_xlsx_has_no_data_sheets(self):
        call = _fake_call([_classify_data("bank_statement"), _Outcome(False, error_kind="parse")])
        result = ocr_bridge.convert_images([b"img"], "scan.jpg", provider_call=call)
        wb = openpyxl.load_workbook(io.BytesIO(build_xlsx(result)))
        self.assertIn("Rejected", wb.sheetnames)
        self.assertEqual(wb["Rejected"]["A1"].value, STATUS_OCR_INCOMPLETE)
        self.assertNotIn("Bank Statement", wb.sheetnames)


class MultiPageChainTests(unittest.TestCase):
    def test_two_pages_chain_links_and_conserves(self):
        page1 = _bank_page(
            [
                _bank_entry("2026-01-01", "in", "100.00", "", "1,100.00"),
                _bank_entry("2026-01-02", "out", "", "50.00", "1,050.00"),
            ],
            opening="1,000.00",
        )
        page2 = _bank_page(
            [_bank_entry("2026-01-03", "in", "25.50", "", "1,075.50")],
            closing="1,075.50",
        )
        call = _fake_call([_classify_data("bank_statement"), page1, page2])
        result = ocr_bridge.convert_images([b"p1", b"p2"], "stmt.pdf", provider_call=call)
        self.assertEqual(result.status, STATUS_OK)
        self.assertEqual(result.doc_type, BANK_STATEMENT)
        self.assertTrue(result.conserved)
        self.assertEqual(result.stats["row_count"], 3)
        self.assertEqual(result.stats["pages"], 2)
        self.assertEqual(result.tables[0].rows[2][-1], Decimal("1075.50"))

    def test_cross_page_break_is_named(self):
        """页 2 首行接不上页 1 末行余额 → validate_ledger 逐行链点名。"""
        page1 = _bank_page(
            [_bank_entry("2026-01-01", "in", "100.00", "", "1,100.00")], opening="1,000.00"
        )
        page2 = _bank_page([_bank_entry("2026-01-03", "in", "25.50", "", "9,999.00")])
        call = _fake_call([_classify_data("bank_statement"), page1, page2])
        result = ocr_bridge.convert_images([b"p1", b"p2"], "stmt.pdf", provider_call=call)
        self.assertFalse(result.conserved)
        self.assertEqual(result.issues[0].kind, ISSUE_GL_BALANCE_CHAIN)
        self.assertEqual(result.issues[0].line_no, 2)


class AdapterTests(unittest.TestCase):
    def test_amount_strings_become_decimal_no_float(self):
        page = _bank_page(
            [_bank_entry("2026-01-01", "in", "1,234.56", "", "1,234.56")], opening="0.00"
        )
        call = _fake_call([_classify_data("bank_statement"), page])
        result = ocr_bridge.convert_images([b"img"], "stmt.jpg", provider_call=call)
        debit = result.tables[0].rows[0][5]
        self.assertIsInstance(debit, Decimal)
        self.assertEqual(debit, Decimal("1234.56"))

    def test_gl_entries_map_debit_credit_and_account(self):
        gl = _Outcome(
            True,
            {
                "document_type": "general_ledger",
                "opening_balance": "0.00",
                "entries": [
                    {
                        "transaction_date": "2026-01-05",
                        "voucher_no": "JV001",
                        "account_code": "1113-01",
                        "description": "pay",
                        "debit": "50.00",
                        "credit": "",
                        "balance": "50.00",
                    }
                ],
            },
        )
        call = _fake_call([_classify_data("general_ledger"), gl])
        result = ocr_bridge.convert_images([b"img"], "gl.jpg", provider_call=call)
        self.assertEqual(result.doc_type, GL_LEDGER)
        self.assertTrue(result.conserved)
        row = result.tables[0].rows[0]
        self.assertEqual(row[0], "1113-01")
        self.assertEqual(row[3], "JV001")

    def test_parenthesised_negative_amount(self):
        self.assertEqual(ocr_bridge._dec("(1,500.00)"), Decimal("-1500.00"))
        self.assertIsNone(ocr_bridge._dec(""))
        self.assertIsNone(ocr_bridge._dec("N/A"))


class ClosingAnchorTests(unittest.TestCase):
    def test_printed_closing_mismatch_is_flagged(self):
        """尾行被"截"但链内自洽 → 印刷期末余额这道独立锚点名。"""
        page = _bank_page(
            [_bank_entry("2026-01-01", "in", "100.00", "", "1,100.00")],
            opening="1,000.00",
            closing="1,200.00",  # 真期末:说明还有行没读到
        )
        call = _fake_call([_classify_data("bank_statement"), page])
        result = ocr_bridge.convert_images([b"img"], "stmt.jpg", provider_call=call)
        self.assertFalse(result.conserved)
        kinds = {i.kind for i in result.issues}
        self.assertIn(ISSUE_CLOSING_ANCHOR, kinds)


class ClassifyAndGenericTests(unittest.TestCase):
    def test_unknown_type_falls_to_generic_grid(self):
        table = _Outcome(
            True,
            {
                "document_type": "generic_table",
                "headers": ["a", "b"],
                "rows": [{"a": "1", "b": "2"}],
            },
        )
        call = _fake_call([_classify_data("mystery"), table])
        result = ocr_bridge.convert_images([b"img"], "x.jpg", provider_call=call)
        self.assertEqual(result.doc_type, GENERIC_TABLE)
        self.assertEqual(result.status, STATUS_OK)
        self.assertEqual(result.tables[0].rows, [["1", "2"]])

    def test_classify_failure_is_generic_not_crash(self):
        table = _Outcome(True, {"document_type": "generic_table", "headers": [], "rows": []})
        call = _fake_call([_Outcome(False, error_kind="timeout"), table])
        result = ocr_bridge.convert_images([b"img"], "x.jpg", provider_call=call)
        self.assertEqual(result.doc_type, GENERIC_TABLE)

    def test_empty_images_rejected(self):
        result = ocr_bridge.convert_images([], "x.jpg", provider_call=_fake_call([]))
        self.assertEqual(result.status, STATUS_OCR_UNAVAILABLE)


class AttributionTests(unittest.TestCase):
    def test_default_call_sets_and_resets_fileconv_attribution(self):
        from services.ai_gateway import attribution

        seen = {}

        def fake_transport(prompt, images, **kwargs):
            seen["attr"] = attribution.current()
            seen["kwargs"] = kwargs
            return _Outcome(True, {})

        with mock.patch("services.ai_gateway.transport.multimodal_to_json", fake_transport):
            ocr_bridge._default_provider_call(
                "p", b"\x89PNG\r\n\x1a\nxx", tenant_id="t-9", api_key=None, max_tokens=64
            )
        self.assertEqual(seen["attr"]["task"], "fileconv_ocr")
        self.assertEqual(seen["attr"]["tenant_id"], "t-9")
        self.assertEqual(seen["kwargs"]["task"], "fileconv_ocr")
        self.assertEqual(seen["kwargs"]["tenant_id"], "t-9")
        self.assertEqual(seen["kwargs"]["tier"], "flash_lite")
        self.assertIsNone(attribution.current())  # finally 里必 reset,不泄漏到后续调用


class ConvertPdfSeamTests(unittest.TestCase):
    def test_no_text_layer_pdf_routes_to_ocr_bridge(self):
        from pypdf import PdfWriter

        from services.fileconv import convert as convert_mod

        writer = PdfWriter()
        writer.add_blank_page(width=200, height=200)
        buf = io.BytesIO()
        writer.write(buf)

        sentinel = ocr_bridge._reject(STATUS_OCR_UNAVAILABLE, "blank.pdf", "x")
        with mock.patch.object(ocr_bridge, "convert_scanned_pdf", return_value=sentinel) as m:
            result = convert_mod.convert_pdf(buf.getvalue(), "blank.pdf", tenant_id="t-1")
        self.assertIs(result, sentinel)
        m.assert_called_once()
        self.assertEqual(m.call_args.kwargs.get("tenant_id"), "t-1")

    def test_rasterize_failure_is_honest_reject(self):
        result = ocr_bridge.convert_scanned_pdf(b"not-a-pdf", "junk.pdf")
        self.assertEqual(result.status, STATUS_OCR_UNAVAILABLE)


if __name__ == "__main__":
    unittest.main()
