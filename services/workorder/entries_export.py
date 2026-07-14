# -*- coding: utf-8 -*-
"""工单影子分录 → Express 导出 xlsx(M1-3KEY 键二 · 只读派生)。

reconcile 步已把工单裁成配平的影子分录(gates.r5_shadow · coa_preset 27 码)。本模块把每条
分录的 coa 码经 coa_erp_bridge(erp_type='express')翻成 Express 实际科目码,渲染成会计可人工
导入 Express 的分录表。桥缺码如实留空 + 标 unmapped(F2 惯例,禁臆造码);桥整个未配置则全列
留空 + 首行提示「科目桥未配置」。

纯读侧:金额取影子原值(Decimal 定点串,不重算一个钱字段)。直推 Express 不在此列(T4b GL
导入通路等料),本件只出人工可导入的文件。
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

_UNMAPPED = "unmapped"

# 列定义:(表头泰中双语, 行字段名)。借/贷分列是 GL 导入通行式,dr_cr 化进两列不再单列。
_COLUMNS = (
    ("ที่มา (来源)", "source"),
    ("รหัส Express (Express 科目)", "express_code"),
    ("รหัส COA (影子科目)", "coa_code"),
    ("ชื่อบัญชี (科目名)", "account_name"),
    ("เดบิต (借方)", "debit"),
    ("เครดิต (贷方)", "credit"),
    ("กฎ (规则)", "rule_key"),
    ("หมายเหตุ (标注)", "note"),
)
_AMOUNT_FIELDS = frozenset({"debit", "credit"})

_HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_UNMAPPED_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_TOTAL_FONT = Font(bold=True)
_RIGHT = Alignment(horizontal="right")

_BRIDGE_MISSING_NOTE = (
    "ยังไม่ได้ตั้งค่าสะพานผังบัญชี Express — คอลัมน์รหัส Express ว่างไว้ กรุณาเติมรหัสเอง "
    "(科目桥未配置 · Express 科目码列留空,请人工补码)"
)


def _dec(value) -> Decimal:
    try:
        return Decimal(str(value))
    except (TypeError, ValueError, ArithmeticError):
        return Decimal("0")


def build_entry_rows(shadow: dict, bridge: dict) -> list[dict]:
    """影子分录 → 导出行(纯函数)。coa 码经桥翻 Express 码;桥缺该码→express_code 留空、
    note 标 unmapped(不臆造)。金额按借/贷分列,Decimal 保精度。"""
    rows = []
    for entry in shadow.get("entries") or []:
        coa = (entry.get("account_code") or "").strip()
        mapped = coa in bridge
        amount = _dec(entry.get("amount"))
        is_debit = entry.get("dr_cr") == "debit"
        rows.append(
            {
                "source": entry.get("source") or "",
                "express_code": bridge.get(coa, ""),
                "coa_code": coa,
                "account_name": entry.get("account_name") or "",
                "debit": amount if is_debit else None,
                "credit": amount if not is_debit else None,
                "rule_key": entry.get("rule_key") or "",
                "note": entry.get("memo") or "",
                "unmapped": not mapped,
            }
        )
    return rows


def _set_cell(ws, r_idx: int, c_idx: int, field: str, row: dict):
    value = row.get(field)
    cell = ws.cell(row=r_idx, column=c_idx)
    if field == "note" and row.get("unmapped"):
        value = _UNMAPPED
    if field in _AMOUNT_FIELDS and isinstance(value, Decimal):
        cell.value = float(value)
        cell.number_format = "#,##0.00"
        cell.alignment = _RIGHT
    else:
        cell.value = "" if value is None else value
    if row.get("unmapped"):
        cell.fill = _UNMAPPED_FILL


def _write_header(ws, r_idx: int) -> None:
    for c_idx, (label, _field) in enumerate(_COLUMNS, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=label)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(c_idx)].width = 20
    ws.freeze_panes = ws.cell(row=r_idx + 1, column=1).coordinate


def _write_totals(ws, r_idx: int, trial_balance: dict) -> None:
    ws.cell(row=r_idx, column=1, value="รวม (合计)").font = _TOTAL_FONT
    for field in ("debit", "credit"):
        c_idx = next(i for i, (_l, f) in enumerate(_COLUMNS, start=1) if f == field)
        cell = ws.cell(row=r_idx, column=c_idx, value=float(_dec(trial_balance.get(field))))
        cell.number_format = "#,##0.00"
        cell.font = _TOTAL_FONT
        cell.alignment = _RIGHT
    balanced = bool(trial_balance.get("balanced"))
    note_idx = next(i for i, (_l, f) in enumerate(_COLUMNS, start=1) if f == "note")
    ws.cell(
        row=r_idx,
        column=note_idx,
        value="สมดุล (配平)" if balanced else "ไม่สมดุล (不平)",
    ).font = _TOTAL_FONT


def build_entries_xlsx(
    shadow: dict, bridge: dict, *, bridge_configured: bool, period: str, client_name: str
) -> bytes:
    """影子分录 → Express 分录 xlsx 字节。bridge_configured=False(本账套未建 express 桥)时
    首行提示 + Express 码列全空;逐码缺映射(桥有但无此码)行标 unmapped。"""
    rows = build_entry_rows(shadow, bridge)
    wb = Workbook()
    ws = wb.active
    ws.title = "Express Entries"

    header_row = 1
    if not bridge_configured:
        ws.cell(row=1, column=1, value=_BRIDGE_MISSING_NOTE).font = _TOTAL_FONT
        header_row = 2
    _write_header(ws, header_row)

    r_idx = header_row + 1
    for row in rows:
        for c_idx, (_label, field) in enumerate(_COLUMNS, start=1):
            _set_cell(ws, r_idx, c_idx, field, row)
        r_idx += 1
    _write_totals(ws, r_idx, shadow.get("trial_balance") or {})

    _write_meta_sheet(wb, period=period, client_name=client_name, row_count=len(rows))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_meta_sheet(wb: Workbook, *, period: str, client_name: str, row_count: int) -> None:
    ws = wb.create_sheet("ข้อมูล (信息)")
    for r_idx, (key, value) in enumerate(
        (
            ("ลูกค้า (客户)", client_name or "-"),
            ("งวด (账期)", period or "-"),
            ("จำนวนรายการ (分录数)", row_count),
            ("รูปแบบ (格式)", "Express journal (人工导入用)"),
        ),
        start=1,
    ):
        ws.cell(row=r_idx, column=1, value=key).font = _TOTAL_FONT
        ws.cell(row=r_idx, column=2, value=value)
