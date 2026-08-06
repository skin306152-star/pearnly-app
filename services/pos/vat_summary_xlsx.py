# -*- coding: utf-8 -*-
"""POS 销项月度汇总 xlsx 排版(POS 项目 · G3)。

v1 拍板:导出文件语言跟请求方 UI 语言走太重,sheet 名/表头固定泰语为主 + 英文括注(事务所
是受众),不接 4 语 i18n——集中在本文件的 _LABELS,不许散落到各 sheet 函数里各写一份。
0 聚合逻辑,只读 services/pos/vat_summary.month_summary 的产出排版,金额口径与那份口径同源。
"""

from __future__ import annotations

import io
from decimal import Decimal
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.pos import sheets_labels

_BORDER = Border(*(Side(style="thin", color="D8DEE8"),) * 4)
_HEAD_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
_HEAD_FONT = Font(name="Tahoma", size=10.5, bold=True, color="FFFFFF")
_CELL_FONT = Font(name="Tahoma", size=10)
_TOTAL_FONT = Font(name="Tahoma", size=10, bold=True)
_NOTE_FONT = Font(name="Tahoma", size=9.5, italic=True)
_RIGHT = Alignment(horizontal="right", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_MONEY_FMT = "#,##0.00"

_LABELS = {
    "sheet_days": "สรุปรายวัน (Daily Summary)",
    "sheet_method": "วิธีชำระเงิน (By Payment)",
    "sheet_invoices": "ใบกำกับภาษีเต็มรูป (Full Inv)",
    "sheet_abb": "เลขที่ ABB (Receipt Range)",
    "col_date": "วันที่ (Date)",
    "col_count": "จำนวนบิล (Count)",
    "col_subtotal": "ยอดก่อนภาษี (Subtotal)",
    "col_discount": "ส่วนลด (Discount)",
    "col_vat": "ภาษีขาย (VAT)",
    "col_gross": "ยอดรวม (Gross)",
    "col_method": "วิธีชำระเงิน (Method)",
    "col_amount": "ยอดเงิน (Amount)",
    "col_doc_number": "เลขที่ใบกำกับภาษี (Doc No.)",
    "col_issued_date": "วันที่ออก (Issue Date)",
    "col_source_receipt": "เลขที่ใบเสร็จเดิม (Source Receipt)",
    "col_buyer_name": "ชื่อผู้ซื้อ (Buyer)",
    "col_buyer_tax_id": "เลขผู้เสียภาษีผู้ซื้อ (Buyer Tax ID)",
    "col_receipt_min": "เลขที่เริ่ม (From)",
    "col_receipt_max": "เลขที่สุด (To)",
    "row_total": "รวม (Total)",
}

# 泰语 + 中文双行(事务所与本地会计各读各的一行),3 条口径都是 vat_summary.py 头注释的原话。
_NOTES = (
    ("ยอดเงินรวมจากรายการขายหน้าร้าน (POS) เท่านั้น", "金额仅按 POS 收银流水聚合。"),
    (
        "บิลที่อัปเกรดเป็นใบกำกับภาษีเต็มรูปแล้วไม่ถูกนับซ้ำ (ดูภาคผนวกใบกำกับภาษีเต็มรูป)",
        "已升级为全式税票的小票不重复计入(见全式税票附录)。",
    ),
    ("ตัดรอบวันตามเวลากรุงเทพ (Asia/Bangkok)", "日切按曼谷时区(Asia/Bangkok)。"),
)


def build_xlsx(data: dict[str, Any]) -> bytes:
    wb = Workbook()
    _days_sheet(wb.active, data)
    _method_sheet(wb.create_sheet(), data)
    _invoices_sheet(wb.create_sheet(), data)
    _abb_sheet(wb.create_sheet(), data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _header(ws, labels: list[str], widths: list[int]) -> None:
    for i, (label, width) in enumerate(zip(labels, widths), start=1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = _HEAD_FONT
        cell.fill = _HEAD_FILL
        cell.alignment = _LEFT
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"


def _row(ws, r: int, values: list, *, money_cols: frozenset = frozenset(), font=_CELL_FONT) -> None:
    for c, v in enumerate(values, start=1):
        money = c in money_cols
        cell = ws.cell(row=r, column=c, value=float(v) if money else v)
        cell.font = font
        cell.border = _BORDER
        cell.alignment = _RIGHT if money else _LEFT
        if money:
            cell.number_format = _MONEY_FMT


def _append_notes(ws, start_row: int) -> None:
    r = start_row + 1
    for th, zh in _NOTES:
        ws.cell(row=r, column=1, value=th).font = _NOTE_FONT
        r += 1
        ws.cell(row=r, column=1, value=zh).font = _NOTE_FONT
        r += 1


def _days_sheet(ws, data: dict) -> None:
    ws.title = _LABELS["sheet_days"]
    _header(
        ws,
        [
            _LABELS[k]
            for k in (
                "col_date",
                "col_count",
                "col_subtotal",
                "col_discount",
                "col_vat",
                "col_gross",
            )
        ],
        [12, 10, 14, 12, 12, 14],
    )
    r = 1
    for day in data["days"]:
        r += 1
        _row(
            ws,
            r,
            [
                day["date"],
                day["sales_count"],
                day["subtotal"],
                day["discount_total"],
                day["vat_amount"],
                day["gross"],
            ],
            money_cols=frozenset({3, 4, 5, 6}),
        )
    totals = data["totals"]
    r += 1
    _row(
        ws,
        r,
        [
            _LABELS["row_total"],
            totals["sales_count"],
            totals["subtotal"],
            totals["discount_total"],
            totals["vat_amount"],
            totals["gross"],
        ],
        money_cols=frozenset({3, 4, 5, 6}),
        font=_TOTAL_FONT,
    )
    _append_notes(ws, r)


def _method_sheet(ws, data: dict) -> None:
    ws.title = _LABELS["sheet_method"]
    _header(ws, [_LABELS[k] for k in ("col_method", "col_count", "col_amount")], [22, 12, 14])
    r = 1
    total_count, total_amount = 0, Decimal("0")
    for method, v in data["by_method"].items():
        r += 1
        total_count += v["count"]
        total_amount += Decimal(str(v["amount"]))
        _row(
            ws,
            r,
            [sheets_labels.method_label(method, "th"), v["count"], v["amount"]],
            money_cols=frozenset({3}),
        )
    r += 1
    _row(
        ws,
        r,
        [_LABELS["row_total"], total_count, f"{total_amount:.2f}"],
        money_cols=frozenset({3}),
        font=_TOTAL_FONT,
    )


def _invoices_sheet(ws, data: dict) -> None:
    ws.title = _LABELS["sheet_invoices"]
    _header(
        ws,
        [
            _LABELS[k]
            for k in (
                "col_doc_number",
                "col_issued_date",
                "col_source_receipt",
                "col_buyer_name",
                "col_buyer_tax_id",
                "col_subtotal",
                "col_discount",
                "col_vat",
                "col_gross",
            )
        ],
        [20, 12, 20, 26, 16, 14, 12, 12, 14],
    )
    r = 1
    for inv in data["full_invoices"]:
        r += 1
        _row(
            ws,
            r,
            [
                inv["doc_number"],
                inv["issued_date"] or "",
                inv["source_receipt_no"] or "",
                inv["buyer_name"] or "",
                inv["buyer_tax_id"] or "",
                inv["subtotal"],
                inv["discount_total"],
                inv["vat_amount"],
                inv["gross"],
            ],
            money_cols=frozenset({6, 7, 8, 9}),
        )


def _abb_sheet(ws, data: dict) -> None:
    ws.title = _LABELS["sheet_abb"]
    _header(
        ws,
        [_LABELS[k] for k in ("col_date", "col_receipt_min", "col_receipt_max", "col_count")],
        [12, 20, 20, 10],
    )
    r = 1
    total_count = 0
    for day in data["abb_ranges"]:
        r += 1
        total_count += day["count"]
        _row(ws, r, [day["date"], day["receipt_min"], day["receipt_max"], day["count"]])
    r += 1
    _row(ws, r, [_LABELS["row_total"], "", "", total_count], font=_TOTAL_FONT)
    _append_notes(ws, r)
