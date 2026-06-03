# -*- coding: utf-8 -*-
"""REFACTOR-D2 守门 · 使用明细报告 usage_report.py 行为契约

usage_report.py(按用户分组的用量报告 · PDF/XLSX · 4 语)此前 0 专属测试。
只测纯/确定性部分(分组 + i18n + XLSX 生成)· 跳过 reportlab PDF 路径(重依赖/字体)。
无 DB/网络/凭证(Wave 0 安全网 · 零冲突)。
"""

import io
import unittest

from openpyxl import load_workbook

from services.usage import usage_report as ur


class I18nTests(unittest.TestCase):
    def test_four_langs_same_keys_nonempty(self):
        langs = ("th", "zh", "en", "ja")
        for lang in langs:
            self.assertIn(lang, ur._I18N)
        base_keys = set(ur._I18N["en"])
        for lang in langs:
            self.assertEqual(set(ur._I18N[lang]), base_keys, f"{lang} 键不齐")
            for k, v in ur._I18N[lang].items():
                self.assertTrue(v, f"{lang}.{k} 空")

    def test_t_accessor(self):
        self.assertEqual(ur._t("zh", "total"), "总计")
        self.assertEqual(ur._t("fr", "total"), ur._I18N["en"]["total"])  # 未知语言 → en
        self.assertEqual(ur._t("en", "no_such_key"), "no_such_key")  # 未知键 → key


class GroupRowsTests(unittest.TestCase):
    def test_groups_by_user_and_sums(self):
        rows = [
            {"user_id": "u1", "user_name": "Alice", "pages": 3, "cost_thb": 1.5},
            {"user_id": "u1", "user_name": "Alice", "pages": 2, "cost_thb": 1.0},
            {"user_id": "u2", "user_name": "Bob", "pages": 5, "cost_thb": 2.5},
        ]
        groups = ur._group_rows(rows)
        self.assertEqual(len(groups), 2)
        g1 = groups[0]
        self.assertEqual(g1["user_label"], "Alice")
        self.assertEqual(g1["pages_sum"], 5)
        self.assertEqual(g1["cost_sum"], 2.5)
        self.assertEqual(len(g1["rows"]), 2)

    def test_preserves_first_seen_order(self):
        rows = [
            {"user_id": "z", "pages": 1},
            {"user_id": "a", "pages": 1},
        ]
        self.assertEqual([g["rows"][0]["user_id"] for g in ur._group_rows(rows)], ["z", "a"])

    def test_user_label_fallback_to_email_prefix(self):
        groups = ur._group_rows([{"user_id": "u", "user_email": "carol@x.com", "pages": 1}])
        self.assertEqual(groups[0]["user_label"], "carol")


class BuildXlsxTests(unittest.TestCase):
    def _rows(self):
        return [
            {
                "user_id": "u1",
                "user_email": "a@b.com",
                "user_name": "Alice",
                "date": "2026-03-15",
                "filename": "a.pdf",
                "pages": 3,
                "cost_thb": 1.5,
            }
        ]

    def test_produces_loadable_xlsx_with_data(self):
        raw = ur.build_xlsx(
            lang="en",
            company="X Co",
            start_date="2026-03-01",
            end_date="2026-03-31",
            rows=self._rows(),
        )
        self.assertIsInstance(raw, bytes)
        ws = load_workbook(io.BytesIO(raw)).active
        self.assertEqual(ws.title, "Usage")
        # 文件名应出现在某个单元格
        seen = {c.value for r in ws.iter_rows() for c in r}
        self.assertIn("a.pdf", seen)

    def test_empty_rows_shows_empty_message(self):
        raw = ur.build_xlsx(
            lang="zh",
            company="X",
            start_date="2026-03-01",
            end_date="2026-03-31",
            rows=[],
        )
        ws = load_workbook(io.BytesIO(raw)).active
        seen = {c.value for r in ws.iter_rows() for c in r}
        self.assertIn(ur._t("zh", "empty"), seen)

    def test_unknown_lang_falls_back(self):
        raw = ur.build_xlsx(lang="fr", company="X", start_date="a", end_date="b", rows=self._rows())
        self.assertIsInstance(raw, bytes)


if __name__ == "__main__":
    unittest.main()
