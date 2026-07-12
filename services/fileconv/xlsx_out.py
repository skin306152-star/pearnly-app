# -*- coding: utf-8 -*-
"""ConvertResult → xlsx。数据表逐张成 sheet,守恒校验结果单独一张 sheet 点名。

金额单元格保留 Decimal 数值 + 千分位格式,便于用户在 Excel 里再核算。
"""

import io
from decimal import Decimal
from typing import Any, List

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from services.fileconv.model import ConvertResult, STATUS_NO_TEXT_LAYER

_HEADER_FILL = PatternFill(start_color="2C5282", end_color="2C5282", fill_type="solid")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_ISSUE_FILL = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
_RIGHT = Alignment(horizontal="right")


def _write_header(ws, columns: List[str]) -> None:
    for col_idx, name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        ws.column_dimensions[get_column_letter(col_idx)].width = 18
    ws.freeze_panes = "A2"


def _write_value(cell, value: Any) -> None:
    if isinstance(value, Decimal):
        cell.value = float(value)
        cell.number_format = "#,##0.00"
        cell.alignment = _RIGHT
    else:
        cell.value = "" if value is None else value


def build_xlsx(result: ConvertResult) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    if result.status == STATUS_NO_TEXT_LAYER:
        ws = wb.create_sheet("Rejected")
        ws["A1"] = "no_text_layer"
        ws["A2"] = "此 PDF 无文字层(疑扫描件),本引擎不做 OCR。请走 OCR 通道。"
    else:
        for i, table in enumerate(result.tables, start=1):
            title = table.name[:28] or f"Sheet{i}"
            ws = wb.create_sheet(title)
            _write_header(ws, table.columns)
            for r_idx, row in enumerate(table.rows, start=2):
                for c_idx, value in enumerate(row, start=1):
                    _write_value(ws.cell(row=r_idx, column=c_idx), value)

    _write_issue_sheet(wb, result)
    _write_summary_sheet(wb, result)

    if not wb.sheetnames:
        wb.create_sheet("Empty")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_issue_sheet(wb: Workbook, result: ConvertResult) -> None:
    ws = wb.create_sheet("Issues")
    _write_header(ws, ["kind", "line", "account", "message", "expected", "actual"])
    if not result.issues:
        ws["A2"] = "OK"
        ws["B2"] = "守恒校验全过 · no discrepancies"
        return
    for r_idx, issue in enumerate(result.issues, start=2):
        for c_idx, value in enumerate(
            [issue.kind, issue.line_no, issue.account, issue.message, issue.expected, issue.actual],
            start=1,
        ):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = _ISSUE_FILL


def _write_summary_sheet(wb: Workbook, result: ConvertResult) -> None:
    ws = wb.create_sheet("Summary")
    ws["A1"] = "field"
    ws["B1"] = "value"
    ws["A1"].font = _HEADER_FONT
    ws["B1"].font = _HEADER_FONT
    ws["A1"].fill = _HEADER_FILL
    ws["B1"].fill = _HEADER_FILL
    fields = [
        ("doc_type", result.doc_type),
        ("status", result.status),
        ("source", result.source_name),
        ("conserved", result.conserved),
        ("issue_count", len(result.issues)),
    ]
    for key, value in result.stats.items():
        fields.append((key, ", ".join(value) if isinstance(value, list) else value))
    for r_idx, (key, value) in enumerate(fields, start=2):
        ws.cell(row=r_idx, column=1, value=key)
        ws.cell(row=r_idx, column=2, value=value)
