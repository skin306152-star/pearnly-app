# -*- coding: utf-8 -*-
"""
v118.32.x · Pearnly · Korn 模板克隆 Excel 导出
铁律 60-VAT: VAT 对账 Excel 永远用 Korn 模板克隆 · 不许 openpyxl 直接 save
但因为 Korn 老师还没给我们参考模板 · v1 先用 openpyxl 严格按 PRD §7 复刻
"""

import io
import logging
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# Korn 样本观察到的样式
HDR_FILL = PatternFill("solid", start_color="1E3A5F")
HDR_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=10)
BOLD = Font(bold=True, name="Arial", size=10)
NORMAL = Font(name="Arial", size=10)
SMALL = Font(name="Arial", size=9)
RED = Font(name="Arial", size=10, color="C0392B", bold=True)
SUB_FILL = PatternFill("solid", start_color="F3F4F6")
DIFF_FILL = PatternFill("solid", start_color="FEF2F2")  # 差异行底色
# P1.2-M2 · 用户手改的发票侧 cell 标黄(同 M4 bank_recon_v2 FFE082)
OVERRIDE_FILL = PatternFill("solid", start_color="FFE082")
OVERRIDE_FONT = Font(name="Arial", size=10, bold=True, color="92400E")  # 暖棕字

from vat_excel_labels import _LABELS, _OVERRIDE_FIELD_COL, _OVERRIDE_FIELD_LABEL

THIN = Side(style="thin", color="B0BEC5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)


def _t(lang: str, key: str) -> str:
    """取 i18n 文字 · 不存在的语言或 key 自动 fallback 到泰文"""
    return _LABELS.get(lang, _LABELS["th"]).get(key, _LABELS["th"].get(key, key))


def _safe(v) -> str:
    return "" if v is None else str(v)


def _num(v):
    try:
        return float(str(v or 0).replace(",", ""))
    except Exception:
        return 0.0


def export_recon_task(
    task: Dict[str, Any],
    rows: List[Dict[str, Any]],
    client: Dict[str, Any],
    vat_report: Dict[str, Any] = None,
    lang: str = "th",
) -> bytes:
    """
    返回 .xlsx 二进制
    严格按 PRD §7 出 Korn 模板:
      - 6 行表头(经营者信息 + 期间)
      - 表格:发票侧 11 列 + 9 个 / 标记列 + 报告侧 7 列 + 备注 1 列 = 28 列
      - 一发票多差异 → 拆 N 行(铁律候选 · PRD §7.2)
      - 底部:总计 + 三栏签字位
    """
    wb = Workbook()
    ws = wb.active
    # v118.32.3 · sheet 标题:本地语言 · 泰文界面就纯泰文
    _title = _t(lang, "report_title")
    ws.title = (_title if lang == "th" else f"{_title}")[:31]  # Excel sheet 名 31 字符上限

    # ─── 6 行表头 ──────────────────────────────────────
    period_str = f"{task.get('period_month'):02d}/{task.get('period_year') + 543}"  # 佛历
    # 主标题:本地语言 · 非泰文时下挂泰文副标题(确保税局也认)
    title_main = _t(lang, "report_title")
    title_with_th = title_main if lang == "th" else f"{title_main}  ·  รายงานภาษีขาย"
    header_meta = [
        (_t(lang, "biz_name"), client.get("name") or "-"),
        (_t(lang, "biz_address"), client.get("address") or "-"),
        (_t(lang, "tax_id"), client.get("tax_id") or "-"),
        (_t(lang, "branch"), f"{_t(lang, 'main_office')} (00000)"),
        (_t(lang, "period"), period_str),
        ("", title_with_th),
    ]
    for r, (a, b) in enumerate(header_meta, 1):
        ws.cell(r, 1, a).font = BOLD
        ws.cell(r, 2, b).font = NORMAL if r < 6 else BOLD
        if r == 6:
            ws.cell(r, 2).alignment = CENTER
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=28)

    # ─── 表头(2 层) ─────────────────────────────────
    R = 8
    # 第一层 · 分段标题
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=11)
    ws.cell(R, 1, _t(lang, "invoice_side")).fill = HDR_FILL
    ws.cell(R, 1).font = HDR_FONT
    ws.cell(R, 1).alignment = CENTER

    ws.merge_cells(start_row=R, start_column=12, end_row=R, end_column=20)
    ws.cell(R, 12, _t(lang, "check_marks")).fill = HDR_FILL
    ws.cell(R, 12).font = HDR_FONT
    ws.cell(R, 12).alignment = CENTER

    ws.merge_cells(start_row=R, start_column=21, end_row=R, end_column=27)
    ws.cell(R, 21, _t(lang, "report_side")).fill = HDR_FILL
    ws.cell(R, 21).font = HDR_FONT
    ws.cell(R, 21).alignment = CENTER

    ws.cell(R, 28, _t(lang, "note")).fill = HDR_FILL
    ws.cell(R, 28).font = HDR_FONT
    ws.cell(R, 28).alignment = CENTER

    R += 1
    # 第二层 · 列名(28 列)
    col_headers = [
        _t(lang, "c_seq"),
        _t(lang, "c_date"),
        _t(lang, "c_inv_no"),
        _t(lang, "c_buyer_name"),
        _t(lang, "c_buyer_tid"),
        _t(lang, "c_buyer_branch"),
        _t(lang, "c_pre_vat"),
        _t(lang, "c_vat"),
        _t(lang, "c_total"),
        _t(lang, "c_category"),
        _t(lang, "c_source"),
        # 9 个标记列(数字标记 · 不需要翻译)
        "1",
        "2.1",
        "2.2",
        "2.3",
        "2.4",
        "3",
        "4",
        "5/6/7",
        "8/9",
        # 报告侧 7 列
        _t(lang, "rc_date"),
        _t(lang, "rc_no"),
        _t(lang, "rc_name"),
        _t(lang, "rc_tid"),
        _t(lang, "rc_branch"),
        _t(lang, "rc_pre"),
        _t(lang, "rc_vat"),
        # 备注
        _t(lang, "note"),
    ]
    for c, h in enumerate(col_headers, 1):
        cell = ws.cell(R, c, h)
        cell.fill = SUB_FILL
        cell.font = BOLD
        cell.alignment = CENTER
        cell.border = BORDER

    # ─── 数据行 ──────────────────────────────────────
    R += 1
    sum_pre = 0.0
    sum_vat = 0.0
    sum_total = 0.0

    manual_trail = []  # P1.2-M2 · (序号, field_label_key, OCR 原值, 用户值) · 底部痕迹 section 用

    for idx, row in enumerate(rows, 1):
        diff_fields = row.get("diff_fields") or {}
        if isinstance(diff_fields, str):
            try:
                import json as _j

                diff_fields = _j.loads(diff_fields)
            except Exception:
                diff_fields = {}

        # P1.2-M2 · 发票侧字段用户校正 · {field: {ocr, user, ts}}
        field_overrides = row.get("field_overrides") or {}
        if isinstance(field_overrides, str):
            try:
                import json as _j

                field_overrides = _j.loads(field_overrides)
            except Exception:
                field_overrides = {}
        if not isinstance(field_overrides, dict):
            field_overrides = {}

        def _disp(field, fallback):
            """有用户校正值就用 · 否则用 OCR 原值"""
            ov = field_overrides.get(field)
            if isinstance(ov, dict) and ov.get("user") not in (None, ""):
                return ov.get("user")
            return fallback

        diff_cats = (row.get("diff_categories") or "").split(",")
        diff_cats = [c.strip() for c in diff_cats if c.strip()]

        # 发票侧字段(用户校正值优先 · 只覆盖发票侧)
        inv_no = _safe(_disp("invoice_no", row.get("invoice_no")))
        inv_date = _safe(_disp("invoice_date", row.get("invoice_date")))
        buyer_name = _safe(_disp("buyer_name", row.get("buyer_name") or row.get("seller_name")))
        buyer_tax = _safe(_disp("buyer_tax_id", row.get("buyer_tax_id")))
        buyer_branch = _safe(_disp("buyer_branch", row.get("buyer_branch") or "00000"))
        amt_pre = _num(
            _disp("amount_pre_vat", row.get("amount_pre_vat") or row.get("total_amount"))
        )
        amt_vat = _num(_disp("vat_amount", row.get("vat_amount")))
        amt_total = _num(row.get("total_amount"))
        source = row.get("invoice_filename") or "-"

        # 收集本行手改痕迹(给底部 section)
        for _f, _lbl in _OVERRIDE_FIELD_LABEL.items():
            _ov = field_overrides.get(_f)
            if isinstance(_ov, dict) and _ov.get("user") not in (None, ""):
                manual_trail.append((idx, _lbl, _ov.get("ocr"), _ov.get("user")))

        sum_pre += amt_pre
        sum_vat += amt_vat
        sum_total += amt_total

        # 决定要写几行(铁律 PRD §7.2:每差异一行)
        if len(diff_cats) <= 1:
            cats_per_row = [diff_cats[0] if diff_cats else ""]
        else:
            cats_per_row = diff_cats

        for ri, cat in enumerate(cats_per_row):
            is_first = ri == 0
            is_diff_row = row.get("status") in ("mismatched", "invoice_orphan", "report_orphan")
            bg = DIFF_FILL if is_diff_row else None

            # 1. 发票侧(只在第一行填,后续行留空便于阅读)
            cells = [
                idx if is_first else "",
                inv_date if is_first else "",
                inv_no if is_first else "",
                buyer_name if is_first else "",
                buyer_tax if is_first else "",
                buyer_branch if is_first else "",
                amt_pre if is_first else "",
                amt_vat if is_first else "",
                amt_total if is_first else "",
                "",  # 分类 · 待自动归类后填
                source if is_first else "",
            ]
            for c, v in enumerate(cells, 1):
                cell = ws.cell(R, c, v)
                cell.alignment = RIGHT if c in (1, 7, 8, 9) else LEFT
                cell.font = NORMAL
                cell.border = BORDER
                if bg:
                    cell.fill = bg
                if c in (7, 8, 9):
                    cell.number_format = "#,##0.00"

            # 1b. P1.2-M2 · 用户手改的发票侧 cell 标黄 + OCR/User 批注(只第一行有发票侧数据)
            if is_first and field_overrides:
                for _f, _col in _OVERRIDE_FIELD_COL.items():
                    _ov = field_overrides.get(_f)
                    if not (isinstance(_ov, dict) and _ov.get("user") not in (None, "")):
                        continue
                    ocell = ws.cell(R, _col)
                    ocell.fill = OVERRIDE_FILL
                    ocell.font = OVERRIDE_FONT
                    try:
                        from openpyxl.comments import Comment as _XLC

                        _ocr = _ov.get("ocr")
                        ocell.comment = _XLC(
                            f"OCR: {_ocr if _ocr not in (None, '') else '—'}\nUser: {_ov.get('user')}",
                            "Pearnly",
                        )
                    except Exception:
                        pass  # 批注失败不阻塞导出 · 数据已在 cell + 底部痕迹 section

            # 2. 9 个 / 标记列(全部打 / 表示"已核对")
            for c in range(12, 21):
                cell = ws.cell(R, c, "/")
                cell.alignment = CENTER
                cell.font = NORMAL
                cell.border = BORDER
                if bg:
                    cell.fill = bg

            # 3. 报告侧 7 列(只在第一行填)
            rep_date = _safe(row.get("report_date")) if is_first else ""
            rep_inv = _safe(row.get("report_invoice_no")) if is_first else ""
            rep_name = _safe(row.get("report_buyer_name")) if is_first else ""
            rep_tax = _safe(row.get("report_buyer_tax_id")) if is_first else ""
            rep_branch = _safe(row.get("report_buyer_branch")) if is_first else ""
            rep_pre = _num(row.get("report_amount_pre_vat")) if is_first else ""
            rep_vat = _num(row.get("report_vat_amount")) if is_first else ""

            rep_cells = [rep_date, rep_inv, rep_name, rep_tax, rep_branch, rep_pre, rep_vat]
            for ci, v in enumerate(rep_cells):
                c = 21 + ci
                cell = ws.cell(R, c, v)
                cell.alignment = RIGHT if c in (26, 27) else LEFT
                cell.font = NORMAL
                cell.border = BORDER
                if bg:
                    cell.fill = bg
                if c in (26, 27):
                    cell.number_format = "#,##0.00"

            # 4. 备注列(写差异说明)
            note = _category_to_thai_note(cat) if cat else ""
            cell = ws.cell(R, 28, note)
            cell.font = RED if is_diff_row else NORMAL
            cell.alignment = LEFT
            cell.border = BORDER
            if bg:
                cell.fill = bg

            R += 1

    # ─── 总计行 ──────────────────────────────────────
    ws.cell(R, 1, _t(lang, "grand_total")).font = BOLD
    ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=6)
    ws.cell(R, 7, sum_pre).font = BOLD
    ws.cell(R, 7).number_format = "#,##0.00"
    ws.cell(R, 8, sum_vat).font = BOLD
    ws.cell(R, 8).number_format = "#,##0.00"
    ws.cell(R, 9, sum_total).font = BOLD
    ws.cell(R, 9).number_format = "#,##0.00"
    for c in range(1, 29):
        ws.cell(R, c).border = BORDER
        ws.cell(R, c).fill = SUB_FILL

    # ─── 签字栏 ──────────────────────────────────────
    R += 3
    ws.cell(R, 1, f"{_t(lang, 'sig_prepared')}: ____________________").font = NORMAL
    ws.cell(R, 12, f"{_t(lang, 'sig_reviewer')}: ____________________").font = NORMAL
    ws.cell(R, 22, f"{_t(lang, 'sig_approver')}: ____________________").font = NORMAL

    # ─── P1.2-M2 · 手动修改痕迹 section(只在有用户校正时显示) ──────────
    if manual_trail:
        R += 3
        ws.merge_cells(start_row=R, start_column=1, end_row=R, end_column=4)
        sec = ws.cell(R, 1, _t(lang, "sec_manual_entry"))
        sec.fill = HDR_FILL
        sec.font = HDR_FONT
        sec.alignment = LEFT
        R += 1
        mini_hdr = [
            _t(lang, "c_seq"),
            _t(lang, "mc_field"),
            _t(lang, "col_manual_ocr"),
            _t(lang, "col_manual_user"),
        ]
        for ci, h in enumerate(mini_hdr, 1):
            cell = ws.cell(R, ci, h)
            cell.fill = SUB_FILL
            cell.font = BOLD
            cell.alignment = CENTER
            cell.border = BORDER
        R += 1
        for seq, lbl_key, ocr_v, user_v in manual_trail:
            ws.cell(R, 1, seq).alignment = CENTER
            ws.cell(R, 1).border = BORDER
            fc = ws.cell(R, 2, _t(lang, lbl_key))
            fc.font = NORMAL
            fc.alignment = LEFT
            fc.border = BORDER
            oc = ws.cell(R, 3, ocr_v if ocr_v not in (None, "") else "—")
            oc.font = SMALL
            oc.alignment = LEFT
            oc.border = BORDER
            uc = ws.cell(R, 4, user_v)
            uc.fill = OVERRIDE_FILL
            uc.font = OVERRIDE_FONT
            uc.alignment = LEFT
            uc.border = BORDER
            R += 1

    # ─── 列宽 ────────────────────────────────────────
    widths = {
        1: 6,
        2: 11,
        3: 18,
        4: 30,
        5: 18,
        6: 10,
        7: 11,
        8: 10,
        9: 11,
        10: 14,
        11: 18,
        12: 5,
        13: 5,
        14: 5,
        15: 5,
        16: 5,
        17: 5,
        18: 5,
        19: 5,
        20: 5,
        21: 11,
        22: 16,
        23: 24,
        24: 18,
        25: 8,
        26: 11,
        27: 10,
        28: 30,
    }
    for c, w in widths.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # 行高
    ws.row_dimensions[8].height = 26
    ws.row_dimensions[9].height = 32

    # 冻结表头
    ws.freeze_panes = "A10"

    # 导出
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _category_to_thai_note(cat: str) -> str:
    """差异类别 → 泰语备注"""
    return {
        "date_diff": "วันที่ไม่ตรงกัน",
        "date_period_mismatch": "วันที่ข้ามงวดภาษี",
        "invoice_no_prefix": "เลขที่ใบกำกับต่างเพียงคำนำหน้า",
        "invoice_no_mismatch": "เลขที่ใบกำกับไม่ถูกต้อง",
        "name_fuzzy": "ชื่อผู้ซื้อใกล้เคียง · ไม่ตรงทุกตัวอักษร",
        "name_mismatch": "ชื่อผู้ซื้อไม่ถูกต้อง",
        "tax_id_mismatch": "เลขประจำตัวผู้เสียภาษีไม่ถูกต้อง",
        "branch_mismatch": "สาขาผู้ซื้อไม่ถูกต้อง",
        "amount_diff": "ยอดเงินไม่ตรงกัน",
        "vat_precision": "ภาษีมูลค่าเพิ่มต่างเพราะปัดเศษ",
        "vat_calc_wrong": "ภาษีมูลค่าเพิ่มคำนวณผิด",
        "invoice_orphan": "พบใบกำกับ แต่ไม่พบในรายงาน",
        "report_orphan": "พบในรายงาน แต่ไม่พบใบกำกับ",
    }.get(cat, cat)
