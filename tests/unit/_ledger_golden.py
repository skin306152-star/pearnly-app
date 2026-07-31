# -*- coding: utf-8 -*-
"""账簿配方的金标语料 —— 三张真 Ocha 小票(Sister Makeup 2569-05)。

写侧测试和「表里算出来的数对不对」那套测试共用这一份,免得两处各摆一套金标然后漂。

金标(合计 3,649.00 / VAT 238.72 / 净额 3,410.28):
  02000138 合计 489.00(内含 VAT 31.99 / 净 457.01)
  02000139 合计 470.00(30.75 / 439.25)
  02000143 合计 2,690.00(175.98 / 2,514.02)
"""

import io
from decimal import Decimal

from openpyxl import load_workbook

from services.ledger.models import parse_sales_docs
from services.ledger.recipes import sales_books

D = Decimal

GOLDEN_LINES = {
    "02000138": [("Cool Betty Inliner", 1, 290), ("Eyelash Curler", 1, 120), ("Tweezers", 1, 79)],
    "02000139": [("Clinicare", 1, 470)],
    "02000143": [
        ("SM powder(01)", 1, 390),
        ("SM powder(02)", 1, 390),
        ("Content Lens", 1, 150),
        ("Case Contact Lens", 2, 100),
        ("Earring", 1, 120),
        ("Giffarine Foundation 04", 1, 200),
        ("Giffarine Foundation 05", 1, 200),
        ("Foundation Cool Betty", 1, 390),
        ("Giffarine FS32", 1, 300),
        ("Focallure Eye base", 1, 150),
        ("Pallets", 1, 300),
    ],
}
GOLDEN_DOC_TOTALS = {
    "02000138": (D("489.00"), D("31.99"), D("457.01")),
    "02000139": (D("470.00"), D("30.75"), D("439.25")),
    "02000143": (D("2690.00"), D("175.98"), D("2514.02")),
}
GOLDEN_GROSS, GOLDEN_VAT, GOLDEN_NET = D("3649.00"), D("238.72"), D("3410.28")

# 02000139(真件 SM-照片语料/A_客户给的原料/IMG_2574.JPG)票面不印开票日期,只印两行,
# 小写 in/out、带冒号、日期用连字符 —— 按真件的印法写,别用理想化的写法糊弄解析器:
#   Time in:  27-05-2569 16:19
#   Time out: 28-05-2569 08:29
# 是三张里唯一跨日的一张;MR.ERP 销项税登记簿(acvatsaled)把 SW02000139 记在 28/05/2569,
# 与「取 Time out」一致 —— 这是本条口径唯一的真值来源。
GOLDEN_SESSION_RAW = "Time in: 27-05-2569 16:19  Time out: 28-05-2569 08:29"

# ⚠️ 这一格是**合成夹具,不是 OCR 的产出**。2026-07-31 拿这张真件跑真 pipeline 量过:
# 模型只把 Time in 那一行抄进 date_raw("27-05-2569"),notes 给的是 "Thank You Ka~",
# 带标签的场次文本一次都没进来过(12 次运行 0 命中)。所以下面这套断言守的是「文本里
# 真有标签时口径对不对」,**不能**读成「产品线上真会按 Time out 入账」——线上目前仍按
# 27 记,差一天(同属 5 月申报期,故未错期)。原因与已试过的修法见 doc_date.py 文件头。
GOLDEN_SESSION_OCR_DATE_RAW = "27-05-2569"  # 线上真拿到的形状,给「现状诚实」那条测试用


def items(invoice):
    return [{"name": n, "qty": str(q), "subtotal": str(a)} for n, q, a in GOLDEN_LINES[invoice]]


def record(invoice, *, dates, payment="transfer", **extra):
    fields = {
        "invoice_number": invoice,
        "payment_method": payment,
        "buyer_name": "ลูกค้าเงินสด",
        "items": items(invoice),
    }
    fields.update(dates)
    fields.update(extra)
    return {"history_id": f"h-{invoice}", "merged_fields": fields}


def golden_records():
    return [
        record("02000138", dates={"date": "27/05/2569"}),
        record("02000139", dates={"date": "27/05/2569", "date_raw": GOLDEN_SESSION_RAW}),
        record("02000143", dates={"date": "28/05/2569"}),
    ]


def build_golden(**kwargs):
    docs = parse_sales_docs(golden_records())
    return docs, sales_books.build(docs, title="Sister Makeup", period_label="2569-05", **kwargs)


def load(result):
    return load_workbook(io.BytesIO(result.content))


def formulas(wb):
    """(sheet, 坐标, 公式) 全集 —— 公式策略的断言都扫这一份。"""
    out = []
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    out.append((ws.title, cell.coordinate, cell.value))
    return out
