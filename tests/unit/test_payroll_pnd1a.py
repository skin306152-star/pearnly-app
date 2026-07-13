# -*- coding: utf-8 -*-
"""ภ.ง.ด.1ก 年度聚合(批次 H 收尾件 · 合成数据)。

金标口径:H1a 真料 7 员工 / Σ78,270(真身份证不入 git,此处用合成的、通过 mod-11 的假
身份证复现同一「7 员工 Σ78,270」口径)——造两个月同一批员工同额度,断言年度聚合
Σ156,540 且逐员工年度额 = 月度额 ×2。
"""

import io
import unittest
from decimal import Decimal

from openpyxl import load_workbook

from services.payroll import model, pnd1a


def _valid_id(prefix12: str) -> str:
    total = sum(int(prefix12[i]) * (13 - i) for i in range(12))
    return prefix12 + str((11 - total % 11) % 10)


# 7 员工合成号 + 金额(Σ=78270,复现 H1a 金标口径 · 见模块顶注)。
_EMP = [
    _valid_id(p)
    for p in (
        "365010069742",
        "110420000252",
        "110080066735",
        "110100234561",
        "365020011119",
        "110330099982",
        "365040055561",
    )
]
_AMOUNTS = [13000, 12040, 14500, 10000, 9500, 11230, 8000]
assert sum(_AMOUNTS) == 78270


def _row(period, seq, emp_id, amount, *, title="นาย", first="ก", last="ข", wht="0"):
    return {
        "period": period,
        "seq": seq,
        "employee_id": emp_id,
        "title": title,
        "first_name": first,
        "last_name": last,
        "income_code": "40(1)",
        "paid_date": None,
        "paid_amount": Decimal(str(amount)),
        "wht_amount": Decimal(str(wht)),
        "condition": "1",
    }


def _two_month_gold_rows():
    rows = []
    for period in ("2569-05", "2569-06"):
        for seq, (emp_id, amount) in enumerate(zip(_EMP, _AMOUNTS), start=1):
            rows.append(_row(period, seq, emp_id, amount))
    return rows


class AggregateGoldTests(unittest.TestCase):
    def test_two_months_doubles_annual_total(self):
        annual_rows, issues = pnd1a.aggregate_year(_two_month_gold_rows())
        self.assertEqual(issues, [])
        self.assertEqual(len(annual_rows), 7)
        totals = pnd1a.annual_totals(annual_rows)
        self.assertEqual(totals["paid_amount"], Decimal("156540"))

    def test_each_employee_annual_amount_is_double_monthly(self):
        annual_rows, _ = pnd1a.aggregate_year(_two_month_gold_rows())
        by_id = {r.employee_id: r for r in annual_rows}
        for emp_id, monthly_amount in zip(_EMP, _AMOUNTS):
            self.assertEqual(by_id[emp_id].paid_amount, Decimal(str(monthly_amount)) * 2)
            self.assertEqual(by_id[emp_id].months_count, 2)
            self.assertEqual(by_id[emp_id].periods, ("2569-05", "2569-06"))


class AggregateSingleMonthTests(unittest.TestCase):
    def test_single_month_matches_monthly_total(self):
        rows = [_row("2569-05", 1, _EMP[0], 13000)]
        annual_rows, issues = pnd1a.aggregate_year(rows)
        self.assertEqual(issues, [])
        self.assertEqual(len(annual_rows), 1)
        self.assertEqual(annual_rows[0].paid_amount, Decimal("13000"))
        self.assertEqual(annual_rows[0].months_count, 1)


class CrossMonthNameChangeTests(unittest.TestCase):
    def test_name_change_across_months_flagged_and_latest_wins(self):
        rows = [
            _row("2569-05", 1, _EMP[0], 13000, first="สมหญิง", last="เดิม"),
            _row("2569-06", 1, _EMP[0], 13000, first="สมหญิง", last="ใหม่"),
        ]
        annual_rows, issues = pnd1a.aggregate_year(rows)
        kinds = [i.kind for i in issues]
        self.assertIn(model.ISSUE_NAME_MISMATCH, kinds)
        self.assertEqual(annual_rows[0].last_name, "ใหม่")  # 最新月姓名出件
        self.assertEqual(annual_rows[0].paid_amount, Decimal("26000"))  # 仍全额计入,不吞


class BadIdRowTests(unittest.TestCase):
    def test_invalid_id_row_flagged_but_still_summed(self):
        bad_id = list(_EMP[0])
        bad_id[-1] = str((int(bad_id[-1]) + 1) % 10)  # 破坏 mod-11 校验位
        rows = [_row("2569-05", 1, "".join(bad_id), 13000)]
        annual_rows, issues = pnd1a.aggregate_year(rows)
        kinds = [i.kind for i in issues]
        self.assertIn(model.ISSUE_INVALID_ID, kinds)
        self.assertEqual(annual_rows[0].paid_amount, Decimal("13000"))  # 只点名不排除


class ConservationCheckTests(unittest.TestCase):
    """独立单测私有断言函数:直接构造一处不守恒的输入,验证点名具体月份(不吞不静默)。"""

    def test_balanced_input_raises_no_issue(self):
        rows = _two_month_gold_rows()
        annual_rows, _ = pnd1a.aggregate_year(rows)
        self.assertEqual(pnd1a._check_conservation(rows, annual_rows), [])

    def test_mismatch_names_the_offending_period(self):
        rows = [_row("2569-05", 1, _EMP[0], 13000), _row("2569-06", 1, _EMP[0], 12000)]
        annual_rows, _ = pnd1a.aggregate_year(rows)
        # 人为破坏其中一月的聚合值,模拟未来实现 bug 悄悄漏计。
        annual_rows[0].paid_by_period["2569-06"] = Decimal("11000")
        issues = pnd1a._check_conservation(rows, annual_rows)
        self.assertEqual(len(issues), 1)
        self.assertEqual(issues[0].kind, model.ISSUE_YEAR_SUM_MISMATCH)
        self.assertIn("2569-06", issues[0].value)
        self.assertNotIn("2569-05", issues[0].value)


class AnnualKeyingWorkbookTests(unittest.TestCase):
    def test_workbook_total_row_matches_annual_totals(self):
        annual_rows, _ = pnd1a.aggregate_year(_two_month_gold_rows())
        wb = load_workbook(io.BytesIO(pnd1a.build_annual_keying_workbook(annual_rows)))
        ws = wb.active
        total_row = len(annual_rows) + 2
        # 列 5 = 年度Σ支付合计(见 pnd1a._COLUMNS 序)。
        self.assertEqual(Decimal(str(ws.cell(row=total_row, column=5).value)), Decimal("156540.00"))

    def test_filename_marks_auxiliary_not_official(self):
        self.assertEqual(
            pnd1a.build_annual_filename(nid="0105548082417", tax_year_be="2569"),
            "PND1A_0105548082417_2569_keying.xlsx",
        )


if __name__ == "__main__":
    unittest.main()
