# -*- coding: utf-8 -*-
"""把会计改过的复核工作簿读回成可重推的记录(确定性解析 · 不过大模型)。

这是闭环的回程。规格见 erp_roundtrip —— 写侧和这里共用同一份列名/键/Sheet 定义,
格式因此不会漂。

三条设计立场:

1. **方向取自 Sheet 名,不从内容猜。** 会计把一行从「ขาย」剪到「ซื้อ」就是在改分类,
   回导时这就是显式方向,压过税号自动判定。留在「รอจำแนก」的 = 还没裁决,不推。
2. **按列名取值,不按列位。** 会计在中间插一列、挪一列都不会让解析错位。
3. **金额按合同自算。** 销项 G/H/I = 数量×单价、J = 税前×7% 是我们自己定的公式合同,
   openpyxl 写公式不带计算结果缓存,原样导出的文件读回来那几格是空的 —— 故读侧按
   合同重算;会计把公式改成死值时,死值优先(那是他的裁决)。
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from services.excel import erp_roundtrip as rt

logger = logging.getLogger(__name__)

_VAT_RATE = 0.07


class RoundtripParseError(ValueError):
    """不是我们导出的工作簿 —— 调用方应回落通用表格路,不要半解析。"""


def _num(raw: Any) -> Optional[float]:
    """单元格 → 数。公式串(以 = 开头)当"没有死值"处理,交给调用方按合同重算。"""
    if raw is None or isinstance(raw, bool):
        return None
    if isinstance(raw, (int, float)):
        return float(raw)
    s = str(raw).strip().replace(",", "").replace("฿", "")
    if not s or s.startswith("="):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _text(raw: Any) -> str:
    if raw is None:
        return ""
    if hasattr(raw, "strftime"):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    return "" if s.lower() in ("none", "null", "nan") else s


def _header_map(ws) -> Dict[str, int]:
    """表头行 → {列名: 列号}。重名列取最左一列(会计复制列时不至于串位)。"""
    out: Dict[str, int] = {}
    for col in range(1, (ws.max_column or 0) + 1):
        name = _text(ws.cell(row=1, column=col).value)
        if name and name not in out:
            out[name] = col
    return out


def _cell(ws, row: int, hmap: Dict[str, int], header: str) -> Any:
    col = hmap.get(header)
    return ws.cell(row=row, column=col).value if col else None


def _row_is_blank(ws, row: int, hmap: Dict[str, int]) -> bool:
    return not any(_text(_cell(ws, row, hmap, h)) for h in hmap)


def _parse_sales_line(ws, row: int, hmap: Dict[str, int]) -> Dict[str, Any]:
    """销项一行 = 一个商品行。金额缺(公式无缓存)时按合同 数量×单价 补。"""
    qty = _num(_cell(ws, row, hmap, rt.SALES_COL_QTY))
    price = _num(_cell(ws, row, hmap, rt.SALES_COL_PRICE))
    amount = _num(_cell(ws, row, hmap, rt.SALES_COL_AMOUNT))
    if amount is None and qty is not None and price is not None:
        amount = round(qty * price, 2)
    return {
        "description": _text(_cell(ws, row, hmap, rt.SALES_COL_ITEM)),
        "qty": qty,
        "unit_price": price,
        "amount": amount,
        "erp_item_code": _text(_cell(ws, row, hmap, rt.COL_ERP_ITEM)),
    }


def _sales_doc_totals(items: List[Dict[str, Any]], pre_vat_cells: List[Optional[float]]) -> Tuple:
    """单据税基/税额/合计。会计把某行税前改成死值 → 用他的;否则按行金额求和再算 7%。"""
    explicit = [v for v in pre_vat_cells if v is not None]
    base = round(sum(explicit), 2) if explicit else round(sum(i["amount"] or 0 for i in items), 2)
    vat = round(base * _VAT_RATE, 2)
    return base, vat, round(base + vat, 2)


def _read_sales_sheet(ws, hmap: Dict[str, int]) -> List[Dict[str, Any]]:
    """销项表按单据聚合(每行一个商品行)。聚合键优先用回导键里的 history_id ——
    会计改票号是这个流程的主要目的之一,按票号聚合会把改过号的行拆散。"""
    groups: Dict[str, Dict[str, Any]] = {}
    order: List[str] = []
    for row in range(2, (ws.max_row or 1) + 1):
        if _row_is_blank(ws, row, hmap):
            continue
        key_raw = _text(_cell(ws, row, hmap, rt.COL_ROW_KEY))
        decoded = rt.decode_row_key(key_raw)
        invoice_no = _text(_cell(ws, row, hmap, rt.SALES_COL_INVOICE))
        # 键 > 票号 > 行号:三级兜底,任何一行都不会被静默并进别人的单据
        gkey = decoded[0] if decoded else (f"inv:{invoice_no}" if invoice_no else f"row:{row}")
        if gkey not in groups:
            groups[gkey] = {
                "history_id": decoded[0] if decoded else "",
                "invoice_number": invoice_no,
                "date": _text(_cell(ws, row, hmap, rt.SALES_COL_DATE)),
                "buyer_name": _text(_cell(ws, row, hmap, rt.SALES_COL_PARTY)),
                "buyer_tax": _text(_cell(ws, row, hmap, rt.COL_PARTY_TAX)),
                "erp_docnum": _text(_cell(ws, row, hmap, rt.COL_ERP_DOCNUM)),
                "erp_party_code": _text(_cell(ws, row, hmap, rt.COL_ERP_PARTY)),
                "items": [],
                "_pre_vat": [],
                "_rows": [],
            }
            order.append(gkey)
        g = groups[gkey]
        g["items"].append(_parse_sales_line(ws, row, hmap))
        g["_pre_vat"].append(_num(_cell(ws, row, hmap, rt.SALES_COL_PRE_VAT)))
        g["_rows"].append(row)

    out: List[Dict[str, Any]] = []
    for gkey in order:
        g = groups.pop(gkey)
        base, vat, total = _sales_doc_totals(g["items"], g.pop("_pre_vat"))
        rows = g.pop("_rows")
        g.update({"amount_before_vat": base, "vat_amount": vat, "total_amount": total})
        out.append({"fields": g, "source_rows": rows})
    return out


def _read_purchase_sheet(ws, hmap: Dict[str, int]) -> List[Dict[str, Any]]:
    """进项表每行一张票(费用票整张进一个科目 · 不拆商品行)。"""
    out: List[Dict[str, Any]] = []
    for row in range(2, (ws.max_row or 1) + 1):
        if _row_is_blank(ws, row, hmap):
            continue
        decoded = rt.decode_row_key(_text(_cell(ws, row, hmap, rt.COL_ROW_KEY)))
        base = _num(_cell(ws, row, hmap, rt.PURCHASE_COL_PRE_VAT))
        vat = _num(_cell(ws, row, hmap, rt.PURCHASE_COL_VAT))
        total = _num(_cell(ws, row, hmap, rt.PURCHASE_COL_TOTAL))
        if total is None and base is not None:
            total = round(base + (vat if vat is not None else base * _VAT_RATE), 2)
        out.append(
            {
                "fields": {
                    "history_id": decoded[0] if decoded else "",
                    "invoice_number": _text(_cell(ws, row, hmap, rt.PURCHASE_COL_INVOICE)),
                    "date": _text(_cell(ws, row, hmap, rt.PURCHASE_COL_DATE)),
                    "seller_name": _text(_cell(ws, row, hmap, rt.PURCHASE_COL_PARTY)),
                    "seller_tax": (
                        _text(_cell(ws, row, hmap, rt.PURCHASE_COL_PARTY_TAX))
                        or _text(_cell(ws, row, hmap, rt.COL_PARTY_TAX))
                    ),
                    "category": _text(_cell(ws, row, hmap, rt.PURCHASE_COL_CATEGORY)),
                    "amount_before_vat": base,
                    "vat_amount": vat,
                    "total_amount": total,
                    "erp_docnum": _text(_cell(ws, row, hmap, rt.COL_ERP_DOCNUM)),
                    "erp_party_code": _text(_cell(ws, row, hmap, rt.COL_ERP_PARTY)),
                    "items": [],
                },
                "source_rows": [row],
            }
        )
    return out


def _read_pending_sheet(ws, hmap: Dict[str, int]) -> List[Dict[str, Any]]:
    """待判表:会计没裁决的票。原样带出来给调用方如实汇报,不推。"""
    out: List[Dict[str, Any]] = []
    for row in range(2, (ws.max_row or 1) + 1):
        if _row_is_blank(ws, row, hmap):
            continue
        decoded = rt.decode_row_key(_text(_cell(ws, row, hmap, rt.COL_ROW_KEY)))
        out.append(
            {
                "fields": {
                    "history_id": decoded[0] if decoded else "",
                    "invoice_number": _text(_cell(ws, row, hmap, rt.PENDING_COL_INVOICE)),
                    "date": _text(_cell(ws, row, hmap, rt.PENDING_COL_DATE)),
                    "items": [],
                },
                "reason": _text(_cell(ws, row, hmap, rt.PENDING_COL_REASON)),
                "source_rows": [row],
            }
        )
    return out


_SHEET_READERS = {
    rt.SHEET_SALES: _read_sales_sheet,
    rt.SHEET_PURCHASE: _read_purchase_sheet,
    rt.SHEET_PENDING: _read_pending_sheet,
}


def looks_like_roundtrip_workbook(file_bytes: bytes) -> bool:
    """便宜的预判:任一 Sheet 名是我们的数据表且表头带全回导列。用于收料口分流。"""
    try:
        return bool(_open_data_sheets(file_bytes))
    except Exception:  # noqa: BLE001 — 预判失败一律当"不是",回落通用路
        return False


def _open_data_sheets(file_bytes: bytes):
    import openpyxl

    # data_only=False:公式格拿到公式串,借此区分「会计手填的死值」和「我们写的公式」。
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    hits = []
    for name in wb.sheetnames:
        if not rt.is_data_sheet(name):
            continue
        ws = wb[name]
        hmap = _header_map(ws)
        if rt.is_roundtrip_sheet(list(hmap.keys())):
            hits.append((name, ws, hmap))
    return hits


def parse_roundtrip_workbook(file_bytes: bytes) -> Dict[str, Any]:
    """读回工作簿 → {"documents": [...], "pending": [...], "sheets": [...]}。

    documents 每条 = {"direction": "sales"|"purchase", "fields": {...}, "source_rows": [...]},
    direction 由所在 Sheet 决定(= 会计的分类裁决)。pending 是留在待判表、还没裁决的票 ——
    调用方必须如实汇报这些没被处理,不能悄悄跳过。
    """
    hits = _open_data_sheets(file_bytes)
    if not hits:
        raise RoundtripParseError("not a Pearnly ERP roundtrip workbook")

    documents: List[Dict[str, Any]] = []
    pending: List[Dict[str, Any]] = []
    for name, ws, hmap in hits:
        reader = _SHEET_READERS[name]
        parsed = reader(ws, hmap)
        direction = rt.sheet_direction(name)
        if direction is None:
            for p in parsed:
                p["sheet"] = name
            pending.extend(parsed)
            continue
        for d in parsed:
            d["direction"] = direction
            d["sheet"] = name
            # 显式方向落进 fields —— express_push.direction.explicit_direction 认这个键,
            # 优先级高于税号自动判定,会计挪行即改分类由此生效。
            d["fields"]["direction"] = direction
            documents.append(d)

    return {
        "documents": documents,
        "pending": pending,
        "sheets": [name for name, _ws, _h in hits],
    }
