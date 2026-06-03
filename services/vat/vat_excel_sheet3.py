# -*- coding: utf-8 -*-
"""Sheet 3 · 一对一对账结果(配对引擎 + KPI 卡 + 状态判定)· vat_excel_build 拆分。"""

from typing import List, Dict

from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from services.recon.field_comparator import parse_date
from services.recon.vat_recon_core import (
    _derive_period,
    _build_recon_pairs,
    _eq_amount,
    _get_inv_total,
    _get_rep_total,
    _diff_dims,
)
from services.vat.vat_excel_styles import (
    FONT_NAME,
    F_HEAD,
    F_NORM,
    F_TITLE,
    F_DIFF_RED,
    FILL_HEAD,
    FILL_OK,
    FILL_DIFF,
    FILL_MISS,
    FILL_FUZZY,
    FILL_OCRMSG,
    BORDER_TH,
    AL_C,
    AL_R,
    AL_L,
    HEAD_HEIGHT,
    ROW_HEIGHT,
)


def _build_sheet3(wb, invoices, report_rows, client_name, period_year, period_month, L):
    # ════════════ Sheet 3 · 对账结果(一对一 15 列 · 后端预算结果)════════════
    ws3 = wb.create_sheet(L["sh3"])
    ws3.sheet_properties.tabColor = "D97706"  # Tab 橙

    # 跑配对(Bug 2/3/4/5 全在这里)
    match_result = _build_recon_pairs(invoices, report_rows)
    pairs = match_result["pairs"]
    unmatched_inv = match_result["unmatched_inv"]
    unmatched_rep = match_result["unmatched_rep"]

    # 计算 KPI
    n_total = len(pairs) + len(unmatched_inv) + len(unmatched_rep)
    n_ok = 0
    for _p in pairs:
        _inv = invoices[_p["inv_idx"]]
        _rep = report_rows[_p["rep_idx"]]
        if not _eq_amount(_get_inv_total(_inv), _get_rep_total(_rep)):
            continue
        if _p["kind"] == "matched_cash":
            n_ok += 1
        elif _p["kind"] == "matched":
            _d = _diff_dims(_inv, _rep)
            if not any(_d.values()):
                n_ok += 1
    n_diff = n_total - n_ok
    diff_amount_total = 0.0
    for p in pairs:
        if p["kind"] in ("matched", "matched_cash"):
            a = _get_inv_total(invoices[p["inv_idx"]])
            b = _get_rep_total(report_rows[p["rep_idx"]])
            if not _eq_amount(a, b):
                diff_amount_total += abs(a - b)
    for ii in unmatched_inv:
        diff_amount_total += _get_inv_total(invoices[ii])
    for ri in unmatched_rep:
        diff_amount_total += _get_rep_total(report_rows[ri])

    # R1 · 标题行(Korn 样式 · 合并全行 · sz=18 · 高 36)
    _n_cols3 = len(L["h_recon"])
    ws3.merge_cells(f"A1:{get_column_letter(_n_cols3)}1")
    _c_title = ws3.cell(row=1, column=1, value=L["title"])
    _c_title.font = F_TITLE
    _c_title.alignment = AL_C
    ws3.row_dimensions[1].height = 36

    # R2 · 客户+期间 meta
    meta_parts = []
    if client_name:
        meta_parts.append(f"{L['client']}: {client_name}")
    if period_year and period_month:
        meta_parts.append(f"{L['period']}: {period_month:02d}/{period_year}")
    if meta_parts:
        ws3.cell(row=2, column=1, value=" · ".join(meta_parts)).font = F_NORM

    # R3 · 空行

    # R4-R5 · KPI 4 大卡(每卡 4 列宽 · 共 16 列 · 彩色底色)
    F_KPI_LBL2 = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
    F_KPI_VAL2 = Font(name=FONT_NAME, size=22, bold=True, color="FFFFFF")
    FILL_KPI_B = PatternFill("solid", fgColor="2563EB")
    FILL_KPI_G = PatternFill("solid", fgColor="16A34A")
    FILL_KPI_R = PatternFill("solid", fgColor="DC2626")
    FILL_KPI_O = PatternFill("solid", fgColor="D97706")
    KPI_ROW_LBL = 4
    KPI_ROW_VAL = 5
    ws3.row_dimensions[KPI_ROW_LBL].height = 22
    ws3.row_dimensions[KPI_ROW_VAL].height = 44

    def _kpi(col_start, label, value, fill):
        ws3.merge_cells(
            start_row=KPI_ROW_LBL,
            start_column=col_start,
            end_row=KPI_ROW_LBL,
            end_column=col_start + 3,
        )
        c1 = ws3.cell(row=KPI_ROW_LBL, column=col_start, value=label)
        c1.font = F_KPI_LBL2
        c1.fill = fill
        c1.alignment = AL_C
        ws3.merge_cells(
            start_row=KPI_ROW_VAL,
            start_column=col_start,
            end_row=KPI_ROW_VAL,
            end_column=col_start + 3,
        )
        c2 = ws3.cell(row=KPI_ROW_VAL, column=col_start, value=value)
        c2.font = F_KPI_VAL2
        c2.fill = fill
        c2.alignment = AL_C

    _kpi(1, L["kpi_total"], n_total, FILL_KPI_B)
    _kpi(5, L["kpi_ok"], n_ok, FILL_KPI_G)
    _kpi(9, L["kpi_diff"], n_diff, FILL_KPI_R)
    _kpi(13, L["kpi_amt"], f"฿ {diff_amount_total:,.2f}", FILL_KPI_O)

    # R6 · 表头(15 列)
    HEADER_ROW = 6
    DATA_START = 7
    headers3 = L["h_recon"]
    for c, h in enumerate(headers3, 1):
        cell = ws3.cell(row=HEADER_ROW, column=c, value=h)
        cell.font = F_HEAD
        cell.fill = FILL_HEAD
        cell.alignment = AL_C
        cell.border = BORDER_TH
    ws3.row_dimensions[HEADER_ROW].height = HEAD_HEIGHT

    # ── 写数据行 · 配对 + 孤儿
    def _status_for(pair, dims) -> tuple:
        """v118.32.4.10.0 · 维度感知状态判定 · 返回 (status_key, text, fill)"""
        kind = pair["kind"]
        if kind == "fuzzy":
            return "st_fuzzy", L["st_fuzzy"], FILL_FUZZY
        if kind == "ocr_missing":
            return "st_ocr_missing", L["st_ocr_missing"], FILL_OCRMSG
        if dims.get("tax_id", "").startswith("~"):
            return "st_fuzzy", L["st_fuzzy"], FILL_FUZZY
        if kind == "matched_cash":
            a = _get_inv_total(invoices[pair["inv_idx"]])
            b = _get_rep_total(report_rows[pair["rep_idx"]])
            if _eq_amount(a, b):
                return "st_cash", L["st_cash"], FILL_OK
            return "st_diff", L["st_diff"], FILL_DIFF
        diff_keys = [k for k, v in dims.items() if v and not (k == "tax_id" and v.startswith("~"))]
        a = _get_inv_total(invoices[pair["inv_idx"]])
        b = _get_rep_total(report_rows[pair["rep_idx"]])
        if not _eq_amount(a, b):
            diff_keys.append("_amt")
        n = len(diff_keys)
        if n == 0:
            return "st_ok", L["st_ok"], FILL_OK
        if n >= 2:
            return "st_multi_diff", L["st_multi_diff"], FILL_DIFF
        key = diff_keys[0]
        if key == "_amt":
            return "st_diff", L["st_diff"], FILL_DIFF
        if key in ("date", "period"):
            return "st_date_diff", L["st_date_diff"], FILL_DIFF
        if key == "branch":
            return "st_branch_diff", L["st_branch_diff"], FILL_DIFF
        if key == "name":
            return "st_name_diff", L["st_name_diff"], FILL_DIFF
        if key == "inv_no":
            return "st_inv_no_diff", L["st_inv_no_diff"], FILL_DIFF
        if key == "tax_id":
            return "st_tax_id_diff", L["st_tax_id_diff"], FILL_DIFF
        return "st_diff", L["st_diff"], FILL_DIFF

    row_cursor = DATA_START
    seq_no = 0
    diff_col_indices = (9, 10, 11, 12, 13, 14)  # 6 维度差异列(含税号差)
    F_FUZZY_BLUE = Font(name=FONT_NAME, size=10, color="2563EB")
    task_rows: List[Dict] = []  # v118.32.4.10.0 · 任务摘要行

    for pair in pairs:
        seq_no += 1
        inv = invoices[pair["inv_idx"]]
        rep = report_rows[pair["rep_idx"]]
        dims = _diff_dims(inv, rep)
        # Bug 2 · 散客:不比较/不显示分公司和税号差
        if pair["kind"] == "matched_cash":
            dims["branch"] = ""
            dims["tax_id"] = ""
        # tax_id fuzzy ~ 前缀处理(显示时去掉 · 用蓝色替换红色)
        tax_id_raw = dims.get("tax_id", "")
        tax_id_is_fuzzy = tax_id_raw.startswith("~")
        tax_id_display = tax_id_raw[1:] if tax_id_is_fuzzy else tax_id_raw
        status_key, status_text, row_fill = _status_for(pair, dims)
        amt_inv = _get_inv_total(inv)
        amt_rep = _get_rep_total(rep)
        amt_diff = round(amt_inv - amt_rep, 2)
        period_inv = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")

        # v4.10.13 · 净额/VAT 分字段差异备注
        amt_pre_inv = float(inv.get("amount_pre_vat") or 0)
        amt_vat_inv = float(inv.get("vat_amount") or 0)
        amt_pre_rep = float((rep.get("report_amount_pre_vat") or rep.get("report_amount")) or 0)
        amt_vat_rep = float(rep.get("report_vat_amount") or 0)
        pre_diff = round(amt_pre_inv - amt_pre_rep, 2)
        vat_diff_ = round(amt_vat_inv - amt_vat_rep, 2)
        if not _eq_amount(amt_inv, amt_rep):
            if not _eq_amount(pre_diff, 0) and not _eq_amount(vat_diff_, 0):
                _amt_note = f"净额差 {pre_diff:+,.2f} · VAT 差 {vat_diff_:+,.2f}"
            elif not _eq_amount(pre_diff, 0):
                _amt_note = f"净额差 {pre_diff:+,.2f} · VAT 一致"
            else:
                _amt_note = f"VAT 差 {vat_diff_:+,.2f} · 净额一致"
        else:
            _amt_note = ""

        # v4.10.13 · ocr_missing · 备注追加提醒(dims 原样保留)
        _base_note = pair.get("note") or ""
        if pair["kind"] == "ocr_missing":
            _base_note = (_base_note + " · OCR 抽取可能不完整 · 请核对原 PDF").lstrip(" · ")
        _note_val = " · ".join(filter(None, [_base_note, _amt_note]))

        values = [
            seq_no,  # 1 #
            status_text,  # 2 status
            inv.get("buyer_name") or rep.get("report_buyer_name") or "",  # 3 客户
            inv.get("invoice_no") or "",  # 4 发票号
            period_inv,  # 5 期间
            amt_inv,  # 6 金额(发)
            amt_rep,  # 7 金额(报)
            amt_diff,  # 8 差异金额
            dims["inv_no"],  # 9 发票号差
            dims["date"],  # 10 日期差
            dims["period"],  # 11 期间差
            tax_id_display,  # 12 税号差(新)
            dims["branch"],  # 13 分公司差
            dims["name"],  # 14 客户名差
            _note_val,  # 15 备注
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = row_fill
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
            # 差异维度列若有内容 · 税号 fuzzy 蓝 · 其余红
            if c in diff_col_indices and v:
                if c == 12 and tax_id_is_fuzzy:
                    cell.font = F_FUZZY_BLUE
                else:
                    cell.font = F_DIFF_RED
        # v118.32.4.10.0 · 收集摘要行
        task_rows.append(
            {
                "status_key": status_key,
                "status": status_text,
                "customer": inv.get("buyer_name") or rep.get("report_buyer_name") or "",
                "invoice_no": inv.get("invoice_no") or "",
                "amount_inv": amt_inv,
                "amount_rep": amt_rep,
                "dims": {
                    k: (v[1:] if k == "tax_id" and v.startswith("~") else v)
                    for k, v in dims.items()
                },
                "kind": pair["kind"],
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 发票孤儿(报告无)
    for ii in unmatched_inv:
        seq_no += 1
        inv = invoices[ii]
        amt_inv = _get_inv_total(inv)
        period_inv = _derive_period(inv.get("invoice_date") or "", inv.get("period") or "")
        values = [
            seq_no,
            L["st_no_rep"],
            inv.get("buyer_name") or "",
            inv.get("invoice_no") or "",
            period_inv,
            amt_inv,
            0,
            amt_inv,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = FILL_MISS
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
        task_rows.append(
            {
                "status_key": "st_no_rep",
                "status": L["st_no_rep"],
                "customer": inv.get("buyer_name") or "",
                "invoice_no": inv.get("invoice_no") or "",
                "amount_inv": amt_inv,
                "amount_rep": 0,
                "dims": {},
                "kind": "invoice_orphan",
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 报告孤儿(发票无)
    for ri in unmatched_rep:
        seq_no += 1
        rep = report_rows[ri]
        amt_rep = _get_rep_total(rep)
        rep_date = rep.get("report_date") or ""
        d = parse_date(rep_date)
        period_rep = f"{d.month:02d}/{d.year}" if d else ""
        values = [
            seq_no,
            L["st_no_inv"],
            rep.get("report_buyer_name") or "",
            rep.get("report_invoice_no") or "",
            period_rep,
            0,
            amt_rep,
            -amt_rep,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
        ]
        for c, v in enumerate(values, 1):
            cell = ws3.cell(row=row_cursor, column=c, value=v)
            cell.font = F_NORM
            cell.border = BORDER_TH
            cell.fill = FILL_MISS
            if c in (6, 7, 8):
                cell.alignment = AL_R
                cell.number_format = "#,##0.00"
            elif c == 1:
                cell.alignment = AL_C
            else:
                cell.alignment = AL_L
        task_rows.append(
            {
                "status_key": "st_no_inv",
                "status": L["st_no_inv"],
                "customer": rep.get("report_buyer_name") or "",
                "invoice_no": rep.get("report_invoice_no") or "",
                "amount_inv": 0,
                "amount_rep": amt_rep,
                "dims": {},
                "kind": "report_orphan",
            }
        )
        ws3.row_dimensions[row_cursor].height = ROW_HEIGHT
        row_cursor += 1

    # 列宽
    widths3 = [5, 22, 26, 16, 10, 14, 14, 13, 22, 14, 16, 22, 18, 26, 22]
    for i, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = f"A{DATA_START}"
    return n_total, n_ok, n_diff, diff_amount_total, task_rows
