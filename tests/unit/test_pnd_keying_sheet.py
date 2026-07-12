# -*- coding: utf-8 -*-
"""PND53/PND3 键入底稿(keying sheet)xlsx 守门(D1-6 · 派单书 §硬约束)。

守恒断言 = 底稿末行合计(Σจำนวนเงินที่จ่าย/Σภาษีที่หัก)与官方 PND53 M 记录(RD Prep txt
HEADER 的 TOT_AMT/TOT_TAX)逐字一致——两者同源同数(都从 pnd_prep._group_by_payee 产出的
同一份 payees 字典算出),不各算一套。金标复用 test_workorder_pnd_prep.py 既有 fixtures,
不另起一份仿真数据。
"""

from __future__ import annotations

import io
import tempfile
import unittest
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from services.tax import pnd_keying_sheet, rdprep
from services.workorder.engine import StepContext
from services.workorder.steps import pnd_prep
from tests.unit.test_workorder_pnd_prep import (
    _INDIVIDUAL_TAX_ID,
    _JURISTIC_TAX_ID,
    _FakeCursor,
    _FakeStore,
    _client_row,
    _pnd_row,
)


def _ctx(store, cur, *, period="2569-05"):
    data = {
        "tax_due": "0",
        "sales_amount": "0",
        "output_vat": "0",
        "purchase_amount": "0",
        "input_vat": "0",
        "period": period,
        "prior_period_check": None,
        "pp30_form": None,
        "gates": {},
    }
    return StepContext(cur=cur, tenant_id="t-1", work_order_id="wo-1", store=store, data=data)


class BuildWorkbookTests(unittest.TestCase):
    """纯渲染层:行字典 → xlsx,不经 pnd_prep。"""

    def test_headers_and_row_values(self):
        rows = [
            {
                "tax_id": _JURISTIC_TAX_ID,
                "title_name": "บริษัท",
                "payee_name": "ทดสอบ จำกัด",
                "address": None,
                "paid_date": date(2026, 5, 10),
                "income_type": "ค่าจ้างทำของ",
                "rate": Decimal("3.00"),
                "paid_amount": Decimal("140.00"),
                "wht_amount": Decimal("4.20"),
                "condition": "หัก ณ ที่จ่าย",
            }
        ]
        wb = load_workbook(io.BytesIO(pnd_keying_sheet.build_workbook(rdprep.PND53, rows)))
        ws = wb.active

        self.assertEqual(ws.cell(row=1, column=1).value, "ลำดับ (序号)")
        self.assertEqual(ws.cell(row=2, column=1).value, 1)
        self.assertEqual(ws.cell(row=2, column=2).value, _JURISTIC_TAX_ID)
        self.assertEqual(ws.cell(row=2, column=3).value, "บริษัท ทดสอบ จำกัด")
        self.assertEqual(ws.cell(row=2, column=5).value, "10/05/2569")
        self.assertEqual(ws.cell(row=2, column=8).value, 140.0)
        self.assertEqual(ws.cell(row=2, column=9).value, 4.2)

        # 合计行紧接末条数据行,第 8/9 列 = Σ 支付金额 / Σ 预扣税额。
        self.assertEqual(ws.cell(row=3, column=3).value, "รวม (合计)")
        self.assertEqual(ws.cell(row=3, column=8).value, 140.0)
        self.assertEqual(ws.cell(row=3, column=9).value, 4.2)

    def test_missing_address_and_date_honestly_blank(self):
        """无结构化地址源/无支付日 → 地址栏诚实标 ไม่มีข้อมูล,日期栏留空,不编造。"""
        rows = [
            {
                "tax_id": _INDIVIDUAL_TAX_ID,
                "title_name": "-",
                "payee_name": "สมชาย ใจดี",
                "address": None,
                "paid_date": None,
                "income_type": "ค่าบริการ",
                "rate": Decimal("5.00"),
                "paid_amount": Decimal("200.00"),
                "wht_amount": Decimal("10.00"),
                "condition": "หัก ณ ที่จ่าย",
            }
        ]
        wb = load_workbook(io.BytesIO(pnd_keying_sheet.build_workbook(rdprep.PND3, rows)))
        ws = wb.active

        self.assertEqual(ws.cell(row=2, column=4).value, "ไม่มีข้อมูล")
        self.assertFalse(ws.cell(row=2, column=5).value)

    def test_totals_decimal_precision_not_float_drift(self):
        rows = [
            {"paid_amount": Decimal("100.10"), "wht_amount": Decimal("3.00")},
            {"paid_amount": Decimal("200.20"), "wht_amount": Decimal("6.01")},
        ]
        t = pnd_keying_sheet.totals(rows)
        self.assertEqual(t["paid_amount"], Decimal("300.30"))
        self.assertEqual(t["wht_amount"], Decimal("9.01"))
        self.assertIsInstance(t["paid_amount"], Decimal)


class PndPrepIntegrationTests(unittest.TestCase):
    """经 pnd_prep.build() 落盘:两 kind 均出,合计与官方 PND53 M 记录逐字一致(守恒断言)。"""

    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self.tmp.cleanup)
        self.out_dir = Path(self.tmp.name)

    def test_pnd53_keying_totals_match_official_m_record(self):
        cur = _FakeCursor(
            ones=[_client_row()],
            alls=[
                [
                    _pnd_row(
                        doc_id="d1",
                        wht_amount="4.20",
                        payee_tax_id=_JURISTIC_TAX_ID,
                        payee_name="บริษัท ทดสอบ จำกัด",
                        base="140.00",
                        rate="3.00",
                    )
                ],
                [{"id": "d1", "doc_date": date(2026, 5, 10)}],
            ],
        )
        store = _FakeStore()
        kinds, _ = pnd_prep.build(_ctx(store, cur), self.out_dir, "2569-05")

        # 官方 M 记录(RD Prep txt)HEADER 的 TOT_AMT(字段 19)/TOT_TAX(字段 20)。
        txt_path, txt_numbers = kinds[pnd_prep.KIND_PND53]
        header = Path(txt_path).read_text(encoding="utf-8").split("\n")[0].split("|")
        m_tot_amt, m_tot_tax = header[18], header[19]

        self.assertIn(pnd_prep.KIND_PND53_KEYING, kinds)
        xlsx_path, xlsx_numbers = kinds[pnd_prep.KIND_PND53_KEYING]
        self.assertTrue(Path(xlsx_path).is_file())
        self.assertEqual(f"{xlsx_numbers['paid_amount_total']:.2f}", m_tot_amt)
        self.assertEqual(f"{xlsx_numbers['wht_total']:.2f}", m_tot_tax)
        self.assertEqual(xlsx_numbers["wht_total"], txt_numbers["wht_total"])

        wb = load_workbook(xlsx_path)
        ws = wb.active
        self.assertEqual(ws.cell(row=3, column=8).value, float(m_tot_amt))
        self.assertEqual(ws.cell(row=3, column=9).value, float(m_tot_tax))

    def test_pnd3_keying_emitted_without_address_no_official_txt(self):
        """个人 payee 缺结构化地址:官方 ภ.ง.ด.3 RD Prep txt 仍不出(既有行为不变),但键入
        底稿 xlsx 照出——辅助件不因缺地址剔除,地址栏诚实留空(D1-6 任务书核心断言)。"""
        cur = _FakeCursor(
            ones=[_client_row()],
            alls=[
                [
                    _pnd_row(
                        doc_id="d2",
                        wht_amount="10.00",
                        payee_tax_id=_INDIVIDUAL_TAX_ID,
                        payee_name="สมชาย ใจดี",
                        base="200.00",
                        rate="5.00",
                    )
                ],
                [],  # d2 doc_date 未预置:替身查不到即空,渲染层需扛住(不崩、不编日期)
            ],
        )
        store = _FakeStore()
        kinds, memo_lines = pnd_prep.build(_ctx(store, cur), self.out_dir, "2569-05")

        self.assertNotIn(pnd_prep.KIND_PND3, kinds)  # 官方件仍不出
        self.assertIn(pnd_prep.KIND_PND3_KEYING, kinds)
        path, numbers = kinds[pnd_prep.KIND_PND3_KEYING]
        self.assertEqual(numbers["paid_amount_total"], Decimal("200.00"))
        self.assertEqual(numbers["wht_total"], Decimal("10.00"))

        wb = load_workbook(path)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=2).value, _INDIVIDUAL_TAX_ID)
        self.assertEqual(ws.cell(row=2, column=4).value, "ไม่มีข้อมูล")
        self.assertFalse(ws.cell(row=2, column=5).value)  # 无 doc_date,日期栏诚实留空
        self.assertTrue(any("คีย์ข้อมูล" in line for line in memo_lines))


if __name__ == "__main__":
    unittest.main()
