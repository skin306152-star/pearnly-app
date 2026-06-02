# -*- coding: utf-8 -*-
"""report_engine · Excel 样式常量(颜色/字体/边框/填充/对齐)leaf。"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

COLOR_HEADER_BG = "1E3A8A"  # 深蓝表头(更专业 · 比老 #2C5282 深一档)
COLOR_HEADER_FG = "FFFFFF"
COLOR_ZEBRA = "F9FAFB"  # 斑马行灰
COLOR_BORDER = "D1D5DB"
COLOR_TOTAL_BG = "DBEAFE"  # 合计行浅蓝
COLOR_TITLE_FG = "111827"  # 顶部大标题深灰
COLOR_INFO_FG = "374151"  # 信息块灰
COLOR_WARN_AMOUNT = "DC2626"  # 大额警告红(>50 万)

# 字体优先级:Tahoma 泰文支持好;退回 Calibri / Arial
FONT_NAME = "Tahoma"
FONT_NAME_NUM = "Calibri"

THIN = Side(style="thin", color=COLOR_BORDER)
THICK = Side(style="medium", color="6B7280")
BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_TOP_THICK = Border(left=THIN, right=THIN, top=THICK, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor=COLOR_HEADER_BG)
ZEBRA_FILL = PatternFill("solid", fgColor=COLOR_ZEBRA)
TOTAL_FILL = PatternFill("solid", fgColor=COLOR_TOTAL_BG)

HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color=COLOR_HEADER_FG)
CELL_FONT = Font(name=FONT_NAME, size=10)
TOTAL_FONT = Font(name=FONT_NAME, size=11, bold=True)
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color=COLOR_TITLE_FG)
INFO_FONT = Font(name=FONT_NAME, size=10, color=COLOR_INFO_FG)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
