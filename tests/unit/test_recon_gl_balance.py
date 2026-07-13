# -*- coding: utf-8 -*-
"""银行对账导出「总账余额列」守门:GL 行运行余额(期初+累计借−贷)透传到导出 · 4 语跟随。
不参与匹配引擎(只新增展示字段)· 见 [[GlRow.balance / BankReconRow.gl_balance]]。"""

import unittest
from datetime import date

import openpyxl

from services.recon.bank_recon_types import GlRow, StatementRow
from services.recon.bank_recon_reconcile import reconcile
from services.recon.bank_recon_excel_i18n import _t


class ReconGlBalanceTests(unittest.TestCase):
    def _rows(self):
        gl = [
            GlRow(
                date(2025, 6, 2),
                "V1",
                "1112",
                "รับ บริษัท แจแปน 111",
                227418.00,
                0.0,
                balance=378232.56,
            ),
            GlRow(date(2025, 6, 5), "V2", "1112", "โอนให้", 0.0, 380000.00, balance=-1632.44),
        ]
        st = [
            StatementRow(
                date=date(2025, 6, 2),
                description="in",
                withdrawal=0.0,
                deposit=227418.00,
                balance=378232.56,
            )
        ]
        return reconcile(st, gl)

    def test_gl_balance_passthrough_to_recon_row(self):
        rows, _ = self._rows()
        matched = [r for r in rows if r.match_status == "matched"]
        self.assertTrue(matched)
        self.assertEqual(matched[0].gl_balance, 378232.56)
        # GL 孤项也带余额
        gl_only = [r for r in rows if r.match_status in ("gl_debit_only", "gl_credit_only")]
        self.assertTrue(any(r.gl_balance == -1632.44 for r in gl_only))

    def test_gl_detail_sheet_has_balance_column_4lang(self):
        # 「总账明细」sheet(รายละเอียดบัญชีแยกประเภท)也要有余额列 · 4 语跟随
        from services.recon.bank_recon_excel import export_bank_recon_excel
        from services.recon.bank_recon_types import BankReconSummary
        import io as _io

        rows, _ = self._rows()
        for lang in ("zh", "en", "th", "ja"):
            blob = export_bank_recon_excel(rows, BankReconSummary(), lang)
            wb = openpyxl.load_workbook(_io.BytesIO(blob))
            ws = wb[_t("sh_gl_detail", lang)]
            hdr = [c.value for c in ws[1]]
            self.assertIn(_t("col_balance", lang), hdr, f"{lang} 总账明细缺余额列")
            ci = hdr.index(_t("col_balance", lang))
            vals = [ws.cell(r, ci + 1).value for r in range(2, ws.max_row + 1)]
            self.assertIn(378232.56, vals, f"{lang} 总账明细余额无值")

    def test_serialize_roundtrip_keeps_gl_balance(self):
        # 序列化/反序列化必须保住 gl_balance(否则存库再读回 → 导出空列)
        from services.recon.bank_recon_serialize import rows_to_json, rows_from_json

        rows, _ = self._rows()
        back = rows_from_json(rows_to_json(rows))
        matched = [r for r in back if r.match_status == "matched"]
        self.assertEqual(matched[0].gl_balance, 378232.56)

    def test_old_job_export_computes_running_balance_fallback(self):
        # 旧任务存量无 gl_balance(早期序列化漏)→ 导出用借贷+期初现算 · 重导出即有余额
        from services.recon.bank_recon_excel import export_bank_recon_excel
        from services.recon.bank_recon_types import BankReconSummary
        import io as _io

        rows, _ = self._rows()
        for r in rows:
            r.gl_balance = 0.0  # 模拟旧任务
        summ = BankReconSummary(gl_opening=150814.56)  # 期初 → 第一笔借 227418 → 378232.56
        blob = export_bank_recon_excel(rows, summ, "th")
        ws = openpyxl.load_workbook(_io.BytesIO(blob))[_t("sh_gl_detail", "th")]
        hdr = [c.value for c in ws[1]]
        ci = hdr.index(_t("col_balance", "th"))
        vals = [ws.cell(r, ci + 1).value for r in range(2, ws.max_row + 1)]
        self.assertIn(378232.56, vals)  # 现算出的运行余额

    def test_running_balance_computed_from_opening(self):
        # parse 层逐行运行余额 = 期初 + 累计(借−贷)· 用 parse_gl_excel 验
        import io
        from services.recon.bank_gl_excel import parse_gl_excel

        wb = openpyxl.Workbook()
        wsx = wb.active
        wsx.title = "Pearnly-GL"
        wsx.append(
            ["วันที่", "เลขที่ใบสำคัญ", "รายละเอียด", "เดบิต=ฝากเงิน", "เครดิต=ถอนเงิน", "คงเหลือ"]
        )
        wsx.append(["ยอดยกมา", None, None, None, None, 100.0])
        wsx.append(["02/05/2026", "V1", "in", 200.0, None, 300.0])
        wsx.append(["02/05/2026", "V2", "out", None, 50.0, 250.0])
        buf = io.BytesIO()
        wb.save(buf)
        res = parse_gl_excel(buf.getvalue(), "gl.xlsx")
        self.assertTrue(res.get("ok"))
        bals = [r.balance for r in res["rows"]]
        self.assertEqual(bals, [300.0, 250.0])  # 100+200=300 · 300-50=250
        # K2 · closing 读自余额列时 closing_printed=True(fileconv GL 路的真锚判据)
        self.assertTrue(res["closing_printed"])
        self.assertEqual(res["closing"], 250.0)

    def test_closing_printed_false_without_balance_column(self):
        # 无余额列 → closing 是借贷衍生值,closing_printed 必须诚实为 False
        # (fileconv K2 据此拒进 GL 路:衍生值当锚校验自己 = 自导自演)
        import io
        from services.recon.bank_gl_excel import parse_gl_excel

        wb = openpyxl.Workbook()
        wsx = wb.active
        wsx.append(["วันที่", "เลขที่ใบสำคัญ", "รายละเอียด", "เดบิต", "เครดิต"])
        wsx.append(["02/05/2026", "V1", "in", 200.0, None])
        wsx.append(["02/05/2026", "V2", "out", None, 50.0])
        buf = io.BytesIO()
        wb.save(buf)
        res = parse_gl_excel(buf.getvalue(), "gl.xlsx")
        self.assertTrue(res.get("ok"))
        self.assertFalse(res["closing_printed"])


if __name__ == "__main__":
    unittest.main()
